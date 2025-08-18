import os
import time
from datetime import datetime
from tkinter import messagebox
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from dotenv import load_dotenv
from openpyxl import load_workbook

from src.core.browser_factory import BrowserFactory
from src.core.logger import log_message
from src.modules.base import BaseModule

load_dotenv()

class MacroscopiaFixacaoModule(BaseModule):
    def __init__(self):
        super().__init__(nome="Macroscopia e Fixação")

    def get_dados_exames(self, file_path: str) -> list:
        """Lê os dados dos exames da planilha Excel"""
        try:
            workbook = load_workbook(file_path)
            sheet = workbook.active
            dados = []
            ultima_mascara = None
            ultimo_macroscopista = None
            
            # Lê da linha 2 em diante (linha 1 é cabeçalho)
            for row in range(2, sheet.max_row + 1):
                codigo = sheet[f'A{row}'].value  # Código do exame
                mascara = sheet[f'B{row}'].value  # Máscara/valor para campo buscaArvore
                macroscopista = sheet[f'C{row}'].value  # Macroscopista
                
                if codigo is not None:
                    codigo = str(codigo).strip()
                    
                    # Se não tem máscara, usa a última válida
                    if mascara is not None and str(mascara).strip():
                        mascara = str(mascara).strip()
                        ultima_mascara = mascara
                    else:
                        mascara = ultima_mascara
                    
                    # Se não tem macroscopista, usa o último válido
                    if macroscopista is not None and str(macroscopista).strip():
                        macroscopista = str(macroscopista).strip()
                        ultimo_macroscopista = macroscopista
                    else:
                        macroscopista = ultimo_macroscopista
                    
                    dados.append({
                        'codigo': codigo,
                        'mascara': mascara,
                        'macroscopista': macroscopista
                    })
            
            workbook.close()
            return dados
        except Exception as e:
            raise Exception(f"Erro ao ler planilha: {e}")

    def verificar_sessao_browser(self, driver) -> bool:
        """Verifica se a sessão do browser ainda está ativa"""
        try:
            driver.current_url
            return True
        except Exception as e:
            if "invalid session id" in str(e).lower():
                log_message("❌ Sessão do browser perdida", "ERROR")
                return False
            return True

    def selecionar_responsavel_macroscopia(self, driver, wait):
        """Seleciona 'Nathalia Fernanda da Silva Lopes' como responsável pela macroscopia"""
        try:
            log_message("🔍 Procurando campo responsável pela macroscopia...", "INFO")
            
            # Aguardar o select estar presente
            select_element = wait.until(
                EC.presence_of_element_located((By.ID, "responsavelMacroscopiaId"))
            )
            
            # Usar Select para interagir com o dropdown
            select = Select(select_element)
            
            # Selecionar "Nathalia Fernanda da Silva Lopes" (value="264417")
            select.select_by_value("264417")
            log_message("✅ Nathalia Fernanda da Silva Lopes selecionada como responsável", "SUCCESS")
            time.sleep(1)
            
        except Exception as e:
            log_message(f"Erro ao selecionar responsável pela macroscopia: {e}", "ERROR")
            raise

    def selecionar_auxiliar_macroscopia(self, driver, wait):
        """Seleciona 'Renata Silva Sevidanis' como auxiliar da macroscopia"""
        try:
            log_message("🔍 Procurando campo auxiliar da macroscopia...", "INFO")
            
            # Aguardar o select estar presente
            select_element = wait.until(
                EC.presence_of_element_located((By.ID, "auxiliarMacroscopiaId"))
            )
            
            # Usar Select para interagir com o dropdown
            select = Select(select_element)
            
            # Selecionar "Renata Silva Sevidanis" (value="241593")
            select.select_by_value("241593")
            log_message("✅ Renata Silva Sevidanis selecionada como auxiliar", "SUCCESS")
            time.sleep(1)
            
        except Exception as e:
            log_message(f"Erro ao selecionar auxiliar da macroscopia: {e}", "ERROR")
            raise

    def digitar_mascara_busca(self, driver, wait, mascara):
        """Digita a máscara no campo buscaArvore"""
        try:
            log_message(f"🔍 Procurando campo buscaArvore para digitar: {mascara}", "INFO")
            
            # Aguardar o campo estar presente e clicável
            campo_busca = wait.until(EC.element_to_be_clickable((By.ID, "buscaArvore")))
            
            # Verificar se o campo está visível
            if not campo_busca.is_displayed():
                log_message("⚠️ Campo buscaArvore não está visível", "WARNING")
                return
            
            # Limpar e digitar a máscara
            campo_busca.clear()
            campo_busca.send_keys(mascara)
            log_message(f"✍️ Máscara '{mascara}' digitada no campo buscaArvore", "SUCCESS")
            time.sleep(0.5)
            
            # Pressionar Enter
            campo_busca.send_keys(Keys.ENTER)
            log_message(f"⌨️ Enter pressionado após digitar máscara", "INFO")
            time.sleep(1)
            
        except Exception as e:
            log_message(f"Erro ao digitar máscara: {e}", "ERROR")
            raise

    def definir_data_fixacao(self, driver, wait):
        """Define a data atual no campo de data de fixação"""
        try:
            log_message("📅 Definindo data atual no campo de fixação...", "INFO")
            
            # Aguardar o campo de data estar presente
            campo_data = wait.until(
                EC.presence_of_element_located((By.XPATH, "//input[@type='date' and @name='dataFixacao']"))
            )
            
            # Obter data atual no formato YYYY-MM-DD
            data_atual = datetime.now().strftime("%Y-%m-%d")
            
            # Limpar e definir a data
            campo_data.clear()
            campo_data.send_keys(data_atual)
            log_message(f"📅 Data de fixação definida para: {data_atual}", "SUCCESS")
            time.sleep(0.5)
            
        except Exception as e:
            log_message(f"Erro ao definir data de fixação: {e}", "ERROR")
            raise

    def definir_hora_fixacao(self, driver, wait):
        """Define 18:00 no campo de hora de fixação"""
        try:
            log_message("🕕 Definindo hora 18:00 no campo de fixação...", "INFO")
            
            # Aguardar o campo de hora estar presente
            campo_hora = wait.until(
                EC.presence_of_element_located((By.XPATH, "//input[@type='time' and @name='dataFixacao']"))
            )
            
            # Limpar e definir a hora
            campo_hora.clear()
            campo_hora.send_keys("18:00")
            log_message("🕕 Hora de fixação definida para: 18:00", "SUCCESS")
            time.sleep(0.5)
            
        except Exception as e:
            log_message(f"Erro ao definir hora de fixação: {e}", "ERROR")
            raise

    def processar_exame(self, driver, wait, codigo, mascara, macroscopista):
        """Processa um exame individual"""
        try:
            log_message(f"\n➡️ Processando exame: {codigo}", "INFO")
            
            # Verificar se a sessão do browser ainda está ativa
            if not self.verificar_sessao_browser(driver):
                raise Exception("Sessão do browser perdida - necessário reiniciar")
            
            # Aguardar e encontrar o campo de código de barras
            log_message("🔍 Procurando campo de código de barras...", "INFO")
            campo_codigo = wait.until(EC.presence_of_element_located((By.ID, "inputSearchCodBarra")))
            
            # Limpar e digitar o código
            campo_codigo.clear()
            campo_codigo.send_keys(codigo)
            log_message(f"✍️ Código '{codigo}' digitado", "INFO")
            time.sleep(1)
            
            # Pressionar Enter
            campo_codigo.send_keys(Keys.ENTER)
            log_message("⌨️ Enter pressionado para buscar exame", "INFO")
            
            # Aguardar carregamento do exame
            log_message("⏳ Aguardando carregamento do exame...", "INFO")
            time.sleep(3)
            
            # Processar os campos específicos do exame
            log_message("🔧 Processando campos do exame...", "INFO")
            
            # 1. Selecionar responsável pela macroscopia
            self.selecionar_responsavel_macroscopia(driver, wait)
            
            # 2. Selecionar auxiliar da macroscopia  
            self.selecionar_auxiliar_macroscopia(driver, wait)
            
            # 3. Digitar máscara no campo buscaArvore
            if mascara:
                self.digitar_mascara_busca(driver, wait, mascara)
            else:
                log_message("⚠️ Nenhuma máscara encontrada, pulando busca", "WARNING")
            
            # 4. Definir data de fixação
            self.definir_data_fixacao(driver, wait)
            
            # 5. Definir hora de fixação
            self.definir_hora_fixacao(driver, wait)
            
            log_message(f"✅ Exame {codigo} processado com sucesso!", "SUCCESS")
            return {'status': 'sucesso', 'detalhes': 'Exame processado com sucesso'}
            
        except Exception as e:
            error_message = str(e)
            log_message(f"❌ Erro ao processar exame {codigo}: {error_message}", "ERROR")
            
            # Verificar se é erro de sessão inválida
            if "invalid session id" in error_message.lower():
                return {'status': 'erro_sessao', 'detalhes': 'Sessão do browser perdida'}
            
            return {'status': 'erro', 'detalhes': error_message}

    def run(self, params: dict):
        """Método principal que executa a automação"""
        username = params.get("username")
        password = params.get("password") 
        excel_file = params.get("excel_file")
        cancel_flag = params.get("cancel_flag")
        
        try:
            # Lê os dados dos exames da planilha
            dados_exames = self.get_dados_exames(excel_file)
            if not dados_exames:
                messagebox.showerror("Erro", "Nenhum dado de exame encontrado na planilha.")
                return
            
            log_message(f"Encontrados {len(dados_exames)} exames para processar", "INFO")
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao ler o Excel: {e}")
            return

        url = os.getenv("SYSTEM_URL", "https://pathoweb.com.br/login/auth")
        driver = None
        resultados = []
        
        try:
            driver = BrowserFactory.create_chrome()
            wait = WebDriverWait(driver, 20)
            
            log_message("Iniciando automação de macroscopia e fixação...", "INFO")
            
            # Login
            log_message("Fazendo login...", "INFO")
            driver.get(url)
            
            # Aguardar página carregar completamente
            wait.until(EC.presence_of_element_located((By.ID, "username")))
            
            username_field = driver.find_element(By.ID, "username")
            username_field.clear()
            username_field.send_keys(username)
            
            password_field = driver.find_element(By.ID, "password")
            password_field.clear()
            password_field.send_keys(password)
            
            submit_button = driver.find_element(By.CSS_SELECTOR, "button[type='submit']")
            submit_button.click()
            
            log_message("Navegando para módulo de exames...", "INFO")
            
            # Navegar para o módulo de exames (módulo 1)
            modulo_link = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "a[href='/site/trocarModulo?modulo=1']")))
            modulo_link.click()
            time.sleep(2)
            
            # Fechar modal se aparecer
            try:
                modal_close_button = driver.find_element(By.CSS_SELECTOR, "#mensagemParaClienteModal .modal-footer button")
                if modal_close_button.is_displayed():
                    modal_close_button.click()
            except Exception:
                pass

            log_message("✅ Login realizado com sucesso. Iniciando processamento dos exames.", "SUCCESS")
            
            # Processar cada exame da planilha
            for i, exame_data in enumerate(dados_exames, 1):
                if cancel_flag and cancel_flag.is_set():
                    log_message("Execução cancelada pelo usuário.", "WARNING")
                    break
                
                codigo = exame_data['codigo']
                mascara = exame_data['mascara']
                macroscopista = exame_data['macroscopista']
                
                log_message(f"\n➡️ Processando exame {i}/{len(dados_exames)}: {codigo}", "INFO")
                log_message(f"   Máscara: {mascara}", "INFO")
                log_message(f"   Macroscopista: {macroscopista}", "INFO")
                
                try:
                    # Verificar se o browser ainda está ativo
                    if not self.verificar_sessao_browser(driver):
                        log_message("🔄 Recriando browser devido à sessão perdida...", "WARNING")
                        try:
                            driver.quit()
                        except:
                            pass
                        
                        # Recriar browser e fazer login novamente
                        driver = BrowserFactory.create_chrome()
                        wait = WebDriverWait(driver, 20)
                        
                        # Fazer login novamente
                        log_message("🔄 Fazendo login novamente...", "INFO")
                        driver.get(url)
                        
                        # Aguardar página carregar completamente
                        wait.until(EC.presence_of_element_located((By.ID, "username")))
                        time.sleep(2)
                        
                        username_field = driver.find_element(By.ID, "username")
                        username_field.clear()
                        username_field.send_keys(username)
                        
                        password_field = driver.find_element(By.ID, "password")
                        password_field.clear()
                        password_field.send_keys(password)
                        
                        submit_button = driver.find_element(By.CSS_SELECTOR, "button[type='submit']")
                        submit_button.click()
                        
                        log_message("🔄 Navegando para módulo de exames novamente...", "INFO")
                        
                        # Navegar para o módulo de exames (módulo 1)
                        modulo_link = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "a[href='/site/trocarModulo?modulo=1']")))
                        modulo_link.click()
                        time.sleep(2.5)
                        
                        # Fechar modal se aparecer
                        try:
                            modal_close_button = driver.find_element(By.CSS_SELECTOR, "#mensagemParaClienteModal .modal-footer button")
                            if modal_close_button.is_displayed():
                                modal_close_button.click()
                                time.sleep(1)
                        except Exception:
                            pass
                        
                        log_message("✅ Browser recriado e login realizado novamente", "SUCCESS")
                    
                    # Processar este exame específico
                    resultado = self.processar_exame(driver, wait, codigo, mascara, macroscopista)
                    resultados.append({
                        'codigo': codigo,
                        'mascara': mascara,
                        'macroscopista': macroscopista,
                        'status': resultado['status'],
                        'detalhes': resultado.get('detalhes', '')
                    })
                    
                except Exception as e:
                    log_message(f"❌ Erro ao processar exame {codigo}: {e}", "ERROR")
                    resultados.append({
                        'codigo': codigo,
                        'mascara': mascara,
                        'macroscopista': macroscopista,
                        'status': 'erro',
                        'detalhes': str(e)
                    })
            
            # Mostrar resumo final
            self.mostrar_resumo_final(resultados)
            
        except Exception as e:
            log_message(f"❌ Erro durante a automação: {e}", "ERROR")
            messagebox.showerror("Erro", f"❌ Erro durante a automação:\n{str(e)[:200]}...")
        finally:
            if driver:
                try:
                    driver.quit()
                    log_message("Browser fechado", "INFO")
                except Exception as quit_error:
                    log_message(f"Erro ao fechar browser: {quit_error}", "WARNING")

    def mostrar_resumo_final(self, resultados):
        """Mostra o resumo final do processamento"""
        total = len(resultados)
        sucesso = len([r for r in resultados if r['status'] == 'sucesso'])
        erro_sessao = len([r for r in resultados if r['status'] == 'erro_sessao'])
        erros = len([r for r in resultados if 'erro' in r['status'] and r['status'] != 'erro_sessao'])
        
        log_message("\n" + "="*50, "INFO")
        log_message("RESUMO FINAL DO PROCESSAMENTO", "INFO")
        log_message("="*50, "INFO")
        log_message(f"Total de exames: {total}", "INFO")
        log_message(f"✅ Processados com sucesso: {sucesso}", "SUCCESS")
        log_message(f"🔄 Erros de sessão (browser perdido): {erro_sessao}", "WARNING")
        log_message(f"❌ Outros erros de processamento: {erros}", "ERROR")
        
        # Mostrar detalhes dos erros se houver
        erros_totais = erro_sessao + erros
        if erros_totais > 0:
            log_message("\nDetalhes dos erros:", "ERROR")
            for r in resultados:
                if 'erro' in r['status']:
                    log_message(f"- {r['codigo']}: {r['detalhes']}", "ERROR")
        
        messagebox.showinfo("Processamento Concluído", 
            f"✅ Processamento finalizado!\n\n"
            f"Total: {total}\n"
            f"Sucesso: {sucesso}\n"
            f"Erros de sessão: {erro_sessao}\n"
            f"Outros erros: {erros}")

def run(params: dict):
    module = MacroscopiaFixacaoModule()
    module.run(params) 