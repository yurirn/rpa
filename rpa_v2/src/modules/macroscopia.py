from tkinter import messagebox
import os
import time
import unicodedata
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC

from dotenv import load_dotenv
from openpyxl import load_workbook

from src.core.browser_factory import BrowserFactory
from src.core.logger import log_message
from src.modules.base import BaseModule

load_dotenv()

DEFAULT_TIMEOUT = 30
SHORT_DELAY = 0.5
MEDIUM_DELAY = 1
LONG_DELAY = 2

class MacroscopiaModule(BaseModule):
    def __init__(self):
        super().__init__(nome="Macroscopia")

    # --- Utilitários Selenium ---
    def aguardar_elemento(self, wait, by, value, timeout=DEFAULT_TIMEOUT):
        """Aguarda um elemento estar presente no DOM."""
        return wait.until(EC.presence_of_element_located((by, value)))

    def aguardar_elemento_clicavel(self, wait, by, value, timeout=DEFAULT_TIMEOUT):
        """Aguarda um elemento estar clicável."""
        return wait.until(EC.element_to_be_clickable((by, value)))

    def clicar_elemento(self, driver, elem):
        """Rola até o elemento e clica nele."""
        driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", elem)
        time.sleep(SHORT_DELAY)
        elem.click()
        time.sleep(SHORT_DELAY)

    def preencher_campo(self, campo, valor):
        """Limpa e preenche um campo de input."""
        campo.clear()
        time.sleep(SHORT_DELAY)
        campo.send_keys(valor)
        time.sleep(SHORT_DELAY)

    def pressionar_enter(self, campo):
        """Pressiona Enter em um campo."""
        campo.send_keys(Keys.ENTER)
        time.sleep(SHORT_DELAY)

    # --- Normalização e busca robusta em selects ---
    def normalizar_nome(self, nome):
        if not nome:
            return ""
        nome = nome.strip().lower()
        nome = unicodedata.normalize('NFKD', nome)
        nome = ''.join([c for c in nome if not unicodedata.combining(c)])
        nome = ' '.join(nome.split())
        return nome

    def buscar_valor_select_por_nome(self, select_elem, nome_busca):
        """Busca o value de uma option em um select pelo nome, normalizando."""
        nome_normalizado = self.normalizar_nome(nome_busca)
        for opt in select_elem.find_elements(By.TAG_NAME, "option"):
            if self.normalizar_nome(opt.text) == nome_normalizado:
                return opt.get_attribute("value")
        return None

    # --- Métodos principais ---
    def get_dados_exames(self, file_path: str) -> list:
        try:
            workbook = load_workbook(file_path)
            sheet = workbook.active
            dados = []
            ultima_mascara = None

            # Lê da linha 2 em diante (linha 1 é cabeçalho)
            for row in range(2, sheet.max_row + 1):
                codigo = sheet[f'A{row}'].value
                mascara = sheet[f'B{row}'].value
                citotecnica = sheet[f'C{row}'].value

                if codigo is not None:
                    codigo = str(codigo).strip()

                    # Se não tem máscara, usa a última válida
                    if mascara is not None and str(mascara).strip():
                        mascara = str(mascara).strip()
                        ultima_mascara = mascara
                    else:
                        mascara = ultima_mascara

                    dados.append({
                        'codigo': codigo,
                        'mascara': mascara,
                        'citotecnica': citotecnica
                    })

            workbook.close()
            return dados
        except Exception as e:
            raise Exception(f"Erro ao ler planilha: {e}")

    def processar_exame(self, driver, wait, codigo, mascara, citotecnica_nome):
        """Processa um exame individual: digita o código, executa rotina de macroscopia."""
        try:
            log_message("Aguardando página carregar...", "INFO")
            time.sleep(SHORT_DELAY)
            campo_codigo = self.aguardar_elemento(wait, By.ID, "inputSearchCodBarra")
            log_message("✅ Campo de código encontrado", "INFO")
            self.preencher_campo(campo_codigo, codigo)
            self.pressionar_enter(campo_codigo)
            return self.aguardar_e_processar_andamento(driver, wait, codigo, mascara, citotecnica_nome)
        except Exception as e:
            log_message(f"Erro ao processar exame {codigo}: {e}", "ERROR")
            return {'status': 'erro', 'detalhes': str(e)}

    def digitar_mascara_e_buscar(self, driver, wait, mascara):
        try:
            campo_busca = self.aguardar_elemento_clicavel(wait, By.ID, "buscaArvore")
            if not campo_busca.is_displayed():
                log_message("⚠️ Campo buscaArvore não está visível", "WARNING")
                return
            self.preencher_campo(campo_busca, mascara)
            self.pressionar_enter(campo_busca)
        except Exception as e:
            log_message(f"Erro ao digitar máscara: {e}", "ERROR")
            raise

    def salvar_macroscopia(self, driver, wait):
        try:
            botoes_onclick = driver.find_elements(By.XPATH, "//a[contains(@onclick, 'ajaxChangeSave')]")
            if botoes_onclick:
                self.clicar_elemento(driver, botoes_onclick[0])
                log_message("💾 Clicou em Salvar usando onclick", "INFO")
            else:
                log_message("❌ Botão Salvar não encontrado.", "ERROR")
                raise Exception("Botão Salvar não encontrado")

            time.sleep(SHORT_DELAY)
        except Exception as e:
            log_message(f"Erro ao salvar: {e}", "ERROR")
            raise

    def selecionar_painel_papanicolau(self, driver, wait):
        """Seleciona a opção 'Papanicolau ( Rotina ) Clone:' no select painel, sem buscar por nome normalizado."""
        try:
            select_elem = self.aguardar_elemento(wait, By.ID, "painel")

            if not select_elem.is_displayed():
                driver.execute_script("$(arguments[0]).val('tecnica_12747').trigger('change');", select_elem)
            else:
                Select(select_elem).select_by_value("tecnica_12747")
            log_message("✅ Painel 'Papanicolau ( Rotina ) Clone:' selecionado", "SUCCESS")
            time.sleep(SHORT_DELAY)
        except Exception as e:
            log_message(f"Erro ao selecionar painel: {e}", "ERROR")
            raise

    def enviar_proxima_etapa(self, driver, wait):
        try:
            botao_enviar = self.aguardar_elemento_clicavel(wait, By.ID, "btn-enviar-proxima-etapa")
            self.clicar_elemento(driver, botao_enviar)
            log_message("➡️ Clicou em Enviar para próxima etapa", "INFO")
        except Exception as e:
            log_message(f"Erro ao enviar para próxima etapa: {e}", "ERROR")
            raise

    def preencher_campo_codigo_novamente(self, driver, wait, codigo):
        try:
            campo_codigo = self.aguardar_elemento(wait, By.ID, "inputSearchCodBarra")
            self.preencher_campo(campo_codigo, codigo)
            self.pressionar_enter(campo_codigo)

            log_message("➡️ Preencheu novamente o campo código.", "INFO")
        except Exception as e:
            log_message(f"Erro ao preencher novamente o campo código: {e}", "ERROR")
            raise

    def selecionar_citotecnica(self, driver, wait, citotecnica_nome):
        """Seleciona a citotécnica pelo nome (do Excel), convertendo para value via dicionário fixo e selecionando pelo value."""
        try:
            select_elem = self.aguardar_elemento(wait, By.ID, "citotecnico")
            time.sleep(1)  # Garante que o select foi populado (caso seja AJAX)
            # Dicionário fixo de nome para value
            nome_para_value = {
                "adriana domiciano fialho": "105789",
                "andrea clementino romero da costa staevie": "105788",
                "administrador do sistema": "514",
                "annai luka vitorino losnak": "519",
                "suporte adrienne intersistemas": "226754",
                "suporte dani intersistemas": "247216",
                "suporte deni intersistemas": "247215",
                "suporte erika intersistemas": "226755",
                "suporte flavia intersistemas": "226759",
                "suporte ingrid intersistemas": "338762",
                "suporte intersistemas": "513",
                "suporte intersistemas (2)": "226760",
                "suporte jose intersistemas": "339195",
                "suporte pedro intersistemas": "226756",
                "suporte priscila intersistemas": "226758",
            }
            nome_normalizado = self.normalizar_nome(citotecnica_nome)
            value = nome_para_value.get(nome_normalizado)
            if not value:
                # Salva o HTML do select para debug
                html = select_elem.get_attribute("outerHTML")
                with open("debug_citotecnico_select.html", "w", encoding="utf-8") as f:
                    f.write(html)
                log_message(f"❌ Citotécnica '{citotecnica_nome}' não encontrada no dicionário. HTML salvo em debug_citotecnico_select.html", "ERROR")
                return
            if not select_elem.is_displayed():
                driver.execute_script(f"$(arguments[0]).val('{value}').trigger('change');", select_elem)
            else:
                Select(select_elem).select_by_value(value)
            log_message(f"✅ Citotécnica selecionada (value: {value})", "SUCCESS")
            time.sleep(MEDIUM_DELAY)
        except Exception as e:
            log_message(f"Erro ao selecionar citotécnica: {e}", "ERROR")
            raise

    def aguardar_e_processar_andamento(self, driver, wait, codigo, mascara, citotecnica_nome):
        log_message("Aguardando div de andamento do exame aparecer...", "INFO")
        inicio = time.time()
        while time.time() - inicio < DEFAULT_TIMEOUT:
            try:
                andamento_div = driver.find_element(By.ID, "divAndamentoExame")
                if andamento_div and andamento_div.is_displayed():
                    log_message("📋 Div de andamento do exame encontrada!", "SUCCESS")
                    break
            except:
                pass
            time.sleep(1)
        else:
            log_message("⚠️ Div de andamento não apareceu no tempo esperado", "WARNING")
            return {'status': 'sem_andamento', 'detalhes': 'Exame não encontrado ou não carregou'}
        time.sleep(SHORT_DELAY)
        if mascara:
            self.digitar_mascara_e_buscar(driver, wait, mascara)
            self.salvar_macroscopia(driver, wait)
            self.selecionar_painel_papanicolau(driver, wait)
            self.enviar_proxima_etapa(driver, wait)
            if codigo:
                self.preencher_campo_codigo_novamente(driver, wait, codigo)
                if citotecnica_nome:
                    self.selecionar_citotecnica(driver, wait, citotecnica_nome)
                self.fechar_exame(driver, wait)
        else:
            log_message("⚠️ Nenhuma máscara encontrada, pulando busca", "WARNING")
        return {'status': 'sucesso'}

    def fechar_exame(self, driver, wait):
        try:
            botao_fechar = wait.until(EC.element_to_be_clickable((By.ID, "fecharExameBarraFerramenta")))
            botao_fechar.click()
            log_message("📁 Exame fechado", "INFO")
        except Exception as e:
            log_message(f"Erro ao fechar exame: {e}", "ERROR")

    def mostrar_resumo_final(self, resultados):
        """Mostra o resumo final do processamento"""
        total = len(resultados)
        sucesso = len([r for r in resultados if r['status'] == 'sucesso'])
        sem_andamento = len([r for r in resultados if r['status'] == 'sem_andamento'])
        erros = len([r for r in resultados if 'erro' in r['status']])

        log_message("\n" + "="*50, "INFO")
        log_message("RESUMO FINAL DO PROCESSAMENTO", "INFO")
        log_message("="*50, "INFO")
        log_message(f"Total de exames: {total}", "INFO")
        log_message(f"✅ Processados com sucesso: {sucesso}", "SUCCESS")
        log_message(f"⚠️ Exames não encontrados: {sem_andamento}", "WARNING")
        log_message(f"❌ Erros de processamento: {erros}", "ERROR")

        # Mostrar detalhes dos erros se houver
        if erros > 0:
            log_message("\nDetalhes dos erros:", "ERROR")
            for r in resultados:
                if 'erro' in r['status']:
                    log_message(f"- {r['codigo']}: {r['detalhes']}", "ERROR")

        messagebox.showinfo("Processamento Concluído",
            f"✅ Processamento finalizado!\n\n"
            f"Total: {total}\n"
            f"Sucesso: {sucesso}\n"
            f"Não encontrados: {sem_andamento}\n"
            f"Erros: {erros}")

    def run(self, params: dict):
        username = params.get("username")
        password = params.get("password")
        excel_file = params.get("excel_file")
        cancel_flag = params.get("cancel_flag")
        headless_mode = params.get("headless_mode")
        try:
            dados_exames = self.get_dados_exames(excel_file)
            if not dados_exames:
                messagebox.showerror("Erro", "Nenhum dado de exame encontrado na planilha.")
                return
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao ler o Excel: {e}")
            return
        url = os.getenv("SYSTEM_URL", "https://pathoweb.com.br/login/auth")
        driver = None
        resultados = []
        try:
            driver = BrowserFactory.create_chrome(headless=headless_mode)
            wait = WebDriverWait(driver, DEFAULT_TIMEOUT)
            log_message("Iniciando automação de macroscopia...", "INFO")
            driver.get(url)
            # Login
            wait.until(EC.presence_of_element_located((By.ID, "username"))).send_keys(username)
            wait.until(EC.presence_of_element_located((By.ID, "password"))).send_keys(password)
            driver.find_element(By.CSS_SELECTOR, "button[type='submit']").click()

            log_message("Verificando se precisa navegar para módulo de exames...", "INFO")
            current_url = driver.current_url
            if current_url == "https://pathoweb.com.br/" or "trocarModulo" in current_url:
                log_message("Detectada tela de seleção de módulos - navegando para módulo de exames...", "INFO")
                try:
                    modulo_link = wait.until(
                        EC.element_to_be_clickable((By.CSS_SELECTOR, "a[href='/site/trocarModulo?modulo=1']")))
                    modulo_link.click()
                    time.sleep(MEDIUM_DELAY)
                    log_message("✅ Navegação para módulo de exames realizada", "SUCCESS")
                except Exception as e:
                    log_message(f"⚠️ Erro ao navegar para módulo: {e}", "WARNING")
                    # Tentar navegar diretamente pela URL como fallback
                    driver.get("https://pathoweb.com.br/moduloExame/index")
                    time.sleep(MEDIUM_DELAY)
                    log_message("🔄 Navegação direta para módulo realizada", "INFO")

            elif "moduloExame" in current_url:
                log_message("✅ Já está no módulo de exames - pulando navegação", "SUCCESS")
            else:
                log_message(f"⚠️ URL inesperada detectada: {current_url}", "WARNING")
                # Tentar navegar diretamente como fallback
                driver.get("https://pathoweb.com.br/moduloExame/index")
                time.sleep(MEDIUM_DELAY)
                log_message("🔄 Navegação direta para módulo realizada (fallback)", "INFO")

            try:
                modal_close_button = driver.find_element(By.CSS_SELECTOR, "#mensagemParaClienteModal .modal-footer button")
                if modal_close_button.is_displayed():
                    self.clicar_elemento(driver, modal_close_button)
            except Exception:
                pass
            for i, exame_data in enumerate(dados_exames, 1):
                if cancel_flag and cancel_flag.is_set():
                    log_message("Execução cancelada pelo usuário.", "WARNING")
                    break
                codigo = exame_data['codigo']
                mascara = exame_data['mascara']
                citotecnica_nome = exame_data.get('citotecnica')
                log_message(f"\n➡️ Processando exame {i}/{len(dados_exames)}: {codigo} (máscara: {mascara}) - Citotécnica: {citotecnica_nome}", "INFO")
                resultado = self.processar_exame(driver, wait, codigo, mascara, citotecnica_nome)
                resultados.append({
                    'codigo': codigo,
                    'mascara': mascara,
                    'citotecnica': citotecnica_nome,
                    'status': resultado['status'],
                    'detalhes': resultado.get('detalhes', '')
                })
        except Exception as e:
            log_message(f"❌ Erro durante a automação: {e}", "ERROR")
        finally:
            if driver:
                try:
                    driver.quit()
                except Exception:
                    pass
            self.mostrar_resumo_final(resultados)

def run(params: dict):
    module = MacroscopiaModule()
    module.run(params)