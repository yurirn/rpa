import os
import time
from tkinter import messagebox
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from dotenv import load_dotenv
from openpyxl import load_workbook

from src.core.browser_factory import BrowserFactory
from src.core.logger import log_message
from src.modules.base import BaseModule

load_dotenv()

class ConclusaoComAlteracaoModule(BaseModule):
    def __init__(self):
        super().__init__(nome="Conclusão com Alteração")

    def get_dados_exames(self, file_path: str) -> list:
        try:
            workbook = load_workbook(file_path)
            sheet = workbook.active
            dados = []

            # Lê da linha 2 em diante (linha 1 é cabeçalho)
            for row in range(2, sheet.max_row + 1):
                codigo = sheet[f'A{row}'].value

                if codigo is not None:
                    codigo = str(codigo).strip()

                    dados.append({'codigo': codigo})

            workbook.close()
            return dados
        except Exception as e:
            raise Exception(f"Erro ao ler planilha: {e}")

    def aguardar_usuario_fechar_exame(self, driver, wait, codigo, timeout=300):
        """Aguarda o usuário fazer alterações e fechar o exame"""
        try:
            log_message(f"⏳ Aguardando usuário processar exame {codigo}...", "INFO")
            log_message(f"⏳ O usuário deve fazer as alterações necessárias e fechar o exame",
                        "WARNING")
            log_message(f"⏳ Timeout: {timeout}s ({timeout // 60} minutos)", "INFO")

            inicio = time.time()
            campo_detectado = False
            contador_log = 0

            while time.time() - inicio < timeout:
                try:
                    # Verificar se o campo de código está visível novamente (sinal que o exame foi fechado)
                    campo_codigo = driver.find_element(By.ID, "inputSearchCodBarra")

                    if campo_codigo.is_displayed():
                        log_message(f"✅ Campo de código detectado - usuário fechou o exame {codigo}!",
                                    "SUCCESS")
                        campo_detectado = True
                        time.sleep(0.5)
                        log_message(f"✅ Exame fechado com sucesso pelo usuário", "SUCCESS")
                        return True

                except Exception:
                    # Elemento não encontrado ou não visível, continuar verificando
                    pass

                # Log informativo a cada 30 segundos para não poluir
                tempo_decorrido = int(time.time() - inicio)
                if tempo_decorrido > contador_log and tempo_decorrido % 30 == 0:
                    minutos = tempo_decorrido // 60
                    segundos = tempo_decorrido % 60
                    log_message(
                        f"⏳ Aguardando usuário fechar o exame... ({minutos}m {segundos}s)",
                        "INFO"
                    )
                    contador_log = tempo_decorrido

                # Intervalo de verificação
                time.sleep(0.5)

            # Timeout atingido
            if not campo_detectado:
                log_message(
                    f"⚠️ Timeout de {timeout}s atingido - usuário não fechou o exame {codigo}",
                    "WARNING"
                )
                return False

        except Exception as e:
            log_message(f"❌ Erro ao aguardar fechamento do exame: {e}", "ERROR")
            import traceback
            log_message(f"❌ Stack trace: {traceback.format_exc()}", "ERROR")
            return False

    def processar_exame(self, driver, wait, codigo):
        """Processa um exame individual"""
        try:
            # Verificar se a sessão do browser ainda está ativa
            if not self.verificar_sessao_browser(driver):
                raise Exception("Sessão do browser perdida - necessário reiniciar")

            campo_codigo = None
            try:
                campo_codigo = wait.until(EC.presence_of_element_located((By.ID, "inputSearchCodBarra")))
                log_message("✅ Campo encontrado pelo ID", "INFO")
            except:
                log_message("❌ Campo não encontrado", "ERROR")
                raise Exception("Campo de código de barras não encontrado")

            # Interagir com o campo
            self.interagir_com_campo_codigo(driver, campo_codigo, codigo)

            # Aguardar usuário fechar o exame
            if not self.aguardar_usuario_fechar_exame(driver, wait, codigo):
                return {'status': 'timeout', 'detalhes': 'Usuário não fechou o exame no tempo esperado'}

            log_message("🎉 Exame processado com sucesso!", "SUCCESS")
            return {'status': 'sucesso', 'detalhes': 'Exame processado'}

        except Exception as e:
            error_message = str(e)
            log_message(f"❌ Erro ao processar exame {codigo}: {error_message}", "ERROR")

            if "invalid session id" in error_message.lower():
                return {'status': 'erro_sessao', 'detalhes': 'Sessão do browser perdida'}

            return {'status': 'erro', 'detalhes': error_message}

    def interagir_com_campo_codigo(self, driver, campo_codigo, codigo):
        """Interage com o campo de código usando os métodos já implementados"""
        log_message("Campo de código encontrado, interagindo...", "INFO")

        try:
            campo_codigo.clear()
            log_message("Campo limpo com sucesso", "INFO")
        except:
            driver.execute_script("arguments[0].value = '';", campo_codigo)
            log_message("Campo limpo com JavaScript", "INFO")

        time.sleep(0.5)

        # Digitar o código
        try:
            campo_codigo.send_keys(codigo)
            log_message(f"Código '{codigo}' digitado com sucesso", "INFO")
        except:
            driver.execute_script(f"arguments[0].value = '{codigo}';", campo_codigo)
            log_message(f"Código '{codigo}' digitado com JavaScript", "INFO")

        time.sleep(1)

        # Pressionar Enter
        try:
            campo_codigo.send_keys(Keys.ENTER)
            log_message("⌨️ Enter pressionado com sucesso", "INFO")
        except:
            driver.execute_script("""
                var element = arguments[0];
                var event = new KeyboardEvent('keydown', {
                    key: 'Enter',
                    code: 'Enter',
                    keyCode: 13,
                    bubbles: true
                });
                element.dispatchEvent(event);
            """, campo_codigo)
            log_message("⌨️ Enter pressionado com JavaScript", "INFO")

        log_message("Aguardando div de andamento do exame aparecer...", "INFO")

        # Aguardar mais tempo para o carregamento após digitar o código
        timeout_andamento = 30
        inicio = time.time()

        while time.time() - inicio < timeout_andamento:
            try:
                # Verificar se a div de andamento apareceu
                andamento_div = driver.find_element(By.ID, "divAndamentoExame")
                if andamento_div and andamento_div.is_displayed():
                    log_message("📋 Div de andamento do exame encontrada!", "SUCCESS")
                    break
            except:
                pass

            time.sleep(1)
            if int(time.time() - inicio) % 5 == 0:  # Log a cada 5 segundos
                log_message(f"⏳ Aguardando carregamento... ({int(time.time() - inicio)}s)", "INFO")
        else:
            log_message("⚠️ Div de andamento não apareceu no tempo esperado", "WARNING")
            return {'status': 'sem_andamento', 'detalhes': 'Exame não encontrado ou não carregou'}


        time.sleep(1)

    def mostrar_resumo_final(self, resultados):
        """Mostra o resumo final do processamento"""
        total = len(resultados)
        sucesso = len([r for r in resultados if r['status'] == 'sucesso'])
        timeout = len([r for r in resultados if r['status'] == 'timeout'])
        erros = len([r for r in resultados if 'erro' in r['status']])

        log_message("\n" + "=" * 50, "INFO")
        log_message("RESUMO FINAL DO PROCESSAMENTO", "INFO")
        log_message("=" * 50, "INFO")
        log_message(f"Total de exames: {total}", "INFO")
        log_message(f"✅ Processados com sucesso: {sucesso}", "SUCCESS")
        log_message(f"⏱️ Timeout (usuário não fechou): {timeout}", "WARNING")
        log_message(f"❌ Erros de processamento: {erros}", "ERROR")

        if erros > 0:
            log_message("\nDetalhes dos erros:", "ERROR")
            for r in resultados:
                if 'erro' in r['status']:
                    log_message(f"- {r['codigo']}: {r['detalhes']}", "ERROR")

        messagebox.showinfo("Processamento Concluído",
                            f"✅ Processamento finalizado!\n\n"
                            f"Total: {total}\n"
                            f"Sucesso: {sucesso}\n"
                            f"Timeout: {timeout}\n"
                            f"Erros: {erros}")

    def verificar_sessao_browser(self, driver) -> bool:
        """Verifica se a sessão do browser ainda está ativa"""
        try:
            driver.current_url
            return True
        except Exception as e:
            if "invalid session id" in str(e).lower():
                log_message("❌ Sessão do browser perdida", "ERROR")
                return False
            return True

    def run(self, params: dict):
        username = params.get("username")
        password = params.get("password")
        excel_file = params.get("excel_file")
        cancel_flag = params.get("cancel_flag")

        try:
            # Lê os dados dos exames da planilha
            dados_exames = self.get_dados_exames(excel_file)
            if not dados_exames:
                messagebox.showerror("Erro", "Nenhum dado de exame encontrado na planilha.")
                return

            log_message(f"Encontrados {len(dados_exames)} exames para processar", "INFO")

        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao ler o Excel: {e}")
            return

        url = os.getenv("SYSTEM_URL", "https://dap.pathoweb.com.br/login/auth")
        driver = None
        resultados = []

        try:
            driver = BrowserFactory.create_chrome()
            wait = WebDriverWait(driver, 20)

            log_message("Iniciando automação de conclusão com alteração...", "INFO")

            # Login
            log_message("Fazendo login...", "INFO")
            driver.get(url)

            wait.until(EC.presence_of_element_located((By.ID, "username")))

            username_field = driver.find_element(By.ID, "username")
            username_field.clear()
            username_field.send_keys(username)

            password_field = driver.find_element(By.ID, "password")
            password_field.clear()
            password_field.send_keys(password)

            submit_button = driver.find_element(By.CSS_SELECTOR, "button[type='submit']")
            submit_button.click()

            # Navegar para módulo de exames
            log_message("Verificando se precisa navegar para módulo de exames...", "INFO")
            current_url = driver.current_url
            if current_url == "https://dap.pathoweb.com.br/" or "trocarModulo" in current_url:
                try:
                    modulo_link = wait.until(
                        EC.element_to_be_clickable((By.CSS_SELECTOR, "a[href='/site/trocarModulo?modulo=1']")))
                    modulo_link.click()
                    time.sleep(2)
                    log_message("✅ Navegação para módulo de exames realizada", "SUCCESS")
                except Exception:
                    driver.get("https://dap.pathoweb.com.br/moduloExame/index")
                    time.sleep(2)

            # Fechar modal se aparecer
            try:
                modal_close_button = driver.find_element(By.CSS_SELECTOR,
                                                         "#mensagemParaClienteModal .modal-footer button")
                if modal_close_button.is_displayed():
                    modal_close_button.click()
            except Exception:
                pass

            log_message("✅ Login realizado com sucesso. Iniciando processamento dos exames.", "SUCCESS")

            # Processar cada exame da planilha
            for i, exame_data in enumerate(dados_exames, 1):
                if cancel_flag and cancel_flag.is_set():
                    log_message("Execução cancelada pelo usuário.", "WARNING")
                    break

                codigo = exame_data['codigo']

                log_message(f"\n➡️ Processando exame {i}/{len(dados_exames)}: {codigo}", "INFO")

                # Processar este exame específico
                resultado = self.processar_exame(driver, wait, codigo)
                resultados.append({
                    'codigo': codigo,
                    'status': resultado['status'],
                    'detalhes': resultado.get('detalhes', '')
                })

            # Mostrar resumo final
            self.mostrar_resumo_final(resultados)

        except Exception as e:
            log_message(f"❌ Erro durante a automação: {e}", "ERROR")
            messagebox.showerror("Erro", f"❌ Erro durante a automação:\n{str(e)[:200]}...")
        finally:
            log_message("✅ Execução finalizada - Browser permanece aberto", "SUCCESS")

def run(params: dict):
    module = ConclusaoComAlteracaoModule()
    module.run(params)
