from tkinter import messagebox
import os
import time
import unicodedata
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC

from dotenv import load_dotenv
from openpyxl import load_workbook

from src.core.browser_factory import BrowserFactory
from src.core.logger import log_message
from src.modules.base import BaseModule

load_dotenv()

DEFAULT_TIMEOUT = 30
SHORT_DELAY = 0.5
MEDIUM_DELAY = 1
LONG_DELAY = 2

class MacroscopiaModule(BaseModule):
    def __init__(self):
        super().__init__(nome="Macroscopia")

    # --- Utilitários Selenium ---
    def aguardar_elemento(self, wait, by, value, timeout=DEFAULT_TIMEOUT):
        """Aguarda um elemento estar presente no DOM."""
        return wait.until(EC.presence_of_element_located((by, value)))

    def aguardar_elemento_clicavel(self, wait, by, value, timeout=DEFAULT_TIMEOUT):
        """Aguarda um elemento estar clicável."""
        return wait.until(EC.element_to_be_clickable((by, value)))

    def clicar_elemento(self, driver, elem):
        """Rola até o elemento e clica nele."""
        log_message("🔹 Rolando até o elemento...", "INFO")
        driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", elem)
        time.sleep(SHORT_DELAY)

        log_message("🔹 Executando clique...", "INFO")
        elem.click()
        time.sleep(SHORT_DELAY)
        log_message("✅ Clique executado com sucesso", "SUCCESS")

    def preencher_campo(self, campo, valor):
        """Limpa e preenche um campo de input."""
        log_message(f"🔹 Limpando campo...", "INFO")
        campo.clear()
        time.sleep(SHORT_DELAY)

        log_message(f"🔹 Digitando valor: '{valor}'", "INFO")
        campo.send_keys(valor)
        time.sleep(SHORT_DELAY)
        log_message("✅ Campo preenchido com sucesso", "SUCCESS")

    def pressionar_enter(self, campo):
        """Pressiona Enter em um campo."""
        log_message("🔹 Enviando tecla ENTER...", "INFO")
        campo.send_keys(Keys.ENTER)
        time.sleep(SHORT_DELAY)
        log_message("✅ ENTER enviado com sucesso", "SUCCESS")

    # --- Normalização e busca robusta em selects ---
    def normalizar_nome(self, nome):
        if not nome:
            return ""
        nome = nome.strip().lower()
        nome = unicodedata.normalize('NFKD', nome)
        nome = ''.join([c for c in nome if not unicodedata.combining(c)])
        nome = ' '.join(nome.split())
        return nome

    def buscar_valor_select_por_nome(self, select_elem, nome_busca):
        """Busca o value de uma option em um select pelo nome, normalizando."""
        nome_normalizado = self.normalizar_nome(nome_busca)
        for opt in select_elem.find_elements(By.TAG_NAME, "option"):
            if self.normalizar_nome(opt.text) == nome_normalizado:
                return opt.get_attribute("value")
        return None

    # --- Métodos principais ---
    def get_dados_exames(self, file_path: str) -> list:
        try:
            workbook = load_workbook(file_path)
            sheet = workbook.active
            dados = []
            ultima_mascara = None

            # Lê da linha 2 em diante (linha 1 é cabeçalho)
            for row in range(2, sheet.max_row + 1):
                codigo = sheet[f'A{row}'].value
                mascara = sheet[f'B{row}'].value
                citotecnica = sheet[f'C{row}'].value

                if codigo is not None:
                    codigo = str(codigo).strip()

                    # Se não tem máscara, usa a última válida
                    if mascara is not None and str(mascara).strip():
                        mascara = str(mascara).strip().upper()  # Sempre maiúsculo
                        ultima_mascara = mascara
                    else:
                        mascara = ultima_mascara

                    dados.append({
                        'codigo': codigo,
                        'mascara': mascara,
                        'citotecnica': citotecnica
                    })

            workbook.close()
            return dados
        except Exception as e:
            raise Exception(f"Erro ao ler planilha: {e}")

    def processar_exame(self, driver, wait, codigo, mascara, citotecnica_nome):
        """Processa um exame individual: digita o código, executa rotina de macroscopia."""
        try:
            log_message(f"🔵 Iniciando processamento do código: {codigo}", "INFO")
            log_message("🔵  Aguardando página carregar...", "INFO")
            time.sleep(SHORT_DELAY)

            log_message("🔵 Buscando campo de código (inputSearchCodBarra)...", "INFO")
            campo_codigo = self.aguardar_elemento(wait, By.ID, "inputSearchCodBarra")
            log_message("✅  Campo de código encontrado", "SUCCESS")

            log_message(f"🔵 Preenchendo campo com código: {codigo}", "INFO")
            self.preencher_campo(campo_codigo, codigo)
            log_message("✅ Campo preenchido com sucesso", "SUCCESS")

            log_message("🔵 Pressionando ENTER no campo...", "INFO")
            self.pressionar_enter(campo_codigo)
            log_message("✅ ENTER pressionado", "SUCCESS")

            log_message("🔵 Chamando aguardar_e_processar_andamento...", "INFO")
            resultado = self.aguardar_e_processar_andamento(driver, wait, codigo, mascara, citotecnica_nome)
            log_message(f"✅ Processamento concluído. Status: {resultado['status']}", "SUCCESS")

            return resultado

        except Exception as e:
            log_message(f"❌ Erro ao processar exame {codigo}: {e}", "ERROR")
            log_message(f"❌ Tipo do erro: {type(e).__name__}", "ERROR")
            import traceback
            log_message(f"❌ Stack trace: {traceback.format_exc()}", "ERROR")
            return {'status': 'erro', 'detalhes': str(e)}

    def digitar_mascara_e_buscar(self, driver, wait, mascara):
        try:
            log_message(f"🟡 Iniciando busca da máscara: '{mascara}'", "INFO")

            log_message("🟡 Aguardando campo 'buscaArvore' estar clicável...", "INFO")
            campo_busca = self.aguardar_elemento_clicavel(wait, By.ID, "buscaArvore")
            log_message("✅ Campo 'buscaArvore' localizado", "SUCCESS")

            if not campo_busca.is_displayed():
                log_message("⚠️ Campo 'buscaArvore' não está visível!", "WARNING")
                return

            log_message(f"🟡 Preenchendo campo com máscara: '{mascara}'", "INFO")
            self.preencher_campo(campo_busca, mascara)
            log_message("✅ Campo preenchido com sucesso", "SUCCESS")

            log_message("🟡 Pressionando ENTER...", "INFO")
            self.pressionar_enter(campo_busca)
            log_message("✅ ENTER pressionado com sucesso", "SUCCESS")

        except Exception as e:
            log_message(f"❌ Erro ao digitar máscara: {e}", "ERROR")
            log_message(f"❌ Tipo do erro: {type(e).__name__}", "ERROR")
            raise

    def salvar_macroscopia(self, driver, wait):
        try:
            log_message("🟠 Iniciando processo de salvamento...", "INFO")

            log_message("🟠 Buscando botões com onclick='ajaxChangeSave'...", "INFO")
            botoes_onclick = driver.find_elements(By.XPATH, "//a[contains(@onclick, 'ajaxChangeSave')]")

            if botoes_onclick:
                log_message(f"✅ Encontrados {len(botoes_onclick)} botão(ões)", "SUCCESS")
                log_message("🟠 Clicando no primeiro botão...", "INFO")
                self.clicar_elemento(driver, botoes_onclick[0])
                log_message("✅ Clique no botão Salvar executado com sucesso", "SUCCESS")
            else:
                log_message("❌ Nenhum botão Salvar encontrado na página!", "ERROR")
                raise Exception("Botão Salvar não encontrado")

            log_message(f"🟠 Aguardando {SHORT_DELAY}s após salvar...", "INFO")
            time.sleep(SHORT_DELAY)
            log_message("✅ Salvamento concluído com sucesso", "SUCCESS")

        except Exception as e:
            log_message(f"❌ Erro ao salvar: {e}", "ERROR")
            log_message(f"❌ Tipo do erro: {type(e).__name__}", "ERROR")
            raise

    def selecionar_painel_papanicolau(self, driver, wait):
        """Seleciona a opção 'Papanicolau ( Rotina ) Clone:' no select painel, sem buscar por nome normalizado."""
        try:
            select_elem = self.aguardar_elemento(wait, By.ID, "painel")

            if not select_elem.is_displayed():
                driver.execute_script("$(arguments[0]).val('tecnica_12747').trigger('change');", select_elem)
            else:
                Select(select_elem).select_by_value("tecnica_12747")
            log_message("✅ Painel 'Papanicolau ( Rotina ) Clone:' selecionado", "SUCCESS")
            time.sleep(SHORT_DELAY)
        except Exception as e:
            log_message(f"Erro ao selecionar painel: {e}", "ERROR")
            raise

    def enviar_proxima_etapa(self, driver, wait):
        try:
            log_message("🟣 Iniciando envio para próxima etapa...", "INFO")

            log_message("🟣 Aguardando botão 'btn-enviar-proxima-etapa' estar clicável...", "INFO")
            botao_enviar = self.aguardar_elemento_clicavel(wait, By.ID, "btn-enviar-proxima-etapa")
            log_message("✅ Botão localizado", "SUCCESS")

            log_message("🟣 Clicando no botão...", "INFO")
            self.clicar_elemento(driver, botao_enviar)
            log_message("✅ Clique executado com sucesso", "SUCCESS")

            log_message(f"🟣 Aguardando {MEDIUM_DELAY}s após envio...", "INFO")
            time.sleep(MEDIUM_DELAY)
            log_message("✅ Envio para próxima etapa concluído", "SUCCESS")

        except Exception as e:
            log_message(f"❌ Erro ao enviar para próxima etapa: {e}", "ERROR")
            log_message(f"❌ Tipo do erro: {type(e).__name__}", "ERROR")
            raise

    def preencher_campo_codigo_novamente(self, driver, wait, codigo):
        try:
            campo_codigo = self.aguardar_elemento(wait, By.ID, "inputSearchCodBarra")
            self.preencher_campo(campo_codigo, codigo)
            self.pressionar_enter(campo_codigo)

            log_message("➡️ Preencheu novamente o campo código.", "INFO")
        except Exception as e:
            log_message(f"Erro ao preencher novamente o campo código: {e}", "ERROR")
            raise

    def selecionar_citotecnica(self, driver, wait, citotecnica_nome):
        """Seleciona a citotécnica pelo primeiro nome (do Excel), convertendo para value via dicionário fixo e selecionando pelo value."""
        try:
            select_elem = self.aguardar_elemento(wait, By.ID, "citotecnico")
            time.sleep(1)  # Garante que o select foi populado (caso seja AJAX)
            # Dicionário fixo de primeiro nome para value
            primeiro_nome_para_value = {
                "adriana": "105789",
                "andrea": "105788",
                # Adicione outros nomes e values conforme necessário
            }
            if not citotecnica_nome:
                log_message(f"❌ Citotécnica não informada.", "ERROR")
                return
            # Extrai o primeiro nome, normaliza
            primeiro_nome = str(citotecnica_nome).strip().split()[0].lower()
            primeiro_nome = unicodedata.normalize('NFKD', primeiro_nome)
            primeiro_nome = ''.join([c for c in primeiro_nome if not unicodedata.combining(c)])
            value = primeiro_nome_para_value.get(primeiro_nome)
            if not value:
                # Salva o HTML do select para debug
                html = select_elem.get_attribute("outerHTML")
                with open("debug_citotecnico_select.html", "w", encoding="utf-8") as f:
                    f.write(html)
                log_message(f"❌ Citotécnica '{citotecnica_nome}' (primeiro nome: '{primeiro_nome}') não encontrada no dicionário. HTML salvo em debug_citotecnico_select.html", "ERROR")
                return
            if not select_elem.is_displayed():
                driver.execute_script(f"$(arguments[0]).val('{value}').trigger('change');", select_elem)
            else:
                Select(select_elem).select_by_value(value)
            log_message(f"✅ Citotécnica selecionada (value: {value})", "SUCCESS")
            time.sleep(MEDIUM_DELAY)
        except Exception as e:
            log_message(f"Erro ao selecionar citotécnica: {e}", "ERROR")
            raise

    def aguardar_e_processar_andamento(self, driver, wait, codigo, mascara, citotecnica_nome):
        log_message("🟢 Iniciando aguardo da div de andamento...", "INFO")
        log_message(f"🟢 Timeout configurado: {DEFAULT_TIMEOUT}s", "INFO")

        inicio = time.time()
        tentativas = 0

        while time.time() - inicio < DEFAULT_TIMEOUT:
            tentativas += 1
            try:
                log_message(f"🟢 Tentativa {tentativas} - buscando divAndamentoExame...", "INFO")
                andamento_div = driver.find_element(By.ID, "divAndamentoExame")

                if andamento_div and andamento_div.is_displayed():
                    tempo_decorrido = time.time() - inicio
                    log_message(f"✅ Div de andamento encontrada após {tempo_decorrido:.2f}s!",
                                "SUCCESS")
                    break
                else:
                    log_message(f"⚠️ Div encontrada mas não está visível (tentativa {tentativas})",
                                "WARNING")

            except Exception as e:
                log_message(
                    f"🟢 Div ainda não encontrada (tentativa {tentativas}): {type(e).__name__}",
                    "INFO")

            time.sleep(1)
        else:
            log_message(f"❌ Timeout de {DEFAULT_TIMEOUT}s atingido após {tentativas} tentativas",
                        "ERROR")
            return {'status': 'sem_andamento', 'detalhes': 'Exame não encontrado ou não carregou'}

        time.sleep(SHORT_DELAY)

        if mascara:
            log_message(f"🟢 Máscara encontrada: '{mascara}' - Iniciando fluxo de processamento",
                        "INFO")

            try:
                log_message("🟢 Chamando digitar_mascara_e_buscar...", "INFO")
                self.digitar_mascara_e_buscar(driver, wait, mascara)
                log_message("✅ digitar_mascara_e_buscar concluído", "SUCCESS")
            except Exception as e:
                log_message(f"❌ Erro em digitar_mascara_e_buscar: {e}", "ERROR")
                raise

            try:
                log_message("🟢 Chamando salvar_macroscopia...", "INFO")
                self.salvar_macroscopia(driver, wait)
                log_message("✅ salvar_macroscopia concluído", "SUCCESS")
            except Exception as e:
                log_message(f"❌ Erro em salvar_macroscopia: {e}", "ERROR")
                raise

            try:
                log_message("🟢 Chamando enviar_proxima_etapa...", "INFO")
                self.enviar_proxima_etapa(driver, wait)
                log_message("✅ enviar_proxima_etapa concluído", "SUCCESS")
            except Exception as e:
                log_message(f"❌ Erro em enviar_proxima_etapa: {e}", "ERROR")
                raise

        else:
            log_message("⚠️ Nenhuma máscara encontrada, pulando busca", "WARNING")

        log_message("✅  Processamento do andamento concluído com sucesso", "SUCCESS")
        return {'status': 'sucesso'}

    def fechar_exame(self, driver, wait):
        try:
            botao_fechar = wait.until(EC.element_to_be_clickable((By.ID, "fecharExameBarraFerramenta")))
            botao_fechar.click()
            log_message("📁 Exame fechado", "INFO")
        except Exception as e:
            log_message(f"Erro ao fechar exame: {e}", "ERROR")

    def mostrar_resumo_final(self, resultados):
        """Mostra o resumo final do processamento"""
        total = len(resultados)
        sucesso = len([r for r in resultados if r['status'] == 'sucesso'])
        sem_andamento = len([r for r in resultados if r['status'] == 'sem_andamento'])
        erros = len([r for r in resultados if 'erro' in r['status']])

        log_message("\n" + "="*50, "INFO")
        log_message("RESUMO FINAL DO PROCESSAMENTO", "INFO")
        log_message("="*50, "INFO")
        log_message(f"Total de exames: {total}", "INFO")
        log_message(f"✅ Processados com sucesso: {sucesso}", "SUCCESS")
        log_message(f"⚠️ Exames não encontrados: {sem_andamento}", "WARNING")
        log_message(f"❌ Erros de processamento: {erros}", "ERROR")

        # Mostrar detalhes dos erros se houver
        if erros > 0:
            log_message("\nDetalhes dos erros:", "ERROR")
            for r in resultados:
                if 'erro' in r['status']:
                    log_message(f"- {r['codigo']}: {r['detalhes']}", "ERROR")

        messagebox.showinfo("Processamento Concluído",
            f"✅ Processamento finalizado!\n\n"
            f"Total: {total}\n"
            f"Sucesso: {sucesso}\n"
            f"Não encontrados: {sem_andamento}\n"
            f"Erros: {erros}")

    def run(self, params: dict):
        username = params.get("username")
        password = params.get("password")
        excel_file = params.get("excel_file")
        cancel_flag = params.get("cancel_flag")
        headless_mode = params.get("headless_mode")
        try:
            dados_exames = self.get_dados_exames(excel_file)
            if not dados_exames:
                messagebox.showerror("Erro", "Nenhum dado de exame encontrado na planilha.")
                return
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao ler o Excel: {e}")
            return
        url = os.getenv("SYSTEM_URL", "https://dap.pathoweb.com.br/login/auth")
        driver = None
        resultados = []
        try:
            driver = BrowserFactory.create_chrome(headless=headless_mode)
            wait = WebDriverWait(driver, DEFAULT_TIMEOUT)
            log_message("Iniciando automação de macroscopia...", "INFO")
            driver.get(url)
            # Login
            wait.until(EC.presence_of_element_located((By.ID, "username"))).send_keys(username)
            wait.until(EC.presence_of_element_located((By.ID, "password"))).send_keys(password)
            driver.find_element(By.CSS_SELECTOR, "button[type='submit']").click()

            log_message("Verificando se precisa navegar para módulo de exames...", "INFO")
            current_url = driver.current_url
            if current_url == "https://dap.pathoweb.com.br/" or "trocarModulo" in current_url:
                log_message("Detectada tela de seleção de módulos - navegando para módulo de exames...", "INFO")
                try:
                    modulo_link = wait.until(
                        EC.element_to_be_clickable((By.CSS_SELECTOR, "a[href='/site/trocarModulo?modulo=1']")))
                    modulo_link.click()
                    time.sleep(MEDIUM_DELAY)
                    log_message("✅ Navegação para módulo de exames realizada", "SUCCESS")
                except Exception as e:
                    log_message(f"⚠️ Erro ao navegar para módulo: {e}", "WARNING")
                    # Tentar navegar diretamente pela URL como fallback
                    driver.get("https://dap.pathoweb.com.br/moduloExame/index")
                    time.sleep(MEDIUM_DELAY)
                    log_message("🔄 Navegação direta para módulo realizada", "INFO")

            elif "moduloExame" in current_url:
                log_message("✅ Já está no módulo de exames - pulando navegação", "SUCCESS")
            else:
                log_message(f"⚠️ URL inesperada detectada: {current_url}", "WARNING")
                # Tentar navegar diretamente como fallback
                driver.get("https://dap.pathoweb.com.br/moduloExame/index")
                time.sleep(MEDIUM_DELAY)
                log_message("🔄 Navegação direta para módulo realizada (fallback)", "INFO")

            try:
                modal_close_button = driver.find_element(By.CSS_SELECTOR, "#mensagemParaClienteModal .modal-footer button")
                if modal_close_button.is_displayed():
                    self.clicar_elemento(driver, modal_close_button)
            except Exception:
                pass
            codigos_processados = []
            for i, exame_data in enumerate(dados_exames, 1):
                if cancel_flag and cancel_flag.is_set():
                    log_message("Execução cancelada pelo usuário.", "WARNING")
                    break
                codigo = exame_data['codigo']
                mascara = exame_data['mascara']
                citotecnica_nome = exame_data.get('citotecnica')
                codigos_processados.append(codigo)
                log_message(f"\n➡️ Processando exame {i}/{len(dados_exames)}: {codigo} (máscara: {mascara}) - Citotécnica: {citotecnica_nome}", "INFO")
                resultado = self.processar_exame(driver, wait, codigo, mascara, citotecnica_nome)
                resultados.append({
                    'codigo': codigo,
                    'mascara': mascara,
                    'citotecnica': citotecnica_nome,
                    'status': resultado['status'],
                    'detalhes': resultado.get('detalhes', '')
                })

            self.mostrar_resumo_final(resultados)
        except Exception as e:
            log_message(f"❌ Erro durante a automação: {e}", "ERROR")
        finally:
            if driver:
                try:
                    driver.quit()
                except Exception:
                    pass

def run(params: dict):
    module = MacroscopiaModule()
    module.run(params)