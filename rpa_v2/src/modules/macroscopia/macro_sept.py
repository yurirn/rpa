import os
import time
from tkinter import messagebox
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from dotenv import load_dotenv
from openpyxl import load_workbook

from src.core.browser_factory import BrowserFactory
from src.core.logger import log_message
from src.modules.base import BaseModule

load_dotenv()


class MacroSeptModule(BaseModule):
    def __init__(self):
        super().__init__(nome="Macro Sept")

    def get_dados_exames(self, file_path: str) -> list:
        try:
            workbook = load_workbook(file_path)
            sheet = workbook.active
            dados = []
            ultima_mascara = None
            data_fixacao = None
            macroscopista_valor = None
            frag_sept_valor = None
            frag_turb_valor = None
            frag_sinu_valor = None

            # Ler cabe√ßalho (linha 1) e criar mapeamento de colunas
            colunas = {}
            for col_idx in range(1, sheet.max_column + 1):
                cell_value = sheet.cell(row=1, column=col_idx).value
                if cell_value:
                    nome_coluna = str(cell_value).strip().lower()
                    colunas[nome_coluna] = col_idx

            log_message(f"üìã Colunas detectadas: {list(colunas.keys())}", "INFO")

            def encontrar_coluna(nomes_possiveis):
                """Encontra a coluna baseado em uma lista de nomes poss√≠veis"""
                for nome in nomes_possiveis:
                    for coluna_nome, col_idx in colunas.items():
                        if nome.lower() in coluna_nome:
                            return col_idx
                return None

            # Encontrar √≠ndices das colunas
            col_data = encontrar_coluna(['data', 'data fixacao', 'data fixa√ß√£o', 'datafixacao'])
            col_num_exame = encontrar_coluna(['num_exame', 'numero', 'n√∫mero', 'codigo', 'c√≥digo', 'cod'])
            col_mascara = encontrar_coluna(['mascara', 'm√°scara', 'mask'])
            col_macroscopista = encontrar_coluna(['macroscopista', 'responsavel', 'respons√°vel', 'resp'])
            col_frag_sept = encontrar_coluna(['frag sept', 'fragsept', 'frag_sept'])
            col_med_sep = encontrar_coluna(['med sep', 'medsep', 'med_sep'])
            col_frag_turb = encontrar_coluna(['frag turb', 'fragturb', 'frag_turb'])
            col_med_turb = encontrar_coluna(['med turb', 'medturb', 'med_turb'])
            col_frag_sinu = encontrar_coluna(['frag sinu', 'fragsinu', 'frag_sinu'])
            col_med_sinu = encontrar_coluna(['med sinu', 'medsinu', 'med_sinu'])
            col_legenda = encontrar_coluna(['legenda', 'leg'])

            # Validar colunas obrigat√≥rias
            if not col_num_exame:
                raise Exception("Coluna de n√∫mero do exame n√£o encontrada!")

            log_message(f"‚úÖ Mapeamento: Num_Exame=col{col_num_exame}, M√°scara=col{col_mascara}, Data=col{col_data}",
                        "INFO")

            # L√™ da linha 2 em diante (linha 1 √© cabe√ßalho)
            for row in range(2, sheet.max_row + 1):
                data = sheet.cell(row=row, column=col_data).value if col_data else None
                num_exame = sheet.cell(row=row, column=col_num_exame).value if col_num_exame else None
                mascara = sheet.cell(row=row, column=col_mascara).value if col_mascara else None
                macroscopista = sheet.cell(row=row, column=col_macroscopista).value if col_macroscopista else None
                frag_sept = sheet.cell(row=row, column=col_frag_sept).value if col_frag_sept else None
                med_sep = sheet.cell(row=row, column=col_med_sep).value if col_med_sep else None
                frag_turb = sheet.cell(row=row, column=col_frag_turb).value if col_frag_turb else None
                med_turb = sheet.cell(row=row, column=col_med_turb).value if col_med_turb else None
                frag_sinu = sheet.cell(row=row, column=col_frag_sinu).value if col_frag_sinu else None
                med_sinu = sheet.cell(row=row, column=col_med_sinu).value if col_med_sinu else None
                legenda = sheet.cell(row=row, column=col_legenda).value if col_legenda else None

                if row == 2 and data:
                    data_fixacao = str(data).strip()

                if macroscopista is not None and str(macroscopista).strip():
                    macroscopista_valor = str(macroscopista).strip().upper()

                if frag_sept is not None and str(frag_sept).strip():
                    frag_sept_valor = str(frag_sept).strip()

                if frag_turb is not None and str(frag_turb).strip():
                    frag_turb_valor = str(frag_turb).strip()

                if frag_sinu is not None and str(frag_sinu).strip():
                    frag_sinu_valor = str(frag_sinu).strip()

                if num_exame is not None:
                    num_exame = str(num_exame).strip()

                    # Se n√£o tem m√°scara, usa a √∫ltima v√°lida
                    if mascara is not None and str(mascara).strip():
                        mascara = str(mascara).strip()
                        ultima_mascara = mascara
                    else:
                        mascara = ultima_mascara

                    frag_sept_extenso = self.converter_numero_para_extenso(frag_sept_valor) if frag_sept_valor else ""
                    frag_turb_extenso = self.converter_numero_para_extenso(frag_turb_valor) if frag_turb_valor else ""
                    frag_sinu_extenso = self.converter_numero_para_extenso(frag_sinu_valor) if frag_sinu_valor else ""

                    legenda_original = str(legenda).strip().lower() if legenda is not None else ""

                    # Regra: se campo_d for 'mult', usar 6
                    if legenda is not None and str(legenda).strip().lower() == 'mult':
                        legenda_valor = '6'
                    else:
                        legenda_valor = str(legenda).strip() if legenda is not None else ""

                    dados.append({
                        'num_exame': num_exame,
                        'mascara': str(mascara).strip() if mascara else "",
                        'macroscopista': macroscopista_valor,
                        'frag_sept': frag_sept_extenso,
                        'med_sep': str(med_sep).strip() if med_sep else "",
                        'frag_turb': frag_turb_extenso,
                        'med_turb': str(med_turb).strip() if med_turb else "",
                        'frag_sinu': frag_sinu_extenso,
                        'med_sinu': str(med_sinu).strip() if med_sinu else "",
                        'legenda': legenda_valor,
                        'legenda_original': legenda_original,
                        'data_fixacao': data_fixacao
                    })

            workbook.close()
            return dados
        except Exception as e:
            raise Exception(f"Erro ao ler planilha: {e}")

    def separar_valores_concatenados(self, valor_concatenado):
        """Separa valores concatenados por 'x' em uma lista
        Exemplo: '0,2x0,3x0,2' -> ['0,2', '0,3', '0,2']
        """
        if not valor_concatenado or valor_concatenado.strip() == "":
            return []

        # Separar por 'x' (case insensitive)
        valores = valor_concatenado.replace('X', 'x').split('x')
        # Limpar espa√ßos e retornar
        return [v.strip() for v in valores if v.strip()]

    def converter_numero_para_extenso(self, valor):
        """Converte n√∫meros para texto por extenso
        Exemplo: '2' -> 'dois', '3' -> 'tr√™s', 'mult' -> 'm√∫ltiplos'
        """
        if not valor or str(valor).strip() == "":
            return ""

        valor_str = str(valor).strip().lower()

        # Mapeamento de n√∫meros para extenso
        numeros_extenso = {
            '1': 'um',
            '2': 'dois',
            '3': 'tr√™s',
            '4': 'quatro',
            '5': 'cinco'
        }

        # Se for 'mult', retorna 'm√∫ltiplos'
        if valor_str == 'mult':
            return 'm√∫ltiplos'

        # Se for um n√∫mero mapeado, retorna por extenso
        if valor_str in numeros_extenso:
            return numeros_extenso[valor_str]

        # Se n√£o encontrar, retorna o valor original
        return valor_str

    def verificar_sessao_browser(self, driver) -> bool:
        """Verifica se a sess√£o do browser ainda est√° ativa"""
        try:
            driver.current_url
            return True
        except Exception as e:
            if "invalid session id" in str(e).lower():
                log_message("‚ùå Sess√£o do browser perdida", "ERROR")
                return False
            return True

    def aguardar_pagina_estavel(self, driver, wait, timeout=10):
        """Aguarda at√© que a p√°gina esteja est√°vel (sem anima√ß√µes ou carregamentos)"""
        try:
            driver.execute_script("""
                return new Promise((resolve) => {
                    if (window.jQuery && window.jQuery.active === 0) {
                        resolve();
                        return;
                    }

                    var checkInterval = setInterval(() => {
                        if (window.jQuery && window.jQuery.active === 0) {
                            clearInterval(checkInterval);
                            resolve();
                        }
                    }, 100);

                    setTimeout(() => {
                        clearInterval(checkInterval);
                        resolve();
                    }, arguments[0]);
                });
            """, timeout * 1000)

            time.sleep(0.5)
            log_message("‚úÖ P√°gina est√°vel", "INFO")

        except Exception as e:
            log_message(f"‚ö†Ô∏è Erro ao aguardar p√°gina est√°vel: {e}", "WARNING")
            time.sleep(1)

    def aguardar_spinner_desaparecer(self, driver, wait, timeout=30):
        """Aguarda at√© que o spinner de loading desapare√ßa"""
        try:
            log_message("‚è≥ Aguardando spinner desaparecer...", "INFO")
            wait.until(EC.invisibility_of_element_located((By.ID, "spinner")))
            time.sleep(1)

            spinners = driver.find_elements(By.CSS_SELECTOR, ".loadModal, .spinner, [class*='loading']")
            for spinner in spinners:
                if spinner.is_displayed():
                    log_message("‚ö†Ô∏è Outro spinner ainda vis√≠vel, aguardando...", "WARNING")
                    time.sleep(2)
                    break

            log_message("‚úÖ Spinner desapareceu", "SUCCESS")

        except Exception as e:
            log_message(f"‚ö†Ô∏è Erro ao aguardar spinner: {e}", "WARNING")
            try:
                driver.execute_script("""
                    var spinners = document.querySelectorAll('.loadModal, .spinner, [class*="loading"]');
                    spinners.forEach(function(spinner) {
                        if (spinner.style.display !== 'none') {
                            spinner.style.display = 'none';
                        }
                    });
                """)
                log_message("üîß Spinner fechado via JavaScript", "INFO")
                time.sleep(1)
            except:
                pass

    def selecionar_responsavel_macroscopia(self, driver, wait, responsavel_macro):
        """Seleciona o respons√°vel pela macroscopia conforme o nome recebido (nome curto)"""
        responsavel_macro_mapper = {
            'BARBARA': 'Barbara Dutra Lopes',
            'NATHALIA': 'Nathalia Fernanda da Silva Lopes',
            'RENATA': 'Renata Silva Sevidanis',
            'HELEN': 'Helen Oliveira dos Santos',
            'CLARA': 'Clara Helena Janz Garcia de Souza',
            'PALOMA': 'Paloma Brenda Silva De Oliveira',
            'ELLEN': 'Ellen Andressa de Alvarenga',
            'VITORIA': 'Vitoria Aquino Nairne Domingues',
            'ANNAI': 'Annai Luk√£ Vitorino Losnak',
            'ANA': 'Ana Carolina Viecele Campos'
        }
        nome_completo = responsavel_macro_mapper.get(responsavel_macro, responsavel_macro)
        select2_container = wait.until(
            EC.element_to_be_clickable(
                (By.XPATH, "//span[@aria-labelledby='select2-responsavelMacroscopiaId-container']"))
        )
        select2_container.click()
        time.sleep(0.3)

        opcao = wait.until(
            EC.element_to_be_clickable((By.XPATH, f"//li[contains(text(), '{nome_completo}')]"))
        )
        opcao.click()
        log_message(f"‚úÖ {nome_completo} selecionado como respons√°vel", "SUCCESS")
        time.sleep(0.2)

    def definir_data_fixacao(self, driver, wait, data_fixacao=None):
        """Define a data de fixa√ß√£o no campo de data de fixa√ß√£o"""
        try:
            if not data_fixacao:
                data_fixacao = '21082025'

            if len(data_fixacao) == 8 and data_fixacao.isdigit():
                data_formatada = f"{data_fixacao[4:8]}-{data_fixacao[2:4]}-{data_fixacao[0:2]}"
            else:
                data_formatada = '2025-08-21'

            campo_data = wait.until(
                EC.presence_of_element_located((By.XPATH, "//input[@type='date' and @name='dataFixacao']"))
            )
            driver.execute_script("""
                var campo = arguments[0];
                campo.value = arguments[1];
                campo.dispatchEvent(new Event('change', { bubbles: true }));
            """, campo_data, data_formatada)
            log_message(f"üìÖ Data de fixa√ß√£o definida para: {data_formatada}", "SUCCESS")
            time.sleep(0.1)
        except Exception as e:
            log_message(f"‚ö†Ô∏è Erro ao definir data de fixa√ß√£o: {e}", "WARNING")

    def definir_hora_fixacao(self, driver, wait):
        """Define 18:00 no campo de hora de fixa√ß√£o"""
        campo_hora = wait.until(
            EC.presence_of_element_located((By.XPATH, "//input[@type='time' and @name='dataFixacao']"))
        )
        campo_hora.clear()
        campo_hora.send_keys("18:00")
        log_message("üïï Hora de fixa√ß√£o definida para: 18:00", "SUCCESS")
        time.sleep(0.1)

    def fechar_exame(self, driver, wait):
        """Clica no bot√£o de fechar exame"""
        try:
            botao_fechar = wait.until(
                EC.element_to_be_clickable((By.ID, "fecharExameBarraFerramenta"))
            )
            botao_fechar.click()
            log_message("üìÅ Exame fechado", "INFO")

            try:
                wait.until(EC.presence_of_element_located((By.ID, "inputSearchCodBarra")))
                log_message("‚úÖ Retornou √† tela principal ap√≥s fechar exame", "INFO")
            except:
                log_message("‚ö†Ô∏è Pode n√£o ter retornado √† tela principal", "WARNING")
                try:
                    current_url = driver.current_url
                    if "modulo=1" not in current_url:
                        modulo_link = driver.find_element(By.CSS_SELECTOR, "a[href='/site/trocarModulo?modulo=1']")
                        modulo_link.click()
                        time.sleep(1.5)
                        log_message("üîÑ Navegou de volta ao m√≥dulo de exames", "INFO")
                except:
                    pass

        except Exception as e:
            log_message(f"Erro ao fechar exame: {e}", "ERROR")

    def digitar_mascara_e_buscar(self, driver, wait, mascara):
        """Digita a m√°scara no campo buscaArvore e pressiona Enter"""
        campo_busca = wait.until(EC.element_to_be_clickable((By.ID, "buscaArvore")))
        campo_busca.send_keys(mascara)
        campo_busca.send_keys(Keys.ENTER)
        log_message(f"‚úçÔ∏è M√°scara '{mascara}' digitada no campo buscaArvore", "SUCCESS")
        time.sleep(0.5)

    def abrir_modal_variaveis_e_preencher(self, driver, wait, mascara, frag_sept, med_sep, frag_turb, med_turb,
                                          frag_sinu, med_sinu, legenda, legenda_original):
        """Abre o modal de vari√°veis e preenche os campos baseado na m√°scara SEPT/ST/TS"""
        try:
            # Clicar no bot√£o "Pesquisar vari√°veis (F7)"
            botao_variaveis = wait.until(
                EC.element_to_be_clickable((By.ID, "cke_70"))
            )
            botao_variaveis.click()
            log_message("üîç Clicou no bot√£o de vari√°veis", "INFO")
            time.sleep(0.8)

            # Verificar se apareceu um alerta
            try:
                alert = driver.switch_to.alert
                alert_text = alert.text
                if "n√£o h√° vari√°veis" in alert_text.lower():
                    log_message(f"‚ö†Ô∏è Alerta detectado: {alert_text}", "WARNING")
                    alert.accept()
                    log_message("‚ö†Ô∏è Pulando preenchimento de vari√°veis - n√£o h√° vari√°veis no texto", "WARNING")
                    return
                else:
                    alert.accept()
            except:
                pass

            # Aguardar o modal aparecer
            wait.until(EC.presence_of_element_located((By.CLASS_NAME, "swal2-popup")))
            log_message("üîç Modal de vari√°veis aberto", "SUCCESS")
            time.sleep(0.3)

            # Preencher os campos usando classe gen√©rica
            campos_input = driver.find_elements(By.CSS_SELECTOR, "input[style*='width: 100px'][style*='color: red']")
            log_message(f"üîç Encontrados {len(campos_input)} campos de input no modal", "INFO")

            # Processar valores concatenados
            valores_med_sep = self.separar_valores_concatenados(med_sep)
            valores_med_turb = self.separar_valores_concatenados(med_turb)
            valores_med_sinu = self.separar_valores_concatenados(med_sinu)

            # Determinar valores baseado na m√°scara
            mascara_upper = mascara.upper() if mascara else ""
            valores = []

            if legenda_original == 'mult':
                legenda = 'M'

            if mascara_upper == 'SEPT':
                valores = [frag_sept] + valores_med_sep + [frag_turb] + valores_med_turb + [frag_sinu] + valores_med_sinu + [legenda]
                log_message(
                    f"üìã SEPT detectado - Frag_sept={frag_sept}, Med_sep={valores_med_sep}, frag_turb={frag_turb}, Med_turb={valores_med_turb}, Frag_sinu={frag_sinu}, Med_sinu={valores_med_sinu}, Legenda={legenda}",
                    "INFO")

            elif mascara_upper == 'ST':
                valores = [frag_sept] + valores_med_sep + [frag_turb] + valores_med_turb + [legenda]
                log_message(
                    f"üìã ST detectado - Frag_sept={frag_sept}, Med_sep={valores_med_sep}, frag_turb={frag_turb}, Med_turb={valores_med_turb}, Legenda={legenda}",
                    "INFO")

            elif mascara_upper == 'TS':
                valores = [frag_turb] + valores_med_turb + [frag_sinu] + valores_med_sinu + [legenda]
                log_message(
                    f"üìã TS detectado - frag_turb={frag_turb}, Med_turb={valores_med_turb}, Frag_sinu={frag_sinu}, Med_sinu={valores_med_sinu}, Legenda={legenda}",
                    "INFO")

            else:
                log_message(f"‚ö†Ô∏è M√°scara '{mascara}' n√£o reconhecida para preenchimento de vari√°veis", "WARNING")
                return

            # Filtrar valores vazios
            valores = [v for v in valores if v and v.strip()]

            log_message(f"üìã Preenchendo {len(valores)} vari√°veis para m√°scara '{mascara}'", "INFO")

            for i, campo in enumerate(campos_input[:len(valores)]):
                if i < len(valores) and valores[i]:
                    try:
                        driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});",
                                              campo)
                        time.sleep(0.1)

                        driver.execute_script("""
                            arguments[0].value = '';
                            arguments[0].value = arguments[1];
                            arguments[0].dispatchEvent(new Event('input', { bubbles: true }));
                            arguments[0].dispatchEvent(new Event('change', { bubbles: true }));
                        """, campo, valores[i])

                        log_message(f"‚úÖ Campo {i + 1} preenchido com: {valores[i]}", "SUCCESS")
                        time.sleep(0.1)
                    except Exception as e:
                        log_message(f"‚ö†Ô∏è Erro ao preencher campo {i + 1}: {e}", "WARNING")

            time.sleep(0.2)

            # Clicar no bot√£o "Inserir"
            botao_inserir = wait.until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, ".swal2-confirm"))
            )
            botao_inserir.click()
            log_message("‚úÖ Campos inseridos no modal", "SUCCESS")

            # Aguardar o modal fechar
            try:
                wait.until(EC.invisibility_of_element_located((By.CLASS_NAME, "swal2-popup")))
                log_message("‚úÖ Modal fechado completamente", "SUCCESS")
            except:
                time.sleep(1)
                log_message("‚è≥ Aguardou fechamento do modal", "INFO")

        except Exception as e:
            log_message(f"‚ö†Ô∏è Erro ao preencher modal de vari√°veis: {e}", "WARNING")
            log_message("‚ö†Ô∏è Continuando sem preencher as vari√°veis", "WARNING")

    def salvar_macroscopia(self, driver, wait):
        """Clica no bot√£o Salvar da macroscopia"""
        try:
            modal = driver.find_element(By.CLASS_NAME, "swal2-popup")
            if modal.is_displayed():
                botao_fechar = driver.find_element(By.CSS_SELECTOR, ".swal2-close")
                botao_fechar.click()
                time.sleep(0.5)
        except:
            pass

        botao_salvar = wait.until(
            EC.element_to_be_clickable((By.ID, "salvarMacro"))
        )
        botao_salvar.click()
        log_message("üíæ Macroscopia salva", "SUCCESS")
        time.sleep(0.3)

    def definir_grupo(self, driver, wait):
        """Define o grupo como 'Seios da Face' usando JavaScript com retry"""
        try:
            grupo_selecionado = 'Seios da Face'
            max_tentativas = 5
            tentativa = 0

            # Verificar se o input existe e qual o valor atual
            try:
                input_grupo = driver.find_element(By.ID, "idRegiao")
                valor_atual = input_grupo.get_attribute("value")

                if valor_atual == grupo_selecionado:
                    log_message(f"‚úÖ Grupo j√° est√° definido como '{grupo_selecionado}'", "SUCCESS")
                    return
                elif valor_atual and valor_atual.strip():
                    log_message(f"‚ö†Ô∏è Valor atual do campo grupo: '{valor_atual}' - ser√° substitu√≠do", "WARNING")
            except:
                log_message("‚ö†Ô∏è Campo grupo n√£o encontrado inicialmente", "WARNING")

            while tentativa < max_tentativas:
                try:
                    tentativa += 1
                    log_message(f"üîÑ Tentativa {tentativa} de {max_tentativas} para clicar no campo de grupo", "INFO")

                    # Procurar a √¢ncora "Vazio" espec√≠fica do campo de grupo
                    script = """
                    var inputRegiao = document.getElementById('idRegiao');
                    if (inputRegiao) {
                        var parentTd = inputRegiao.closest('td');
                        if (parentTd) {
                            var ancora = parentTd.querySelector('a.table-editable-ancora');
                            if (ancora && ancora.offsetParent !== null) {
                                return ancora;
                            }
                        }
                    }
                    return null;
                    """
                    campo_grupo_ancora = driver.execute_script(script)

                    if not campo_grupo_ancora:
                        log_message(f"‚ö†Ô∏è √Çncora de grupo n√£o encontrada na tentativa {tentativa}", "WARNING")
                        time.sleep(0.5)
                        continue

                    # Scroll at√© o elemento e aguardar
                    driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});",
                                          campo_grupo_ancora)
                    time.sleep(0.3)

                    # Clicar usando diferentes m√©todos
                    try:
                        # M√©todo 1: Click JavaScript direto
                        driver.execute_script("arguments[0].click();", campo_grupo_ancora)
                        log_message(f"üñ±Ô∏è Clicou na √¢ncora via JavaScript (tentativa {tentativa})", "INFO")
                    except:
                        # M√©todo 2: Click Selenium tradicional
                        campo_grupo_ancora.click()
                        log_message(f"üñ±Ô∏è Clicou na √¢ncora via Selenium (tentativa {tentativa})", "INFO")

                    time.sleep(0.5)

                    # Validar se o input ficou vis√≠vel ap√≥s o clique
                    input_grupo = driver.find_element(By.ID, "idRegiao")
                    is_visible = driver.execute_script("""
                        var input = arguments[0];
                        var style = window.getComputedStyle(input);
                        return style.display !== 'none' && style.visibility !== 'hidden' && input.offsetParent !== null;
                    """, input_grupo)

                    if not is_visible:
                        log_message(f"‚ö†Ô∏è Input ainda n√£o est√° vis√≠vel ap√≥s clique (tentativa {tentativa})", "WARNING")
                        time.sleep(0.3)
                        continue

                    log_message(f"‚úÖ Input de grupo est√° vis√≠vel e pronto para preenchimento", "SUCCESS")

                    # Limpar e preencher o campo
                    driver.execute_script("""
                        var input = arguments[0];
                        input.value = '';
                        input.focus();
                        input.value = arguments[1];
                        input.dispatchEvent(new Event('input', { bubbles: true }));
                        input.dispatchEvent(new Event('change', { bubbles: true }));
                        input.dispatchEvent(new KeyboardEvent('keyup', { bubbles: true }));
                    """, input_grupo, grupo_selecionado)

                    log_message(f"‚úÖ Grupo '{grupo_selecionado}' preenchido no campo", "SUCCESS")
                    time.sleep(0.5)

                    # Tentar selecionar da lista de autocomplete
                    try:
                        # Aguardar o dropdown aparecer
                        dropdown = wait.until(
                            EC.presence_of_element_located(
                                (By.CSS_SELECTOR, "ul.typeahead.dropdown-menu[style*='display: block']"))
                        )

                        # Procurar a op√ß√£o no dropdown
                        opcao_autocomplete = wait.until(
                            EC.element_to_be_clickable((By.XPATH,
                                                        f"//ul[@class='typeahead dropdown-menu']//a[contains(text(), '{grupo_selecionado}')]"))
                        )
                        opcao_autocomplete.click()
                        log_message(f"‚úÖ Op√ß√£o '{grupo_selecionado}' selecionada do autocomplete", "SUCCESS")
                        time.sleep(0.3)

                        # Validar se o valor foi realmente preenchido
                        valor_final = input_grupo.get_attribute("value")
                        if valor_final == grupo_selecionado:
                            log_message(f"‚úÖ Valida√ß√£o final: Grupo definido como '{grupo_selecionado}'", "SUCCESS")
                            return
                        else:
                            log_message(f"‚ö†Ô∏è Valor final n√£o corresponde: '{valor_final}' != '{grupo_selecionado}'",
                                        "WARNING")
                            continue

                    except Exception as autocomplete_error:
                        log_message(f"‚ö†Ô∏è Autocomplete n√£o apareceu ou n√£o foi clic√°vel: {autocomplete_error}",
                                    "WARNING")
                        # Tentar confirmar com Enter
                        try:
                            input_grupo.send_keys(Keys.ENTER)
                            time.sleep(0.3)
                            log_message("‚úÖ Confirmado com Enter", "SUCCESS")

                            # Validar se o valor foi preenchido
                            valor_final = input_grupo.get_attribute("value")
                            if valor_final == grupo_selecionado:
                                return
                        except:
                            # Clicar fora para fechar o dropdown
                            driver.execute_script("document.body.click();")
                            time.sleep(0.3)
                        continue

                except Exception as e:
                    log_message(f"‚ö†Ô∏è Erro na tentativa {tentativa}: {e}", "WARNING")
                    time.sleep(0.5)
                    continue

            # Se chegou aqui, esgotou todas as tentativas
            raise Exception(f"N√£o foi poss√≠vel definir o grupo ap√≥s {max_tentativas} tentativas")

        except Exception as e:
            log_message(f"‚ùå Erro ao definir grupo: {e}", "ERROR")
            raise

    def definir_representacao_secao(self, driver, wait):
        """Define a representa√ß√£o como 'Se√ß√£o' usando JavaScript"""
        try:
            # Verificar se o select existe e qual o valor atual
            try:
                select_representacao = driver.find_element(By.ID, "representacao")
                valor_atual = select_representacao.get_attribute("value")

                if valor_atual == "S":
                    log_message("‚úÖ Representa√ß√£o j√° est√° definida como 'Se√ß√£o'", "SUCCESS")
                    return
                elif valor_atual != "S":
                    log_message(f"‚ö†Ô∏è Representa√ß√£o atual √© '{valor_atual}', mas precisa ser 'S' (Se√ß√£o)", "WARNING")
            except:
                log_message("‚ö†Ô∏è Campo representacao n√£o encontrado", "WARNING")
                return

            # Procurar especificamente pelo campo de representa√ß√£o
            script = """
            // Procurar especificamente pelo campo de representa√ß√£o que tem o select com id="representacao"
            var selectRepresentacao = document.getElementById('representacao');
            if (selectRepresentacao) {
                // Encontrar a √¢ncora que est√° no mesmo td que o select representacao
                var parentTd = selectRepresentacao.closest('td');
                if (parentTd) {
                    var ancora = parentTd.querySelector('a[class*="table-editable-ancora"]');
                    if (ancora && ancora.offsetParent !== null) {
                        return ancora;
                    }
                }
            }

            // Fallback: procurar por √¢ncoras que estejam pr√≥ximas a selects de representa√ß√£o
            var selectsRepresentacao = document.querySelectorAll('select[id*="representacao"], select[name*="representacao"]');
            for (var i = 0; i < selectsRepresentacao.length; i++) {
                var select = selectsRepresentacao[i];
                var parentTd = select.closest('td');
                if (parentTd) {
                    var ancora = parentTd.querySelector('a[class*="table-editable-ancora"]');
                    if (ancora && ancora.offsetParent !== null) {
                        return ancora;
                    }
                }
            }

            // √öltimo fallback: procurar por texto "representa√ß√£o" ou "-- representa√ß√£o --" que n√£o seja de procedimento
            var fragmentosContainer = document.getElementById('fragmentosContainer');
            if (fragmentosContainer) {
                var elementos = fragmentosContainer.querySelectorAll('a[class*="table-editable-ancora"]');
                for (var i = 0; i < elementos.length; i++) {
                    var elemento = elementos[i];
                    if ((elemento.textContent.toLowerCase().includes('representa√ß√£o') || elemento.textContent.includes('-- representa√ß√£o --')) && elemento.offsetParent !== null) {
                        var parentTd = elemento.closest('td');
                        if (parentTd && !parentTd.querySelector('input[id*="procedimento"]')) {
                            return elemento;
                        }
                    }
                }
            }
            return null;
            """
            campo_representacao = driver.execute_script(script)

            if not campo_representacao:
                log_message("‚ö†Ô∏è Campo de representa√ß√£o n√£o encontrado", "WARNING")
                return

            # Verificar o texto da √¢ncora para log
            if "Se√ß√£o" in campo_representacao.text:
                log_message("‚úÖ Representa√ß√£o j√° mostra 'Se√ß√£o', mas vamos garantir", "INFO")
            elif "-- representa√ß√£o --" in campo_representacao.text:
                log_message("üìù Campo de representa√ß√£o encontrado, precisa ser preenchido", "INFO")
            else:
                log_message(f"‚ö†Ô∏è Texto inesperado no campo de representa√ß√£o: '{campo_representacao.text}'", "WARNING")

            # Clicar via JavaScript
            driver.execute_script("arguments[0].click();", campo_representacao)
            log_message("üîç Clicou no campo de representa√ß√£o via JS", "INFO")
            time.sleep(0.5)

            # Aguardar o select aparecer e selecionar via JavaScript
            select_representacao = wait.until(
                EC.presence_of_element_located((By.ID, "representacao"))
            )

            # Selecionar "Se√ß√£o" (valor "S") via JavaScript
            driver.execute_script("""
                arguments[0].value = 'S';
                arguments[0].dispatchEvent(new Event('change', { bubbles: true }));
            """, select_representacao)

            log_message("‚úÖ Representa√ß√£o definida como 'Se√ß√£o' via JS", "SUCCESS")
            time.sleep(0.5)

            # Clicar fora para confirmar a sele√ß√£o
            driver.execute_script("document.body.click();")
            time.sleep(0.3)

        except Exception as e:
            log_message(f"‚ö†Ô∏è Erro ao definir representa√ß√£o: {e}", "WARNING")

    def definir_regiao(self, driver, wait, mascara=None):
        """Define a regi√£o de acordo com a m√°scara usando JavaScript"""
        try:
            if not mascara:
                log_message("‚ö†Ô∏è Nenhuma m√°scara fornecida para definir regi√£o", "WARNING")
                return

            # Regras de m√°scara para regi√£o
            mascara_regiao = {
                'SEPT': 'STS: Septo/Turbina/Sinus',
                'ST': 'ST: Septo/Turbina',
                'TS': 'TS: Turbina/Sinus'
            }

            mascara_upper = mascara.upper()
            regiao_valor = mascara_regiao.get(mascara_upper)

            if not regiao_valor:
                log_message(f"‚ö†Ô∏è M√°scara '{mascara}' n√£o possui regi√£o configurada", "WARNING")
                return

            # Verificar se j√° existe um campo de regi√£o preenchido com o valor correto
            try:
                inputs_regiao = driver.find_elements(By.XPATH, "//input[contains(@name, 'regiao_')]")
                for input_reg in inputs_regiao:
                    valor_atual = input_reg.get_attribute("value")
                    if valor_atual == regiao_valor:
                        log_message(f"‚úÖ Regi√£o j√° est√° definida como '{regiao_valor}' - pulando", "SUCCESS")
                        return
                    elif valor_atual and valor_atual != regiao_valor:
                        log_message(f"‚ö†Ô∏è Regi√£o atual √© '{valor_atual}', precisa mudar para '{regiao_valor}'",
                                    "WARNING")
                        break
            except:
                pass

            # Procurar e clicar no campo de regi√£o para edit√°-lo
            script = """
            // Procurar especificamente por campos de regi√£o na tabela de fragmentos
            var tbody = document.getElementById('tdRegiao');
            if (tbody) {
                var inputs = tbody.querySelectorAll('input[name*="regiao_"]');
                for (var i = 0; i < inputs.length; i++) {
                    var input = inputs[i];
                    var parentTd = input.closest('td');
                    if (parentTd) {
                        var ancora = parentTd.querySelector('a[class*="table-editable-ancora"]');
                        if (ancora && ancora.offsetParent !== null) {
                            return {element: ancora, input: input};
                        }
                    }
                }
            }
            return null;
            """
            resultado_regiao = driver.execute_script(script)

            if resultado_regiao:
                campo_regiao = resultado_regiao['element']
                input_regiao = resultado_regiao['input']

                # Clicar na √¢ncora para abrir o campo de edi√ß√£o
                driver.execute_script("arguments[0].click();", campo_regiao)
                log_message("üîç Clicou no campo de regi√£o para editar", "INFO")
                time.sleep(0.5)

                # Aguardar o input ficar vis√≠vel e preencher
                try:
                    # Aguardar o input aparecer
                    wait.until(lambda d: input_regiao.is_displayed() or input_regiao.get_attribute("style") != "display: none;")

                    # Limpar e preencher o campo
                    driver.execute_script("""
                        arguments[0].value = '';
                        arguments[0].value = arguments[1];
                        arguments[0].dispatchEvent(new Event('input', { bubbles: true }));
                        arguments[0].dispatchEvent(new Event('change', { bubbles: true }));
                    """, input_regiao, regiao_valor)

                    log_message(f"‚úçÔ∏è Definiu regi√£o como '{regiao_valor}' via JS", "SUCCESS")
                    time.sleep(0.5)

                    # Clicar fora para confirmar a edi√ß√£o
                    driver.execute_script("document.body.click();")
                    time.sleep(0.5)

                    # Verificar se o valor foi realmente definido
                    valor_definido = input_regiao.get_attribute("value")
                    if valor_definido == regiao_valor:
                        log_message(f"‚úÖ Valor de regi√£o confirmado: '{valor_definido}'", "SUCCESS")
                    else:
                        log_message(
                            f"‚ö†Ô∏è Valor n√£o foi definido corretamente. Esperado: '{regiao_valor}', Atual: '{valor_definido}'",
                            "WARNING")

                except Exception as input_error:
                    log_message(f"‚ö†Ô∏è Erro ao preencher input de regi√£o: {input_error}", "WARNING")
            else:
                log_message("‚ö†Ô∏è Campo de regi√£o n√£o encontrado ou n√£o vis√≠vel", "WARNING")

        except Exception as e:
            log_message(f"‚ö†Ô∏è Erro ao definir regi√£o: {e}", "WARNING")

    def definir_quantidade_fragmentos(self, driver, wait, legenda):
        """Define a quantidade de fragmentos baseado no valor do campo legenda"""
        try:
            if not legenda or not legenda.strip():
                log_message("‚ö†Ô∏è Campo legenda est√° vazio, n√£o definindo quantidade de fragmentos", "WARNING")
                return

            quantidade_valor = legenda.strip()
            log_message(f"üìù Usando quantidade da legenda: {quantidade_valor}", "INFO")

            # Procurar pelos campos de quantidade na tabela de fragmentos
            script = """
            // Procurar especificamente por campos de quantidade na tabela de fragmentos
            var tbody = document.getElementById('tdRegiao');
            if (tbody) {
                var inputs = tbody.querySelectorAll('input[name*="quantidade_"]');
                for (var i = 0; i < inputs.length; i++) {
                    var input = inputs[i];
                    var parentTd = input.closest('td');
                    if (parentTd) {
                        var ancora = parentTd.querySelector('a[class*="table-editable-ancora"]');
                        if (ancora && ancora.offsetParent !== null) {
                            return {element: ancora, input: input};
                        }
                    }
                }
            }
            return null;
            """
            resultado_quantidade = driver.execute_script(script)

            if resultado_quantidade:
                campo_quantidade = resultado_quantidade['element']
                input_quantidade = resultado_quantidade['input']

                # Verificar se j√° tem o valor correto
                valor_atual = input_quantidade.get_attribute("value")
                if valor_atual == quantidade_valor:
                    log_message(f"‚úÖ Quantidade j√° est√° definida como '{quantidade_valor}' - pulando", "SUCCESS")
                    return

                # Clicar na √¢ncora para abrir o campo
                driver.execute_script("arguments[0].click();", campo_quantidade)
                log_message("üîç Clicou no campo de quantidade para editar", "INFO")
                time.sleep(0.5)

                # Aguardar o input ficar vis√≠vel e preencher
                try:
                    # Aguardar o input aparecer
                    wait.until(lambda d: input_quantidade.is_displayed() or input_quantidade.get_attribute("style") != "display: none;")

                    # Limpar e preencher o campo
                    driver.execute_script("""
                        arguments[0].value = '';
                        arguments[0].value = arguments[1];
                        arguments[0].dispatchEvent(new Event('input', { bubbles: true }));
                        arguments[0].dispatchEvent(new Event('change', { bubbles: true }));
                    """, input_quantidade, quantidade_valor)

                    log_message(f"‚úçÔ∏è Definiu quantidade como '{quantidade_valor}' via JS", "SUCCESS")
                    time.sleep(0.5)

                    # Clicar fora para confirmar a edi√ß√£o
                    driver.execute_script("document.body.click();")
                    time.sleep(0.3)

                    # Verificar se o valor foi definido
                    valor_definido = input_quantidade.get_attribute("value")
                    if valor_definido == quantidade_valor:
                        log_message(f"‚úÖ Valor de quantidade confirmado: '{valor_definido}'", "SUCCESS")
                    else:
                        log_message(
                            f"‚ö†Ô∏è Valor n√£o foi definido corretamente. Esperado: '{quantidade_valor}', Atual: '{valor_definido}'",
                            "WARNING")

                except Exception as input_error:
                    log_message(f"‚ö†Ô∏è Erro ao preencher input de quantidade: {input_error}", "WARNING")
            else:
                log_message("‚ö†Ô∏è Campo de quantidade n√£o encontrado ou n√£o vis√≠vel", "WARNING")

        except Exception as e:
            log_message(f"‚ö†Ô∏è Erro ao definir quantidade de fragmentos: {e}", "WARNING")

    def definir_quantidade_blocos(self, driver, wait):
        """Define a quantidade de blocos como 1 (valor fixo)"""
        try:
            log_message("üìù Definindo quantidade de blocos como: 1", "INFO")

            # Procurar pelos campos de quantidade de blocos na tabela de fragmentos
            script = """
            // Procurar especificamente por campos de quantidade de blocos na tabela de fragmentos
            var tbody = document.getElementById('tdRegiao');
            if (tbody) {
                var inputs = tbody.querySelectorAll('input[name*="quantidadeBlocos_"]');
                for (var i = 0; i < inputs.length; i++) {
                    var input = inputs[i];
                    var parentTd = input.closest('td');
                    if (parentTd) {
                        var ancora = parentTd.querySelector('a[class*="table-editable-ancora"]');
                        if (ancora && ancora.offsetParent !== null) {
                            return {element: ancora, input: input};
                        }
                    }
                }
            }
            return null;
            """
            resultado_blocos = driver.execute_script(script)

            if resultado_blocos:
                campo_blocos = resultado_blocos['element']
                input_blocos = resultado_blocos['input']

                # Verificar se j√° tem o valor correto
                valor_atual = input_blocos.get_attribute("value")
                if valor_atual == "1":
                    log_message("‚úÖ Quantidade de blocos j√° est√° definida como '1' - pulando", "SUCCESS")
                    return

                # Clicar na √¢ncora para abrir o campo
                driver.execute_script("arguments[0].click();", campo_blocos)
                log_message("üîç Clicou no campo de quantidade de blocos para editar", "INFO")
                time.sleep(0.5)

                # Aguardar o input ficar vis√≠vel e preencher
                try:
                    # Aguardar o input aparecer
                    wait.until(lambda d: input_blocos.is_displayed() or input_blocos.get_attribute(
                        "style") != "display: none;")

                    # Limpar e preencher o campo
                    driver.execute_script("""
                        arguments[0].value = '';
                        arguments[0].value = '1';
                        arguments[0].dispatchEvent(new Event('input', { bubbles: true }));
                        arguments[0].dispatchEvent(new Event('change', { bubbles: true }));
                    """, input_blocos)

                    log_message("‚úçÔ∏è Definiu quantidade de blocos como '1' via JS", "SUCCESS")
                    time.sleep(0.5)

                    # Clicar fora para confirmar a edi√ß√£o
                    driver.execute_script("document.body.click();")
                    time.sleep(0.3)

                    # Verificar se o valor foi definido
                    valor_definido = input_blocos.get_attribute("value")
                    if valor_definido == "1":
                        log_message("‚úÖ Valor de quantidade de blocos confirmado: '1'", "SUCCESS")
                    else:
                        log_message(f"‚ö†Ô∏è Valor n√£o foi definido corretamente. Esperado: '1', Atual: '{valor_definido}'",
                                    "WARNING")

                except Exception as input_error:
                    log_message(f"‚ö†Ô∏è Erro ao preencher input de quantidade de blocos: {input_error}", "WARNING")
            else:
                log_message("‚ö†Ô∏è Campo de quantidade de blocos n√£o encontrado ou n√£o vis√≠vel", "WARNING")

        except Exception as e:
            log_message(f"‚ö†Ô∏è Erro ao definir quantidade de blocos: {e}", "WARNING")

    def salvar_fragmentos(self, driver, wait):
        """Clica no bot√£o Salvar dos fragmentos"""
        try:
            # Aguardar o bot√£o estar presente e clic√°vel
            botao_salvar_fragmentos = wait.until(
                EC.element_to_be_clickable((By.XPATH,
                                            "//a[contains(@class, 'btn-primary') and contains(@data-url, '/macroscopia/saveMacroscopiaFragAjax')]"))
            )

            # Verificar se o bot√£o est√° vis√≠vel
            if not botao_salvar_fragmentos.is_displayed():
                log_message("‚ö†Ô∏è Bot√£o salvar fragmentos n√£o est√° vis√≠vel", "WARNING")
                return

            # Rolar at√© o bot√£o para garantir visibilidade
            driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});",
                                  botao_salvar_fragmentos)
            time.sleep(0.5)

            # Clicar no bot√£o
            botao_salvar_fragmentos.click()
            log_message("üíæ Clicou em Salvar fragmentos", "SUCCESS")

            # Aguardar que o spinner desapare√ßa ap√≥s salvar
            self.aguardar_spinner_desaparecer(driver, wait, timeout=15)

        except Exception as e:
            log_message(f"‚ö†Ô∏è Erro ao salvar fragmentos: {e}", "WARNING")
            # Tentar encontrar o bot√£o por outras formas
            try:
                # Tentar por t√≠tulo
                botao_titulo = wait.until(
                    EC.element_to_be_clickable((By.XPATH, "//a[@title='Salvar' and contains(@class, 'btn-primary')]"))
                )
                driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});",
                                      botao_titulo)
                time.sleep(0.5)
                botao_titulo.click()
                log_message("üíæ Clicou em Salvar fragmentos (por t√≠tulo)", "SUCCESS")
                self.aguardar_spinner_desaparecer(driver, wait, timeout=15)
                return
            except:
                pass

            try:
                # Tentar por texto do bot√£o
                botao_texto = wait.until(
                    EC.element_to_be_clickable(
                        (By.XPATH, "//a[contains(@class, 'btn-primary') and contains(text(), 'Salvar')]"))
                )
                driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});",
                                      botao_texto)
                time.sleep(0.5)
                botao_texto.click()
                log_message("üíæ Clicou em Salvar fragmentos (por texto)", "SUCCESS")
                self.aguardar_spinner_desaparecer(driver, wait, timeout=15)
                return
            except:
                pass

            log_message(f"‚ùå N√£o foi poss√≠vel encontrar o bot√£o Salvar fragmentos: {e}", "ERROR")
            raise

    def preencher_campos_pre_envio(self, driver, wait, mascara, legenda):
        """Preenche os campos necess√°rios antes do envio para pr√≥xima etapa"""
        try:
            log_message("üìù Preenchendo campos pr√©-envio...", "INFO")

            # Definir grupo
            log_message(f"üìù Definindo grupo para m√°scara: {mascara}", "INFO")
            try:
                self.definir_grupo(driver, wait)
                self.aguardar_pagina_estavel(driver, wait, timeout=3)
            except Exception as e:
                log_message(f"‚ö†Ô∏è Erro ao definir grupo: {e}", "WARNING")

            # Definir representa√ß√£o como "Se√ß√£o"
            log_message("üìù Definindo representa√ß√£o como Se√ß√£o", "INFO")
            try:
                self.definir_representacao_secao(driver, wait)
                self.aguardar_pagina_estavel(driver, wait, timeout=3)
            except Exception as e:
                log_message(f"‚ö†Ô∏è Erro ao definir representa√ß√£o: {e}", "WARNING")

            # Definir regi√£o baseada na m√°scara
            log_message(f"üìù Definindo regi√£o para m√°scara: {mascara}", "INFO")
            try:
                self.definir_regiao(driver, wait, mascara)
                self.aguardar_pagina_estavel(driver, wait, timeout=3)
            except Exception as e:
                log_message(f"‚ö†Ô∏è Erro ao definir regi√£o: {e}", "WARNING")

            # Definir quantidade de fragmentos baseado na legenda
            log_message(f"üìù Definindo quantidade de fragmentos baseado na legenda: {legenda}", "INFO")
            try:
                self.definir_quantidade_fragmentos(driver, wait, legenda)
                self.aguardar_pagina_estavel(driver, wait, timeout=3)
            except Exception as e:
                log_message(f"‚ö†Ô∏è Erro ao definir quantidade de fragmentos: {e}", "WARNING")

            # Definir quantidade de blocos como 1
            log_message("üìù Definindo quantidade de blocos como 1", "INFO")
            try:
                self.definir_quantidade_blocos(driver, wait)
                self.aguardar_pagina_estavel(driver, wait, timeout=3)
            except Exception as e:
                log_message(f"‚ö†Ô∏è Erro ao definir quantidade de blocos: {e}", "WARNING")

            log_message("‚úÖ Campos pr√©-envio preenchidos", "SUCCESS")

        except Exception as e:
            log_message(f"‚ö†Ô∏è Erro ao preencher campos pr√©-envio: {e}", "WARNING")

    def enviar_proxima_etapa(self, driver, wait):
        """Clica no bot√£o de enviar para pr√≥xima etapa"""
        try:
            self.aguardar_pagina_estavel(driver, wait)
            self.aguardar_spinner_desaparecer(driver, wait)

            botao_enviar = wait.until(
                EC.element_to_be_clickable((By.ID, "btn-enviar-proxima-etapa"))
            )

            if not botao_enviar.is_displayed() or not botao_enviar.is_enabled():
                log_message("‚ö†Ô∏è Bot√£o n√£o est√° vis√≠vel ou habilitado", "WARNING")
                raise Exception("Bot√£o n√£o est√° interativo")

            try:
                driver.execute_script("arguments[0].click();", botao_enviar)
                log_message("‚û°Ô∏è Clicou em Enviar para pr√≥xima etapa via JS", "INFO")
            except:
                botao_enviar.click()
                log_message("‚û°Ô∏è Clicou em Enviar para pr√≥xima etapa", "INFO")

            time.sleep(2)

            # Verificar modal de assinatura
            try:
                modal_assinatura = driver.find_element(By.ID, "assinatura")
                if modal_assinatura.is_displayed():
                    log_message("üìã Modal de assinatura detectado", "INFO")
                    return {'status': 'aguardando_assinatura'}
            except:
                pass

            # Verificar erros
            try:
                erros = driver.find_elements(By.CSS_SELECTOR, ".alert-danger, .error-message")
                if erros:
                    mensagem_erro = erros[0].text
                    log_message(f"‚ùå Erro detectado: {mensagem_erro}", "ERROR")
                    return {'status': 'erro', 'detalhes': mensagem_erro}
            except:
                pass

            log_message("‚úÖ Envio para pr√≥xima etapa realizado com sucesso", "SUCCESS")
            return {'status': 'sucesso', 'detalhes': 'Enviado para pr√≥xima etapa'}

        except Exception as e:
            log_message(f"Erro ao enviar para pr√≥xima etapa: {e}", "ERROR")

            try:
                driver.execute_script("""
                    var spinners = document.querySelectorAll('.loadModal, .spinner, [class*="loading"]');
                    spinners.forEach(function(spinner) {
                        spinner.style.display = 'none';
                    });
                """)
                log_message("üîß Spinners fechados via JavaScript", "INFO")
            except:
                pass

            raise

    def assinar_com_george(self, driver, wait):
        """Faz o processo de assinatura com Dr. George"""
        try:
            wait.until(EC.presence_of_element_located((By.ID, "assinatura")))
            log_message("üìã Modal de assinatura aberto", "INFO")

            checkbox_george = wait.until(
                EC.element_to_be_clickable((By.XPATH, "//input[@type='checkbox' and @value='2173']"))
            )
            checkbox_george.click()
            log_message("‚úÖ Checkbox do Dr. George marcado", "INFO")
            time.sleep(1)

            campo_senha = wait.until(
                EC.presence_of_element_located((By.NAME, "senha_2173"))
            )
            campo_senha.send_keys("1323")
            log_message("üîê Senha digitada", "INFO")
            time.sleep(1)

            botao_assinar = wait.until(
                EC.element_to_be_clickable((By.ID, "salvarAss"))
            )
            botao_assinar.click()
            log_message("‚úçÔ∏è Clicou em Assinar", "INFO")
            time.sleep(1.5)

        except Exception as e:
            log_message(f"Erro no processo de assinatura: {e}", "ERROR")
            raise

    def processar_exame(self, driver, wait, num_exame, mascara, macroscopista, frag_sept, med_sep, frag_turb, med_turb,
                        frag_sinu, med_sinu, legenda, legenda_original, data_fixacao):
        """Processa um exame individual"""
        try:
            if not self.verificar_sessao_browser(driver):
                raise Exception("Sess√£o do browser perdida - necess√°rio reiniciar")

            try:
                campo_codigo = wait.until(
                    EC.presence_of_element_located((By.XPATH, "//input[@placeholder='Leitor de c√≥digo de barras']")))
                log_message("‚úÖ Campo de c√≥digo encontrado", "INFO")
            except:
                campo_codigo = wait.until(EC.presence_of_element_located((By.ID, "inputSearchCodBarra")))
                log_message("‚úÖ Campo de c√≥digo encontrado pelo ID", "INFO")

            campo_codigo.clear()
            campo_codigo.send_keys(num_exame)
            log_message(f"‚úçÔ∏è C√≥digo '{num_exame}' digitado no campo", "SUCCESS")

            try:
                botao_pesquisar = wait.until(EC.element_to_be_clickable((By.ID, "consultarExameBarraAbrirPorBarCode")))
                botao_pesquisar.click()
                log_message("üîç Clicou no bot√£o de pesquisar exame", "SUCCESS")
            except Exception as e:
                log_message(f"‚ö†Ô∏è N√£o foi poss√≠vel clicar no bot√£o de pesquisar: {e}", "WARNING")
                raise

            return self.aguardar_e_processar_andamento(driver, wait, mascara, macroscopista, frag_sept, med_sep,
                                                       frag_turb, med_turb, frag_sinu, med_sinu, legenda,
                                                       legenda_original, data_fixacao)

        except Exception as e:
            error_message = str(e)
            log_message(f"Erro ao processar exame {num_exame}: {error_message}", "ERROR")

            if "invalid session id" in error_message.lower():
                log_message("‚ùå Erro de sess√£o inv√°lida detectado", "ERROR")
                return {'status': 'erro_sessao', 'detalhes': 'Sess√£o do browser perdida'}

            try:
                screenshot_path = f"erro_exame_{num_exame}_{int(time.time())}.png"
                driver.save_screenshot(screenshot_path)
                log_message(f"Screenshot do erro salvo em: {screenshot_path}", "INFO")
            except:
                pass
            return {'status': 'erro', 'detalhes': error_message}

    def aguardar_e_processar_andamento(self, driver, wait, mascara, macroscopista, frag_sept, med_sep, frag_turb,
                                       med_turb, frag_sinu, med_sinu, legenda, legenda_original, data_fixacao):
        """Aguarda a div de andamento e processa o exame"""
        try:
            wait.until(EC.presence_of_element_located((By.ID, "divAndamentoExame")))
            log_message("üìã Div de andamento do exame encontrada!", "SUCCESS")
            time.sleep(0.5)
        except:
            log_message("‚ö†Ô∏è Div de andamento n√£o apareceu no tempo esperado", "WARNING")
            return {'status': 'sem_andamento', 'detalhes': 'Exame n√£o encontrado ou n√£o carregou'}

        log_message("‚úÖ Exame carregado - iniciando processo de conclus√£o", "SUCCESS")
        return self.processar_conclusao_completa(driver, wait, mascara, macroscopista, frag_sept, med_sep, frag_turb,
                                                 med_turb, frag_sinu, med_sinu, legenda, legenda_original, data_fixacao)

    def processar_conclusao_completa(self, driver, wait, mascara, macroscopista, frag_sept, med_sep, frag_turb,
                                     med_turb, frag_sinu, med_sinu, legenda, legenda_original, data_fixacao):
        try:

            # 1. Selecionar respons√°vel pela macroscopia
            log_message("üìã Selecionando respons√°vel pela macroscopia!", "INFO")
            self.selecionar_responsavel_macroscopia(driver, wait, macroscopista)

            # 2. Definir data de fixa√ß√£o
            self.definir_data_fixacao(driver, wait, data_fixacao)

            # 3. Definir hora 18:00
            self.definir_hora_fixacao(driver, wait)

            # 4. Digitar a m√°scara e buscar
            if mascara:
                self.digitar_mascara_e_buscar(driver, wait, mascara)
            else:
                log_message("‚ö†Ô∏è Nenhuma m√°scara encontrada, pulando busca", "WARNING")

            # 5. Abrir modal de vari√°veis e preencher campos
            try:
                self.abrir_modal_variaveis_e_preencher(driver, wait, mascara, frag_sept, med_sep, frag_turb, med_turb,
                                                       frag_sinu, med_sinu, legenda, legenda_original)
            except Exception as var_error:
                log_message(f"‚ö†Ô∏è Erro no modal de vari√°veis: {var_error}", "WARNING")
                log_message("‚ö†Ô∏è Continuando o processo sem as vari√°veis", "WARNING")

            # 6. Salvar macroscopia
            self.salvar_macroscopia(driver, wait)

            # 7. Preencher campos pr√©-envio (grupo, representa√ß√£o, regi√£o, quantidade de fragmentos e blocos)
            try:
                self.preencher_campos_pre_envio(driver, wait, mascara, legenda)
            except Exception as campos_error:
                log_message(f"‚ö†Ô∏è Erro ao preencher campos pr√©-envio: {campos_error}", "WARNING")

            # 8. Salvar fragmentos
            self.salvar_fragmentos(driver, wait)

            # 9. Enviar para pr√≥xima etapa
            resultado_envio = self.enviar_proxima_etapa(driver, wait)

            if resultado_envio.get('status') == 'aguardando_assinatura':
                log_message("üìã Modal de assinatura aberto - iniciando processo de assinatura", "INFO")
                try:
                    self.assinar_com_george(driver, wait)
                    log_message("‚úÖ Assinatura realizada com sucesso", "SUCCESS")
                    return {'status': 'sucesso', 'detalhes': 'Macroscopia assinada com sucesso'}
                except Exception as assinatura_error:
                    log_message(f"‚ùå Erro na assinatura: {assinatura_error}", "ERROR")
                    return {'status': 'erro_assinatura', 'detalhes': str(assinatura_error)}
            elif resultado_envio.get('status') == 'erro':
                log_message(f"‚ö†Ô∏è Erro no envio para pr√≥xima etapa: {resultado_envio.get('detalhes')}", "WARNING")
                return {'status': 'erro_envio', 'detalhes': resultado_envio.get('detalhes')}
            else:
                log_message("üéâ Processo de macroscopia finalizado com sucesso!", "SUCCESS")
                return {'status': 'sucesso', 'detalhes': 'Macroscopia processada com sucesso'}

        except Exception as e:
            log_message(f"Erro durante processo de macroscopia: {e}", "ERROR")
            return {'status': 'erro_macroscopia', 'detalhes': str(e)}

    def mostrar_resumo_final(self, resultados):
        """Mostra o resumo final do processamento"""
        total = len(resultados)
        sucesso = len([r for r in resultados if r['status'] == 'sucesso'])
        sem_andamento = len([r for r in resultados if r['status'] == 'sem_andamento'])
        erro_sessao = len([r for r in resultados if r['status'] == 'erro_sessao'])
        erros = len([r for r in resultados if 'erro' in r['status'] and r['status'] != 'erro_sessao'])

        log_message("\n" + "=" * 50, "INFO")
        log_message("RESUMO FINAL DO PROCESSAMENTO", "INFO")
        log_message("=" * 50, "INFO")
        log_message(f"Total de exames: {total}", "INFO")
        log_message(f"‚úÖ Processados com sucesso: {sucesso}", "SUCCESS")
        log_message(f"‚ö†Ô∏è Exames n√£o encontrados: {sem_andamento}", "WARNING")
        log_message(f"üîÑ Erros de sess√£o (browser perdido): {erro_sessao}", "WARNING")
        log_message(f"‚ùå Outros erros de processamento: {erros}", "ERROR")

        if erro_sessao + erros > 0:
            log_message("\nDetalhes dos erros:", "ERROR")
            for r in resultados:
                if 'erro' in r['status']:
                    log_message(f"- {r.get('detalhes', 'Erro desconhecido')}", "ERROR")

        messagebox.showinfo("Processamento Conclu√≠do",
                            f"‚úÖ Processamento finalizado!\n\n"
                            f"Total: {total}\n"
                            f"Sucesso: {sucesso}\n"
                            f"N√£o encontrados: {sem_andamento}\n"
                            f"Erros de sess√£o: {erro_sessao}\n"
                            f"Outros erros: {erros}")

    def run(self, params: dict):
        username = params.get("username")
        password = params.get("password")
        excel_file = params.get("excel_file")
        cancel_flag = params.get("cancel_flag")
        headless_mode = params.get("headless_mode")

        try:
            dados_exames = self.get_dados_exames(excel_file)
            if not dados_exames:
                messagebox.showerror("Erro", "Nenhum dado de exame encontrado na planilha.")
                return

            log_message(f"Encontrados {len(dados_exames)} exames para processar", "INFO")

        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao ler o Excel: {e}")
            return

        url = os.getenv("SYSTEM_URL", "https://pathoweb.com.br/login/auth")
        driver = None
        resultados = []

        try:
            driver = BrowserFactory.create_chrome(headless=headless_mode)
            wait = WebDriverWait(driver, 10)

            log_message("Iniciando automa√ß√£o de macroscopia septoplastia...", "INFO")

            # Login
            log_message("Fazendo login...", "INFO")
            driver.get(url)

            wait.until(EC.presence_of_element_located((By.ID, "username")))

            username_field = driver.find_element(By.ID, "username")
            username_field.clear()
            username_field.send_keys(username)

            password_field = driver.find_element(By.ID, "password")
            password_field.clear()
            password_field.send_keys(password)

            submit_button = driver.find_element(By.CSS_SELECTOR, "button[type='submit']")
            submit_button.click()

            log_message("Verificando se precisa navegar para m√≥dulo de exames...", "INFO")
            current_url = driver.current_url

            if current_url == "https://pathoweb.com.br/" or "trocarModulo" in current_url:
                log_message("Detectada tela de sele√ß√£o de m√≥dulos - navegando para m√≥dulo de exames...", "INFO")
                try:
                    modulo_link = wait.until(
                        EC.element_to_be_clickable((By.CSS_SELECTOR, "a[href='/site/trocarModulo?modulo=1']")))
                    modulo_link.click()
                    time.sleep(2)
                    log_message("‚úÖ Navegou para o m√≥dulo de exames", "SUCCESS")
                except Exception as e:
                    log_message(f"‚ö†Ô∏è Erro ao navegar para m√≥dulo: {e}", "WARNING")
                    driver.get("https://pathoweb.com.br/moduloExame/index")
                    time.sleep(2)

            elif "moduloExame" in current_url:
                log_message("‚úÖ J√° est√° no m√≥dulo de exames - pulando navega√ß√£o", "SUCCESS")
            else:
                log_message(f"‚ö†Ô∏è URL inesperada detectada: {current_url}", "WARNING")
                driver.get("https://pathoweb.com.br/moduloExame/index")
                time.sleep(2)
                log_message("üîÑ Navega√ß√£o direta para m√≥dulo realizada (fallback)", "INFO")

            # Fechar modal se aparecer
            try:
                modal_close_button = driver.find_element(By.CSS_SELECTOR,
                                                         "#mensagemParaClienteModal .modal-footer button")
                if modal_close_button.is_displayed():
                    modal_close_button.click()
                    time.sleep(0.5)
            except Exception:
                pass

            log_message("‚úÖ Login realizado com sucesso. Iniciando processamento dos exames.", "SUCCESS")

            # Processar cada exame da planilha
            for i, exame_data in enumerate(dados_exames, 1):
                if cancel_flag and cancel_flag.is_set():
                    log_message("‚ö†Ô∏è Processo cancelado pelo usu√°rio", "WARNING")
                    break

                num_exame = exame_data['num_exame']
                mascara = exame_data['mascara']
                macroscopista = exame_data['macroscopista']
                frag_sept = exame_data['frag_sept']
                med_sep = exame_data['med_sep']
                frag_turb = exame_data['frag_turb']
                med_turb = exame_data['med_turb']
                frag_sinu = exame_data['frag_sinu']
                med_sinu = exame_data['med_sinu']
                legenda = exame_data['legenda']
                legenda_original = exame_data['legenda_original']
                data_fixacao = exame_data['data_fixacao']

                log_message(f"\n‚û°Ô∏è Processando exame {i}/{len(dados_exames)}: {num_exame} (m√°scara: {mascara})", "INFO")

                try:
                    resultado = self.processar_exame(driver, wait, num_exame, mascara, macroscopista,
                                                     frag_sept, med_sep, frag_turb, med_turb,
                                                     frag_sinu, med_sinu, legenda, legenda_original, data_fixacao)
                    resultados.append(resultado)

                    if resultado['status'] == 'erro_sessao':
                        log_message("‚ùå Sess√£o perdida - abortando processamento", "ERROR")
                        break

                except Exception as e:
                    log_message(f"‚ùå Erro ao processar exame {num_exame}: {e}", "ERROR")
                    resultados.append({'status': 'erro', 'detalhes': str(e)})

            # Mostrar resumo final
            self.mostrar_resumo_final(resultados)

        except Exception as e:
            log_message(f"‚ùå Erro durante a automa√ß√£o: {e}", "ERROR")
            messagebox.showerror("Erro", f"‚ùå Erro durante a automa√ß√£o:\n{str(e)[:200]}...")
        finally:
            if driver:
                try:
                    driver.quit()
                    log_message("‚úÖ Browser fechado com sucesso", "SUCCESS")
                except Exception as quit_error:
                    log_message(f"‚ö†Ô∏è Erro ao fechar browser: {quit_error}", "WARNING")


def run(params: dict):
    module = MacroSeptModule()
    module.run(params)