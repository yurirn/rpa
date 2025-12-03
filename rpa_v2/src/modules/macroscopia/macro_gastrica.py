import os
import time
from tkinter import messagebox
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.action_chains import ActionChains
from dotenv import load_dotenv
from openpyxl import load_workbook

from src.core.browser_factory import BrowserFactory
from src.core.logger import log_message
from src.modules.base import BaseModule

load_dotenv()

class MacroGastricaModule(BaseModule):
    def __init__(self):
        super().__init__(nome="Macro Gástrica")

    def get_dados_exames(self, file_path: str) -> list:
        try:
            workbook = load_workbook(file_path)
            sheet = workbook.active
            dados = []
            ultima_mascara = None
            data_fixacao = None
            responsavel_macro_valor = None

            # Ler cabeçalho (linha 1) e criar mapeamento de colunas
            colunas = {}
            for col_idx in range(1, sheet.max_column + 1):
                cell_value = sheet.cell(row=1, column=col_idx).value
                if cell_value:
                    # Normalizar nome da coluna (minúsculo, sem espaços extras)
                    nome_coluna = str(cell_value).strip().lower()
                    colunas[nome_coluna] = col_idx
            
            log_message(f"📋 Colunas detectadas: {list(colunas.keys())}", "INFO")
            
            # Mapear nomes possíveis para cada campo (flexível)
            def encontrar_coluna(nomes_possiveis):
                """Encontra a coluna baseado em uma lista de nomes possíveis"""
                for nome in nomes_possiveis:
                    for coluna_nome, col_idx in colunas.items():
                        if nome.lower() in coluna_nome:
                            return col_idx
                return None
            
            # Encontrar índices das colunas
            col_codigo = encontrar_coluna(['codigo', 'código', 'cod', 'num_exame', 'numero', 'número'])
            col_mascara = encontrar_coluna(['mascara', 'máscara', 'mask'])
            col_responsavel = encontrar_coluna(['responsavel', 'responsável', 'resp', 'macroscopista'])
            col_qtd_frag = encontrar_coluna(['qtd_frag', 'qtd frag', 'fragmentos', 'quantidade', 'qtd'])
            col_qtd_frag2 = encontrar_coluna(['qtd_frag_2', 'qtd frag 2', 'fragmentos2', 'quantidade2', 'qtd2'])
            col_md1 = encontrar_coluna(['md1', 'medida 1', 'med1', 'medida1', 'campo e', 'e'])
            col_md2 = encontrar_coluna(['md2', 'medida 2', 'med2', 'medida2', 'campo f', 'f'])
            col_md3 = encontrar_coluna(['md3', 'medida 3', 'med3', 'medida3', 'campo g', 'g'])
            col_md4 = encontrar_coluna(['md4', 'medida 4', 'med4', 'medida4'])
            col_md5 = encontrar_coluna(['md5', 'medida 5', 'med5', 'medida5'])
            col_md6 = encontrar_coluna(['md6', 'medida 6', 'med6', 'medida6'])
            col_data = encontrar_coluna(['data', 'data fixacao', 'data fixação', 'datafixacao'])
            
            # Validar colunas obrigatórias
            if not col_codigo:
                raise Exception("Coluna de código não encontrada! Use um nome como 'Código' ou 'Codigo'")
            
            log_message(
                f"✅ Mapeamento: Código=col{col_codigo}, Máscara=col{col_mascara}, Data=col{col_data}, "
                f"QtdFrag=col{col_qtd_frag}, QtdFrag2=col{col_qtd_frag2}, "
                f"Md1=col{col_md1}, Md2=col{col_md2}, Md3=col{col_md3}, Md4=col{col_md4}, Md5=col{col_md5}, Md6=col{col_md6}",
                "INFO"
            )

            # Lê da linha 2 em diante (linha 1 é cabeçalho)
            for row in range(2, sheet.max_row + 1):
                codigo = sheet.cell(row=row, column=col_codigo).value if col_codigo else None
                mascara = sheet.cell(row=row, column=col_mascara).value if col_mascara else None
                responsavel_macro = sheet.cell(row=row, column=col_responsavel).value if col_responsavel else None
                qtd_frag = sheet.cell(row=row, column=col_qtd_frag).value if col_qtd_frag else None
                md1 = sheet.cell(row=row, column=col_md1).value if col_md1 else None
                md2 = sheet.cell(row=row, column=col_md2).value if col_md2 else None
                md3 = sheet.cell(row=row, column=col_md3).value if col_md3 else None
                qtd_frag2 = sheet.cell(row=row, column=col_qtd_frag2).value if col_qtd_frag2 else None
                md4 = sheet.cell(row=row, column=col_md4).value if col_md4 else None
                md5 = sheet.cell(row=row, column=col_md5).value if col_md5 else None
                md6 = sheet.cell(row=row, column=col_md6).value if col_md6 else None
                data_col = sheet.cell(row=row, column=col_data).value if col_data else None

                if row == 2 and data_col:
                    data_fixacao = str(data_col).strip()

                if responsavel_macro is not None and str(responsavel_macro).strip():
                    responsavel_macro_valor = str(responsavel_macro).strip().upper()

                if codigo is not None:
                    codigo = str(codigo).strip()
                    # Se não tem máscara, usa a última válida
                    if mascara is not None and str(mascara).strip():
                        mascara = str(mascara).strip()
                        ultima_mascara = mascara
                    else:
                        mascara = ultima_mascara

                    # Preservar o valor original de qtd_frag antes de converter
                    qtd_frag_original = str(qtd_frag).strip().lower() if qtd_frag is not None else ""
                    # novo campo: preservar original do frasco 2
                    qtd_frag2_original = str(qtd_frag2).strip().lower() if qtd_frag2 is not None else ""

                    if qtd_frag is not None and str(qtd_frag).strip().lower() == 'mult':
                        qtd_frag_valor = '6'
                    else:
                        qtd_frag_valor = str(qtd_frag).strip() if qtd_frag is not None else ""

                    # normalização simples para frasco 2 (sem regra de 'mult' por enquanto)
                    qtd_frag2_valor = str(qtd_frag2).strip() if qtd_frag2 is not None else ""

                    dados.append({
                        'codigo': codigo,
                        'mascara': mascara,
                        'responsavel_macro': responsavel_macro_valor,
                        'qtd_frag': qtd_frag_valor,
                        'qtd_frag_original': qtd_frag_original,
                        'md1': str(md1).strip() if md1 is not None else "",
                        'md2': str(md2).strip() if md2 is not None else "",
                        'md3': str(md3).strip() if md3 is not None else "",
                        'qtd_frag2': qtd_frag2_valor,
                        'qtd_frag2_original': qtd_frag2_original,
                        'md4': str(md4).strip() if md4 is not None else "",
                        'md5': str(md5).strip() if md5 is not None else "",
                        'md6': str(md6).strip() if md6 is not None else "",
                        'data_fixacao': data_fixacao
                    })
            workbook.close()
            return dados
        except Exception as e:
            raise Exception(f"Erro ao ler planilha: {e}")

    def verificar_sessao_browser(self, driver) -> bool:
        """Verifica se a sessão do browser ainda está ativa"""
        try:
            driver.current_url
            return True
        except Exception as e:
            if "invalid session id" in str(e).lower():
                log_message("❌ Sessão do browser perdida", "ERROR")
                return False
            return True

    def verificar_elemento_interativo(self, driver, elemento):
        """Verifica se um elemento está realmente interativo"""
        try:
            # Verificar se o elemento está visível e habilitado
            if not elemento.is_displayed() or not elemento.is_enabled():
                return False
            
            # Verificar se o elemento não está sobreposto por outros elementos
            rect = elemento.rect
            center_x = rect['x'] + rect['width'] / 2
            center_y = rect['y'] + rect['height'] / 2
            
            # Usar JavaScript para verificar se o elemento está realmente clicável
            is_clickable = driver.execute_script("""
                var elem = arguments[0];
                var rect = elem.getBoundingClientRect();
                var centerX = rect.left + rect.width / 2;
                var centerY = rect.top + rect.height / 2;
                
                // Verificar se há algum elemento sobrepondo
                var elementAtPoint = document.elementFromPoint(centerX, centerY);
                return elementAtPoint === elem || elem.contains(elementAtPoint);
            """, elemento)
            
            return is_clickable
        except:
            return False

    def aguardar_pagina_estavel(self, driver, wait, timeout=10):
        """Aguarda até que a página esteja estável (sem animações ou carregamentos)"""
        try:
            # Aguardar até que não haja requisições AJAX em andamento
            driver.execute_script("""
                return new Promise((resolve) => {
                    if (window.jQuery && window.jQuery.active === 0) {
                        resolve();
                        return;
                    }
                    
                    var checkInterval = setInterval(() => {
                        if (window.jQuery && window.jQuery.active === 0) {
                            clearInterval(checkInterval);
                            resolve();
                        }
                    }, 100);
                    
                    // Timeout de segurança
                    setTimeout(() => {
                        clearInterval(checkInterval);
                        resolve();
                    }, arguments[0]);
                });
            """, timeout * 1000)
            
            # Aguardar um pouco mais para garantir estabilidade
            time.sleep(0.5)
            log_message("✅ Página estável", "INFO")
            
        except Exception as e:
            log_message(f"⚠️ Erro ao aguardar página estável: {e}", "WARNING")
            time.sleep(1)  # Fallback

    def aguardar_spinner_desaparecer(self, driver, wait, timeout=30):
        """Aguarda até que o spinner de loading desapareça"""
        try:
            log_message("⏳ Aguardando spinner desaparecer...", "INFO")
            
            # Aguardar até que o spinner não esteja mais visível
            wait.until(EC.invisibility_of_element_located((By.ID, "spinner")))
            
            # Aguardar um pouco mais para garantir que não há outros spinners
            time.sleep(1)
            
            # Verificar se há outros spinners ou modais de loading
            spinners = driver.find_elements(By.CSS_SELECTOR, ".loadModal, .spinner, [class*='loading']")
            for spinner in spinners:
                if spinner.is_displayed():
                    log_message("⚠️ Outro spinner ainda visível, aguardando...", "WARNING")
                    time.sleep(2)
                    break
            
            log_message("✅ Spinner desapareceu", "SUCCESS")
            
        except Exception as e:
            log_message(f"⚠️ Erro ao aguardar spinner: {e}", "WARNING")
            # Tentar fechar o spinner via JavaScript se necessário
            try:
                driver.execute_script("""
                    var spinners = document.querySelectorAll('.loadModal, .spinner, [class*="loading"]');
                    spinners.forEach(function(spinner) {
                        if (spinner.style.display !== 'none') {
                            spinner.style.display = 'none';
                        }
                    });
                """)
                log_message("🔧 Spinner fechado via JavaScript", "INFO")
                time.sleep(1)
            except:
                pass

    def clicar_elemento_robusto(self, driver, wait, elemento, nome_elemento="elemento"):
        """Clica em um elemento de forma robusta, lidando com elementos interceptados"""
        try:
            # Rolar até o elemento para garantir visibilidade
            driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", elemento)
            time.sleep(1)
            
            # Verificar se há elementos sobrepostos e aguardar eles desaparecerem
            try:
                # Aguardar elementos sobrepostos desaparecerem (como dropdowns, tooltips, etc.)
                WebDriverWait(driver, 3).until_not(
                    EC.presence_of_element_located((By.XPATH, "//li[contains(@class, 'dropdown-menu') or contains(@class, 'show')]"))
                )
            except:
                pass  # Se não houver elementos sobrepostos, continua
            
            # Tentar clicar normalmente primeiro
            try:
                elemento.click()
                log_message(f"✅ Clicou em {nome_elemento}", "SUCCESS")
                return True
            except Exception as click_error:
                log_message(f"⚠️ Erro no clique normal em {nome_elemento}: {click_error}", "WARNING")
                # Se falhar, tentar clique via JavaScript
                try:
                    driver.execute_script("arguments[0].click();", elemento)
                    log_message(f"✅ Clicou em {nome_elemento} (via JavaScript)", "SUCCESS")
                    return True
                except Exception as js_error:
                    log_message(f"❌ Erro no clique JavaScript em {nome_elemento}: {js_error}", "ERROR")
                    return False
                    
        except Exception as e:
            log_message(f"❌ Erro geral ao clicar em {nome_elemento}: {e}", "ERROR")
            return False

    def selecionar_responsavel_macroscopia(self, driver, wait, responsavel_macro):
        """Seleciona o responsável pela macroscopia conforme o nome recebido (nome curto)"""
        # Mapper de nomes: primeiro nome em caixa alta -> nome completo
        responsavel_macro_mapper = {
            'BARBARA': 'Barbara Dutra Lopes',
            'NATHALIA': 'Nathalia Fernanda da Silva Lopes',
            'RENATA': 'Renata Silva Sevidanis',
            'HELEN': 'Helen Oliveira dos Santos',
            'CLARA': 'Clara Helena Janz Garcia de Souza',
            'PALOMA': 'Paloma Brenda Silva De Oliveira',
            'ELLEN': 'Ellen Andressa de Alvarenga',
            'VITORIA': 'Vitoria Aquino Nairne Domingues',
            'ANNAI': 'Annai Lukã Vitorino Losnak',
            'ANA' : 'Ana Carolina Viecele Campos',
            'LEILA': 'Leila dos Santos Martins',
            'PAOLA': 'Paola Galdino Sousa'
        }
        nome_completo = responsavel_macro_mapper.get(responsavel_macro, responsavel_macro)
        select2_container = wait.until(
            EC.element_to_be_clickable((By.XPATH, "//span[@aria-labelledby='select2-responsavelMacroscopiaId-container']"))
        )
        select2_container.click()
        time.sleep(0.3)
        # Seleciona a opção pelo nome completo
        opcao = wait.until(
            EC.element_to_be_clickable((By.XPATH, f"//li[contains(text(), '{nome_completo}')]") )
        )
        opcao.click()
        log_message(f"✅ {nome_completo} selecionado como responsável", "SUCCESS")
        time.sleep(0.2)

    def selecionar_auxiliar_macroscopia(self, driver, wait):
        """Seleciona 'Renata Silva Sevidanis' como auxiliar da macroscopia"""
        # Aguardar o componente Select2 estar presente e clicar
        select2_container = wait.until(
            EC.element_to_be_clickable((By.XPATH, "//span[@aria-labelledby='select2-auxiliarMacroscopiaId-container']"))
        )
        select2_container.click()
        time.sleep(0.2)
        
        # Aguardar e clicar na opção "Renata Silva Sevidanis"
        opcao_renata = wait.until(
            EC.element_to_be_clickable((By.XPATH, "//li[contains(text(), 'Renata Silva Sevidanis')]"))
        )
        opcao_renata.click()
        log_message("✅ Renata Silva Sevidanis selecionada como auxiliar", "SUCCESS")
        time.sleep(0.2)

    def definir_data_fixacao(self, driver, wait, data_fixacao=None):
        """Define a data de fixação no campo de data de fixação"""
        try:
            if not data_fixacao:
                data_fixacao = '21082025'  # fallback para data padrão se não vier da planilha
            # Converter 21082025 para 2025-08-21
            if len(data_fixacao) == 8 and data_fixacao.isdigit():
                data_formatada = f"{data_fixacao[4:8]}-{data_fixacao[2:4]}-{data_fixacao[0:2]}"
            else:
                data_formatada = '2025-08-21'
            campo_data = wait.until(
                EC.presence_of_element_located((By.XPATH, "//input[@type='date' and @name='dataFixacao']"))
            )
            driver.execute_script("""
                var campo = arguments[0];
                campo.value = arguments[1];
                campo.dispatchEvent(new Event('change', { bubbles: true }));
            """, campo_data, data_formatada)
            log_message(f"📅 Data de fixação definida para: {data_formatada}", "SUCCESS")
            time.sleep(0.1)
        except Exception as e:
            log_message(f"⚠️ Erro ao definir data de fixação: {e}", "WARNING")

    def definir_hora_fixacao(self, driver, wait):
        """Define 18:00 no campo de hora de fixação"""
        # Aguardar o campo de hora estar presente
        campo_hora = wait.until(
            EC.presence_of_element_located((By.XPATH, "//input[@type='time' and @name='dataFixacao']"))
        )
        
        # Limpar e definir a hora
        campo_hora.clear()
        campo_hora.send_keys("18:00")
        log_message("🕕 Hora de fixação definida para: 18:00", "SUCCESS")
        time.sleep(0.1)

    def fechar_exame(self, driver, wait):
        """Clica no botão de fechar exame"""
        try:
            botao_fechar = wait.until(
                EC.element_to_be_clickable((By.ID, "fecharExameBarraFerramenta"))
            )
            botao_fechar.click()
            log_message("📁 Exame fechado", "INFO")
            
            # Aguardar retornar à tela principal
            try:
                # Verificar se voltou à tela principal aguardando o campo de código aparecer
                wait.until(EC.presence_of_element_located((By.ID, "inputSearchCodBarra")))
                log_message("✅ Retornou à tela principal após fechar exame", "INFO")
            except:
                log_message("⚠️ Pode não ter retornado à tela principal", "WARNING")
                # Tentar navegar de volta ao módulo se necessário
                try:
                    current_url = driver.current_url
                    if "modulo=1" not in current_url:
                        modulo_link = driver.find_element(By.CSS_SELECTOR, "a[href='/site/trocarModulo?modulo=1']")
                        modulo_link.click()
                        time.sleep(1.5)
                        log_message("🔄 Navegou de volta ao módulo de exames", "INFO")
                except:
                    pass
                    
        except Exception as e:
            log_message(f"Erro ao fechar exame: {e}", "ERROR")

    def digitar_mascara_e_buscar(self, driver, wait, mascara):
        """Digita a máscara no campo buscaArvore e pressiona Enter"""
        # Aguardar o campo estar presente e clicável
        campo_busca = wait.until(EC.element_to_be_clickable((By.ID, "buscaArvore")))
        
        # Digitar a máscara e pressionar Enter
        campo_busca.send_keys(mascara)
        campo_busca.send_keys(Keys.ENTER)
        log_message(f"✍️ Máscara '{mascara}' digitada no campo buscaArvore", "SUCCESS")
        time.sleep(0.5)

    def abrir_modal_variaveis_e_preencher(self, driver, wait, mascara, qtd_frag, qtd_frag_original, md1, md2, md3, qtd_frag2, qtd_frag2_original, md4, md5, md6):
        """Abre o modal de variáveis e preenche os campos baseado na máscara"""
        try:
            # Clicar no botão "Pesquisar variáveis (F7)"
            botao_variaveis = wait.until(
                EC.element_to_be_clickable((By.ID, "cke_70"))
            )
            botao_variaveis.click()
            log_message("🔍 Clicou no botão de variáveis", "INFO")
            
            # Aguardar um pouco para o sistema processar
            time.sleep(0.8)
            
            # Verificar se apareceu um alerta
            try:
                alert = driver.switch_to.alert
                alert_text = alert.text
                if "não há variáveis" in alert_text.lower():
                    log_message(f"⚠️ Alerta detectado: {alert_text}", "WARNING")
                    alert.accept()  # Aceitar o alerta
                    log_message("⚠️ Pulando preenchimento de variáveis - não há variáveis no texto", "WARNING")
                    return
                else:
                    alert.accept()  # Aceitar qualquer outro alerta
            except:
                # Não há alerta, continuar normalmente
                pass
            
            # Aguardar o modal aparecer
            wait.until(EC.presence_of_element_located((By.CLASS_NAME, "swal2-popup")))
            log_message("🔍 Modal de variáveis aberto", "SUCCESS")
            time.sleep(0.3)
        
            # Preencher os campos usando classe genérica (IDs podem mudar)
            campos_input = driver.find_elements(By.CSS_SELECTOR, "input[style*='width: 100px'][style*='color: red']")
            log_message(f"🔍 Encontrados {len(campos_input)} campos de input no modal", "INFO")
            
            # Determinar valores baseado na máscara
            mascara_upper = mascara.upper() if mascara else ""
            valores = []
            
            if mascara_upper in ['VBSEM', 'VBCOM']:
                # med1, med2, med3 e tamanho da parede (na quantidade de fragmentos)
                valores = [md1, md2, md3, qtd_frag]

            elif mascara_upper == 'APC':
                # med1 e med2 sem med3
                valores = [md1, md2]

            elif mascara_upper == 'COLO':
                # Ordem correta: quantidade fragmentos, med1, med2, med3, quantidade legenda
                # Usar qtd_frag_original para verificar se era 'mult' na planilha
                if qtd_frag_original == 'mult':
                    valores = ["Múltiplos", md1, md2, md3, "M"]
                else:
                    valores = [qtd_frag, md1, md2, md3, qtd_frag]

            elif mascara_upper in ['RTU-FIT', 'RTU-FIP']:
                # peso (qtd_frag), med1, med2, med3 - PESO VEM PRIMEIRO!
                valores = [qtd_frag, md1, md2, md3]

            elif mascara_upper in ['HEMO-FIT', 'HEMO-FIP']:
                if qtd_frag_original == 'mult':
                    valores = ["Múltiplos", md1, md2, md3, "M"]
                else:
                    valores = [qtd_frag, md1, md2, md3, qtd_frag]

            elif mascara_upper in ['A/C2F', 'A/I2F', 'A/P2F', 'G/E2F', 'G/P2F']:
                if qtd_frag_original == 'mult' and qtd_frag2_original == 'mult':
                    valores = ["Múltiplos", md1, md2, md3, "Múltiplos", md4, md5, md6, "M", "M"]
                elif qtd_frag_original == 'mult' and not qtd_frag2_original == 'mult':
                    valores = ["Múltiplos", md1, md2, md3, qtd_frag2, md4, md5, md6, "M", qtd_frag2]
                elif not qtd_frag_original == 'mult' and qtd_frag2_original == 'mult':
                    valores = [qtd_frag, md1, md2, md3, "Múltiplos", md4, md5, md6, qtd_frag, "M"]
                else:
                    valores = [qtd_frag, md1, md2, md3, qtd_frag2, md4, md5, md6, qtd_frag, qtd_frag2]

            else:
                # Padrão original (máscaras antigas)
                # Usar qtd_frag_original para verificar se era 'mult' na planilha
                if qtd_frag_original == 'mult':
                    valores = ["Múltiplos", md1, md2, md3, "M"]
                else:
                    valores = [qtd_frag, md1, md2, md3, qtd_frag]

            log_message(f"📋 Preenchendo variáveis para máscara '{mascara}': {valores}", "INFO")
            
            for i, campo in enumerate(campos_input[:len(valores)]):  # Limitar ao número de valores
                if i < len(valores) and valores[i]:
                    try:
                        campo.clear()
                        campo.send_keys(str(valores[i]))
                        log_message(f"✍️ Campo {i+1} preenchido com: {valores[i]}", "SUCCESS")
                    except Exception as e:
                        log_message(f"⚠️ Erro ao preencher campo {i+1}: {e}", "WARNING")
            
            time.sleep(0.2)
            
            # Clicar no botão "Inserir"
            botao_inserir = wait.until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, ".swal2-confirm"))
            )
            botao_inserir.click()
            log_message("✅ Campos inseridos no modal", "SUCCESS")
            
            # Aguardar o modal fechar completamente
            try:
                wait.until(EC.invisibility_of_element_located((By.CLASS_NAME, "swal2-popup")))
                log_message("✅ Modal fechado completamente", "SUCCESS")
            except:
                # Se não conseguir detectar fechamento, aguardar um tempo fixo
                time.sleep(1)
                log_message("⏳ Aguardou fechamento do modal", "INFO")
            
        except Exception as e:
            log_message(f"⚠️ Erro ao preencher modal de variáveis: {e}", "WARNING")
            log_message("⚠️ Continuando sem preencher as variáveis", "WARNING")

    def salvar_macroscopia(self, driver, wait):
        """Clica no botão Salvar da macroscopia"""
        # Verificar se ainda há modal aberto e fechar se necessário
        try:
            modal = driver.find_element(By.CLASS_NAME, "swal2-popup")
            if modal.is_displayed():
                log_message("⚠️ Modal ainda aberto, tentando fechar...", "WARNING")
                # Tentar fechar o modal
                try:
                    botao_cancelar = driver.find_element(By.CSS_SELECTOR, ".swal2-cancel")
                    botao_cancelar.click()
                    time.sleep(0.5)
                except:
                    # Se não conseguir fechar, pressionar ESC
                    driver.find_element(By.TAG_NAME, "body").send_keys(Keys.ESCAPE)
                    time.sleep(0.5)
        except:
            # Não há modal, continuar normalmente
            pass
        
        # Aguardar o botão estar presente e clicável
        botao_salvar = wait.until(
            EC.element_to_be_clickable((By.ID, "salvarMacro"))
        )
        botao_salvar.click()
        log_message("💾 Macroscopia salva", "SUCCESS")
        time.sleep(0.3)

    def definir_grupo_baseado_mascara(self, driver, wait, mascara):
        """Define o grupo baseado na máscara (Estômago ou Intestino) - versão melhorada com JavaScript."""
        if not mascara:
            log_message("⚠️ Nenhuma máscara fornecida para definir grupo", "WARNING")
            return

        mascaras_estomago = [
            'A/C', 'A/I', 'AIC', 'AIF', 'ANTRO', 'COTO', 'DUO', 'DUO ', 'ESOFF',
            'GASTRICA', 'POLIPO', 'G/POLIPO', 'ULCERA',
            'A/C2F', 'A/I2F', 'A/P2F', 'G/E2F', 'G/P2F'
        ]
        mascaras_intestino = ['B/COLON', 'ICR', 'P/COLON']
        mascaras_vesicula = ['VBSEM', 'VBCOM']
        mascaras_apendice = ['APC']
        mascaras_prostata = ['RTU-FIT', 'RTU-FIP']
        mascaras_geral = ['HEMO-FIT', 'HEMO-FIP']
        mascaras_utero = ['COLO']
        # Máscaras mistas estômago/intestino
        mascaras_estomago_intestino = ['G/PCOLON', 'G/BCOLON']

        grupo_selecionado = None
        mascara_upper = mascara.upper()
        
        if mascara_upper in mascaras_estomago:
            grupo_selecionado = "Estomago"
        elif mascara_upper in mascaras_intestino:
            grupo_selecionado = "Intestino"
        elif mascara_upper in mascaras_vesicula:
            grupo_selecionado = "Vesicula biliar"
        elif mascara_upper in mascaras_apendice:
            grupo_selecionado = "Apendice"
        elif mascara_upper in mascaras_prostata:
            grupo_selecionado = "Prostata"
        elif mascara_upper in mascaras_geral:
            grupo_selecionado = "Geral"
        elif mascara_upper in mascaras_utero:
            grupo_selecionado = "Utero"
        elif mascara_upper in mascaras_estomago_intestino:
            grupo_selecionado = "Estomago e intestino"
        else:
            log_message(f"⚠️ Máscara '{mascara}' não encontrada nas regras definidas", "WARNING")
            return

        try:
            # Verificar se o input existe e qual o valor atual
            try:
                input_grupo = driver.find_element(By.ID, "idRegiao")
                valor_atual = input_grupo.get_attribute("value")
                
                if valor_atual == grupo_selecionado:
                    log_message(f"✅ Grupo já está definido como '{grupo_selecionado}' - pulando", "SUCCESS")
                    return
                elif valor_atual and valor_atual != grupo_selecionado:
                    log_message(f"⚠️ Grupo atual é '{valor_atual}', precisa mudar para '{grupo_selecionado}'", "WARNING")
                else:
                    log_message(f"📝 Campo de grupo vazio, definindo como '{grupo_selecionado}'", "INFO")
            except:
                log_message("⚠️ Campo idRegiao não encontrado", "WARNING")
                return
            
            # Tentar encontrar especificamente o campo de grupo pelo ID idRegiao
            script = """
            // Procurar especificamente pelo campo de grupo que tem o input com id="idRegiao"
            var inputGrupo = document.getElementById('idRegiao');
            if (inputGrupo) {
                // Encontrar a âncora que está no mesmo td que o input idRegiao
                var parentTd = inputGrupo.closest('td');
                if (parentTd) {
                    var ancora = parentTd.querySelector('a[class*="table-editable-ancora"]');
                    if (ancora && ancora.offsetParent !== null) {
                        return ancora;
                    }
                }
            }
            
            // Fallback: procurar por âncoras que estejam próximas a inputs de grupo
            var inputsGrupo = document.querySelectorAll('input[id*="Regiao"], input[data-autocompleteurl*="consultarRegiao"]');
            for (var i = 0; i < inputsGrupo.length; i++) {
                var input = inputsGrupo[i];
                var parentTd = input.closest('td');
                if (parentTd) {
                    var ancora = parentTd.querySelector('a[class*="table-editable-ancora"]');
                    if (ancora && ancora.offsetParent !== null) {
                        return ancora;
                    }
                }
            }
            
            // Último fallback: procurar por âncoras que não sejam de procedimento
            var fragmentosContainer = document.getElementById('fragmentosContainer');
            if (fragmentosContainer) {
                var elementos = fragmentosContainer.querySelectorAll('a[class*="table-editable-ancora"]');
                for (var i = 0; i < elementos.length; i++) {
                    var elemento = elementos[i];
                    if (elemento.textContent.includes('Vazio') && elemento.offsetParent !== null) {
                        var parentTd = elemento.closest('td');
                        if (parentTd && !parentTd.querySelector('input[id*="procedimento"]')) {
                            return elemento;
                        }
                    }
                }
            }
            return null;
            """
            campo_grupo = driver.execute_script(script)
                
            if campo_grupo:
                # Usar JavaScript para clicar no elemento
                driver.execute_script("arguments[0].click();", campo_grupo)
                log_message(f"🔍 Clicou no campo de grupo via JS", "INFO")
                time.sleep(0.5)

                # Aguardar o campo de input aparecer e preencher via JavaScript
                input_grupo = wait.until(
                    EC.presence_of_element_located((By.ID, "idRegiao"))
                )
                
                # Limpar o campo primeiro
                driver.execute_script("arguments[0].value = '';", input_grupo)
                
                # Preencher via JavaScript
                driver.execute_script("""
                    arguments[0].value = arguments[1];
                    arguments[0].dispatchEvent(new Event('input', { bubbles: true }));
                    arguments[0].dispatchEvent(new Event('change', { bubbles: true }));
                """, input_grupo, grupo_selecionado)
                
                # Aguardar um pouco para o dropdown aparecer e tentar clicar na opção
                time.sleep(0.5)
                
                # Tentar clicar na opção do dropdown com timeout menor
                try:
                    # Aguardar até 3 segundos pela opção aparecer
                    wait_dropdown = WebDriverWait(driver, 3)
                    opcao_dropdown = wait_dropdown.until(
                        EC.element_to_be_clickable((By.XPATH, f"//li[contains(@class, 'active')]//a[contains(text(), '{grupo_selecionado}')]"))
                    )
                    opcao_dropdown.click()
                    log_message(f"✅ Selecionou '{grupo_selecionado}' no dropdown", "SUCCESS")
                except:
                    # Se não conseguir clicar no dropdown rapidamente, pressionar Enter
                    try:
                        input_grupo.send_keys(Keys.ENTER)
                        log_message(f"✍️ Pressionou Enter para confirmar '{grupo_selecionado}' (dropdown não apareceu)", "SUCCESS")
                    except:
                        # Último recurso: clicar fora para fechar o dropdown
                        driver.execute_script("document.body.click();")
                        log_message(f"🔍 Clicou fora para fechar dropdown de '{grupo_selecionado}'", "INFO")
                
                time.sleep(0.5)
            else:
                log_message("⚠️ Campo de grupo não encontrado ou não visível", "WARNING")
                
        except Exception as e:
            log_message(f"⚠️ Erro ao definir grupo: {e}", "WARNING")

    def definir_representacao_secao(self, driver, wait):
        """Define a representação como 'Seção' usando JavaScript"""
        try:
            # Verificar se o select existe e qual o valor atual
            try:
                select_representacao = driver.find_element(By.ID, "representacao")
                valor_atual = select_representacao.get_attribute("value")
                
                if valor_atual == "S":
                    log_message("✅ Representação já está definida como 'Seção'", "SUCCESS")
                    return
                elif valor_atual != "S":
                    log_message(f"⚠️ Representação atual é '{valor_atual}', mas precisa ser 'S' (Seção)", "WARNING")
            except:
                log_message("⚠️ Campo representacao não encontrado", "WARNING")
                return
            
            # Procurar especificamente pelo campo de representação
            script = """
            // Procurar especificamente pelo campo de representação que tem o select com id="representacao"
            var selectRepresentacao = document.getElementById('representacao');
            if (selectRepresentacao) {
                // Encontrar a âncora que está no mesmo td que o select representacao
                var parentTd = selectRepresentacao.closest('td');
                if (parentTd) {
                    var ancora = parentTd.querySelector('a[class*="table-editable-ancora"]');
                    if (ancora && ancora.offsetParent !== null) {
                        return ancora;
                    }
                }
            }
            
            // Fallback: procurar por âncoras que estejam próximas a selects de representação
            var selectsRepresentacao = document.querySelectorAll('select[id*="representacao"], select[name*="representacao"]');
            for (var i = 0; i < selectsRepresentacao.length; i++) {
                var select = selectsRepresentacao[i];
                var parentTd = select.closest('td');
                if (parentTd) {
                    var ancora = parentTd.querySelector('a[class*="table-editable-ancora"]');
                    if (ancora && ancora.offsetParent !== null) {
                        return ancora;
                    }
                }
            }
            
            // Último fallback: procurar por texto "representação" ou "-- representação --" que não seja de procedimento
            var fragmentosContainer = document.getElementById('fragmentosContainer');
            if (fragmentosContainer) {
                var elementos = fragmentosContainer.querySelectorAll('a[class*="table-editable-ancora"]');
                for (var i = 0; i < elementos.length; i++) {
                    var elemento = elementos[i];
                    if ((elemento.textContent.toLowerCase().includes('representação') || elemento.textContent.includes('-- representação --')) && elemento.offsetParent !== null) {
                        var parentTd = elemento.closest('td');
                        if (parentTd && !parentTd.querySelector('input[id*="procedimento"]')) {
                            return elemento;
                        }
                    }
                }
            }
            return null;
            """
            campo_representacao = driver.execute_script(script)
            
            if not campo_representacao:
                log_message("⚠️ Campo de representação não encontrado", "WARNING")
                return

            # Verificar o texto da âncora para log
            if "Seção" in campo_representacao.text:
                log_message("✅ Representação já mostra 'Seção', mas vamos garantir", "INFO")
            elif "-- representação --" in campo_representacao.text:
                log_message("📝 Campo de representação encontrado, precisa ser preenchido", "INFO")
            else:
                log_message(f"⚠️ Texto inesperado no campo de representação: '{campo_representacao.text}'", "WARNING")

            # Clicar via JavaScript
            driver.execute_script("arguments[0].click();", campo_representacao)
            log_message("🔍 Clicou no campo de representação via JS", "INFO")
            time.sleep(0.5)

            # Aguardar o select aparecer e selecionar via JavaScript
            select_representacao = wait.until(
                EC.presence_of_element_located((By.ID, "representacao"))
            )

            # Selecionar "Seção" (valor "S") via JavaScript
            driver.execute_script("""
                arguments[0].value = 'S';
                arguments[0].dispatchEvent(new Event('change', { bubbles: true }));
            """, select_representacao)

            log_message("✅ Representação definida como 'Seção' via JS", "SUCCESS")
            time.sleep(0.5)
            
            # Clicar fora para confirmar a seleção
            driver.execute_script("document.body.click();")
            time.sleep(0.3)

        except Exception as e:
            log_message(f"⚠️ Erro ao definir representação: {e}", "WARNING")

    def definir_regiao_gastrica(self, driver, wait, mascara=None):
        """Define a região de acordo com a máscara usando JavaScript"""
        try:
            if not mascara:
                log_message("⚠️ Nenhuma máscara fornecida para definir região", "WARNING")
                return

            # Regras de máscara para região
            mascara_regiao = {
                'A/C': 'AC: Antro/Corpo',
                'A/I': 'AI: Antro/Incisura',
                'AIC': 'AIC: Antro/Incisura/Corpo',
                'AIF': 'AIF: Antro/Incisura/Fundo',
                'ANTRO': 'AN: Antro',
                'COTO': 'Coto: Coto',
                'ESOFF': 'Esofago: Esôfago',
                'GASTRICA': 'GA: Gastrica',
                'G/POLIPO': 'POL/GASTRICA: Pólipo e biópsia gástrica',
                'POLIPO': 'Pólipo: Pólipo',
                'ICR': 'ICR: Íleo/Cólon/Reto',
                'DUO': 'Duodeno: Duodeno',
                'ULCERA': 'UG: Úlcera Gastrica',
                'VBSEM': 'VB: Vesicula biliar',
                'VBCOM': 'VB: Vesicula biliar',
                'APC': 'APC: Apendice cecal',
                'RTU-FIT': 'RTU: Resseccao transuretral',
                'RTU-FIP': 'RTU: Resseccao transuretral',
                'HEMO-FIT': 'HEMO: Hemorroida',
                'HEMO-FIP': 'HEMO: Hemorroida',
            }
            # COLO e outras máscaras sem região definida (região em branco ou manual)
            mascaras_sem_regiao = ['B/COLON', 'P/COLON', 'COLO']

            mascara_upper = mascara.upper().replace('Ó', 'O').replace('Ô', 'O')
            mascara_map = {k.upper().replace('Ó', 'O').replace('Ô', 'O'): v for k, v in mascara_regiao.items()}
            mascaras_sem_regiao_norm = [m.upper().replace('Ó', 'O').replace('Ô', 'O') for m in mascaras_sem_regiao]

            if mascara_upper in mascaras_sem_regiao_norm:
                log_message(f"⚠️ Máscara '{mascara}' não exige preenchimento de região (manual)", "WARNING")
                return

            regiao_valor = mascara_map.get(mascara_upper)
            if not regiao_valor:
                log_message(f"⚠️ Máscara '{mascara}' não encontrada nas regras de região", "WARNING")
                log_message(f"🔍 Máscaras disponíveis: {list(mascara_map.keys())}", "INFO")
                return
            
            log_message(f"📝 Máscara '{mascara}' → Região '{regiao_valor}'", "INFO")

            # Verificar se já existe um campo de região preenchido com o valor correto
            try:
                inputs_regiao = driver.find_elements(By.XPATH, "//input[contains(@name, 'regiao_')]")
                for input_reg in inputs_regiao:
                    valor_atual = input_reg.get_attribute("value")
                    if valor_atual == regiao_valor:
                        log_message(f"✅ Região já está definida como '{regiao_valor}' - pulando", "SUCCESS")
                        return
                    elif valor_atual and valor_atual != regiao_valor:
                        log_message(f"⚠️ Região atual é '{valor_atual}', precisa mudar para '{regiao_valor}'", "WARNING")
                        break
            except:
                pass

            # Procurar e clicar no campo de região para editá-lo
            script = """
            // Procurar especificamente por campos de região na tabela de fragmentos
            var tbody = document.getElementById('tdRegiao');
            if (tbody) {
                var inputs = tbody.querySelectorAll('input[name*="regiao_"]');
                for (var i = 0; i < inputs.length; i++) {
                    var input = inputs[i];
                    var parentTd = input.closest('td');
                    if (parentTd) {
                        var ancora = parentTd.querySelector('a[class*="table-editable-ancora"]');
                        if (ancora && ancora.offsetParent !== null) {
                            return {element: ancora, input: input};
                        }
                    }
                }
            }
            return null;
            """
            resultado_regiao = driver.execute_script(script)
            
            if resultado_regiao:
                campo_regiao = resultado_regiao['element']
                input_regiao = resultado_regiao['input']
                
                # Clicar na âncora para abrir o campo de edição
                driver.execute_script("arguments[0].click();", campo_regiao)
                log_message("🔍 Clicou no campo de região para editar", "INFO")
                time.sleep(0.5)

                # Aguardar o input ficar visível e preencher via JavaScript
                try:
                    # Aguardar o input aparecer
                    wait.until(lambda d: input_regiao.is_displayed() or input_regiao.get_attribute("style") != "display: none;")
                    
                    # Limpar e preencher o campo
                    driver.execute_script("""
                        arguments[0].value = '';
                        arguments[0].value = arguments[1];
                        arguments[0].dispatchEvent(new Event('input', { bubbles: true }));
                        arguments[0].dispatchEvent(new Event('change', { bubbles: true }));
                    """, input_regiao, regiao_valor)
                    
                    log_message(f"✍️ Definiu região como '{regiao_valor}' via JS", "SUCCESS")
                    time.sleep(0.5)
                    
                    # Clicar fora para confirmar a edição
                    driver.execute_script("document.body.click();")
                    time.sleep(0.5)
                    
                    # Verificar se o valor foi realmente definido
                    valor_definido = input_regiao.get_attribute("value")
                    if valor_definido == regiao_valor:
                        log_message(f"✅ Valor de região confirmado: '{valor_definido}'", "SUCCESS")
                    else:
                        log_message(f"⚠️ Valor não foi definido corretamente. Esperado: '{regiao_valor}', Atual: '{valor_definido}'", "WARNING")
                        
                except Exception as input_error:
                    log_message(f"⚠️ Erro ao preencher input de região: {input_error}", "WARNING")
            else:
                log_message("⚠️ Campo de região não encontrado ou não visível", "WARNING")

        except Exception as e:
            log_message(f"⚠️ Erro ao definir região: {e}", "WARNING")

    def obter_padrao_fragmentos_blocos(self, mascara):
        """Retorna os padrões de fragmentos e blocos para cada tipo de máscara"""
        mascara_upper = mascara.upper() if mascara else ""
        
        # Padrões: (fragmentos_padrao, blocos_padrao, usar_sempre_padrao)
        # usar_sempre_padrao=True significa que IGNORA o valor da planilha
        padroes = {
            'VBSEM': (3, 1, True),      # Sempre 3F1B
            'VBCOM': (3, 1, True),      # Sempre 3F1B
            'APC': (3, 1, True),        # Sempre 3F1B
            'COLO': (None, 1, False),   # Quantidade variável da planilha, 1 bloco
            'RTU-FIT': (6, 1, True),    # Sempre 6 (múltiplos) 1B - campo_d vai para peso
            'RTU-FIP': (6, 1, True),    # Sempre 6 (múltiplos) 1B - campo_d vai para peso
            'HEMO-FIT': (None, 1, False),  # Quantidade variável da planilha, 1 bloco
            'HEMO-FIP': (None, 1, False),  # Quantidade variável da planilha, 1 bloco
        }
        
        return padroes.get(mascara_upper, (None, 1, False))
    
    # ========================= HELPERS 2 FRASCOS =========================
    def _buscar_proxima_regiao_vazia(self, driver):
        """Retorna dict {element, input} da PRÓXIMA região vazia na tabela de fragmentos."""
        script = """
        var tbody = document.getElementById('tdRegiao');
        if (tbody) {
            var inputs = tbody.querySelectorAll('input[name*="regiao_"]');
            for (var i = 0; i < inputs.length; i++) {
                var input = inputs[i];
                if (!input.value || input.value.trim() === '') {
                    var parentTd = input.closest('td');
                    if (parentTd) {
                        var ancora = parentTd.querySelector('a[class*="table-editable-ancora"]');
                        if (ancora && ancora.offsetParent !== null) {
                            return {element: ancora, input: input};
                        }
                    }
                }
            }
        }
        return null;
        """
        return driver.execute_script(script)

    def _buscar_proxima_quantidade_vazia(self, driver):
        """Retorna dict {element, input} da PRÓXIMA quantidade "livre" na tabela.
        Prioriza primeiro campos com valor '0' e depois campos vazios.
        """
        script = """
        var tbody = document.getElementById('tdRegiao');
        if (!tbody) { return null; }

        var candidatosZero = [];
        var candidatosVazio = [];
        var inputs = tbody.querySelectorAll('input[name*="quantidade_"]');

        for (var i = 0; i < inputs.length; i++) {
            var input = inputs[i];
            var valor = (input.value || '').trim();
            var parentTd = input.closest('td');
            if (!parentTd) { continue; }
            var ancora = parentTd.querySelector('a[class*="table-editable-ancora"]');
            if (!ancora || ancora.offsetParent === null) { continue; }

            if (valor === '0') {
                candidatosZero.push({ element: ancora, input: input });
            } else if (valor === '') {
                candidatosVazio.push({ element: ancora, input: input });
            }
        }

        if (candidatosZero.length > 0) {
            return candidatosZero[0];
        }
        if (candidatosVazio.length > 0) {
            return candidatosVazio[0];
        }
        return null;
        """
        return driver.execute_script(script)

    def _buscar_proxima_qtd_blocos_vazia(self, driver):
        """Retorna dict {element, input} da PRÓXIMA quantidade de blocos vazia na tabela de fragmentos."""
        script = """
        var tbody = document.getElementById('tdRegiao');
        if (tbody) {
            var inputs = tbody.querySelectorAll('input[name*="quantidadeBlocos_"]');
            for (var i = 0; i < inputs.length; i++) {
                var input = inputs[i];
                if (!input.value || input.value.trim() === '') {
                    var parentTd = input.closest('td');
                    if (parentTd) {
                        var ancora = parentTd.querySelector('a[class*="table-editable-ancora"]');
                        if (ancora && ancora.offsetParent !== null) {
                            return {element: ancora, input: input};
                        }
                    }
                }
            }
        }
        return null;
        """
        return driver.execute_script(script)

    def _preencher_regiao_linha(self, driver, wait, regiao_valor):
        """Preenche UMA linha de região com o valor informado, usando o próximo campo vazio."""
        if not regiao_valor:
            return
        resultado = self._buscar_proxima_regiao_vazia(driver)
        if not resultado:
            log_message("⚠️ Nenhum campo de região vazio encontrado para preencher", "WARNING")
            return

        campo_regiao = resultado['element']
        input_regiao = resultado['input']

        driver.execute_script("arguments[0].click();", campo_regiao)
        log_message("🔍 Clicou no próximo campo de região vazio", "INFO")
        time.sleep(0.4)

        try:
            wait.until(lambda d: input_regiao.is_displayed() or input_regiao.get_attribute("style") != "display: none;")
            driver.execute_script(
                """
                arguments[0].value = '';
                arguments[0].value = arguments[1];
                arguments[0].dispatchEvent(new Event('input', { bubbles: true }));
                arguments[0].dispatchEvent(new Event('change', { bubbles: true }));
                """,
                input_regiao,
                regiao_valor,
            )
            log_message(f"✍️ Região preenchida com '{regiao_valor}'", "SUCCESS")
            time.sleep(0.3)
            driver.execute_script("document.body.click();")
            time.sleep(0.3)
        except Exception as e:
            log_message(f"⚠️ Erro ao preencher linha de região: {e}", "WARNING")

    def _preencher_quantidade_linha(self, driver, wait, quantidade_valor):
        """Preenche UMA linha de quantidade de fragmentos no próximo campo vazio."""
        if not quantidade_valor:
            return
        resultado = self._buscar_proxima_quantidade_vazia(driver)
        if not resultado:
            log_message("⚠️ Nenhum campo de quantidade vazio encontrado para preencher", "WARNING")
            return

        campo_qtd = resultado['element']
        input_qtd = resultado['input']

        driver.execute_script("arguments[0].click();", campo_qtd)
        log_message("🔍 Clicou no próximo campo de quantidade vazio", "INFO")
        time.sleep(0.4)

        try:
            wait.until(lambda d: input_qtd.is_displayed() or input_qtd.get_attribute("style") != "display: none;")
            driver.execute_script(
                """
                arguments[0].value = '';
                arguments[0].value = arguments[1];
                arguments[0].dispatchEvent(new Event('input', { bubbles: true }));
                arguments[0].dispatchEvent(new Event('change', { bubbles: true }));
                """,
                input_qtd,
                str(quantidade_valor),
            )
            log_message(f"✍️ Quantidade de fragmentos preenchida com '{quantidade_valor}'", "SUCCESS")
            time.sleep(0.3)
            driver.execute_script("document.body.click();")
            time.sleep(0.3)
        except Exception as e:
            log_message(f"⚠️ Erro ao preencher linha de quantidade: {e}", "WARNING")

    def _preencher_blocos_linha(self, driver, wait, blocos_valor="1"):
        """Preenche UMA linha de quantidade de blocos no próximo campo vazio."""
        if not blocos_valor:
            return
        resultado = self._buscar_proxima_qtd_blocos_vazia(driver)
        if not resultado:
            log_message("⚠️ Nenhum campo de quantidade de blocos vazio encontrado para preencher", "WARNING")
            return

        campo_blocos = resultado['element']
        input_blocos = resultado['input']

        driver.execute_script("arguments[0].click();", campo_blocos)
        log_message("🔍 Clicou no próximo campo de blocos vazio", "INFO")
        time.sleep(0.4)

        try:
            wait.until(lambda d: input_blocos.is_displayed() or input_blocos.get_attribute("style") != "display: none;")
            driver.execute_script(
                """
                arguments[0].value = '';
                arguments[0].value = arguments[1];
                arguments[0].dispatchEvent(new Event('input', { bubbles: true }));
                arguments[0].dispatchEvent(new Event('change', { bubbles: true }));
                """,
                input_blocos,
                str(blocos_valor),
            )
            log_message(f"✍️ Quantidade de blocos preenchida com '{blocos_valor}'", "SUCCESS")
            time.sleep(0.3)
            driver.execute_script("document.body.click();")
            time.sleep(0.3)
        except Exception as e:
            log_message(f"⚠️ Erro ao preencher linha de blocos: {e}", "WARNING")

    # ====================== MÉTODOS 2 FRASCOS ==========================
    def definir_regiao_2frascos(self, driver, wait, mascara):
        """Define as regiões para máscaras de 2 frascos (preenche 2 linhas na tabela)."""
        if not mascara:
            log_message("⚠️ Nenhuma máscara fornecida para definir região (2 frascos)", "WARNING")
            return

        mascara_upper = mascara.upper()
        mapa_regioes = {
            'A/C2F': ("AN: Antro", "CP: Corpo"),
            'A/I2F': ("AN: Antro", "INC: Incisura"),
            'A/P2F': ("AN: Antro", "POLG: Pólipo gástrico"),
            'G/E2F': ("GA: Gastrica", "Esofago: Esôfago"),
            'G/P2F': ("GA: Gastrica", "POLG: Pólipo gástrico"),
        }

        regioes = mapa_regioes.get(mascara_upper)
        if not regioes:
            log_message(f"⚠️ Máscara '{mascara}' não está mapeada para 2 frascos", "WARNING")
            return

        regiao1, regiao2 = regioes
        log_message(f"📝 Definindo regiões 2 frascos: Frasco1='{regiao1}', Frasco2='{regiao2}'", "INFO")

        self._preencher_regiao_linha(driver, wait, regiao1)
        self._preencher_regiao_linha(driver, wait, regiao2)

    def definir_quantidade_fragmentos_2frascos(self, driver, wait, mascara, qtd_frag, qtd_frag2):
        """Define quantidade de fragmentos para máscaras de 2 frascos (linha 1 e 2)."""
        mascara_upper = mascara.upper() if mascara else ""

        # Padrão para o frasco 1 reaproveitando a regra existente
        frag_padrao, _, usar_sempre_padrao = self.obter_padrao_fragmentos_blocos(mascara_upper)

        # Frasco 1
        if usar_sempre_padrao and frag_padrao:
            qtd1 = str(frag_padrao)
            log_message(f"📝 [2F] Frasco 1 usando padrão FIXO de {frag_padrao} fragmentos para '{mascara_upper}'", "INFO")
        elif qtd_frag and str(qtd_frag).strip():
            qtd1 = str(qtd_frag).strip()
            log_message(f"📝 [2F] Frasco 1 usando quantidade da planilha: {qtd1}", "INFO")
        elif frag_padrao:
            qtd1 = str(frag_padrao)
            log_message(f"📝 [2F] Frasco 1 sem valor na planilha, usando padrão {qtd1}", "INFO")
        else:
            qtd1 = ""
            log_message("⚠️ [2F] Frasco 1 sem quantidade definida e sem padrão", "WARNING")

        # Frasco 2 – para essas máscaras, sempre virá da planilha (sem padrão específico por enquanto)
        if qtd_frag2 and str(qtd_frag2).strip():
            qtd2 = str(qtd_frag2).strip()
            log_message(f"📝 [2F] Frasco 2 usando quantidade da planilha: {qtd2}", "INFO")
        else:
            qtd2 = ""
            log_message("⚠️ [2F] Frasco 2 sem quantidade na planilha (mantendo vazio)", "WARNING")

        if qtd1:
            self._preencher_quantidade_linha(driver, wait, qtd1)
        if qtd2:
            self._preencher_quantidade_linha(driver, wait, qtd2)

    def definir_quantidade_blocos_2frascos(self, driver, wait):
        """Define quantidade de blocos para máscaras de 2 frascos (1 bloco em cada frasco)."""
        log_message("📝 [2F] Definindo quantidade de blocos: 1 para cada frasco", "INFO")
        self._preencher_blocos_linha(driver, wait, "1")
        self._preencher_blocos_linha(driver, wait, "1")

    # ====================== MÉTODOS 1 FRASCO ============================
    def definir_quantidade_fragmentos(self, driver, wait, mascara, qtd_frag):
        """Define a quantidade de fragmentos usando JavaScript melhorado"""
        try:
            # Obter padrão da máscara
            fragmentos_padrao, _, usar_sempre_padrao = self.obter_padrao_fragmentos_blocos(mascara)
            
            # Determinar quantidade a usar
            if usar_sempre_padrao and fragmentos_padrao:
                # Para RTU, VBSEM, VBCOM, APC: SEMPRE usar o padrão, ignorar planilha
                quantidade_valor = str(fragmentos_padrao)
                log_message(f"📝 Usando padrão FIXO de {fragmentos_padrao} fragmentos para '{mascara}' (ignora planilha)", "INFO")
            elif qtd_frag and qtd_frag.strip():
                # Para outras máscaras: usar valor da planilha se existir
                quantidade_valor = qtd_frag.strip()
                log_message(f"📝 Usando quantidade da planilha: {quantidade_valor}", "INFO")
            elif fragmentos_padrao:
                # Fallback: usar padrão se planilha estiver vazia
                quantidade_valor = str(fragmentos_padrao)
                log_message(f"📝 Campo qtd_frag vazio, usando padrão de {fragmentos_padrao} fragmentos para '{mascara}'", "INFO")
            else:
                log_message("⚠️ qtd_frag está vazio e não há padrão, não definindo quantidade", "WARNING")
                return

            log_message(f"✅ Definindo quantidade de fragmentos como: {quantidade_valor}", "INFO")

            # Procurar pelos campos de quantidade na tabela de fragmentos
            script = """
            // Procurar especificamente por campos de quantidade na tabela de fragmentos
            var tbody = document.getElementById('tdRegiao');
            if (tbody) {
                var inputs = tbody.querySelectorAll('input[name*="quantidade_"]');
                for (var i = 0; i < inputs.length; i++) {
                    var input = inputs[i];
                    var parentTd = input.closest('td');
                    if (parentTd) {
                        var ancora = parentTd.querySelector('a[class*="table-editable-ancora"]');
                        if (ancora && ancora.offsetParent !== null) {
                            return {element: ancora, input: input};
                        }
                    }
                }
            }
            return null;
            """

            resultado_quantidade = driver.execute_script(script)
            
            if resultado_quantidade:
                campo_quantidade = resultado_quantidade['element']
                input_quantidade = resultado_quantidade['input']
                
                # Verificar se já tem o valor correto
                valor_atual = input_quantidade.get_attribute("value")
                if valor_atual == quantidade_valor:
                    log_message(f"✅ Quantidade já está definida como '{quantidade_valor}' - pulando", "SUCCESS")
                    return
                
                # Clicar na âncora para abrir o campo
                driver.execute_script("arguments[0].click();", campo_quantidade)
                log_message("🔍 Clicou no campo de quantidade para editar", "INFO")
                time.sleep(0.5)

                # Aguardar o input ficar visível e preencher via JavaScript
                try:
                    # Aguardar o input aparecer
                    wait.until(lambda d: input_quantidade.is_displayed() or input_quantidade.get_attribute("style") != "display: none;")
                    
                    # Limpar e preencher o campo
                    driver.execute_script("""
                        arguments[0].value = '';
                        arguments[0].value = arguments[1];
                        arguments[0].dispatchEvent(new Event('input', { bubbles: true }));
                        arguments[0].dispatchEvent(new Event('change', { bubbles: true }));
                    """, input_quantidade, quantidade_valor)
                    
                    log_message(f"✍️ Definiu quantidade como '{quantidade_valor}' via JS", "SUCCESS")
                    time.sleep(0.5)
                    
                    # Clicar fora para confirmar a edição
                    driver.execute_script("document.body.click();")
                    time.sleep(0.3)
                    
                    # Verificar se o valor foi definido
                    valor_definido = input_quantidade.get_attribute("value")
                    if valor_definido == quantidade_valor:
                        log_message(f"✅ Valor de quantidade confirmado: '{valor_definido}'", "SUCCESS")
                    else:
                        log_message(f"⚠️ Valor não foi definido corretamente. Esperado: '{quantidade_valor}', Atual: '{valor_definido}'", "WARNING")
                        
                except Exception as input_error:
                    log_message(f"⚠️ Erro ao preencher input de quantidade: {input_error}", "WARNING")
            else:
                log_message("⚠️ Campo de quantidade não encontrado ou não visível", "WARNING")

        except Exception as e:
            log_message(f"⚠️ Erro ao definir quantidade de fragmentos: {e}", "WARNING")

    def definir_quantidade_blocos(self, driver, wait):
        """Define a quantidade de blocos usando JavaScript melhorado"""
        try:
            log_message("📝 Definindo quantidade de blocos como: 1", "INFO")
            
            # Procurar pelos campos de quantidade de blocos na tabela de fragmentos
            script = """
            // Procurar especificamente por campos de quantidade de blocos na tabela de fragmentos
            var tbody = document.getElementById('tdRegiao');
            if (tbody) {
                var inputs = tbody.querySelectorAll('input[name*="quantidadeBlocos_"]');
                for (var i = 0; i < inputs.length; i++) {
                    var input = inputs[i];
                    var parentTd = input.closest('td');
                    if (parentTd) {
                        var ancora = parentTd.querySelector('a[class*="table-editable-ancora"]');
                        if (ancora && ancora.offsetParent !== null) {
                            return {element: ancora, input: input};
                        }
                    }
                }
            }
            return null;
            """
            resultado_blocos = driver.execute_script(script)
            
            if resultado_blocos:
                campo_blocos = resultado_blocos['element']
                input_blocos = resultado_blocos['input']
                
                # Verificar se já tem o valor correto
                valor_atual = input_blocos.get_attribute("value")
                if valor_atual == "1":
                    log_message("✅ Quantidade de blocos já está definida como '1' - pulando", "SUCCESS")
                    return
                
                # Clicar na âncora para abrir o campo
                driver.execute_script("arguments[0].click();", campo_blocos)
                log_message("🔍 Clicou no campo de quantidade de blocos para editar", "INFO")
                time.sleep(0.5)

                # Aguardar o input ficar visível e preencher via JavaScript
                try:
                    # Aguardar o input aparecer
                    wait.until(lambda d: input_blocos.is_displayed() or input_blocos.get_attribute("style") != "display: none;")
                    
                    # Limpar e preencher o campo
                    driver.execute_script("""
                        arguments[0].value = '';
                        arguments[0].value = '1';
                        arguments[0].dispatchEvent(new Event('input', { bubbles: true }));
                        arguments[0].dispatchEvent(new Event('change', { bubbles: true }));
                    """, input_blocos)
                    
                    log_message("✍️ Definiu quantidade de blocos como '1' via JS", "SUCCESS")
                    time.sleep(0.5)
                    
                    # Clicar fora para confirmar a edição
                    driver.execute_script("document.body.click();")
                    time.sleep(0.3)
                    
                    # Verificar se o valor foi definido
                    valor_definido = input_blocos.get_attribute("value")
                    if valor_definido == "1":
                        log_message("✅ Valor de quantidade de blocos confirmado: '1'", "SUCCESS")
                    else:
                        log_message(f"⚠️ Valor não foi definido corretamente. Esperado: '1', Atual: '{valor_definido}'", "WARNING")
                        
                except Exception as input_error:
                    log_message(f"⚠️ Erro ao preencher input de quantidade de blocos: {input_error}", "WARNING")
            else:
                log_message("⚠️ Campo de quantidade de blocos não encontrado ou não visível", "WARNING")

        except Exception as e:
            log_message(f"⚠️ Erro ao definir quantidade de blocos: {e}", "WARNING")

    def definir_descricao_auxiliar(self, driver, wait):
        """Define uma descrição auxiliar padrão se necessário"""
        try:
            log_message("📝 Verificando descrição auxiliar", "INFO")
            
            # Procurar pelos campos de descrição auxiliar na tabela de fragmentos
            script = """
            // Procurar especificamente por campos de descrição auxiliar na tabela de fragmentos
            var tbody = document.getElementById('tdRegiao');
            if (tbody) {
                var inputs = tbody.querySelectorAll('input[name*="descricaoAuxiliar_"]');
                for (var i = 0; i < inputs.length; i++) {
                    var input = inputs[i];
                    var parentTd = input.closest('td');
                    if (parentTd) {
                        var ancora = parentTd.querySelector('a[class*="table-editable-ancora"]');
                        if (ancora && ancora.offsetParent !== null) {
                            return {element: ancora, input: input, text: ancora.textContent};
                        }
                    }
                }
            }
            return null;
            """
            resultado_descricao = driver.execute_script(script)
            
            if resultado_descricao:
                campo_descricao = resultado_descricao['element']
                input_descricao = resultado_descricao['input']
                texto_atual = resultado_descricao['text']
                
                # Se já tem uma descrição (não é "Vazio"), manter
                if texto_atual and texto_atual.strip() != "Vazio":
                    log_message(f"✅ Descrição auxiliar já preenchida: '{texto_atual}' - mantendo", "SUCCESS")
                    return
                
                # Se está vazio, pode deixar vazio mesmo (é opcional)
                log_message("✅ Descrição auxiliar está vazia - mantendo vazio (opcional)", "SUCCESS")
            else:
                log_message("⚠️ Campo de descrição auxiliar não encontrado", "WARNING")

        except Exception as e:
            log_message(f"⚠️ Erro ao verificar descrição auxiliar: {e}", "WARNING")

    def salvar_fragmentos(self, driver, wait):
        """Clica no botão Salvar dos fragmentos"""
        try:
            # Aguardar o botão estar presente e clicável
            botao_salvar_fragmentos = wait.until(
                EC.element_to_be_clickable((By.XPATH, "//a[contains(@class, 'btn-primary') and contains(@data-url, '/macroscopia/saveMacroscopiaFragAjax')]"))
            )
            
            # Verificar se o botão está visível
            if not botao_salvar_fragmentos.is_displayed():
                log_message("⚠️ Botão salvar fragmentos não está visível", "WARNING")
                return
            
            # Rolar até o botão para garantir visibilidade
            driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", botao_salvar_fragmentos)
            time.sleep(1)  # Aumentar tempo de espera
            
            # Verificar se há elementos sobrepostos e aguardar eles desaparecerem
            try:
                # Aguardar elementos sobrepostos desaparecerem (como dropdowns, tooltips, etc.)
                WebDriverWait(driver, 3).until_not(
                    EC.presence_of_element_located((By.XPATH, "//li[contains(@class, 'dropdown-menu') or contains(@class, 'show')]"))
                )
            except:
                pass  # Se não houver elementos sobrepostos, continua
            
            # Usar função robusta para clicar
            if not self.clicar_elemento_robusto(driver, wait, botao_salvar_fragmentos, "Salvar fragmentos"):
                raise Exception("Não foi possível clicar no botão Salvar fragmentos")
            
            # Aguardar que o spinner desapareça após salvar
            self.aguardar_spinner_desaparecer(driver, wait, timeout=15)
            
        except Exception as e:
            log_message(f"⚠️ Erro ao salvar fragmentos: {e}", "WARNING")
            # Tentar encontrar o botão por outras formas
            try:
                # Tentar por título
                botao_titulo = wait.until(
                    EC.element_to_be_clickable((By.XPATH, "//a[@title='Salvar' and contains(@class, 'btn-primary')]"))
                )
                # Usar função robusta para clicar
                if self.clicar_elemento_robusto(driver, wait, botao_titulo, "Salvar fragmentos (por título)"):
                    log_message("💾 Clicou em Salvar fragmentos (por título)", "SUCCESS")
                
                self.aguardar_spinner_desaparecer(driver, wait, timeout=15)
                return
            except:
                pass
            
            try:
                # Tentar por texto do botão
                botao_texto = wait.until(
                    EC.element_to_be_clickable((By.XPATH, "//a[contains(@class, 'btn-primary') and contains(text(), 'Salvar')]"))
                )
                # Usar função robusta para clicar
                if self.clicar_elemento_robusto(driver, wait, botao_texto, "Salvar fragmentos (por texto)"):
                    log_message("💾 Clicou em Salvar fragmentos (por texto)", "SUCCESS")
                
                self.aguardar_spinner_desaparecer(driver, wait, timeout=15)
                return
            except:
                pass
            
            log_message(f"❌ Não foi possível encontrar o botão Salvar fragmentos: {e}", "ERROR")
            raise

    def preencher_campos_pre_envio(self, driver, wait, mascara, qtd_frag, qtd_frag2):
        """Preenche todos os campos necessários antes de enviar para próxima etapa"""
        try:
            log_message("📝 Iniciando preenchimento dos campos pré-envio...", "INFO")
            
            # Aguardar que a página esteja estável
            self.aguardar_pagina_estavel(driver, wait)
            
            # Verificar se estamos na página correta
            try:
                fragmentos_container = driver.find_element(By.ID, "fragmentosContainer")
                if not fragmentos_container.is_displayed():
                    log_message("⚠️ Container de fragmentos não está visível", "WARNING")
                    return
            except:
                log_message("⚠️ Container de fragmentos não encontrado", "WARNING")
                return
            
            # Verificar se há elementos interativos antes de prosseguir
            try:
                elementos_interativos = driver.find_elements(By.XPATH, "//a[contains(@class, 'table-editable-ancora')]")
                if not elementos_interativos:
                    log_message("⚠️ Nenhum elemento interativo encontrado", "WARNING")
                    return
                log_message(f"🔍 Encontrados {len(elementos_interativos)} elementos interativos", "INFO")
            except:
                log_message("⚠️ Erro ao verificar elementos interativos", "WARNING")
                return
            
            # Debug: mostrar estado atual da tabela
            try:
                debug_script = """
                var tbody = document.getElementById('tdRegiao');
                if (tbody) {
                    var inputs = tbody.querySelectorAll('input[name*="_"]');
                    var result = [];
                    for (var i = 0; i < inputs.length; i++) {
                        var input = inputs[i];
                        result.push({
                            name: input.name,
                            value: input.value,
                            type: input.type
                        });
                    }
                    return result;
                }
                return [];
                """
                campos_debug = driver.execute_script(debug_script)
                log_message(f"🔍 DEBUG - Campos na tabela: {len(campos_debug)}", "INFO")
                for campo in campos_debug:
                    log_message(f"  - {campo['name']}: '{campo['value']}'", "INFO")
            except Exception as debug_error:
                log_message(f"⚠️ Erro no debug da tabela: {debug_error}", "WARNING")
            
            # 1. Definir grupo baseado na máscara - SEMPRE EXECUTAR
            log_message(f"📝 Definindo grupo para máscara: {mascara}", "INFO")
            try:
                self.definir_grupo_baseado_mascara(driver, wait, mascara)
                self.aguardar_pagina_estavel(driver, wait, timeout=3)
            except Exception as e:
                log_message(f"⚠️ Erro ao definir grupo: {e}", "WARNING")
            
            # 2. Definir representação como "Seção" - SEMPRE EXECUTAR
            log_message("📝 Definindo representação como Seção", "INFO")
            try:
                self.definir_representacao_secao(driver, wait)
                self.aguardar_pagina_estavel(driver, wait, timeout=3)
            except Exception as e:
                log_message(f"⚠️ Erro ao definir representação: {e}", "WARNING")

            if mascara and mascara.upper() in ['A/C2F', 'A/I2F', 'A/P2F', 'G/E2F', 'G/P2F']:
                log_message("📝 Máscara de 2 frascos detectada ", "INFO")

                log_message("⌨️ Executando ALT + M para adicionar nova linha", "INFO")
                try:
                    actions = ActionChains(driver)
                    actions.key_down(Keys.ALT).send_keys('m').key_up(Keys.ALT).perform()
                    log_message("✅ Atalho ALT + M executado", "SUCCESS")
                    time.sleep(0.5)
                    self.aguardar_pagina_estavel(driver, wait, timeout=3)
                except Exception as e:
                    log_message(f"⚠️ Erro ao executar ALT + M: {e}", "WARNING")

                try:
                    self.definir_regiao_2frascos(driver, wait, mascara)
                    self.aguardar_pagina_estavel(driver, wait, timeout=3)
                except Exception as e:
                    log_message(f"⚠️ Erro ao definir região (2 frascos): {e}", "WARNING")

                try:
                    self.definir_quantidade_fragmentos_2frascos(driver, wait, mascara, qtd_frag, qtd_frag2)
                    self.aguardar_pagina_estavel(driver, wait, timeout=3)
                except Exception as e:
                    log_message(f"⚠️ Erro ao definir quantidade (2 frascos): {e}", "WARNING")

                try:
                    self.definir_quantidade_blocos_2frascos(driver, wait)
                    self.aguardar_pagina_estavel(driver, wait, timeout=3)
                except Exception as e:
                    log_message(f"⚠️ Erro ao definir quantidade de blocos (2 frascos): {e}", "WARNING")

            else:
                log_message("📝 Máscara de 1 frasco detectada ", "INFO")
                try:
                    self.definir_regiao_gastrica(driver, wait, mascara)
                    self.aguardar_pagina_estavel(driver, wait, timeout=3)
                except Exception as e:
                    log_message(f"⚠️ Erro ao definir região: {e}", "WARNING")

                try:
                    self.definir_quantidade_fragmentos(driver, wait, mascara, qtd_frag)
                    self.aguardar_pagina_estavel(driver, wait, timeout=3)
                except Exception as e:
                    log_message(f"⚠️ Erro ao definir quantidade: {e}", "WARNING")

                try:
                    self.definir_quantidade_blocos(driver, wait)
                    self.aguardar_pagina_estavel(driver, wait, timeout=3)
                except Exception as e:
                    log_message(f"⚠️ Erro ao definir quantidade de blocos: {e}", "WARNING")
            
            # 6. Verificar descrição auxiliar (opcional)
            try:
                self.definir_descricao_auxiliar(driver, wait)
                self.aguardar_pagina_estavel(driver, wait, timeout=2)
            except Exception as e:
                log_message(f"⚠️ Erro ao verificar descrição auxiliar: {e}", "WARNING")
            
            log_message("✅ Campos pré-envio preenchidos com sucesso!", "SUCCESS")
            
        except Exception as e:
            log_message(f"⚠️ Erro no preenchimento dos campos pré-envio: {e}", "WARNING")
            log_message("⚠️ Continuando com o envio para próxima etapa", "WARNING")

    def enviar_proxima_etapa(self, driver, wait):
        """Clica no botão de enviar para próxima etapa"""
        try:
            # Aguardar que a página esteja estável primeiro
            self.aguardar_pagina_estavel(driver, wait)
            
            # Aguardar que o spinner desapareça
            self.aguardar_spinner_desaparecer(driver, wait)
            
            # Tentar encontrar o botão
            botao_enviar = wait.until(
                EC.element_to_be_clickable((By.ID, "btn-enviar-proxima-etapa"))
            )
            
            # Verificar se o botão está realmente clicável
            if not botao_enviar.is_displayed() or not botao_enviar.is_enabled():
                log_message("⚠️ Botão não está visível ou habilitado", "WARNING")
                raise Exception("Botão não está interativo")
            
            # Tentar clicar via JavaScript primeiro
            try:
                driver.execute_script("arguments[0].click();", botao_enviar)
                log_message("➡️ Clicou em Enviar para próxima etapa via JS", "INFO")
            except:
                # Se JavaScript falhar, tentar clique normal
                botao_enviar.click()
                log_message("➡️ Clicou em Enviar para próxima etapa", "INFO")
            
            # Aguardar processamento
            time.sleep(2)
            
            # Verificar se apareceu algum modal ou erro
            try:
                # Verificar se apareceu modal de assinatura
                modal_assinatura = driver.find_element(By.ID, "assinatura")
                if modal_assinatura.is_displayed():
                    log_message("📋 Modal de assinatura detectado", "INFO")
                    return {'status': 'aguardando_assinatura', 'detalhes': 'Modal de assinatura aberto'}
            except:
                pass
            
            # Verificar se há erros
            try:
                erros = driver.find_elements(By.CSS_SELECTOR, ".alert-danger, .error-message")
                if erros:
                    erro_texto = erros[0].text
                    log_message(f"⚠️ Erro detectado: {erro_texto}", "WARNING")
                    return {'status': 'erro', 'detalhes': erro_texto}
            except:
                pass
            
            log_message("✅ Envio para próxima etapa realizado com sucesso", "SUCCESS")
            return {'status': 'sucesso', 'detalhes': 'Enviado para próxima etapa'}
            
        except Exception as e:
            log_message(f"Erro ao enviar para próxima etapa: {e}", "ERROR")
            
            # Tentar fechar spinners que possam estar bloqueando
            try:
                driver.execute_script("""
                    var spinners = document.querySelectorAll('.loadModal, .spinner, [class*="loading"]');
                    spinners.forEach(function(spinner) {
                        spinner.style.display = 'none';
                    });
                """)
                log_message("🔧 Spinners fechados via JavaScript", "INFO")
            except:
                pass
            
            raise

    def assinar_com_george(self, driver, wait):
        """Faz o processo de assinatura com Dr. George"""
        try:
            # Aguardar o modal de assinatura aparecer
            wait.until(EC.presence_of_element_located((By.ID, "assinatura")))
            log_message("📋 Modal de assinatura aberto", "INFO")
            
            # Encontrar e clicar no checkbox do Dr. George (value="2173")
            checkbox_george = wait.until(
                EC.element_to_be_clickable((By.XPATH, "//input[@type='checkbox' and @value='2173']"))
            )
            checkbox_george.click()
            log_message("✅ Checkbox do Dr. George marcado", "INFO")
            time.sleep(1)
            
            # Aguardar o campo de senha aparecer e digitar a senha
            campo_senha = wait.until(
                EC.presence_of_element_located((By.NAME, "senha_2173"))
            )
            campo_senha.send_keys("1323")
            log_message("🔐 Senha digitada", "INFO")
            time.sleep(1)
            
            # Clicar no botão Assinar
            botao_assinar = wait.until(
                EC.element_to_be_clickable((By.ID, "salvarAss"))
            )
            botao_assinar.click()
            log_message("✍️ Clicou em Assinar", "INFO")
            time.sleep(1.5)
            
        except Exception as e:
            log_message(f"Erro no processo de assinatura: {e}", "ERROR")
            raise

    def run(self, params: dict):
        username = params.get("username")
        password = params.get("password")
        excel_file = params.get("excel_file")
        cancel_flag = params.get("cancel_flag")
        headless_mode = params.get("headless_mode")

        try:
            # Lê os dados dos exames da planilha (código e máscara)
            dados_exames = self.get_dados_exames(excel_file)
            if not dados_exames:
                messagebox.showerror("Erro", "Nenhum dado de exame encontrado na planilha.")
                return
            
            log_message(f"Encontrados {len(dados_exames)} exames para processar", "INFO")
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao ler o Excel: {e}")
            return

        url = os.getenv("SYSTEM_URL", "https://dap.pathoweb.com.br/login/auth")
        driver = None
        resultados = []
        
        try:
            driver = BrowserFactory.create_chrome(headless=headless_mode)
            wait = WebDriverWait(driver, 10)
            
            log_message("Iniciando automação de macroscopia gástrica...", "INFO")
            
            # Login
            log_message("Fazendo login...", "INFO")
            driver.get(url)
            
            # Aguardar página carregar completamente
            wait.until(EC.presence_of_element_located((By.ID, "username")))
            
            username_field = driver.find_element(By.ID, "username")
            username_field.clear()
            username_field.send_keys(username)
            
            password_field = driver.find_element(By.ID, "password")
            password_field.clear()
            password_field.send_keys(password)
            
            submit_button = driver.find_element(By.CSS_SELECTOR, "button[type='submit']")
            submit_button.click()

            log_message("Verificando se precisa navegar para módulo de exames...", "INFO")
            current_url = driver.current_url

            if current_url == "https://dap.pathoweb.com.br/" or "trocarModulo" in current_url:
                log_message("Detectada tela de seleção de módulos - navegando para módulo de exames...", "INFO")
                try:
                    modulo_link = wait.until(
                        EC.element_to_be_clickable((By.CSS_SELECTOR, "a[href='/site/trocarModulo?modulo=1']")))
                    modulo_link.click()
                    time.sleep(2)
                    log_message("✅ Navegação para módulo de exames realizada", "SUCCESS")
                except Exception as e:
                    log_message(f"⚠️ Erro ao navegar para módulo: {e}", "WARNING")
                    # Tentar navegar diretamente pela URL como fallback
                    driver.get("https://dap.pathoweb.com.br/moduloExame/index")
                    time.sleep(2)
                    log_message("🔄 Navegação direta para módulo realizada", "INFO")

            elif "moduloExame" in current_url:
                log_message("✅ Já está no módulo de exames - pulando navegação", "SUCCESS")
            else:
                log_message(f"⚠️ URL inesperada detectada: {current_url}", "WARNING")
                # Tentar navegar diretamente como fallback
                driver.get("https://dap.pathoweb.com.br/moduloExame/index")
                time.sleep(2)
                log_message("🔄 Navegação direta para módulo realizada (fallback)", "INFO")
            
            # Fechar modal se aparecer
            try:
                modal_close_button = driver.find_element(By.CSS_SELECTOR, "#mensagemParaClienteModal .modal-footer button")
                if modal_close_button.is_displayed():
                    modal_close_button.click()
            except Exception:
                pass

            log_message("✅ Login realizado com sucesso. Iniciando processamento dos exames.", "SUCCESS")
            
            # Processar cada exame da planilha
            for i, exame_data in enumerate(dados_exames, 1):
                if cancel_flag and cancel_flag.is_set():
                    log_message("Execução cancelada pelo usuário.", "WARNING")
                    break
                
                codigo = exame_data['codigo']
                mascara = exame_data['mascara']
                qtd_frag = exame_data['qtd_frag']
                qtd_frag_original = exame_data['qtd_frag_original']
                md1 = exame_data['md1']
                md2 = exame_data['md2']
                md3 = exame_data['md3']
                qtd_frag2 = exame_data['qtd_frag2']
                qtd_frag2_original = exame_data['qtd_frag2_original']
                md4 = exame_data['md4']
                md5 = exame_data['md5']
                md6 = exame_data['md6']
                responsavel_macro = exame_data['responsavel_macro']
                data_fixacao = exame_data['data_fixacao']

                log_message(f"\n➡️ Processando exame {i}/{len(dados_exames)}: {codigo} (máscara: {mascara})", "INFO")
                
                try:
                    # Verificar se o browser ainda está ativo
                    if not self.verificar_sessao_browser(driver):
                        log_message("🔄 Recriando browser devido à sessão perdida...", "WARNING")
                        try:
                            driver.quit()
                        except:
                            pass
                        
                        # Recriar browser e fazer login novamente
                        driver = BrowserFactory.create_chrome(headless=headless_mode)
                        wait = WebDriverWait(driver, 10)
                        
                        # Fazer login novamente
                        log_message("🔄 Fazendo login novamente...", "INFO")
                        driver.get(url)
                        
                        # Aguardar página carregar completamente
                        wait.until(EC.presence_of_element_located((By.ID, "username")))
                        time.sleep(2)
                        
                        username_field = driver.find_element(By.ID, "username")
                        username_field.clear()
                        username_field.send_keys(username)
                        
                        password_field = driver.find_element(By.ID, "password")
                        password_field.clear()
                        password_field.send_keys(password)
                        
                        submit_button = driver.find_element(By.CSS_SELECTOR, "button[type='submit']")
                        submit_button.click()
                        
                        log_message("🔄 Navegando para módulo de exames novamente...", "INFO")
                        
                        # Navegar para o módulo de exames (módulo 1)
                        modulo_link = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "a[href='/site/trocarModulo?modulo=1']")))
                        modulo_link.click()
                        time.sleep(2.5)
                        
                        # Fechar modal se aparecer
                        try:
                            modal_close_button = driver.find_element(By.CSS_SELECTOR, "#mensagemParaClienteModal .modal-footer button")
                            if modal_close_button.is_displayed():
                                modal_close_button.click()
                                time.sleep(1)
                        except Exception:
                            pass
                        
                        log_message("✅ Browser recriado e login realizado novamente", "SUCCESS")
                    
                    # Processar este exame específico
                    resultado = self.processar_exame(
                        driver, wait,
                        codigo, mascara,
                        qtd_frag, qtd_frag_original,
                        md1, md2, md3,
                        qtd_frag2, qtd_frag2_original,
                        md4, md5, md6,
                        responsavel_macro, data_fixacao
                    )
                    resultados.append({
                        'codigo': codigo,
                        'mascara': mascara,
                        'qtd_frag': qtd_frag,
                        'md1': md1,
                        'md2': md2,
                        'md3': md3,
                        'qtd_frag2': qtd_frag2,
                        'md4': md4,
                        'md5': md5,
                        'md6': md6,
                        'status': resultado['status'],
                        'detalhes': resultado.get('detalhes', '')
                    })
                    
                except Exception as e:
                    log_message(f"❌ Erro ao processar exame {codigo}: {e}", "ERROR")
                    resultados.append({
                        'codigo': codigo,
                        'mascara': mascara,
                        'qtd_frag': qtd_frag,
                        'md1': md1,
                        'md2': md2,
                        'md3': md3,
                        'qtd_frag2': qtd_frag2,
                        'md4': md4,
                        'md5': md5,
                        'md6': md6,
                        'status': 'erro',
                        'detalhes': str(e)
                    })
            
            # Mostrar resumo final
            self.mostrar_resumo_final(resultados)
            
        except Exception as e:
            log_message(f"❌ Erro durante a automação: {e}", "ERROR")
            messagebox.showerror("Erro", f"❌ Erro durante a automação:\n{str(e)[:200]}...")
        finally:
            if driver:
                try:
                    driver.quit()
                    log_message("Browser fechado", "INFO")
                except Exception as quit_error:
                    log_message(f"Erro ao fechar browser: {quit_error}", "WARNING")

    def processar_exame(self, driver, wait, codigo, mascara, qtd_frag, qtd_frag_original, md1, md2, md3, qtd_frag2, qtd_frag2_original, md4, md5, md6, responsavel_macro, data_fixacao):
        """Processa um exame individual"""
        try:
            # Verificar se a sessão do browser ainda está ativa
            if not self.verificar_sessao_browser(driver):
                raise Exception("Sessão do browser perdida - necessário reiniciar")
            
            # Aguardar e encontrar o campo de código de barras diretamente pelo placeholder (mais confiável)
            try:
                campo_codigo = wait.until(EC.presence_of_element_located((By.XPATH, "//input[@placeholder='Leitor de código de barras']")))
                log_message("✅ Campo de código encontrado", "INFO")
            except:
                # Fallback para ID se placeholder não funcionar
                campo_codigo = wait.until(EC.presence_of_element_located((By.ID, "inputSearchCodBarra")))
                log_message("✅ Campo de código encontrado pelo ID", "INFO")

            campo_codigo.clear()
            campo_codigo.send_keys(codigo)
            log_message(f"✍️ Código '{codigo}' digitado no campo", "SUCCESS")

            # Clicar no botão de pesquisar (consultarExameBarraAbrirPorBarCode)
            try:
                botao_pesquisar = wait.until(EC.element_to_be_clickable((By.ID, "consultarExameBarraAbrirPorBarCode")))
                botao_pesquisar.click()
                log_message("🔍 Clicou no botão de pesquisar exame", "SUCCESS")
            except Exception as e:
                log_message(f"⚠️ Não foi possível clicar no botão de pesquisar: {e}", "WARNING")
                raise

            # Aguardar div de andamento aparecer
            return self.aguardar_e_processar_andamento(
                driver, wait,
                mascara,
                qtd_frag, qtd_frag_original,
                md1, md2, md3,
                qtd_frag2, qtd_frag2_original,
                md4, md5, md6,
                responsavel_macro, data_fixacao
            )

        except Exception as e:
            error_message = str(e)
            log_message(f"Erro ao processar exame {codigo}: {error_message}", "ERROR")
            
            # Verificar se é erro de sessão inválida
            if "invalid session id" in error_message.lower():
                log_message("❌ Erro de sessão inválida detectado", "ERROR")
                return {'status': 'erro_sessao', 'detalhes': 'Sessão do browser perdida'}
            
            # Screenshot do erro para outros tipos de erro
            try:
                screenshot_path = f"erro_exame_{codigo}_{int(time.time())}.png"
                driver.save_screenshot(screenshot_path)
                log_message(f"Screenshot do erro salvo em: {screenshot_path}", "INFO")
            except:
                pass
            return {'status': 'erro', 'detalhes': error_message}

    def aguardar_e_processar_andamento(self, driver, wait, mascara, qtd_frag, qtd_frag_original, md1, md2, md3, qtd_frag2, qtd_frag2_original, md4, md5, md6, responsavel_macro, data_fixacao):
        """Aguarda a div de andamento e processa o exame"""
        # Aguardar div de andamento aparecer (otimizado)
        try:
            wait.until(EC.presence_of_element_located((By.ID, "divAndamentoExame")))
            log_message("📋 Div de andamento do exame encontrada!", "SUCCESS")
            time.sleep(0.5)  # Reduzido de 2 para 0.5
        except:
            log_message("⚠️ Div de andamento não apareceu no tempo esperado", "WARNING")
            return {'status': 'sem_andamento', 'detalhes': 'Exame não encontrado ou não carregou'}
        
        # Processar conclusão diretamente sem verificar SVG
        log_message("✅ Exame carregado - iniciando processo de conclusão", "SUCCESS")
        return self.processar_conclusao_completa(
            driver, wait,
            mascara,
            qtd_frag, qtd_frag_original,
            md1, md2, md3,
            qtd_frag2, qtd_frag2_original,
            md4, md5, md6,
            responsavel_macro, data_fixacao
        )

    def processar_conclusao_completa(self, driver, wait, mascara, qtd_frag, qtd_frag_original, md1, md2, md3, qtd_frag2, qtd_frag2_original, md4, md5, md6, responsavel_macro, data_fixacao):
        try:
            # 1. Selecionar responsável pela macroscopia
            self.selecionar_responsavel_macroscopia(driver, wait, responsavel_macro)

            # 2. Auxiliar da Macroscopia - não precisa alterar, já vem preenchido no login
            # self.selecionar_auxiliar_macroscopia(driver, wait)  # COMENTADO - campo já vem preenchido automaticamente
            
            # 3. Definir data de fixação correta
            self.definir_data_fixacao(driver, wait, data_fixacao)

            # 4. Definir hora 18:00
            self.definir_hora_fixacao(driver, wait)
            
            # 5. Digitar a máscara e buscar (se houver)
            if mascara:
                self.digitar_mascara_e_buscar(driver, wait, mascara)
            else:
                log_message("⚠️ Nenhuma máscara encontrada, pulando busca", "WARNING")
            
            # 6. Abrir modal de variáveis e preencher campos (opcional)
            try:
                self.abrir_modal_variaveis_e_preencher(
                    driver, wait,
                    mascara,
                    qtd_frag, qtd_frag_original,
                    md1, md2, md3,
                    qtd_frag2, qtd_frag2_original,
                    md4, md5, md6
                )
            except Exception as var_error:
                log_message(f"⚠️ Erro no modal de variáveis: {var_error}", "WARNING")
                log_message("⚠️ Continuando o processo sem as variáveis", "WARNING")
            
            # 7. Salvar macroscopia
            self.salvar_macroscopia(driver, wait)
            
            # 8. Preencher campos necessários antes de enviar para próxima etapa
            self.preencher_campos_pre_envio(driver, wait, mascara, qtd_frag, qtd_frag2)

            # 9. Salvar fragmentos
            self.salvar_fragmentos(driver, wait)
            
            # 10. Enviar para próxima etapa
            resultado_envio = self.enviar_proxima_etapa(driver, wait)
            
            # Verificar o resultado do envio
            if resultado_envio.get('status') == 'aguardando_assinatura':
                log_message("📋 Modal de assinatura aberto - iniciando processo de assinatura", "INFO")
                try:
                    self.assinar_com_george(driver, wait)
                    log_message("🎉 Processo de macroscopia e assinatura finalizado com sucesso!", "SUCCESS")
                    return {'status': 'sucesso', 'detalhes': 'Macroscopia e assinatura processadas com sucesso'}
                except Exception as assinatura_error:
                    log_message(f"⚠️ Erro na assinatura: {assinatura_error}", "WARNING")
                    return {'status': 'erro_assinatura', 'detalhes': str(assinatura_error)}
            elif resultado_envio.get('status') == 'erro':
                log_message(f"⚠️ Erro no envio para próxima etapa: {resultado_envio.get('detalhes')}", "WARNING")
                return {'status': 'erro_envio', 'detalhes': resultado_envio.get('detalhes')}
            else:
                log_message("🎉 Processo de macroscopia finalizado com sucesso!", "SUCCESS")
                return {'status': 'sucesso', 'detalhes': 'Macroscopia processada com sucesso'}
            
        except Exception as e:
            log_message(f"Erro durante processo de macroscopia: {e}", "ERROR")
            return {'status': 'erro_macroscopia', 'detalhes': str(e)}

    def mostrar_resumo_final(self, resultados):
        """Mostra o resumo final do processamento"""
        total = len(resultados)
        sucesso = len([r for r in resultados if r['status'] == 'sucesso'])
        sem_andamento = len([r for r in resultados if r['status'] == 'sem_andamento'])
        erro_sessao = len([r for r in resultados if r['status'] == 'erro_sessao'])
        erros = len([r for r in resultados if 'erro' in r['status'] and r['status'] != 'erro_sessao'])
        
        log_message("\n" + "="*50, "INFO")
        log_message("RESUMO FINAL DO PROCESSAMENTO", "INFO")
        log_message("="*50, "INFO")
        log_message(f"Total de exames: {total}", "INFO")
        log_message(f"✅ Processados com sucesso: {sucesso}", "SUCCESS")
        log_message(f"⚠️ Exames não encontrados: {sem_andamento}", "WARNING")
        log_message(f"🔄 Erros de sessão (browser perdido): {erro_sessao}", "WARNING")
        log_message(f"❌ Outros erros de processamento: {erros}", "ERROR")
        
        # Mostrar detalhes dos erros se houver
        erros_totais = erro_sessao + erros
        if erros_totais > 0:
            log_message("\nDetalhes dos erros:", "ERROR")
            for r in resultados:
                if 'erro' in r['status']:
                    log_message(f"- {r['codigo']}: {r['detalhes']}", "ERROR")
        
        messagebox.showinfo("Processamento Concluído", 
            f"✅ Processamento finalizado!\n\n"
            f"Total: {total}\n"
            f"Sucesso: {sucesso}\n"
            f"Não encontrados: {sem_andamento}\n"
            f"Erros de sessão: {erro_sessao}\n"
            f"Outros erros: {erros}")

def run(params: dict):
    module = MacroGastricaModule()
    module.run(params)
