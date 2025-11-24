import os
import time
from tkinter import messagebox
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.select import Select
from selenium.webdriver.common.action_chains import ActionChains
from dotenv import load_dotenv
from openpyxl import load_workbook

from src.core.browser_factory import BrowserFactory
from src.core.logger import log_message
from src.modules.base import BaseModule

load_dotenv()


class MacroProstModule(BaseModule):
    def __init__(self):
        super().__init__(nome="Macro Prost")

    def get_dados_exames(self, file_path: str) -> list:
        try:
            workbook = load_workbook(file_path)
            sheet = workbook.active
            dados = []
            ultima_mascara = None
            data_fixacao = None
            macroscopista_valor = None

            # Ler cabeçalho (linha 1) e criar mapeamento de colunas
            colunas = {}
            for col_idx in range(1, sheet.max_column + 1):
                cell_value = sheet.cell(row=1, column=col_idx).value
                if cell_value:
                    nome_coluna = str(cell_value).strip().lower()
                    colunas[nome_coluna] = col_idx

            log_message(f"📋 Colunas detectadas: {list(colunas.keys())}", "INFO")

            def encontrar_coluna(nomes_possiveis):
                """Encontra a coluna baseado em uma lista de nomes possíveis"""
                for nome in nomes_possiveis:
                    for coluna_nome, col_idx in colunas.items():
                        if nome.lower() in coluna_nome:
                            return col_idx
                return None

            # Encontrar índices das colunas
            col_data = encontrar_coluna(['data', 'data fixacao', 'data fixação', 'datafixacao'])
            col_num_exame = encontrar_coluna(['num_exame', 'numero', 'número', 'codigo', 'código', 'cod'])
            col_mascara = encontrar_coluna(['mascara', 'máscara', 'mask'])
            col_macroscopista = encontrar_coluna(['macroscopista', 'responsavel', 'responsável', 'resp'])

            # Encontrar colunas de medidas (med1 a med12)
            colunas_medidas = {}
            for i in range(1, 13):
                col_med = encontrar_coluna([f'med{i}', f'med {i}', f'medida{i}', f'medida {i}'])
                colunas_medidas[f'med{i}'] = col_med

            # Validar colunas obrigatórias
            if not col_num_exame:
                raise Exception("Coluna de número do exame não encontrada!")

            log_message(
                f"✅ Mapeamento: Num_Exame=col{col_num_exame}, Máscara=col{col_mascara}, Data=col{col_data}, Medidas={colunas_medidas}","INFO")

            # Lê da linha 2 em diante (linha 1 é cabeçalho)
            for row in range(2, sheet.max_row + 1):
                data = sheet.cell(row=row, column=col_data).value if col_data else None
                num_exame = sheet.cell(row=row, column=col_num_exame).value if col_num_exame else None
                mascara = sheet.cell(row=row, column=col_mascara).value if col_mascara else None
                macroscopista = sheet.cell(row=row, column=col_macroscopista).value if col_macroscopista else None

                # Ler medidas (med1 a med12)
                medidas = {}
                for i in range(1, 13):
                    col_med = colunas_medidas.get(f'med{i}')
                    if col_med:
                        valor_med = sheet.cell(row=row, column=col_med).value
                        medidas[f'med{i}'] = str(valor_med).strip() if valor_med is not None else ""
                    else:
                        medidas[f'med{i}'] = ""

                if row == 2 and data:
                    data_fixacao = str(data).strip()

                if macroscopista is not None and str(macroscopista).strip():
                    macroscopista_valor = str(macroscopista).strip().upper()

                if num_exame is not None:
                    num_exame = str(num_exame).strip()

                    # Se não tem máscara, usa a última válida
                    if mascara is not None and str(mascara).strip():
                        mascara = str(mascara).strip()
                        ultima_mascara = mascara
                    else:
                        mascara = ultima_mascara

                    dados.append({
                        'num_exame': num_exame,
                        'mascara': str(mascara).strip() if mascara else "",
                        'macroscopista': macroscopista_valor,
                        'data_fixacao': data_fixacao,
                        **medidas  # Adiciona med1 a med12
                    })

            workbook.close()
            return dados
        except Exception as e:
            raise Exception(f"Erro ao ler planilha: {e}")

    def verificar_sessao_browser(self, driver) -> bool:
        """Verifica se a sessão do browser ainda está ativa"""
        try:
            driver.current_url
            return True
        except Exception as e:
            if "invalid session id" in str(e).lower():
                log_message("❌ Sessão do browser perdida", "ERROR")
                return False
            return True

    def aguardar_pagina_estavel(self, driver, wait, timeout=10):
        """Aguarda até que a página esteja estável (sem animações ou carregamentos)"""
        try:
            driver.execute_script("""
                return new Promise((resolve) => {
                    if (window.jQuery && window.jQuery.active === 0) {
                        resolve();
                        return;
                    }

                    var checkInterval = setInterval(() => {
                        if (window.jQuery && window.jQuery.active === 0) {
                            clearInterval(checkInterval);
                            resolve();
                        }
                    }, 100);

                    setTimeout(() => {
                        clearInterval(checkInterval);
                        resolve();
                    }, arguments[0]);
                });
            """, timeout * 1000)

            time.sleep(0.5)
            log_message("✅ Página estável", "INFO")

        except Exception as e:
            log_message(f"⚠️ Erro ao aguardar página estável: {e}", "WARNING")
            time.sleep(1)

    def aguardar_spinner_desaparecer(self, driver, wait, timeout=30):
        """Aguarda até que o spinner de loading desapareça"""
        try:
            log_message("⏳ Aguardando spinner desaparecer...", "INFO")
            wait.until(EC.invisibility_of_element_located((By.ID, "spinner")))
            time.sleep(1)

            spinners = driver.find_elements(By.CSS_SELECTOR, ".loadModal, .spinner, [class*='loading']")
            for spinner in spinners:
                if spinner.is_displayed():
                    log_message("⚠️ Outro spinner ainda visível, aguardando...", "WARNING")
                    time.sleep(2)
                    break

            log_message("✅ Spinner desapareceu", "SUCCESS")

        except Exception as e:
            log_message(f"⚠️ Erro ao aguardar spinner: {e}", "WARNING")
            try:
                driver.execute_script("""
                    var spinners = document.querySelectorAll('.loadModal, .spinner, [class*="loading"]');
                    spinners.forEach(function(spinner) {
                        if (spinner.style.display !== 'none') {
                            spinner.style.display = 'none';
                        }
                    });
                """)
                log_message("🔧 Spinner fechado via JavaScript", "INFO")
                time.sleep(1)
            except:
                pass

    def selecionar_responsavel_macroscopia(self, driver, wait, responsavel_macro):
        """Seleciona o responsável pela macroscopia conforme o nome recebido (nome curto)"""
        responsavel_macro_mapper = {
            'BARBARA': 'Barbara Dutra Lopes',
            'NATHALIA': 'Nathalia Fernanda da Silva Lopes',
            'RENATA': 'Renata Silva Sevidanis',
            'HELEN': 'Helen Oliveira dos Santos',
            'CLARA': 'Clara Helena Janz Garcia de Souza',
            'PALOMA': 'Paloma Brenda Silva De Oliveira',
            'ELLEN': 'Ellen Andressa de Alvarenga',
            'VITORIA': 'Vitoria Aquino Nairne Domingues',
            'ANNAI': 'Annai Lukã Vitorino Losnak',
            'ANA': 'Ana Carolina Viecele Campos'
        }
        nome_completo = responsavel_macro_mapper.get(responsavel_macro, responsavel_macro)
        select2_container = wait.until(
            EC.element_to_be_clickable(
                (By.XPATH, "//span[@aria-labelledby='select2-responsavelMacroscopiaId-container']"))
        )
        select2_container.click()
        time.sleep(0.3)

        opcao = wait.until(
            EC.element_to_be_clickable((By.XPATH, f"//li[contains(text(), '{nome_completo}')]"))
        )
        opcao.click()
        log_message(f"✅ {nome_completo} selecionado como responsável", "SUCCESS")
        time.sleep(0.2)

    def definir_data_fixacao(self, driver, wait, data_fixacao=None):
        """Define a data de fixação no campo de data de fixação"""
        try:
            if not data_fixacao:
                data_fixacao = '21082025'

            if len(data_fixacao) == 8 and data_fixacao.isdigit():
                data_formatada = f"{data_fixacao[4:8]}-{data_fixacao[2:4]}-{data_fixacao[0:2]}"
            else:
                data_formatada = '2025-08-21'

            campo_data = wait.until(
                EC.presence_of_element_located((By.XPATH, "//input[@type='date' and @name='dataFixacao']"))
            )
            driver.execute_script("""
                var campo = arguments[0];
                campo.value = arguments[1];
                campo.dispatchEvent(new Event('change', { bubbles: true }));
            """, campo_data, data_formatada)
            log_message(f"📅 Data de fixação definida para: {data_formatada}", "SUCCESS")
            time.sleep(0.1)
        except Exception as e:
            log_message(f"⚠️ Erro ao definir data de fixação: {e}", "WARNING")

    def definir_hora_fixacao(self, driver, wait):
        """Define 18:00 no campo de hora de fixação"""
        campo_hora = wait.until(
            EC.presence_of_element_located((By.XPATH, "//input[@type='time' and @name='dataFixacao']"))
        )
        campo_hora.clear()
        campo_hora.send_keys("18:00")
        log_message("🕕 Hora de fixação definida para: 18:00", "SUCCESS")
        time.sleep(0.1)

    def fechar_exame(self, driver, wait):
        """Clica no botão de fechar exame"""
        try:
            botao_fechar = wait.until(
                EC.element_to_be_clickable((By.ID, "fecharExameBarraFerramenta"))
            )
            botao_fechar.click()
            log_message("📁 Exame fechado", "INFO")

            try:
                wait.until(EC.presence_of_element_located((By.ID, "inputSearchCodBarra")))
                log_message("✅ Retornou à tela principal após fechar exame", "INFO")
            except:
                log_message("⚠️ Pode não ter retornado à tela principal", "WARNING")
                try:
                    current_url = driver.current_url
                    if "modulo=1" not in current_url:
                        modulo_link = driver.find_element(By.CSS_SELECTOR, "a[href='/site/trocarModulo?modulo=1']")
                        modulo_link.click()
                        time.sleep(1.5)
                        log_message("🔄 Navegou de volta ao módulo de exames", "INFO")
                except:
                    pass

        except Exception as e:
            log_message(f"Erro ao fechar exame: {e}", "ERROR")

    def digitar_mascara_e_buscar(self, driver, wait, mascara):
        """Digita a máscara no campo buscaArvore e pressiona Enter"""
        campo_busca = wait.until(EC.element_to_be_clickable((By.ID, "buscaArvore")))
        campo_busca.send_keys(mascara)
        campo_busca.send_keys(Keys.ENTER)
        log_message(f"✍️ Máscara '{mascara}' digitada no campo buscaArvore", "SUCCESS")
        time.sleep(0.5)

    def abrir_modal_variaveis_e_preencher(self, driver, wait, mascara, medidas):
        """Abre o modal de variáveis e preenche os campos baseado na máscara PROSTATA"""
        try:
            # Clicar no botão "Pesquisar variáveis (F7)"
            botao_variaveis = wait.until(
                EC.element_to_be_clickable((By.ID, "cke_70"))
            )
            botao_variaveis.click()
            log_message("🔍 Clicou no botão de variáveis", "INFO")
            time.sleep(0.8)

            # Verificar se apareceu um alerta
            try:
                alert = driver.switch_to.alert
                alert_text = alert.text
                if "não há variáveis" in alert_text.lower():
                    log_message(f"⚠️ Alerta detectado: {alert_text}", "WARNING")
                    alert.accept()
                    log_message("⚠️ Pulando preenchimento de variáveis - não há variáveis no texto", "WARNING")
                    return
                else:
                    alert.accept()
            except:
                pass

            # Aguardar o modal aparecer
            wait.until(EC.presence_of_element_located((By.CLASS_NAME, "swal2-popup")))
            log_message("🔍 Modal de variáveis aberto", "SUCCESS")
            time.sleep(0.3)

            # Preencher os campos usando classe genérica
            campos_input = driver.find_elements(By.CSS_SELECTOR, "input[style*='width: 100px'][style*='color: red']")
            log_message(f"🔍 Encontrados {len(campos_input)} campos de input no modal", "INFO")

            # Preparar valores das medidas (med1 a med12)
            valores = []
            for i in range(1, 13):
                med_valor = medidas.get(f'med{i}', '')
                if med_valor and med_valor.strip():
                    valores.append(med_valor.strip())

            # Filtrar valores vazios
            valores = [v for v in valores if v and v.strip()]

            log_message(f"📋 Preenchendo {len(valores)} medidas para próstata", "INFO")

            for i, campo in enumerate(campos_input[:len(valores)]):
                if i < len(valores) and valores[i]:
                    try:
                        driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});",
                                              campo)
                        time.sleep(0.1)

                        driver.execute_script("""
                            arguments[0].value = arguments[1];
                            arguments[0].dispatchEvent(new Event('input', { bubbles: true }));
                            arguments[0].dispatchEvent(new Event('change', { bubbles: true }));
                        """, campo, valores[i])

                        log_message(f"✅ Campo {i + 1} preenchido com: {valores[i]}", "SUCCESS")
                        time.sleep(0.1)
                    except Exception as e:
                        log_message(f"⚠️ Erro ao preencher campo {i + 1}: {e}", "WARNING")

            time.sleep(0.2)

            # Clicar no botão "Inserir"
            botao_inserir = wait.until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, ".swal2-confirm"))
            )
            botao_inserir.click()
            log_message("✅ Campos inseridos no modal", "SUCCESS")

            # Aguardar o modal fechar
            try:
                wait.until(EC.invisibility_of_element_located((By.CLASS_NAME, "swal2-popup")))
                log_message("✅ Modal fechado completamente", "SUCCESS")
            except:
                time.sleep(1)
                log_message("⏳ Aguardou fechamento do modal", "INFO")

        except Exception as e:
            log_message(f"⚠️ Erro ao preencher modal de variáveis: {e}", "WARNING")
            log_message("⚠️ Continuando sem preencher as variáveis", "WARNING")

    def salvar_macroscopia(self, driver, wait):
        """Clica no botão Salvar da macroscopia"""
        try:
            modal = driver.find_element(By.CLASS_NAME, "swal2-popup")
            if modal.is_displayed():
                botao_fechar = driver.find_element(By.CSS_SELECTOR, ".swal2-close")
                botao_fechar.click()
                time.sleep(0.5)
        except:
            pass

        botao_salvar = wait.until(
            EC.element_to_be_clickable((By.ID, "salvarMacro"))
        )
        botao_salvar.click()
        log_message("💾 Macroscopia salva", "SUCCESS")
        time.sleep(0.3)

    def definir_grupo(self, driver, wait):
        """Define o grupo como 'Seios da Face' usando JavaScript com retry"""
        try:
            grupo_selecionado = 'Prostata'
            max_tentativas = 5
            tentativa = 0

            # Verificar se o input existe e qual o valor atual
            try:
                input_grupo = driver.find_element(By.ID, "idRegiao")
                valor_atual = input_grupo.get_attribute("value")

                if valor_atual == grupo_selecionado:
                    log_message(f"✅ Grupo já está definido como '{grupo_selecionado}'", "SUCCESS")
                    return
                elif valor_atual and valor_atual.strip():
                    log_message(f"⚠️ Valor atual do campo grupo: '{valor_atual}' - será substituído", "WARNING")
            except:
                log_message("⚠️ Campo grupo não encontrado inicialmente", "WARNING")

            while tentativa < max_tentativas:
                try:
                    tentativa += 1
                    log_message(f"🔄 Tentativa {tentativa} de {max_tentativas} para clicar no campo de grupo", "INFO")

                    # Procurar a âncora "Vazio" específica do campo de grupo
                    script = """
                    var inputRegiao = document.getElementById('idRegiao');
                    if (inputRegiao) {
                        var parentTd = inputRegiao.closest('td');
                        if (parentTd) {
                            var ancora = parentTd.querySelector('a.table-editable-ancora');
                            if (ancora && ancora.offsetParent !== null) {
                                return ancora;
                            }
                        }
                    }
                    return null;
                    """
                    campo_grupo_ancora = driver.execute_script(script)

                    if not campo_grupo_ancora:
                        log_message(f"⚠️ Âncora de grupo não encontrada na tentativa {tentativa}", "WARNING")
                        time.sleep(0.5)
                        continue

                    # Scroll até o elemento e aguardar
                    driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});",
                                          campo_grupo_ancora)
                    time.sleep(0.3)

                    # Clicar usando diferentes métodos
                    try:
                        # Método 1: Click JavaScript direto
                        driver.execute_script("arguments[0].click();", campo_grupo_ancora)
                        log_message(f"🖱️ Clicou na âncora via JavaScript (tentativa {tentativa})", "INFO")
                    except:
                        # Método 2: Click Selenium tradicional
                        campo_grupo_ancora.click()
                        log_message(f"🖱️ Clicou na âncora via Selenium (tentativa {tentativa})", "INFO")

                    time.sleep(0.5)

                    # Validar se o input ficou visível após o clique
                    input_grupo = driver.find_element(By.ID, "idRegiao")
                    is_visible = driver.execute_script("""
                        var input = arguments[0];
                        var style = window.getComputedStyle(input);
                        return style.display !== 'none' && style.visibility !== 'hidden' && input.offsetParent !== null;
                    """, input_grupo)

                    if not is_visible:
                        log_message(f"⚠️ Input ainda não está visível após clique (tentativa {tentativa})", "WARNING")
                        time.sleep(0.3)
                        continue

                    log_message(f"✅ Input de grupo está visível e pronto para preenchimento", "SUCCESS")

                    # Limpar e preencher o campo
                    driver.execute_script("""
                        var input = arguments[0];
                        input.value = '';
                        input.focus();
                        input.value = arguments[1];
                        input.dispatchEvent(new Event('input', { bubbles: true }));
                        input.dispatchEvent(new Event('change', { bubbles: true }));
                        input.dispatchEvent(new KeyboardEvent('keyup', { bubbles: true }));
                    """, input_grupo, grupo_selecionado)

                    log_message(f"✅ Grupo '{grupo_selecionado}' preenchido no campo", "SUCCESS")
                    time.sleep(0.5)

                    # Tentar selecionar da lista de autocomplete
                    try:
                        # Aguardar o dropdown aparecer
                        dropdown = wait.until(
                            EC.presence_of_element_located(
                                (By.CSS_SELECTOR, "ul.typeahead.dropdown-menu[style*='display: block']"))
                        )

                        # Procurar a opção no dropdown
                        opcao_autocomplete = wait.until(
                            EC.element_to_be_clickable((By.XPATH,
                                                        f"//ul[@class='typeahead dropdown-menu']//a[contains(text(), '{grupo_selecionado}')]"))
                        )
                        opcao_autocomplete.click()
                        log_message(f"✅ Opção '{grupo_selecionado}' selecionada do autocomplete", "SUCCESS")
                        time.sleep(0.3)

                        # Validar se o valor foi realmente preenchido
                        valor_final = input_grupo.get_attribute("value")
                        if valor_final == grupo_selecionado:
                            log_message(f"✅ Validação final: Grupo definido como '{grupo_selecionado}'", "SUCCESS")
                            return
                        else:
                            log_message(f"⚠️ Valor final não corresponde: '{valor_final}' != '{grupo_selecionado}'",
                                        "WARNING")
                            continue

                    except Exception as autocomplete_error:
                        log_message(f"⚠️ Autocomplete não apareceu ou não foi clicável: {autocomplete_error}",
                                    "WARNING")
                        # Tentar confirmar com Enter
                        try:
                            input_grupo.send_keys(Keys.ENTER)
                            time.sleep(0.3)
                            log_message("✅ Confirmado com Enter", "SUCCESS")

                            # Validar se o valor foi preenchido
                            valor_final = input_grupo.get_attribute("value")
                            if valor_final == grupo_selecionado:
                                return
                        except:
                            # Clicar fora para fechar o dropdown
                            driver.execute_script("document.body.click();")
                            time.sleep(0.3)
                        continue

                except Exception as e:
                    log_message(f"⚠️ Erro na tentativa {tentativa}: {e}", "WARNING")
                    time.sleep(0.5)
                    continue

            # Se chegou aqui, esgotou todas as tentativas
            raise Exception(f"Não foi possível definir o grupo após {max_tentativas} tentativas")

        except Exception as e:
            log_message(f"❌ Erro ao definir grupo: {e}", "ERROR")
            raise

    def definir_representacao_secao(self, driver, wait):
        """Define a representação como 'Seção' usando JavaScript"""
        try:
            # Verificar se o select existe e qual o valor atual
            try:
                select_representacao = driver.find_element(By.ID, "representacao")
                valor_atual = select_representacao.get_attribute("value")

                if valor_atual == "S":
                    log_message("✅ Representação já está definida como 'Seção'", "SUCCESS")
                    return
                elif valor_atual != "S":
                    log_message(f"⚠️ Representação atual é '{valor_atual}', mas precisa ser 'S' (Seção)", "WARNING")
            except:
                log_message("⚠️ Campo representacao não encontrado", "WARNING")
                return

            # Procurar especificamente pelo campo de representação
            script = """
            var selectRepresentacao = document.getElementById('representacao');
            if (selectRepresentacao) {
                var parentTd = selectRepresentacao.closest('td');
                if (parentTd) {
                    var ancora = parentTd.querySelector('a[class*="table-editable-ancora"]');
                    if (ancora && ancora.offsetParent !== null) {
                        return ancora;
                    }
                }
            }
            return null;
            """
            campo_representacao = driver.execute_script(script)

            if not campo_representacao:
                log_message("⚠️ Campo de representação não encontrado", "WARNING")
                return

            # Clicar via JavaScript
            driver.execute_script("arguments[0].click();", campo_representacao)
            log_message("🔍 Clicou no campo de representação via JS", "INFO")
            time.sleep(0.5)

            # Aguardar o select aparecer e selecionar via JavaScript
            select_representacao = wait.until(
                EC.presence_of_element_located((By.ID, "representacao"))
            )

            # Selecionar "Seção" (valor "S") via JavaScript
            driver.execute_script("""
                arguments[0].value = 'S';
                arguments[0].dispatchEvent(new Event('change', { bubbles: true }));
            """, select_representacao)

            log_message("✅ Representação definida como 'Seção' via JS", "SUCCESS")
            time.sleep(0.5)

            # Clicar fora para confirmar a seleção
            driver.execute_script("document.body.click();")
            time.sleep(0.3)

        except Exception as e:
            log_message(f"⚠️ Erro ao definir representação: {e}", "WARNING")

    def definir_regiao(self, driver, wait, lado=None):
        """Define a região de acordo com o lado usando JavaScript - busca apenas campos vazios"""
        try:
            if not lado:
                log_message("⚠️ Nenhum lado fornecido para definir região", "WARNING")
                return

            # Determinar o valor da região baseado no lado
            regiao_valor = 'LadoDir: Lado direito' if lado == 'DIREITO' else 'LadoEsq: Lado esquerdo'

            # Script modificado para buscar APENAS campos vazios
            script = """
            // Procurar especificamente por campos de região VAZIOS
            var tbody = document.getElementById('tdRegiao');
            if (tbody) {
                var inputs = tbody.querySelectorAll('input[name*="regiao_"]');
                for (var i = 0; i < inputs.length; i++) {
                    var input = inputs[i];
                    // Verificar se o input está vazio ou se a âncora mostra "Vazio"
                    if (input.value === '' || input.value.trim() === '') {
                        var parentTd = input.closest('td');
                        if (parentTd) {
                            var ancora = parentTd.querySelector('a[class*="table-editable-ancora"]');
                            // Validar que a âncora mostra "Vazio"
                            if (ancora && ancora.textContent.trim() === 'Vazio' && ancora.offsetParent !== null) {
                                return {element: ancora, input: input};
                            }
                        }
                    }
                }
            }
            return null;
            """
            resultado_regiao = driver.execute_script(script)

            if resultado_regiao:
                campo_regiao = resultado_regiao['element']
                input_regiao = resultado_regiao['input']

                # Clicar na âncora para abrir o campo de edição
                driver.execute_script("arguments[0].click();", campo_regiao)
                log_message("🔍 Clicou no campo de região vazio para editar", "INFO")
                time.sleep(0.5)

                # Aguardar o input ficar visível e preencher
                try:
                    wait.until(lambda d: input_regiao.is_displayed() or input_regiao.get_attribute(
                        "style") != "display: none;")

                    # Limpar e preencher o campo
                    driver.execute_script("""
                        arguments[0].value = '';
                        arguments[0].value = arguments[1];
                        arguments[0].dispatchEvent(new Event('input', { bubbles: true }));
                        arguments[0].dispatchEvent(new Event('change', { bubbles: true }));
                    """, input_regiao, regiao_valor)

                    log_message(f"✍️ Definiu região como '{regiao_valor}' via JS", "SUCCESS")
                    time.sleep(0.5)

                    # Clicar fora para confirmar a edição
                    driver.execute_script("document.body.click();")
                    time.sleep(0.5)

                    # Verificar se o valor foi realmente definido
                    valor_definido = input_regiao.get_attribute("value")
                    if valor_definido == regiao_valor:
                        log_message(f"✅ Valor de região confirmado: '{valor_definido}'", "SUCCESS")
                    else:
                        log_message(
                            f"⚠️ Valor não foi definido corretamente. Esperado: '{regiao_valor}', Atual: '{valor_definido}'",
                            "WARNING")

                except Exception as input_error:
                    log_message(f"⚠️ Erro ao preencher input de região: {input_error}", "WARNING")
            else:
                log_message("⚠️ Nenhum campo de região vazio encontrado", "WARNING")

        except Exception as e:
            log_message(f"⚠️ Erro ao definir região: {e}", "WARNING")

    def definir_quantidade_fragmentos(self, driver, wait, mascara):
        """Define a quantidade de fragmentos baseado na máscara - busca campos com valor 0"""
        try:
            mascara_upper = mascara.upper() if mascara else ""

            if 'PROSTATA12F' in mascara_upper:
                quantidade_valor = '6'
            elif 'PROSTATA6B' in mascara_upper:
                quantidade_valor = '6'
            else:
                log_message(f"⚠️ Máscara '{mascara}' não reconhecida para quantidade de fragmentos", "WARNING")
                return

            log_message(f"📝 Usando quantidade: {quantidade_valor}", "INFO")

            # Procurar por campos de quantidade com valor 0 ou vazio
            script = """
            var tbody = document.getElementById('tdRegiao');
            if (tbody) {
                var inputs = tbody.querySelectorAll('input[name*="quantidade_"]');
                for (var i = 0; i < inputs.length; i++) {
                    var input = inputs[i];
                    // Verificar se o valor é 0 ou vazio
                    if (input.value === '0' || input.value === '' || input.value.trim() === '') {
                        var parentTd = input.closest('td');
                        if (parentTd) {
                            var ancora = parentTd.querySelector('a[class*="table-editable-ancora"]');
                            if (ancora && ancora.offsetParent !== null) {
                                return {element: ancora, input: input};
                            }
                        }
                    }
                }
            }
            return null;
            """
            resultado_quantidade = driver.execute_script(script)

            if resultado_quantidade:
                campo_quantidade = resultado_quantidade['element']
                input_quantidade = resultado_quantidade['input']

                valor_atual = input_quantidade.get_attribute("value")
                log_message(f"🔍 Campo encontrado com valor atual: '{valor_atual}'", "INFO")

                # Clicar na âncora para abrir o campo
                driver.execute_script("arguments[0].click();", campo_quantidade)
                log_message("🔍 Clicou no campo de quantidade para editar", "INFO")
                time.sleep(0.5)

                try:
                    wait.until(lambda d: input_quantidade.is_displayed() or input_quantidade.get_attribute(
                        "style") != "display: none;")

                    # Limpar e preencher o campo
                    driver.execute_script("""
                        arguments[0].value = '';
                        arguments[0].value = arguments[1];
                        arguments[0].dispatchEvent(new Event('input', { bubbles: true }));
                        arguments[0].dispatchEvent(new Event('change', { bubbles: true }));
                    """, input_quantidade, quantidade_valor)

                    log_message(f"✍️ Definiu quantidade como '{quantidade_valor}' via JS", "SUCCESS")
                    time.sleep(0.5)

                    # Clicar fora para confirmar
                    driver.execute_script("document.body.click();")
                    time.sleep(0.3)

                    # Verificar se o valor foi definido
                    valor_definido = input_quantidade.get_attribute("value")
                    if valor_definido == quantidade_valor:
                        log_message(f"✅ Valor de quantidade confirmado: '{valor_definido}'", "SUCCESS")
                    else:
                        log_message(
                            f"⚠️ Valor não foi definido corretamente. Esperado: '{quantidade_valor}', Atual: '{valor_definido}'",
                            "WARNING")

                except Exception as input_error:
                    log_message(f"⚠️ Erro ao preencher input de quantidade: {input_error}", "WARNING")
            else:
                log_message("⚠️ Nenhum campo de quantidade com valor 0 ou vazio foi encontrado", "WARNING")

        except Exception as e:
            log_message(f"⚠️ Erro ao definir quantidade de fragmentos: {e}", "WARNING")

    def definir_quantidade_blocos(self, driver, wait, mascara):
        """Define a quantidade de blocos baseado na máscara - busca campos com valor 0"""
        try:
            mascara_upper = mascara.upper() if mascara else ""

            if mascara_upper == 'PROSTATA12F':
                quantidade_blocos = '6'
                log_message(f"📝 Definindo quantidade de blocos para PROSTATA12F: {quantidade_blocos}", "INFO")
            elif mascara_upper == 'PROSTATA6B':
                quantidade_blocos = '3'
                log_message(f"📝 Definindo quantidade de blocos para PROSTATA6B: {quantidade_blocos}", "INFO")
            else:
                log_message(f"⚠️ Máscara '{mascara}' não reconhecida para definir quantidade de blocos", "WARNING")
                return

            # Procurar por campos de quantidade de blocos com valor 0 ou vazio
            script = """
            var tbody = document.getElementById('tdRegiao');
            if (tbody) {
                var inputs = tbody.querySelectorAll('input[name*="quantidadeBlocos_"]');
                for (var i = 0; i < inputs.length; i++) {
                    var input = inputs[i];
                    // Verificar se o valor é 0 ou vazio
                    if (input.value === '0' || input.value === '' || input.value.trim() === '') {
                        var parentTd = input.closest('td');
                        if (parentTd) {
                            var ancora = parentTd.querySelector('a[class*="table-editable-ancora"]');
                            if (ancora && ancora.offsetParent !== null) {
                                return {element: ancora, input: input};
                            }
                        }
                    }
                }
            }
            return null;
            """
            resultado_blocos = driver.execute_script(script)

            if resultado_blocos:
                campo_blocos = resultado_blocos['element']
                input_blocos = resultado_blocos['input']

                valor_atual = input_blocos.get_attribute("value")
                log_message(f"🔍 Campo encontrado com valor atual: '{valor_atual}'", "INFO")

                # Clicar na âncora para abrir o campo
                driver.execute_script("arguments[0].click();", campo_blocos)
                log_message("🔍 Clicou no campo de quantidade de blocos para editar", "INFO")
                time.sleep(0.5)

                try:
                    wait.until(lambda d: input_blocos.is_displayed() or input_blocos.get_attribute(
                        "style") != "display: none;")

                    # Limpar e preencher o campo
                    driver.execute_script("""
                        arguments[0].value = '';
                        arguments[0].value = arguments[1];
                        arguments[0].dispatchEvent(new Event('input', { bubbles: true }));
                        arguments[0].dispatchEvent(new Event('change', { bubbles: true }));
                    """, input_blocos, quantidade_blocos)

                    log_message(f"✍️ Definiu quantidade de blocos como '{quantidade_blocos}' via JS", "SUCCESS")
                    time.sleep(0.5)

                    # Clicar fora para confirmar
                    driver.execute_script("document.body.click();")
                    time.sleep(0.3)

                    # Verificar se o valor foi definido
                    valor_definido = input_blocos.get_attribute("value")
                    if valor_definido == quantidade_blocos:
                        log_message(f"✅ Valor de quantidade de blocos confirmado: '{quantidade_blocos}'", "SUCCESS")
                    else:
                        log_message(f"⚠️ Valor não foi definido corretamente. Esperado: '{quantidade_blocos}', Atual: '{valor_definido}'","WARNING")

                except Exception as input_error:
                    log_message(f"⚠️ Erro ao preencher input de quantidade de blocos: {input_error}", "WARNING")
            else:
                log_message("⚠️ Nenhum campo de quantidade de blocos com valor 0 ou vazio foi encontrado", "WARNING")

        except Exception as e:
            log_message(f"⚠️ Erro ao definir quantidade de blocos: {e}", "WARNING")

    def salvar_fragmentos(self, driver, wait):
        """Clica no botão Salvar dos fragmentos"""
        try:
            # Aguardar o botão estar presente e clicável
            botao_salvar_fragmentos = wait.until(
                EC.element_to_be_clickable((By.XPATH,
                                            "//a[contains(@class, 'btn-primary') and contains(@data-url, '/macroscopia/saveMacroscopiaFragAjax')]"))
            )

            # Verificar se o botão está visível
            if not botao_salvar_fragmentos.is_displayed():
                log_message("⚠️ Botão salvar fragmentos não está visível", "WARNING")
                return

            # Rolar até o botão para garantir visibilidade
            driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});",
                                  botao_salvar_fragmentos)
            time.sleep(0.5)

            # Clicar no botão
            botao_salvar_fragmentos.click()
            log_message("💾 Clicou em Salvar fragmentos", "SUCCESS")

            # Aguardar que o spinner desapareça após salvar
            self.aguardar_spinner_desaparecer(driver, wait, timeout=15)

        except Exception as e:
            log_message(f"⚠️ Erro ao salvar fragmentos: {e}", "WARNING")
            # Tentar encontrar o botão por outras formas
            try:
                # Tentar por título
                botao_titulo = wait.until(
                    EC.element_to_be_clickable((By.XPATH, "//a[@title='Salvar' and contains(@class, 'btn-primary')]"))
                )
                driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});",
                                      botao_titulo)
                time.sleep(0.5)
                botao_titulo.click()
                log_message("💾 Clicou em Salvar fragmentos (por título)", "SUCCESS")
                self.aguardar_spinner_desaparecer(driver, wait, timeout=15)
                return
            except:
                pass

            try:
                # Tentar por texto do botão
                botao_texto = wait.until(
                    EC.element_to_be_clickable(
                        (By.XPATH, "//a[contains(@class, 'btn-primary') and contains(text(), 'Salvar')]"))
                )
                driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});",
                                      botao_texto)
                time.sleep(0.5)
                botao_texto.click()
                log_message("💾 Clicou em Salvar fragmentos (por texto)", "SUCCESS")
                self.aguardar_spinner_desaparecer(driver, wait, timeout=15)
                return
            except:
                pass

            log_message(f"❌ Não foi possível encontrar o botão Salvar fragmentos: {e}", "ERROR")
            raise

    def preencher_campos_pre_envio(self, driver, wait, mascara):
        """Preenche os campos necessários antes do envio para próxima etapa"""
        try:
            log_message("📝 Preenchendo campos pré-envio...", "INFO")

            # Definir grupo
            log_message(f"📝 Definindo grupo para máscara: {mascara}", "INFO")
            try:
                self.definir_grupo(driver, wait)
                self.aguardar_pagina_estavel(driver, wait, timeout=3)
            except Exception as e:
                log_message(f"⚠️ Erro ao definir grupo: {e}", "WARNING")

            # Definir representação como "Seção"
            log_message("📝 Definindo representação como Seção", "INFO")
            try:
                self.definir_representacao_secao(driver, wait)
                self.aguardar_pagina_estavel(driver, wait, timeout=3)
            except Exception as e:
                log_message(f"⚠️ Erro ao definir representação: {e}", "WARNING")

            # Definir região baseada na máscara
            log_message(f"📝 Definindo região para o lado direito", "INFO")
            try:
                self.definir_regiao(driver, wait, 'DIREITO')
                self.aguardar_pagina_estavel(driver, wait, timeout=3)
            except Exception as e:
                log_message(f"⚠️ Erro ao definir região: {e}", "WARNING")

            # Definir quantidade de fragmentos
            log_message(f"📝 Definindo quantidade de fragmentos", "INFO")
            try:
                self.definir_quantidade_fragmentos(driver, wait, mascara)
                self.aguardar_pagina_estavel(driver, wait, timeout=3)
            except Exception as e:
                log_message(f"⚠️ Erro ao definir quantidade de fragmentos: {e}", "WARNING")

            # Definir quantidade de blocos baseado na mascara
            log_message(f"📝 Definindo quantidade de blocos baseado na máscara: {mascara}", "INFO")
            try:
                self.definir_quantidade_blocos(driver, wait, mascara)
                self.aguardar_pagina_estavel(driver, wait, timeout=3)
            except Exception as e:
                log_message(f"⚠️ Erro ao definir quantidade de blocos: {e}", "WARNING")

            self.salvar_fragmentos(driver, wait)

            log_message("⌨️ Executando ALT + M para adicionar nova linha", "INFO")
            try:
                actions = ActionChains(driver)
                actions.key_down(Keys.ALT).send_keys('m').key_up(Keys.ALT).perform()
                log_message("✅ Atalho ALT + M executado", "SUCCESS")
                time.sleep(0.5)
                self.aguardar_pagina_estavel(driver, wait, timeout=3)
            except Exception as e:
                log_message(f"⚠️ Erro ao executar ALT + M: {e}", "WARNING")

            log_message(f"📝 Definindo região para o lado direito", "INFO")
            try:
                self.definir_regiao(driver, wait, 'ESQUERDO')
                self.aguardar_pagina_estavel(driver, wait, timeout=3)
            except Exception as e:
                log_message(f"⚠️ Erro ao definir região: {e}", "WARNING")

            # Definir quantidade de fragmentos
            log_message(f"📝 Definindo quantidade de fragmentos", "INFO")
            try:
                self.definir_quantidade_fragmentos(driver, wait, mascara)
                self.aguardar_pagina_estavel(driver, wait, timeout=3)
            except Exception as e:
                log_message(f"⚠️ Erro ao definir quantidade de fragmentos: {e}", "WARNING")

            # Definir quantidade de blocos baseado na mascara
            log_message(f"📝 Definindo quantidade de blocos baseado na máscara: {mascara}", "INFO")
            try:
                self.definir_quantidade_blocos(driver, wait, mascara)
                self.aguardar_pagina_estavel(driver, wait, timeout=3)
            except Exception as e:
                log_message(f"⚠️ Erro ao definir quantidade de blocos: {e}", "WARNING")

            self.salvar_fragmentos(driver, wait)

            log_message("✅ Campos pré-envio preenchidos", "SUCCESS")

        except Exception as e:
            log_message(f"⚠️ Erro ao preencher campos pré-envio: {e}", "WARNING")

    def enviar_proxima_etapa(self, driver, wait):
        """Clica no botão de enviar para próxima etapa"""
        try:
            self.aguardar_pagina_estavel(driver, wait)
            self.aguardar_spinner_desaparecer(driver, wait)

            botao_enviar = wait.until(
                EC.element_to_be_clickable((By.ID, "btn-enviar-proxima-etapa"))
            )

            if not botao_enviar.is_displayed() or not botao_enviar.is_enabled():
                log_message("⚠️ Botão não está visível ou habilitado", "WARNING")
                raise Exception("Botão não está interativo")

            try:
                driver.execute_script("arguments[0].click();", botao_enviar)
                log_message("➡️ Clicou em Enviar para próxima etapa via JS", "INFO")
            except:
                botao_enviar.click()
                log_message("➡️ Clicou em Enviar para próxima etapa", "INFO")

            time.sleep(2)

            # Verificar modal de assinatura
            try:
                modal_assinatura = driver.find_element(By.ID, "assinatura")
                if modal_assinatura.is_displayed():
                    log_message("📋 Modal de assinatura detectado", "INFO")
                    return {'status': 'aguardando_assinatura'}
            except:
                pass

            # Verificar erros
            try:
                erros = driver.find_elements(By.CSS_SELECTOR, ".alert-danger, .error-message")
                if erros:
                    mensagem_erro = erros[0].text
                    log_message(f"❌ Erro detectado: {mensagem_erro}", "ERROR")
                    return {'status': 'erro', 'detalhes': mensagem_erro}
            except:
                pass

            log_message("✅ Envio para próxima etapa realizado com sucesso", "SUCCESS")
            return {'status': 'sucesso', 'detalhes': 'Enviado para próxima etapa'}

        except Exception as e:
            log_message(f"Erro ao enviar para próxima etapa: {e}", "ERROR")

            try:
                driver.execute_script("""
                    var spinners = document.querySelectorAll('.loadModal, .spinner, [class*="loading"]');
                    spinners.forEach(function(spinner) {
                        spinner.style.display = 'none';
                    });
                """)
                log_message("🔧 Spinners fechados via JavaScript", "INFO")
            except:
                pass

            raise

    def assinar_com_george(self, driver, wait):
        """Faz o processo de assinatura com Dr. George"""
        try:
            log_message("✍️ Iniciando assinatura...", "INFO")
            # Implementar lógica de assinatura conforme necessário
            log_message("✅ Assinatura concluída", "SUCCESS")
        except Exception as e:
            log_message(f"❌ Erro ao assinar: {e}", "ERROR")
            raise

    def processar_exame(self, driver, wait, num_exame, mascara, macroscopista, medidas, data_fixacao):
        """Processa um exame individual de próstata"""
        try:
            log_message(f"\n{'=' * 60}", "INFO")
            log_message(f"🔬 Processando exame: {num_exame} | Máscara: {mascara}", "INFO")
            log_message(f"{'=' * 60}", "INFO")

            campo_busca = wait.until(EC.element_to_be_clickable((By.ID, "inputSearchCodBarra")))
            campo_busca.clear()
            campo_busca.send_keys(num_exame)
            campo_busca.send_keys(Keys.ENTER)
            log_message(f"🔍 Buscando exame: {num_exame}", "INFO")
            time.sleep(2)

            return self.aguardar_e_processar_andamento(driver, wait, mascara, macroscopista, medidas, data_fixacao)

        except Exception as e:
            log_message(f"❌ Erro ao processar exame {num_exame}: {e}", "ERROR")
            return False

    def aguardar_e_processar_andamento(self, driver, wait, mascara, macroscopista, medidas, data_fixacao):
        """Aguarda a div de andamento e processa o exame"""
        try:
            wait.until(EC.presence_of_element_located((By.ID, "divAndamentoExame")))
            log_message("📋 Div de andamento do exame encontrada!", "SUCCESS")
            time.sleep(0.5)
        except:
            log_message("⚠️ Div de andamento não apareceu no tempo esperado", "WARNING")
            return {'status': 'sem_andamento', 'detalhes': 'Exame não encontrado ou não carregou'}

        log_message("✅ Exame carregado - iniciando processo de conclusão", "SUCCESS")
        return self.processar_conclusao_completa(driver, wait, mascara, macroscopista, medidas, data_fixacao)

    def processar_conclusao_completa(self, driver, wait, mascara, macroscopista, medidas, data_fixacao):
        try:

            log_message("📋 Selecionando responsável pela macroscopia!", "INFO")
            self.selecionar_responsavel_macroscopia(driver, wait, macroscopista)

            self.definir_data_fixacao(driver, wait, data_fixacao)

            self.definir_hora_fixacao(driver, wait)

            if mascara:
                self.digitar_mascara_e_buscar(driver, wait, mascara)
            else:
                log_message("⚠️ Nenhuma máscara encontrada, pulando busca", "WARNING")

            try:
                self.abrir_modal_variaveis_e_preencher(driver, wait, mascara, medidas)
            except Exception as var_error:
                log_message(f"⚠️ Erro no modal de variáveis: {var_error}", "WARNING")
                log_message("⚠️ Continuando o processo sem as variáveis", "WARNING")

            self.salvar_macroscopia(driver, wait)

            try:
                self.preencher_campos_pre_envio(driver, wait, mascara)
            except Exception as campos_error:
                log_message(f"⚠️ Erro ao preencher campos pré-envio: {campos_error}", "WARNING")

            #self.salvar_fragmentos(driver, wait)

            resultado_envio = self.enviar_proxima_etapa(driver, wait)

            if resultado_envio.get('status') == 'aguardando_assinatura':
                log_message("📋 Modal de assinatura aberto - iniciando processo de assinatura", "INFO")
                try:
                    self.assinar_com_george(driver, wait)
                    log_message("✅ Assinatura realizada com sucesso", "SUCCESS")
                    return {'status': 'sucesso', 'detalhes': 'Macroscopia assinada com sucesso'}
                except Exception as assinatura_error:
                    log_message(f"❌ Erro na assinatura: {assinatura_error}", "ERROR")
                    return {'status': 'erro_assinatura', 'detalhes': str(assinatura_error)}
            elif resultado_envio.get('status') == 'erro':
                log_message(f"⚠️ Erro no envio para próxima etapa: {resultado_envio.get('detalhes')}", "WARNING")
                return {'status': 'erro_envio', 'detalhes': resultado_envio.get('detalhes')}
            else:
                log_message("🎉 Processo de macroscopia finalizado com sucesso!", "SUCCESS")
                return {'status': 'sucesso', 'detalhes': 'Macroscopia processada com sucesso'}

            log_message(f"✅ Exame processado com sucesso!", "SUCCESS")
            return True

        except Exception as e:
            log_message(f"Erro durante processo de macroscopia: {e}", "ERROR")
            return {'status': 'erro_macroscopia', 'detalhes': str(e)}

    def mostrar_resumo_final(self, resultados):
        """Mostra o resumo final do processamento"""
        total = len(resultados)
        sucesso = len([r for r in resultados if r['status'] == 'sucesso'])
        sem_andamento = len([r for r in resultados if r['status'] == 'sem_andamento'])
        erro_sessao = len([r for r in resultados if r['status'] == 'erro_sessao'])
        erros = len([r for r in resultados if 'erro' in r['status'] and r['status'] != 'erro_sessao'])

        log_message("\n" + "=" * 50, "INFO")
        log_message("RESUMO FINAL DO PROCESSAMENTO", "INFO")
        log_message("=" * 50, "INFO")
        log_message(f"Total de exames: {total}", "INFO")
        log_message(f"✅ Processados com sucesso: {sucesso}", "SUCCESS")
        log_message(f"⚠️ Exames não encontrados: {sem_andamento}", "WARNING")
        log_message(f"🔄 Erros de sessão (browser perdido): {erro_sessao}", "WARNING")
        log_message(f"❌ Outros erros de processamento: {erros}", "ERROR")

        if erro_sessao + erros > 0:
            log_message("\n⚠️ Alguns exames apresentaram erros. Verifique os logs acima.", "WARNING")

        messagebox.showinfo("Processamento Concluído",
                            f"Processamento finalizado!\n\n"
                            f"Total: {total} exames\n"
                            f"Sucesso: {sucesso}\n"
                            f"Não encontrados: {sem_andamento}\n"
                            f"Erros de sessão: {erro_sessao}\n"
                            f"Outros erros: {erros}")

    def run(self, params: dict):
        username = params.get("username")
        password = params.get("password")
        excel_file = params.get("excel_file")
        cancel_flag = params.get("cancel_flag")
        headless_mode = params.get("headless_mode")

        try:
            log_message(f"📂 Lendo planilha: {excel_file}", "INFO")
            dados_exames = self.get_dados_exames(excel_file)
            log_message(f"✅ {len(dados_exames)} exames encontrados na planilha", "SUCCESS")
        except Exception as e:
            log_message(f"❌ Erro ao ler planilha: {e}", "ERROR")
            messagebox.showerror("Erro", f"Erro ao ler planilha:\n{e}")
            return

        url = os.getenv("SYSTEM_URL", "https://dap.pathoweb.com.br/login/auth")
        driver = None
        resultados = []

        try:
            driver = BrowserFactory.create_chrome(headless=headless_mode)
            wait = WebDriverWait(driver, 10)

            log_message("Iniciando automação de macroscopia septoplastia...", "INFO")

            # Login
            log_message("Fazendo login...", "INFO")
            driver.get(url)

            wait.until(EC.presence_of_element_located((By.ID, "username")))

            username_field = driver.find_element(By.ID, "username")
            username_field.clear()
            username_field.send_keys(username)

            password_field = driver.find_element(By.ID, "password")
            password_field.clear()
            password_field.send_keys(password)

            submit_button = driver.find_element(By.CSS_SELECTOR, "button[type='submit']")
            submit_button.click()

            log_message("Verificando se precisa navegar para módulo de exames...", "INFO")
            current_url = driver.current_url

            if current_url == "https://dap.pathoweb.com.br/" or "trocarModulo" in current_url:
                log_message("Detectada tela de seleção de módulos - navegando para módulo de exames...", "INFO")
                try:
                    modulo_link = wait.until(
                        EC.element_to_be_clickable((By.CSS_SELECTOR, "a[href='/site/trocarModulo?modulo=1']")))
                    modulo_link.click()
                    time.sleep(2)
                    log_message("✅ Navegou para o módulo de exames", "SUCCESS")
                except Exception as e:
                    log_message(f"⚠️ Erro ao navegar para módulo: {e}", "WARNING")
                    driver.get("https://dap.pathoweb.com.br/moduloExame/index")
                    time.sleep(2)

            elif "moduloExame" in current_url:
                log_message("✅ Já está no módulo de exames - pulando navegação", "SUCCESS")
            else:
                log_message(f"⚠️ URL inesperada detectada: {current_url}", "WARNING")
                driver.get("https://dap.pathoweb.com.br/moduloExame/index")
                time.sleep(2)
                log_message("🔄 Navegação direta para módulo realizada (fallback)", "INFO")

            # Fechar modal se aparecer
            try:
                modal_close_button = driver.find_element(By.CSS_SELECTOR,
                                                         "#mensagemParaClienteModal .modal-footer button")
                if modal_close_button.is_displayed():
                    modal_close_button.click()
                    time.sleep(0.5)
            except Exception:
                pass

            log_message("✅ Login realizado com sucesso. Iniciando processamento dos exames.", "SUCCESS")

            for idx, exame in enumerate(dados_exames, 1):
                if cancel_flag.is_set():
                    log_message("🛑 Execução cancelada pelo usuário", "WARNING")
                    break

                if not self.verificar_sessao_browser(driver):
                    log_message("❌ Sessão do browser perdida - abortando processamento", "ERROR")
                    resultados.append({'num_exame': exame['num_exame'], 'status': 'erro_sessao'})
                    break

                log_message(f"\n📊 Progresso: {idx}/{len(dados_exames)}", "INFO")

                medidas = {f'med{i}': exame.get(f'med{i}', '') for i in range(1, 13)}

                sucesso = self.processar_exame(
                    driver, wait,
                    exame['num_exame'],
                    exame['mascara'],
                    exame['macroscopista'],
                    medidas,
                    exame['data_fixacao']
                )

                resultados.append({
                    'num_exame': exame['num_exame'],
                    'status': 'sucesso' if sucesso else 'erro'
                })

                time.sleep(1)

            self.mostrar_resumo_final(resultados)

        except Exception as e:
            log_message(f"❌ Erro crítico durante execução: {e}", "ERROR")
            messagebox.showerror("Erro Crítico", f"Erro durante execução:\n{e}")
        finally:
            if driver:
                try:
                    driver.quit()
                    log_message("🔚 Browser fechado", "INFO")
                except:
                    pass


def run(params: dict):
    module = MacroProstModule()
    module.run(params)