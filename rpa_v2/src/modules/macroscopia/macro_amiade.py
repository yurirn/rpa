import os
import time
from tkinter import messagebox
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from dotenv import load_dotenv
from openpyxl import load_workbook

from src.core.browser_factory import BrowserFactory
from src.core.logger import log_message
from src.modules.base import BaseModule

load_dotenv()


class MacroAmiadeModule(BaseModule):
    def __init__(self):
        super().__init__(nome="Macro Amiade")

    def get_dados_exames(self, file_path: str) -> list:
        try:
            workbook = load_workbook(file_path)
            sheet = workbook.active
            dados = []
            ultima_mascara = None
            data_fixacao = None
            macroscopista_valor = None
            frag_ade_valor = None

            # Ler cabeçalho (linha 1) e criar mapeamento de colunas
            colunas = {}
            for col_idx in range(1, sheet.max_column + 1):
                cell_value = sheet.cell(row=1, column=col_idx).value
                if cell_value:
                    nome_coluna = str(cell_value).strip().lower()
                    colunas[nome_coluna] = col_idx

            log_message(f"📋 Colunas detectadas: {list(colunas.keys())}", "INFO")

            def encontrar_coluna(nomes_possiveis):
                """Encontra a coluna baseado em uma lista de nomes possíveis"""
                for nome in nomes_possiveis:
                    for coluna_nome, col_idx in colunas.items():
                        if nome.lower() in coluna_nome:
                            return col_idx
                return None

            # Encontrar índices das colunas
            col_data = encontrar_coluna(['data', 'data fixacao', 'data fixação', 'datafixacao'])
            col_num_exame = encontrar_coluna(['num_exame', 'numero', 'número', 'codigo', 'código', 'cod'])
            col_mascara = encontrar_coluna(['mascara', 'máscara', 'mask'])
            col_macroscopista = encontrar_coluna(['macroscopista', 'responsavel', 'responsável', 'resp'])
            col_amg_maior = encontrar_coluna(['amg>', 'amg maior'])
            col_amg_menor = encontrar_coluna(['amg<', 'amg menor'])
            col_frag_ade = encontrar_coluna(['frag ade', 'fragade', 'frag_ade'])
            col_ade = encontrar_coluna(['med ade', 'medade', 'med_ade'])
            col_legenda = encontrar_coluna(['legenda', 'leg'])

            # Validar colunas obrigatórias
            if not col_num_exame:
                raise Exception("Coluna de número do exame não encontrada!")

            log_message(f"✅ Mapeamento: Num_Exame=col{col_num_exame}, Máscara=col{col_mascara}, Data=col{col_data}",
                        "INFO")

            # Lê da linha 2 em diante (linha 1 é cabeçalho)
            for row in range(2, sheet.max_row + 1):
                data = sheet.cell(row=row, column=col_data).value if col_data else None
                num_exame = sheet.cell(row=row, column=col_num_exame).value if col_num_exame else None
                mascara = sheet.cell(row=row, column=col_mascara).value if col_mascara else None
                macroscopista = sheet.cell(row=row, column=col_macroscopista).value if col_macroscopista else None
                amg_maior = sheet.cell(row=row, column=col_amg_maior).value if col_amg_maior else None
                amg_menor = sheet.cell(row=row, column=col_amg_menor).value if col_amg_menor else None
                frag_ade = sheet.cell(row=row, column=col_frag_ade).value if col_frag_ade else None
                ade = sheet.cell(row=row, column=col_ade).value if col_ade else None
                legenda = sheet.cell(row=row, column=col_legenda).value if col_legenda else None

                if row == 2 and data:
                    data_fixacao = str(data).strip()

                if macroscopista is not None and str(macroscopista).strip():
                    macroscopista_valor = str(macroscopista).strip().upper()

                if frag_ade is not None and str(frag_ade).strip():
                    frag_ade_valor = str(frag_ade).strip()

                if num_exame is not None:
                    num_exame = str(num_exame).strip()

                    if mascara is not None and str(mascara).strip():
                        mascara = str(mascara).strip()
                        ultima_mascara = mascara
                    else:
                        mascara = ultima_mascara

                    frag_ade_extenso = self.converter_numero_para_extenso(frag_ade_valor) if frag_ade_valor else ""

                    legenda_original = str(legenda).strip().lower() if legenda is not None else ""

                    # Regra: se campo_d for 'mult', usar 6
                    if legenda is not None and str(legenda).strip().lower() == 'mult':
                        legenda_valor = '6'
                    else:
                        legenda_valor = str(legenda).strip() if legenda is not None else ""

                    dados.append({
                        'num_exame': num_exame,
                        'mascara': str(mascara).strip() if mascara else "",
                        'macroscopista': macroscopista_valor,
                        'amg_maior': str(amg_maior).strip() if amg_maior else "",
                        'amg_menor': str(amg_menor).strip() if amg_menor else "",
                        'frag_ade': frag_ade_extenso,
                        'ade': str(ade).strip() if ade else "",
                        'legenda': legenda_valor,
                        'legenda_original': legenda_original,
                        'data_fixacao': data_fixacao
                    })

            workbook.close()
            return dados
        except Exception as e:
            raise Exception(f"Erro ao ler planilha: {e}")

    def converter_numero_para_extenso(self, valor):
        """Converte números para texto por extenso
        Exemplo: '2' -> 'dois', '3' -> 'três', 'mult' -> 'múltiplos'
        """
        if not valor or str(valor).strip() == "":
            return ""

        valor_str = str(valor).strip().lower()

        # Mapeamento de números para extenso
        numeros_extenso = {
            '1': 'um',
            '2': 'dois',
            '3': 'três',
            '4': 'quatro',
            '5': 'cinco'
        }

        # Se for 'mult', retorna 'múltiplos'
        if valor_str == 'mult':
            return 'múltiplos'

        # Se for um número mapeado, retorna por extenso
        if valor_str in numeros_extenso:
            return numeros_extenso[valor_str]

        # Se não encontrar, retorna o valor original
        return valor_str

    def separar_valores_concatenados(self, valor_concatenado):
        """Separa valores concatenados por 'x' em uma lista
        Exemplo: '0,2x0,3x0,2' -> ['0,2', '0,3', '0,2']
        """
        if not valor_concatenado or valor_concatenado.strip() == "":
            return []

        # Separar por 'x' (case insensitive)
        valores = valor_concatenado.replace('X', 'x').split('x')
        # Limpar espaços e retornar
        return [v.strip() for v in valores if v.strip()]

    def verificar_sessao_browser(self, driver) -> bool:
        """Verifica se a sessão do browser ainda está ativa"""
        try:
            driver.current_url
            return True
        except Exception as e:
            if "invalid session id" in str(e).lower():
                log_message("❌ Sessão do browser perdida", "ERROR")
                return False
            return True

    def aguardar_pagina_estavel(self, driver, wait, timeout=10):
        """Aguarda até que a página esteja estável (sem animações ou carregamentos)"""
        try:
            driver.execute_script("""
                return new Promise((resolve) => {
                    if (window.jQuery && window.jQuery.active === 0) {
                        resolve();
                        return;
                    }

                    var checkInterval = setInterval(() => {
                        if (window.jQuery && window.jQuery.active === 0) {
                            clearInterval(checkInterval);
                            resolve();
                        }
                    }, 100);

                    setTimeout(() => {
                        clearInterval(checkInterval);
                        resolve();
                    }, arguments[0]);
                });
            """, timeout * 1000)

            time.sleep(0.5)
            log_message("✅ Página estável", "INFO")

        except Exception as e:
            log_message(f"⚠️ Erro ao aguardar página estável: {e}", "WARNING")
            time.sleep(1)

    def aguardar_spinner_desaparecer(self, driver, wait, timeout=30):
        """Aguarda até que o spinner de loading desapareça"""
        try:
            log_message("⏳ Aguardando spinner desaparecer...", "INFO")
            wait.until(EC.invisibility_of_element_located((By.ID, "spinner")))
            time.sleep(1)

            spinners = driver.find_elements(By.CSS_SELECTOR, ".loadModal, .spinner, [class*='loading']")
            for spinner in spinners:
                if spinner.is_displayed():
                    log_message("⚠️ Outro spinner ainda visível, aguardando...", "WARNING")
                    time.sleep(2)
                    break

            log_message("✅ Spinner desapareceu", "SUCCESS")

        except Exception as e:
            log_message(f"⚠️ Erro ao aguardar spinner: {e}", "WARNING")
            try:
                driver.execute_script("""
                    var spinners = document.querySelectorAll('.loadModal, .spinner, [class*="loading"]');
                    spinners.forEach(function(spinner) {
                        if (spinner.style.display !== 'none') {
                            spinner.style.display = 'none';
                        }
                    });
                """)
                log_message("🔧 Spinner fechado via JavaScript", "INFO")
                time.sleep(1)
            except:
                pass

    def selecionar_responsavel_macroscopia(self, driver, wait, responsavel_macro):
        """Seleciona o responsável pela macroscopia conforme o nome recebido (nome curto)"""
        responsavel_macro_mapper = {
            'BARBARA': 'Barbara Dutra Lopes',
            'NATHALIA': 'Nathalia Fernanda da Silva Lopes',
            'RENATA': 'Renata Silva Sevidanis',
            'HELEN': 'Helen Oliveira dos Santos',
            'CLARA': 'Clara Helena Janz Garcia de Souza',
            'PALOMA': 'Paloma Brenda Silva De Oliveira',
            'ELLEN': 'Ellen Andressa de Alvarenga',
            'VITORIA': 'Vitoria Aquino Nairne Domingues',
            'ANNAI': 'Annai Lukã Vitorino Losnak',
            'ANA': 'Ana Carolina Viecele Campos'
        }
        nome_completo = responsavel_macro_mapper.get(responsavel_macro, responsavel_macro)
        select2_container = wait.until(
            EC.element_to_be_clickable(
                (By.XPATH, "//span[@aria-labelledby='select2-responsavelMacroscopiaId-container']"))
        )
        select2_container.click()
        time.sleep(0.3)

        opcao = wait.until(
            EC.element_to_be_clickable((By.XPATH, f"//li[contains(text(), '{nome_completo}')]"))
        )
        opcao.click()
        log_message(f"✅ {nome_completo} selecionado como responsável", "SUCCESS")
        time.sleep(0.2)

    def definir_data_fixacao(self, driver, wait, data_fixacao=None):
        """Define a data de fixação no campo de data de fixação"""
        try:
            if not data_fixacao:
                data_fixacao = '21082025'

            if len(data_fixacao) == 8 and data_fixacao.isdigit():
                data_formatada = f"{data_fixacao[4:8]}-{data_fixacao[2:4]}-{data_fixacao[0:2]}"
            else:
                data_formatada = '2025-08-21'

            campo_data = wait.until(
                EC.presence_of_element_located((By.XPATH, "//input[@type='date' and @name='dataFixacao']"))
            )
            driver.execute_script("""
                var campo = arguments[0];
                campo.value = arguments[1];
                campo.dispatchEvent(new Event('change', { bubbles: true }));
            """, campo_data, data_formatada)
            log_message(f"📅 Data de fixação definida para: {data_formatada}", "SUCCESS")
            time.sleep(0.1)
        except Exception as e:
            log_message(f"⚠️ Erro ao definir data de fixação: {e}", "WARNING")

    def definir_hora_fixacao(self, driver, wait):
        """Define 18:00 no campo de hora de fixação"""
        campo_hora = wait.until(
            EC.presence_of_element_located((By.XPATH, "//input[@type='time' and @name='dataFixacao']"))
        )
        campo_hora.clear()
        campo_hora.send_keys("18:00")
        log_message("🕕 Hora de fixação definida para: 18:00", "SUCCESS")
        time.sleep(0.1)

    def fechar_exame(self, driver, wait):
        """Clica no botão de fechar exame"""
        try:
            botao_fechar = wait.until(
                EC.element_to_be_clickable((By.ID, "fecharExameBarraFerramenta"))
            )
            botao_fechar.click()
            log_message("📁 Exame fechado", "INFO")

            try:
                wait.until(EC.presence_of_element_located((By.ID, "inputSearchCodBarra")))
                log_message("✅ Retornou à tela principal após fechar exame", "INFO")
            except:
                log_message("⚠️ Pode não ter retornado à tela principal", "WARNING")
                try:
                    current_url = driver.current_url
                    if "modulo=1" not in current_url:
                        modulo_link = driver.find_element(By.CSS_SELECTOR, "a[href='/site/trocarModulo?modulo=1']")
                        modulo_link.click()
                        time.sleep(1.5)
                        log_message("🔄 Navegou de volta ao módulo de exames", "INFO")
                except:
                    pass

        except Exception as e:
            log_message(f"Erro ao fechar exame: {e}", "ERROR")

    def digitar_mascara_e_buscar(self, driver, wait, mascara):
        """Digita a máscara no campo buscaArvore e pressiona Enter"""
        campo_busca = wait.until(EC.element_to_be_clickable((By.ID, "buscaArvore")))
        campo_busca.send_keys(mascara)
        campo_busca.send_keys(Keys.ENTER)
        log_message(f"✍️ Máscara '{mascara}' digitada no campo buscaArvore", "SUCCESS")
        time.sleep(0.5)

    def abrir_modal_variaveis_e_preencher(self, driver, wait, mascara, amg_maior, amg_menor, frag_ade, ade, legenda, legenda_original):
        """Abre o modal de variáveis e preenche os campos baseado na máscara AMIADE"""
        try:
            # Clicar no botão "Pesquisar variáveis (F7)"
            botao_variaveis = wait.until(
                EC.element_to_be_clickable((By.ID, "cke_70"))
            )
            botao_variaveis.click()
            log_message("🔍 Clicou no botão de variáveis", "INFO")
            time.sleep(0.8)

            # Verificar se apareceu um alerta
            try:
                alert = driver.switch_to.alert
                alert_text = alert.text
                if "não há variáveis" in alert_text.lower():
                    log_message(f"⚠️ Alerta detectado: {alert_text}", "WARNING")
                    alert.accept()
                    log_message("⚠️ Pulando preenchimento de variáveis - não há variáveis no texto", "WARNING")
                    return
                else:
                    alert.accept()
            except:
                pass

            # Aguardar o modal aparecer
            wait.until(EC.presence_of_element_located((By.CLASS_NAME, "swal2-popup")))
            log_message("🔍 Modal de variáveis aberto", "SUCCESS")
            time.sleep(0.3)

            # Preencher os campos usando classe genérica
            campos_input = driver.find_elements(By.CSS_SELECTOR, "input[style*='width: 100px'][style*='color: red']")
            log_message(f"🔍 Encontrados {len(campos_input)} campos de input no modal", "INFO")

            # Processar valores concatenados
            valores_amg_maior = self.separar_valores_concatenados(amg_maior)
            valores_amg_menor = self.separar_valores_concatenados(amg_menor)
            valores_ade = self.separar_valores_concatenados(ade)

            # Determinar valores baseado na máscara
            mascara_upper = mascara.upper() if mascara else ""
            valores = []

            if legenda_original == 'mult':
                legenda = 'M'

            if mascara_upper == 'AMIADE':
                # AMIADE: Amg> (3 campos), Amg< (3 campos), Frag ade, Ade (3 campos), legenda
                valores = valores_amg_maior + valores_amg_menor + [frag_ade] + valores_ade + [legenda]
                log_message(f"📋 AMIADE detectado - valores: Amg>={valores_amg_maior}, Amg<={valores_amg_menor}, Frag_ade={frag_ade}, Ade={valores_ade}, Legenda={legenda}","INFO")

            elif mascara_upper == 'AMI':
                # AMI: Amg> (3 campos), Amg< (3 campos), legenda
                valores = valores_amg_maior + valores_amg_menor + [legenda]
                log_message(f"📋 AMI detectado - valores: Amg>={valores_amg_maior}, Amg<={valores_amg_menor}", "INFO")
            else:
                log_message(f"⚠️ Máscara '{mascara}' não reconhecida para preenchimento de variáveis", "WARNING")
                return

            # Filtrar valores vazios
            valores = [v for v in valores if v and v.strip()]

            log_message(f"📋 Preenchendo {len(valores)} variáveis para máscara '{mascara}'", "INFO")

            for i, campo in enumerate(campos_input[:len(valores)]):
                if i < len(valores) and valores[i]:
                    try:
                        driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});",
                                              campo)
                        time.sleep(0.1)

                        driver.execute_script("""
                            arguments[0].value = '';
                            arguments[0].value = arguments[1];
                            arguments[0].dispatchEvent(new Event('input', { bubbles: true }));
                            arguments[0].dispatchEvent(new Event('change', { bubbles: true }));
                        """, campo, valores[i])

                        log_message(f"✅ Campo {i + 1} preenchido com: {valores[i]}", "SUCCESS")
                        time.sleep(0.1)
                    except Exception as e:
                        log_message(f"⚠️ Erro ao preencher campo {i + 1}: {e}", "WARNING")

            time.sleep(0.2)

            # Clicar no botão "Inserir"
            botao_inserir = wait.until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, ".swal2-confirm"))
            )
            botao_inserir.click()
            log_message("✅ Campos inseridos no modal", "SUCCESS")

            # Aguardar o modal fechar
            try:
                wait.until(EC.invisibility_of_element_located((By.CLASS_NAME, "swal2-popup")))
                log_message("✅ Modal fechado completamente", "SUCCESS")
            except:
                time.sleep(1)
                log_message("⏳ Aguardou fechamento do modal", "INFO")

        except Exception as e:
            log_message(f"⚠️ Erro ao preencher modal de variáveis: {e}", "WARNING")
            log_message("⚠️ Continuando sem preencher as variáveis", "WARNING")

    def salvar_macroscopia(self, driver, wait):
        """Clica no botão Salvar da macroscopia"""
        try:
            modal = driver.find_element(By.CLASS_NAME, "swal2-popup")
            if modal.is_displayed():
                botao_fechar = driver.find_element(By.CSS_SELECTOR, ".swal2-close")
                botao_fechar.click()
                time.sleep(0.5)
        except:
            pass

        botao_salvar = wait.until(
            EC.element_to_be_clickable((By.ID, "salvarMacro"))
        )
        botao_salvar.click()
        log_message("💾 Macroscopia salva", "SUCCESS")
        time.sleep(0.3)

    def definir_grupo(self, driver, wait):
        """Define o grupo como 'Seios da Face' usando JavaScript"""
        try:
            grupo_selecionado = 'Seios da Face'

            # Verificar se o input existe e qual o valor atual
            try:
                input_grupo = driver.find_element(By.ID, "grupo")
                valor_atual = input_grupo.get_attribute("value")

                if valor_atual == grupo_selecionado:
                    log_message(f"✅ Grupo já está definido como '{grupo_selecionado}'", "SUCCESS")
                    return
                elif valor_atual and valor_atual.strip():
                    log_message(f"⚠️ Valor atual do campo grupo: '{valor_atual}' - será substituído", "WARNING")
            except:
                log_message("⚠️ Campo grupo não encontrado", "WARNING")

            # Procurar especificamente pelo campo de grupo
            script = """
            // Procurar especificamente pelo campo de grupo que tem o input com id="idRegiao"
            var inputGrupo = document.getElementById('idRegiao');
            if (inputGrupo) {
                // Encontrar a âncora que está no mesmo td que o input idRegiao
                var parentTd = inputGrupo.closest('td');
                if (parentTd) {
                    var ancora = parentTd.querySelector('a[class*="table-editable-ancora"]');
                    if (ancora && ancora.offsetParent !== null) {
                        return ancora;
                    }
                }
            }

            // Fallback: procurar por âncoras que estejam próximas a inputs de grupo
            var inputsGrupo = document.querySelectorAll('input[id*="Regiao"], input[data-autocompleteurl*="consultarRegiao"]');
            for (var i = 0; i < inputsGrupo.length; i++) {
                var input = inputsGrupo[i];
                var parentTd = input.closest('td');
                if (parentTd) {
                    var ancora = parentTd.querySelector('a[class*="table-editable-ancora"]');
                    if (ancora && ancora.offsetParent !== null) {
                        return ancora;
                    }
                }
            }

            // Último fallback: procurar por âncoras que não sejam de procedimento
            var fragmentosContainer = document.getElementById('fragmentosContainer');
            if (fragmentosContainer) {
                var elementos = fragmentosContainer.querySelectorAll('a[class*="table-editable-ancora"]');
                for (var i = 0; i < elementos.length; i++) {
                    var elemento = elementos[i];
                    if (elemento.textContent.includes('Vazio') && elemento.offsetParent !== null) {
                        var parentTd = elemento.closest('td');
                        if (parentTd && !parentTd.querySelector('input[id*="procedimento"]')) {
                            return elemento;
                        }
                    }
                }
            }
            return null;
            """
            campo_grupo = driver.execute_script(script)

            log_message(campo_grupo, "SUCCESS")

            if campo_grupo:
                # Usar JavaScript para clicar no elemento
                driver.execute_script("arguments[0].click();", campo_grupo)
                log_message(f"🔍 Clicou no campo de grupo via JS", "INFO")
                time.sleep(0.5)

                # Aguardar o campo de input aparecer e preencher via JavaScript
                input_grupo = wait.until(
                    EC.presence_of_element_located((By.ID, "idRegiao"))
                )

                # Limpar o campo primeiro
                driver.execute_script("arguments[0].value = '';", input_grupo)

                # Preencher via JavaScript
                driver.execute_script("""
                    arguments[0].value = arguments[1];
                    arguments[0].dispatchEvent(new Event('input', { bubbles: true }));
                    arguments[0].dispatchEvent(new Event('change', { bubbles: true }));
                """, input_grupo, grupo_selecionado)

                # Aguardar um pouco para o dropdown aparecer e tentar clicar na opção
                time.sleep(0.5)

                # Tentar clicar na opção do dropdown com timeout menor
                try:
                    # Aguardar até 3 segundos pela opção aparecer
                    wait_dropdown = WebDriverWait(driver, 3)
                    opcao_dropdown = wait_dropdown.until(
                        EC.element_to_be_clickable(
                            (By.XPATH, f"//li[contains(@class, 'active')]//a[contains(text(), '{grupo_selecionado}')]"))
                    )
                    opcao_dropdown.click()
                    log_message(f"✅ Selecionou '{grupo_selecionado}' no dropdown", "SUCCESS")
                except:
                    # Se não conseguir clicar no dropdown rapidamente, pressionar Enter
                    try:
                        input_grupo.send_keys(Keys.ENTER)
                        log_message(f"✍️ Pressionou Enter para confirmar '{grupo_selecionado}' (dropdown não apareceu)",
                                    "SUCCESS")
                    except:
                        # Último recurso: clicar fora para fechar o dropdown
                        driver.execute_script("document.body.click();")
                        log_message(f"🔍 Clicou fora para fechar dropdown de '{grupo_selecionado}'", "INFO")

                time.sleep(0.5)
            else:
                log_message("⚠️ Campo de grupo não encontrado ou não visível", "WARNING")

        except Exception as e:
            log_message(f"⚠️ Erro ao definir grupo: {e}", "WARNING")
            raise

    def definir_representacao_secao(self, driver, wait):
        """Define a representação como 'Seção' usando JavaScript"""
        try:
            # Verificar se o select existe e qual o valor atual
            try:
                select_representacao = driver.find_element(By.ID, "representacao")
                valor_atual = select_representacao.get_attribute("value")

                if valor_atual == "S":
                    log_message("✅ Representação já está definida como 'Seção'", "SUCCESS")
                    return
                elif valor_atual != "S":
                    log_message(f"⚠️ Representação atual é '{valor_atual}', mas precisa ser 'S' (Seção)", "WARNING")
            except:
                log_message("⚠️ Campo representacao não encontrado", "WARNING")
                return

            # Procurar especificamente pelo campo de representação
            script = """
            // Procurar especificamente pelo campo de representação que tem o select com id="representacao"
            var selectRepresentacao = document.getElementById('representacao');
            if (selectRepresentacao) {
                // Encontrar a âncora que está no mesmo td que o select representacao
                var parentTd = selectRepresentacao.closest('td');
                if (parentTd) {
                    var ancora = parentTd.querySelector('a[class*="table-editable-ancora"]');
                    if (ancora && ancora.offsetParent !== null) {
                        return ancora;
                    }
                }
            }

            // Fallback: procurar por âncoras que estejam próximas a selects de representação
            var selectsRepresentacao = document.querySelectorAll('select[id*="representacao"], select[name*="representacao"]');
            for (var i = 0; i < selectsRepresentacao.length; i++) {
                var select = selectsRepresentacao[i];
                var parentTd = select.closest('td');
                if (parentTd) {
                    var ancora = parentTd.querySelector('a[class*="table-editable-ancora"]');
                    if (ancora && ancora.offsetParent !== null) {
                        return ancora;
                    }
                }
            }

            // Último fallback: procurar por texto "representação" ou "-- representação --" que não seja de procedimento
            var fragmentosContainer = document.getElementById('fragmentosContainer');
            if (fragmentosContainer) {
                var elementos = fragmentosContainer.querySelectorAll('a[class*="table-editable-ancora"]');
                for (var i = 0; i < elementos.length; i++) {
                    var elemento = elementos[i];
                    if ((elemento.textContent.toLowerCase().includes('representação') || elemento.textContent.includes('-- representação --')) && elemento.offsetParent !== null) {
                        var parentTd = elemento.closest('td');
                        if (parentTd && !parentTd.querySelector('input[id*="procedimento"]')) {
                            return elemento;
                        }
                    }
                }
            }
            return null;
            """
            campo_representacao = driver.execute_script(script)

            if not campo_representacao:
                log_message("⚠️ Campo de representação não encontrado", "WARNING")
                return

            # Verificar o texto da âncora para log
            if "Seção" in campo_representacao.text:
                log_message("✅ Representação já mostra 'Seção', mas vamos garantir", "INFO")
            elif "-- representação --" in campo_representacao.text:
                log_message("📝 Campo de representação encontrado, precisa ser preenchido", "INFO")
            else:
                log_message(f"⚠️ Texto inesperado no campo de representação: '{campo_representacao.text}'", "WARNING")

            # Clicar via JavaScript
            driver.execute_script("arguments[0].click();", campo_representacao)
            log_message("🔍 Clicou no campo de representação via JS", "INFO")
            time.sleep(0.5)

            # Aguardar o select aparecer e selecionar via JavaScript
            select_representacao = wait.until(
                EC.presence_of_element_located((By.ID, "representacao"))
            )

            # Selecionar "Seção" (valor "S") via JavaScript
            driver.execute_script("""
                arguments[0].value = 'S';
                arguments[0].dispatchEvent(new Event('change', { bubbles: true }));
            """, select_representacao)

            log_message("✅ Representação definida como 'Seção' via JS", "SUCCESS")
            time.sleep(0.5)

            # Clicar fora para confirmar a seleção
            driver.execute_script("document.body.click();")
            time.sleep(0.3)

        except Exception as e:
            log_message(f"⚠️ Erro ao definir representação: {e}", "WARNING")

    def definir_regiao(self, driver, wait, mascara=None):
        """Define a região de acordo com a máscara usando JavaScript"""
        try:
            if not mascara:
                log_message("⚠️ Nenhuma máscara fornecida para definir região", "WARNING")
                return

            # Regras de máscara para região
            mascara_regiao = {
                'AMIADE': 'AAA: Amig>/Amig</Adenoide',
                'AMI': 'AA: Amig>/Amig<'
            }

            mascara_upper = mascara.upper()
            regiao_valor = mascara_regiao.get(mascara_upper)

            if not regiao_valor:
                log_message(f"⚠️ Máscara '{mascara}' não possui região configurada", "WARNING")
                return

            # Verificar se já existe um campo de região preenchido com o valor correto
            try:
                inputs_regiao = driver.find_elements(By.XPATH, "//input[contains(@name, 'regiao_')]")
                for input_reg in inputs_regiao:
                    valor_atual = input_reg.get_attribute("value")
                    if valor_atual == regiao_valor:
                        log_message(f"✅ Região já está definida como '{regiao_valor}' - pulando", "SUCCESS")
                        return
                    elif valor_atual and valor_atual != regiao_valor:
                        log_message(f"⚠️ Região atual é '{valor_atual}', precisa mudar para '{regiao_valor}'",
                                    "WARNING")
                        break
            except:
                pass

            # Procurar e clicar no campo de região para editá-lo
            script = """
                        // Procurar especificamente por campos de região na tabela de fragmentos
                        var tbody = document.getElementById('tdRegiao');
                        if (tbody) {
                            var inputs = tbody.querySelectorAll('input[name*="regiao_"]');
                            for (var i = 0; i < inputs.length; i++) {
                                var input = inputs[i];
                                var parentTd = input.closest('td');
                                if (parentTd) {
                                    var ancora = parentTd.querySelector('a[class*="table-editable-ancora"]');
                                    if (ancora && ancora.offsetParent !== null) {
                                        return {element: ancora, input: input};
                                    }
                                }
                            }
                        }
                        return null;
                        """
            resultado_regiao = driver.execute_script(script)

            if resultado_regiao:
                campo_regiao = resultado_regiao['element']
                input_regiao = resultado_regiao['input']

                # Clicar na âncora para abrir o campo de edição
                driver.execute_script("arguments[0].click();", campo_regiao)
                log_message("🔍 Clicou no campo de região para editar", "INFO")
                time.sleep(0.5)

                # Aguardar o input ficar visível e preencher
                try:
                    # Aguardar o input aparecer
                    wait.until(lambda d: input_regiao.is_displayed() or input_regiao.get_attribute(
                        "style") != "display: none;")

                    # Limpar e preencher o campo
                    driver.execute_script("""
                                    arguments[0].value = '';
                                    arguments[0].value = arguments[1];
                                    arguments[0].dispatchEvent(new Event('input', { bubbles: true }));
                                    arguments[0].dispatchEvent(new Event('change', { bubbles: true }));
                                """, input_regiao, regiao_valor)

                    log_message(f"✍️ Definiu região como '{regiao_valor}' via JS", "SUCCESS")
                    time.sleep(0.5)

                    # Clicar fora para confirmar a edição
                    driver.execute_script("document.body.click();")
                    time.sleep(0.5)

                    # Verificar se o valor foi realmente definido
                    valor_definido = input_regiao.get_attribute("value")
                    if valor_definido == regiao_valor:
                        log_message(f"✅ Valor de região confirmado: '{valor_definido}'", "SUCCESS")
                    else:
                        log_message(
                            f"⚠️ Valor não foi definido corretamente. Esperado: '{regiao_valor}', Atual: '{valor_definido}'",
                            "WARNING")

                except Exception as input_error:
                    log_message(f"⚠️ Erro ao preencher input de região: {input_error}", "WARNING")
            else:
                log_message("⚠️ Campo de região não encontrado ou não visível", "WARNING")

        except Exception as e:
            log_message(f"⚠️ Erro ao definir região: {e}", "WARNING")

    def definir_quantidade_fragmentos(self, driver, wait, legenda):
        """Define a quantidade de fragmentos baseado no valor do campo legenda"""
        try:
            if not legenda or not legenda.strip():
                log_message("⚠️ Campo legenda está vazio, não definindo quantidade de fragmentos", "WARNING")
                return

            quantidade_valor = legenda.strip()
            log_message(f"📝 Usando quantidade da legenda: {quantidade_valor}", "INFO")

            # Procurar pelos campos de quantidade na tabela de fragmentos
            script = """
            // Procurar especificamente por campos de quantidade na tabela de fragmentos
            var tbody = document.getElementById('tdRegiao');
            if (tbody) {
                var inputs = tbody.querySelectorAll('input[name*="quantidade_"]');
                for (var i = 0; i < inputs.length; i++) {
                    var input = inputs[i];
                    var parentTd = input.closest('td');
                    if (parentTd) {
                        var ancora = parentTd.querySelector('a[class*="table-editable-ancora"]');
                        if (ancora && ancora.offsetParent !== null) {
                            return {element: ancora, input: input};
                        }
                    }
                }
            }
            return null;
            """
            resultado_quantidade = driver.execute_script(script)

            if resultado_quantidade:
                campo_quantidade = resultado_quantidade['element']
                input_quantidade = resultado_quantidade['input']

                # Verificar se já tem o valor correto
                valor_atual = input_quantidade.get_attribute("value")
                if valor_atual == quantidade_valor:
                    log_message(f"✅ Quantidade já está definida como '{quantidade_valor}' - pulando", "SUCCESS")
                    return

                # Clicar na âncora para abrir o campo
                driver.execute_script("arguments[0].click();", campo_quantidade)
                log_message("🔍 Clicou no campo de quantidade para editar", "INFO")
                time.sleep(0.5)

                # Aguardar o input ficar visível e preencher
                try:
                    # Aguardar o input aparecer
                    wait.until(lambda d: input_quantidade.is_displayed() or input_quantidade.get_attribute(
                        "style") != "display: none;")

                    # Limpar e preencher o campo
                    driver.execute_script("""
                        arguments[0].value = '';
                        arguments[0].value = arguments[1];
                        arguments[0].dispatchEvent(new Event('input', { bubbles: true }));
                        arguments[0].dispatchEvent(new Event('change', { bubbles: true }));
                    """, input_quantidade, quantidade_valor)

                    log_message(f"✍️ Definiu quantidade como '{quantidade_valor}' via JS", "SUCCESS")
                    time.sleep(0.5)

                    # Clicar fora para confirmar a edição
                    driver.execute_script("document.body.click();")
                    time.sleep(0.3)

                    # Verificar se o valor foi definido
                    valor_definido = input_quantidade.get_attribute("value")
                    if valor_definido == quantidade_valor:
                        log_message(f"✅ Valor de quantidade confirmado: '{valor_definido}'", "SUCCESS")
                    else:
                        log_message(
                            f"⚠️ Valor não foi definido corretamente. Esperado: '{quantidade_valor}', Atual: '{valor_definido}'",
                            "WARNING")

                except Exception as input_error:
                    log_message(f"⚠️ Erro ao preencher input de quantidade: {input_error}", "WARNING")
            else:
                log_message("⚠️ Campo de quantidade não encontrado ou não visível", "WARNING")

        except Exception as e:
            log_message(f"⚠️ Erro ao definir quantidade de fragmentos: {e}", "WARNING")

    def definir_quantidade_blocos(self, driver, wait):
        """Define a quantidade de blocos como 1 (valor fixo)"""
        try:
            log_message("📝 Definindo quantidade de blocos como: 1", "INFO")

            # Procurar pelos campos de quantidade de blocos na tabela de fragmentos
            script = """
            // Procurar especificamente por campos de quantidade de blocos na tabela de fragmentos
            var tbody = document.getElementById('tdRegiao');
            if (tbody) {
                var inputs = tbody.querySelectorAll('input[name*="quantidadeBlocos_"]');
                for (var i = 0; i < inputs.length; i++) {
                    var input = inputs[i];
                    var parentTd = input.closest('td');
                    if (parentTd) {
                        var ancora = parentTd.querySelector('a[class*="table-editable-ancora"]');
                        if (ancora && ancora.offsetParent !== null) {
                            return {element: ancora, input: input};
                        }
                    }
                }
            }
            return null;
            """
            resultado_blocos = driver.execute_script(script)

            if resultado_blocos:
                campo_blocos = resultado_blocos['element']
                input_blocos = resultado_blocos['input']

                # Verificar se já tem o valor correto
                valor_atual = input_blocos.get_attribute("value")
                if valor_atual == "1":
                    log_message("✅ Quantidade de blocos já está definida como '1' - pulando", "SUCCESS")
                    return

                # Clicar na âncora para abrir o campo
                driver.execute_script("arguments[0].click();", campo_blocos)
                log_message("🔍 Clicou no campo de quantidade de blocos para editar", "INFO")
                time.sleep(0.5)

                # Aguardar o input ficar visível e preencher
                try:
                    # Aguardar o input aparecer
                    wait.until(lambda d: input_blocos.is_displayed() or input_blocos.get_attribute(
                        "style") != "display: none;")

                    # Limpar e preencher o campo
                    driver.execute_script("""
                        arguments[0].value = '';
                        arguments[0].value = '1';
                        arguments[0].dispatchEvent(new Event('input', { bubbles: true }));
                        arguments[0].dispatchEvent(new Event('change', { bubbles: true }));
                    """, input_blocos)

                    log_message("✍️ Definiu quantidade de blocos como '1' via JS", "SUCCESS")
                    time.sleep(0.5)

                    # Clicar fora para confirmar a edição
                    driver.execute_script("document.body.click();")
                    time.sleep(0.3)

                    # Verificar se o valor foi definido
                    valor_definido = input_blocos.get_attribute("value")
                    if valor_definido == "1":
                        log_message("✅ Valor de quantidade de blocos confirmado: '1'", "SUCCESS")
                    else:
                        log_message(f"⚠️ Valor não foi definido corretamente. Esperado: '1', Atual: '{valor_definido}'",
                                    "WARNING")

                except Exception as input_error:
                    log_message(f"⚠️ Erro ao preencher input de quantidade de blocos: {input_error}", "WARNING")
            else:
                log_message("⚠️ Campo de quantidade de blocos não encontrado ou não visível", "WARNING")

        except Exception as e:
            log_message(f"⚠️ Erro ao definir quantidade de blocos: {e}", "WARNING")

    def salvar_fragmentos(self, driver, wait):
        """Clica no botão Salvar dos fragmentos"""
        try:
            # Aguardar o botão estar presente e clicável
            botao_salvar_fragmentos = wait.until(
                EC.element_to_be_clickable((By.XPATH,
                                            "//a[contains(@class, 'btn-primary') and contains(@data-url, '/macroscopia/saveMacroscopiaFragAjax')]"))
            )

            # Verificar se o botão está visível
            if not botao_salvar_fragmentos.is_displayed():
                log_message("⚠️ Botão salvar fragmentos não está visível", "WARNING")
                return

            # Rolar até o botão para garantir visibilidade
            driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});",
                                  botao_salvar_fragmentos)
            time.sleep(0.5)

            # Clicar no botão
            botao_salvar_fragmentos.click()
            log_message("💾 Clicou em Salvar fragmentos", "SUCCESS")

            # Aguardar que o spinner desapareça após salvar
            self.aguardar_spinner_desaparecer(driver, wait, timeout=15)

        except Exception as e:
            log_message(f"⚠️ Erro ao salvar fragmentos: {e}", "WARNING")
            # Tentar encontrar o botão por outras formas
            try:
                # Tentar por título
                botao_titulo = wait.until(
                    EC.element_to_be_clickable((By.XPATH, "//a[@title='Salvar' and contains(@class, 'btn-primary')]"))
                )
                driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});",
                                      botao_titulo)
                time.sleep(0.5)
                botao_titulo.click()
                log_message("💾 Clicou em Salvar fragmentos (por título)", "SUCCESS")
                self.aguardar_spinner_desaparecer(driver, wait, timeout=15)
                return
            except:
                pass

            try:
                # Tentar por texto do botão
                botao_texto = wait.until(
                    EC.element_to_be_clickable(
                        (By.XPATH, "//a[contains(@class, 'btn-primary') and contains(text(), 'Salvar')]"))
                )
                driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});",
                                      botao_texto)
                time.sleep(0.5)
                botao_texto.click()
                log_message("💾 Clicou em Salvar fragmentos (por texto)", "SUCCESS")
                self.aguardar_spinner_desaparecer(driver, wait, timeout=15)
                return
            except:
                pass

            log_message(f"❌ Não foi possível encontrar o botão Salvar fragmentos: {e}", "ERROR")
            raise

    def preencher_campos_pre_envio(self, driver, wait, mascara, legenda):
        """Preenche os campos necessários antes do envio para próxima etapa"""
        try:
            log_message("📝 Preenchendo campos pré-envio...", "INFO")

            # Definir grupo
            log_message(f"📝 Definindo grupo para máscara: {mascara}", "INFO")
            try:
                self.definir_grupo(driver, wait)
                self.aguardar_pagina_estavel(driver, wait, timeout=3)
            except Exception as e:
                log_message(f"⚠️ Erro ao definir grupo: {e}", "WARNING")

            # Definir representação como "Seção"
            log_message("📝 Definindo representação como Seção", "INFO")
            try:
                self.definir_representacao_secao(driver, wait)
                self.aguardar_pagina_estavel(driver, wait, timeout=3)
            except Exception as e:
                log_message(f"⚠️ Erro ao definir representação: {e}", "WARNING")

            # Definir região baseada na máscara
            log_message(f"📝 Definindo região para máscara: {mascara}", "INFO")
            try:
                self.definir_regiao(driver, wait, mascara)
                self.aguardar_pagina_estavel(driver, wait, timeout=3)
            except Exception as e:
                log_message(f"⚠️ Erro ao definir região: {e}", "WARNING")

            # Definir quantidade de fragmentos baseado na legenda
            log_message(f"📝 Definindo quantidade de fragmentos baseado na legenda: {legenda}", "INFO")
            try:
                self.definir_quantidade_fragmentos(driver, wait, legenda)
                self.aguardar_pagina_estavel(driver, wait, timeout=3)
            except Exception as e:
                log_message(f"⚠️ Erro ao definir quantidade de fragmentos: {e}", "WARNING")

            # Definir quantidade de blocos como 1
            log_message("📝 Definindo quantidade de blocos como 1", "INFO")
            try:
                self.definir_quantidade_blocos(driver, wait)
                self.aguardar_pagina_estavel(driver, wait, timeout=3)
            except Exception as e:
                log_message(f"⚠️ Erro ao definir quantidade de blocos: {e}", "WARNING")

            log_message("✅ Campos pré-envio preenchidos", "SUCCESS")

        except Exception as e:
            log_message(f"⚠️ Erro ao preencher campos pré-envio: {e}", "WARNING")

    def enviar_proxima_etapa(self, driver, wait):
        """Clica no botão de enviar para próxima etapa"""
        try:
            self.aguardar_pagina_estavel(driver, wait)
            self.aguardar_spinner_desaparecer(driver, wait)

            botao_enviar = wait.until(
                EC.element_to_be_clickable((By.ID, "btn-enviar-proxima-etapa"))
            )

            if not botao_enviar.is_displayed() or not botao_enviar.is_enabled():
                log_message("⚠️ Botão não está visível ou habilitado", "WARNING")
                raise Exception("Botão não está interativo")

            try:
                driver.execute_script("arguments[0].click();", botao_enviar)
                log_message("➡️ Clicou em Enviar para próxima etapa via JS", "INFO")
            except:
                botao_enviar.click()
                log_message("➡️ Clicou em Enviar para próxima etapa", "INFO")

            time.sleep(2)

            # Verificar modal de assinatura
            try:
                modal_assinatura = driver.find_element(By.ID, "assinatura")
                if modal_assinatura.is_displayed():
                    log_message("📋 Modal de assinatura detectado", "INFO")
                    return {'status': 'aguardando_assinatura'}
            except:
                pass

            # Verificar erros
            try:
                erros = driver.find_elements(By.CSS_SELECTOR, ".alert-danger, .error-message")
                if erros:
                    mensagem_erro = erros[0].text
                    log_message(f"❌ Erro detectado: {mensagem_erro}", "ERROR")
                    return {'status': 'erro', 'detalhes': mensagem_erro}
            except:
                pass

            log_message("✅ Envio para próxima etapa realizado com sucesso", "SUCCESS")
            return {'status': 'sucesso', 'detalhes': 'Enviado para próxima etapa'}

        except Exception as e:
            log_message(f"Erro ao enviar para próxima etapa: {e}", "ERROR")

            try:
                driver.execute_script("""
                    var spinners = document.querySelectorAll('.loadModal, .spinner, [class*="loading"]');
                    spinners.forEach(function(spinner) {
                        spinner.style.display = 'none';
                    });
                """)
                log_message("🔧 Spinners fechados via JavaScript", "INFO")
            except:
                pass

            raise

    def assinar_com_george(self, driver, wait):
        """Faz o processo de assinatura com Dr. George"""
        try:
            wait.until(EC.presence_of_element_located((By.ID, "assinatura")))
            log_message("📋 Modal de assinatura aberto", "INFO")

            checkbox_george = wait.until(
                EC.element_to_be_clickable((By.XPATH, "//input[@type='checkbox' and @value='2173']"))
            )
            checkbox_george.click()
            log_message("✅ Checkbox do Dr. George marcado", "INFO")
            time.sleep(1)

            campo_senha = wait.until(
                EC.presence_of_element_located((By.NAME, "senha_2173"))
            )
            campo_senha.send_keys("1323")
            log_message("🔐 Senha digitada", "INFO")
            time.sleep(1)

            botao_assinar = wait.until(
                EC.element_to_be_clickable((By.ID, "salvarAss"))
            )
            botao_assinar.click()
            log_message("✍️ Clicou em Assinar", "INFO")
            time.sleep(1.5)

        except Exception as e:
            log_message(f"Erro no processo de assinatura: {e}", "ERROR")
            raise

    def processar_exame(self, driver, wait, num_exame, mascara, macroscopista, amg_maior, amg_menor, frag_ade, ade,
                        legenda, legenda_original, data_fixacao):
        """Processa um exame individual"""
        try:
            if not self.verificar_sessao_browser(driver):
                raise Exception("Sessão do browser perdida - necessário reiniciar")

            try:
                campo_codigo = wait.until(
                    EC.presence_of_element_located((By.XPATH, "//input[@placeholder='Leitor de código de barras']")))
                log_message("✅ Campo de código encontrado", "INFO")
            except:
                campo_codigo = wait.until(EC.presence_of_element_located((By.ID, "inputSearchCodBarra")))
                log_message("✅ Campo de código encontrado pelo ID", "INFO")

            campo_codigo.clear()
            campo_codigo.send_keys(num_exame)
            log_message(f"✍️ Código '{num_exame}' digitado no campo", "SUCCESS")

            try:
                botao_pesquisar = wait.until(EC.element_to_be_clickable((By.ID, "consultarExameBarraAbrirPorBarCode")))
                botao_pesquisar.click()
                log_message("🔍 Clicou no botão de pesquisar exame", "SUCCESS")
            except Exception as e:
                log_message(f"⚠️ Não foi possível clicar no botão de pesquisar: {e}", "WARNING")
                raise

            return self.aguardar_e_processar_andamento(driver, wait, mascara, macroscopista, amg_maior, amg_menor,
                                                       frag_ade, ade, legenda, legenda_original, data_fixacao)

        except Exception as e:
            error_message = str(e)
            log_message(f"Erro ao processar exame {num_exame}: {error_message}", "ERROR")

            if "invalid session id" in error_message.lower():
                log_message("❌ Erro de sessão inválida detectado", "ERROR")
                return {'status': 'erro_sessao', 'detalhes': 'Sessão do browser perdida'}

            try:
                screenshot_path = f"erro_exame_{num_exame}_{int(time.time())}.png"
                driver.save_screenshot(screenshot_path)
                log_message(f"Screenshot do erro salvo em: {screenshot_path}", "INFO")
            except:
                pass
            return {'status': 'erro', 'detalhes': error_message}

    def aguardar_e_processar_andamento(self, driver, wait, mascara, macroscopista, amg_maior, amg_menor, frag_ade, ade,
                                       legenda, legenda_original, data_fixacao):
        """Aguarda a div de andamento e processa o exame"""
        try:
            wait.until(EC.presence_of_element_located((By.ID, "divAndamentoExame")))
            log_message("📋 Div de andamento do exame encontrada!", "SUCCESS")
            time.sleep(0.5)
        except:
            log_message("⚠️ Div de andamento não apareceu no tempo esperado", "WARNING")
            return {'status': 'sem_andamento', 'detalhes': 'Exame não encontrado ou não carregou'}

        log_message("✅ Exame carregado - iniciando processo de conclusão", "SUCCESS")
        return self.processar_conclusao_completa(driver, wait, mascara, macroscopista, amg_maior, amg_menor, frag_ade,
                                                 ade, legenda, legenda_original, data_fixacao)

    def processar_conclusao_completa(self, driver, wait, mascara, macroscopista, amg_maior, amg_menor, frag_ade, ade, legenda, legenda_original, data_fixacao):
        try:
            # 1. Selecionar responsável pela macroscopia
            self.selecionar_responsavel_macroscopia(driver, wait, macroscopista)

            # 2. Definir data de fixação
            self.definir_data_fixacao(driver, wait, data_fixacao)

            # 3. Definir hora 18:00
            self.definir_hora_fixacao(driver, wait)

            # 4. Digitar a máscara e buscar
            if mascara:
                self.digitar_mascara_e_buscar(driver, wait, mascara)
            else:
                log_message("⚠️ Nenhuma máscara encontrada, pulando busca", "WARNING")

            # 5. Abrir modal de variáveis e preencher campos
            try:
                self.abrir_modal_variaveis_e_preencher(driver, wait, mascara, amg_maior, amg_menor, frag_ade, ade, legenda, legenda_original)
            except Exception as var_error:
                log_message(f"⚠️ Erro no modal de variáveis: {var_error}", "WARNING")
                log_message("⚠️ Continuando o processo sem as variáveis", "WARNING")

            # 6. Salvar macroscopia
            self.salvar_macroscopia(driver, wait)

            # 7. Preencher campos pré-envio (grupo, etc)
            try:
                self.preencher_campos_pre_envio(driver, wait, mascara, legenda)
            except Exception as campos_error:
                log_message(f"⚠️ Erro ao preencher campos pré-envio: {campos_error}", "WARNING")

            # 8. Salvar fragmentos
            self.salvar_fragmentos(driver, wait)

            # 9. Enviar para próxima etapa
            resultado_envio = self.enviar_proxima_etapa(driver, wait)

            if resultado_envio.get('status') == 'aguardando_assinatura':
                log_message("📋 Modal de assinatura aberto - iniciando processo de assinatura", "INFO")
                try:
                    self.assinar_com_george(driver, wait)
                    log_message("✅ Assinatura realizada com sucesso", "SUCCESS")
                    return {'status': 'sucesso', 'detalhes': 'Macroscopia assinada com sucesso'}
                except Exception as assinatura_error:
                    log_message(f"❌ Erro na assinatura: {assinatura_error}", "ERROR")
                    return {'status': 'erro_assinatura', 'detalhes': str(assinatura_error)}
            elif resultado_envio.get('status') == 'erro':
                log_message(f"⚠️ Erro no envio para próxima etapa: {resultado_envio.get('detalhes')}", "WARNING")
                return {'status': 'erro_envio', 'detalhes': resultado_envio.get('detalhes')}
            else:
                log_message("🎉 Processo de macroscopia finalizado com sucesso!", "SUCCESS")
                return {'status': 'sucesso', 'detalhes': 'Macroscopia processada com sucesso'}

        except Exception as e:
            log_message(f"Erro durante processo de macroscopia: {e}", "ERROR")
            return {'status': 'erro_macroscopia', 'detalhes': str(e)}

    def mostrar_resumo_final(self, resultados):
        """Mostra o resumo final do processamento"""
        total = len(resultados)
        sucesso = len([r for r in resultados if r['status'] == 'sucesso'])
        sem_andamento = len([r for r in resultados if r['status'] == 'sem_andamento'])
        erro_sessao = len([r for r in resultados if r['status'] == 'erro_sessao'])
        erros = len([r for r in resultados if 'erro' in r['status'] and r['status'] != 'erro_sessao'])

        log_message("\n" + "=" * 50, "INFO")
        log_message("RESUMO FINAL DO PROCESSAMENTO", "INFO")
        log_message("=" * 50, "INFO")
        log_message(f"Total de exames: {total}", "INFO")
        log_message(f"✅ Processados com sucesso: {sucesso}", "SUCCESS")
        log_message(f"⚠️ Exames não encontrados: {sem_andamento}", "WARNING")
        log_message(f"🔄 Erros de sessão (browser perdido): {erro_sessao}", "WARNING")
        log_message(f"❌ Outros erros de processamento: {erros}", "ERROR")

        if erro_sessao + erros > 0:
            log_message("\nDetalhes dos erros:", "ERROR")
            for r in resultados:
                if 'erro' in r['status']:
                    log_message(f"- {r.get('detalhes', 'Erro desconhecido')}", "ERROR")

        messagebox.showinfo("Processamento Concluído",
                            f"✅ Processamento finalizado!\n\n"
                            f"Total: {total}\n"
                            f"Sucesso: {sucesso}\n"
                            f"Não encontrados: {sem_andamento}\n"
                            f"Erros de sessão: {erro_sessao}\n"
                            f"Outros erros: {erros}")

    def run(self, params: dict):
        username = params.get("username")
        password = params.get("password")
        excel_file = params.get("excel_file")
        cancel_flag = params.get("cancel_flag")
        headless_mode = params.get("headless_mode")

        try:
            dados_exames = self.get_dados_exames(excel_file)
            if not dados_exames:
                messagebox.showerror("Erro", "Nenhum dado de exame encontrado na planilha.")
                return

            log_message(f"Encontrados {len(dados_exames)} exames para processar", "INFO")

        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao ler o Excel: {e}")
            return

        url = os.getenv("SYSTEM_URL", "https://pathoweb.com.br/login/auth")
        driver = None
        resultados = []

        try:
            driver = BrowserFactory.create_chrome(headless=headless_mode)
            wait = WebDriverWait(driver, 10)

            log_message("Iniciando automação de macroscopia amiade...", "INFO")

            # Login
            log_message("Fazendo login...", "INFO")
            driver.get(url)

            wait.until(EC.presence_of_element_located((By.ID, "username")))

            username_field = driver.find_element(By.ID, "username")
            username_field.clear()
            username_field.send_keys(username)

            password_field = driver.find_element(By.ID, "password")
            password_field.clear()
            password_field.send_keys(password)

            submit_button = driver.find_element(By.CSS_SELECTOR, "button[type='submit']")
            submit_button.click()

            log_message("Verificando se precisa navegar para módulo de exames...", "INFO")
            current_url = driver.current_url

            if current_url == "https://pathoweb.com.br/" or "trocarModulo" in current_url:
                log_message("Detectada tela de seleção de módulos - navegando para módulo de exames...", "INFO")
                try:
                    modulo_link = wait.until(
                        EC.element_to_be_clickable((By.CSS_SELECTOR, "a[href='/site/trocarModulo?modulo=1']")))
                    modulo_link.click()
                    time.sleep(2)
                    log_message("✅ Navegou para o módulo de exames", "SUCCESS")
                except Exception as e:
                    log_message(f"⚠️ Erro ao navegar para módulo: {e}", "WARNING")
                    driver.get("https://pathoweb.com.br/moduloExame/index")
                    time.sleep(2)

            elif "moduloExame" in current_url:
                log_message("✅ Já está no módulo de exames - pulando navegação", "SUCCESS")
            else:
                log_message(f"⚠️ URL inesperada detectada: {current_url}", "WARNING")
                driver.get("https://pathoweb.com.br/moduloExame/index")
                time.sleep(2)
                log_message("🔄 Navegação direta para módulo realizada (fallback)", "INFO")

            # Fechar modal se aparecer
            try:
                modal_close_button = driver.find_element(By.CSS_SELECTOR,
                                                         "#mensagemParaClienteModal .modal-footer button")
                if modal_close_button.is_displayed():
                    modal_close_button.click()
                    time.sleep(0.5)
            except Exception:
                pass

            log_message("✅ Login realizado com sucesso. Iniciando processamento dos exames.", "SUCCESS")

            # Processar cada exame da planilha
            for i, exame_data in enumerate(dados_exames, 1):
                if cancel_flag and cancel_flag.is_set():
                    log_message("⚠️ Processo cancelado pelo usuário", "WARNING")
                    break

                num_exame = exame_data['num_exame']
                mascara = exame_data['mascara']
                macroscopista = exame_data['macroscopista']
                amg_maior = exame_data['amg_maior']
                amg_menor = exame_data['amg_menor']
                frag_ade = exame_data['frag_ade']
                ade = exame_data['ade']
                legenda = exame_data['legenda']
                legenda_original = exame_data['legenda_original']
                data_fixacao = exame_data['data_fixacao']

                log_message(f"\n➡️ Processando exame {i}/{len(dados_exames)}: {num_exame} (máscara: {mascara})", "INFO")

                try:
                    resultado = self.processar_exame(driver, wait, num_exame, mascara, macroscopista,
                                                     amg_maior, amg_menor, frag_ade, ade, legenda, legenda_original, data_fixacao)
                    resultados.append(resultado)

                    if resultado['status'] == 'erro_sessao':
                        log_message("❌ Sessão perdida - abortando processamento", "ERROR")
                        break

                except Exception as e:
                    log_message(f"❌ Erro ao processar exame {num_exame}: {e}", "ERROR")
                    resultados.append({'status': 'erro', 'detalhes': str(e)})

            # Mostrar resumo final
            self.mostrar_resumo_final(resultados)

        except Exception as e:
            log_message(f"❌ Erro durante a automação: {e}", "ERROR")
            messagebox.showerror("Erro", f"❌ Erro durante a automação:\n{str(e)[:200]}...")
        finally:
            if driver:
                try:
                    driver.quit()
                    log_message("✅ Browser fechado com sucesso", "SUCCESS")
                except Exception as quit_error:
                    log_message(f"⚠️ Erro ao fechar browser: {quit_error}", "WARNING")


def run(params: dict):
    module = MacroAmiadeModule()
    module.run(params)