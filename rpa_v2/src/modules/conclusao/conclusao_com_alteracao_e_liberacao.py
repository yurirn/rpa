import os
import time
from tkinter import messagebox
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from dotenv import load_dotenv
from openpyxl import load_workbook

from src.core.browser_factory import BrowserFactory
from src.core.logger import log_message
from src.modules.base import BaseModule

load_dotenv()

class ConclusaoComAlteracaoELiberacaoModule(BaseModule):
    def __init__(self):
        super().__init__(nome="Conclusão com Alteração e Liberação")

    def get_dados_exames(self, file_path: str) -> list:
        """Lê do Excel apenas: código do exame, patologista e unimed.
        Se patologista/unimed estiverem vazios, herda o último valor não-vazio.
        """
        try:
            workbook = load_workbook(file_path)
            sheet = workbook.active
            dados: list[dict] = []

            ultimo_patologista = ""
            ultima_unimed = ""

            # Lê da linha 2 em diante (linha 1 é cabeçalho)
            for row in range(2, sheet.max_row + 1):
                valores_herdados = []

                codigo = sheet[f"A{row}"].value
                patologista = sheet[f"B{row}"].value
                unimed = sheet[f"C{row}"].value

                if codigo is None:
                    continue

                codigo = str(codigo).strip()
                if not codigo:
                    continue

                # Patologista: salva último não-vazio, senão herda
                if patologista is not None and str(patologista).strip():
                    patologista = str(patologista).strip()
                    ultimo_patologista = patologista
                else:
                    patologista = ultimo_patologista
                    if patologista:
                        valores_herdados.append(f"patologista='{patologista}'")

                # Unimed: salva último não-vazio, senão herda
                if unimed is not None and str(unimed).strip():
                    unimed = str(unimed).strip()
                    ultima_unimed = unimed
                else:
                    unimed = ultima_unimed
                    if unimed:
                        valores_herdados.append(f"unimed='{unimed}'")

                # Log quando valores são herdados
                if valores_herdados:
                    log_message(
                        f"📋 Linha {row}: Exame {codigo} herdou valores: {', '.join(valores_herdados)}",
                        "INFO",
                    )

                dados.append(
                    {
                        "codigo": codigo,
                        "patologista": patologista,
                        "unimed": unimed,
                    }
                )

            workbook.close()
            return dados
        except Exception as e:
            raise Exception(f"Erro ao ler planilha: {e}")

    def verificar_sessao_browser(self, driver) -> bool:
        try:
            driver.current_url
            return True
        except Exception as e:
            if "invalid session id" in str(e).lower():
                log_message("❌ Sessão do browser perdida", "ERROR")
                return False
            return True

    def enviar_proxima_etapa(self, driver, wait):
        """Clica no botão de enviar para próxima etapa"""
        try:
            botao_enviar = wait.until(
                EC.element_to_be_clickable((By.ID, "btn-enviar-proxima-etapa"))
            )
            botao_enviar.click()
            log_message("➡️ Clicou em Enviar para próxima etapa", "INFO")
            time.sleep(1.5)
        except Exception as e:
            log_message(f"Erro ao enviar para próxima etapa: {e}", "ERROR")
            raise

    def assinar_com_george(self, driver, wait):
        """Faz o processo de assinatura com Dr. George"""
        try:
            # Aguardar o modal de assinatura aparecer
            wait.until(EC.presence_of_element_located((By.ID, "assinatura")))
            log_message("📋 Modal de assinatura aberto", "INFO")

            # Aguardar o campo de senha aparecer e digitar a senha
            campo_senha = wait.until(
                EC.presence_of_element_located((By.NAME, "senha_2173"))
            )
            campo_senha.send_keys("1323")
            log_message("🔐 Senha digitada", "INFO")
            time.sleep(1)

            # Clicar no botão Assinar
            botao_assinar = wait.until(
                EC.element_to_be_clickable((By.ID, "salvarAss"))
            )
            botao_assinar.click()
            log_message("✍️ Clicou em Assinar", "INFO")
            time.sleep(1.5)

        except Exception as e:
            log_message(f"Erro no processo de assinatura: {e}", "ERROR")
            raise

    def processar_exame(self, driver, wait, codigo, patologista, unimed):
        """
        Novo fluxo:
        - digitar código e abrir exame
        - aguardar usuário clicar "Enviar para próxima etapa"
        - quando abrir modal de assinatura, assinar conforme regra
        """
        try:
            if not self.verificar_sessao_browser(driver):
                raise Exception("Sessão do browser perdida - necessário reiniciar")

            campo_codigo = wait.until(
                EC.presence_of_element_located((By.ID, "inputSearchCodBarra"))
            )
            self.interagir_com_campo_codigo(driver, campo_codigo, codigo)

            # Esperar usuário enviar para próxima etapa (não clicar automaticamente)
            if not self.aguardar_usuario_enviar_para_proxima_etapa(driver, wait, codigo, timeout=600):
                return {"status": "timeout", "detalhes": "Usuário não enviou para próxima etapa no tempo esperado"}

            is_unimed = (unimed or "").strip().lower() == "sim"
            self.processar_assinatura(driver, wait, patologista, is_unimed)

            log_message("🎉 Exame processado com sucesso!", "SUCCESS")
            return {"status": "sucesso", "detalhes": "Exame enviado pelo usuário e assinado"}

        except Exception as e:
            error_message = str(e)
            log_message(f"❌ Erro ao processar exame {codigo}: {error_message}", "ERROR")

            if "invalid session id" in error_message.lower():
                return {"status": "erro_sessao", "detalhes": "Sessão do browser perdida"}

            return {"status": "erro", "detalhes": error_message}

    def interagir_com_campo_codigo(self, driver, campo_codigo, codigo):
        log_message("Campo de código encontrado, interagindo...", "INFO")

        try:
            campo_codigo.clear()
            log_message("Campo limpo com sucesso", "INFO")
        except Exception:
            driver.execute_script("arguments[0].value = '';", campo_codigo)
            log_message("Campo limpo com JavaScript", "INFO")

        time.sleep(0.5)

        try:
            campo_codigo.send_keys(codigo)
            log_message(f"Código '{codigo}' digitado com sucesso", "INFO")
        except Exception:
            driver.execute_script(f"arguments[0].value = '{codigo}';", campo_codigo)
            driver.execute_script(
                """
                var element = arguments[0];
                var event = new Event('input', { bubbles: true });
                element.dispatchEvent(event);
                """,
                campo_codigo,
            )
            log_message(f"Código '{codigo}' digitado com JavaScript", "INFO")

        time.sleep(1)

        try:
            campo_codigo.send_keys(Keys.ENTER)
            log_message("⌨️ Enter pressionado com sucesso", "INFO")
        except Exception:
            driver.execute_script(
                """
                var element = arguments[0];
                var event = new KeyboardEvent('keydown', {
                    key: 'Enter',
                    code: 'Enter',
                    keyCode: 13,
                    bubbles: true
                });
                element.dispatchEvent(event);
                """,
                campo_codigo,
            )
            log_message("⌨️ Enter pressionado com JavaScript", "INFO")

        # Aguardar div de andamento aparecer (mantém seu comportamento)
        log_message("Aguardando div de andamento do exame aparecer...", "INFO")
        timeout_andamento = 30
        inicio = time.time()

        while time.time() - inicio < timeout_andamento:
            try:
                andamento_div = driver.find_element(By.ID, "divAndamentoExame")
                if andamento_div and andamento_div.is_displayed():
                    log_message("📋 Div de andamento do exame encontrada!", "SUCCESS")
                    break
            except Exception:
                pass

            time.sleep(1)
            if int(time.time() - inicio) % 5 == 0:
                log_message(
                    f"⏳ Aguardando carregamento... ({int(time.time() - inicio)}s)",
                    "INFO",
                )
        else:
            log_message("⚠️ Div de andamento não apareceu no tempo esperado", "WARNING")
            return {"status": "sem_andamento", "detalhes": "Exame não encontrado ou não carregou"}

        time.sleep(1)
        return {"status": "ok"}

    def aguardar_usuario_enviar_para_proxima_etapa(self, driver, wait, codigo, timeout=600):
        """
        Aguarda o usuário clicar manualmente em "Enviar para próxima etapa".
        Critério robusto: aparecer o modal de assinatura (#assinatura).
        """
        try:
            log_message(f"⏳ Aguardando usuário clicar em Enviar para próxima etapa (exame {codigo})...", "WARNING")
            log_message(f"⏳ Timeout: {timeout}s ({timeout // 60} minutos)", "INFO")

            inicio = time.time()
            contador_log = 0

            while time.time() - inicio < timeout:
                try:
                    assinatura_modal = driver.find_element(By.ID, "assinatura")
                    if assinatura_modal.is_displayed():
                        log_message("📋 Modal de assinatura detectado (usuário enviou para próxima etapa)", "SUCCESS")
                        return True
                except Exception:
                    pass

                tempo_decorrido = int(time.time() - inicio)
                if tempo_decorrido > contador_log and tempo_decorrido % 30 == 0:
                    minutos = tempo_decorrido // 60
                    segundos = tempo_decorrido % 60
                    log_message(f"⏳ Aguardando envio... ({minutos}m {segundos}s)", "INFO")
                    contador_log = tempo_decorrido

                time.sleep(0.2)

            log_message(
                f"⚠️ Timeout de {timeout}s atingido - usuário não enviou o exame {codigo} para próxima etapa",
                "WARNING",
            )
            return False

        except Exception as e:
            log_message(f"❌ Erro ao aguardar envio: {e}", "ERROR")
            return False

    def get_patologista_info(self, nome_patologista):
        """
        Mapeamento: NOME -> (checkbox_value, senha)
        Reaproveite do `conclusao.py` e complete os values corretos.
        """
        patologistas = {
            "GEORGE": ("2173", "1323"),
            "LEANDRO": ("73069", "1308"),
            "MIRELLA": ("269762", "6523"),
            "MARINA": ("269765", "1404"),
            "ARYELA": ("306997", "1209"),
        }

        nome_upper = (nome_patologista or "").upper().strip()
        if nome_upper in patologistas:
            return patologistas[nome_upper]

        log_message(f"⚠️ Patologista '{nome_patologista}' não encontrado no mapeamento", "WARNING")
        return None

    def assinar_com_patologista(self, driver, wait, nome_patologista, checkbox_value, senha):
        try:
            log_message(f"📝 Assinando com {nome_patologista}...", "INFO")

            checkbox = wait.until(
                EC.element_to_be_clickable(
                    (By.XPATH, f"//input[@type='checkbox' and @value='{checkbox_value}']")
                )
            )
            checkbox.click()
            log_message(f"✅ Checkbox de {nome_patologista} marcado", "INFO")
            time.sleep(0.5)

            campo_senha = wait.until(
                EC.presence_of_element_located((By.NAME, f"senha_{checkbox_value}"))
            )
            campo_senha.clear()
            campo_senha.send_keys(senha)
            log_message(f"🔐 Senha de {nome_patologista} digitada", "INFO")
            time.sleep(0.5)

        except Exception as e:
            log_message(f"Erro ao assinar com {nome_patologista}: {e}", "ERROR")
            raise

    def processar_assinatura(self, driver, wait, patologista, is_unimed: bool):
        """
        Regra:
        - sempre assina com o patologista da planilha
        - se UNIMED=sim, assina também com George
        - depois clica em "Assinar" (salvarAss)
        """
        try:
            wait.until(EC.presence_of_element_located((By.ID, "assinatura")))
            log_message("📋 Modal de assinatura aberto", "INFO")

            info_patologista = self.get_patologista_info(patologista)
            if not info_patologista:
                raise Exception(f"Patologista '{patologista}' não encontrado no sistema")

            checkbox_patologista, senha_patologista = info_patologista
            self.assinar_com_patologista(driver, wait, patologista, checkbox_patologista, senha_patologista)

            if is_unimed:
                info_george = self.get_patologista_info("GEORGE")
                if not info_george:
                    raise Exception("George não encontrado no mapeamento")
                checkbox_george, senha_george = info_george
                self.assinar_com_patologista(driver, wait, "Dr. George", checkbox_george, senha_george)

            botao_assinar = wait.until(EC.element_to_be_clickable((By.ID, "salvarAss")))
            botao_assinar.click()
            log_message("✍️ Clicou em Assinar", "INFO")
            time.sleep(1.5)

        except Exception as e:
            log_message(f"Erro no processo de assinatura: {e}", "ERROR")
            raise

    def mostrar_resumo_final(self, resultados):
        total = len(resultados)
        sucesso = len([r for r in resultados if r["status"] == "sucesso"])
        timeout = len([r for r in resultados if r["status"] == "timeout"])
        erros = len([r for r in resultados if "erro" in r["status"]])

        log_message("\n" + "=" * 50, "INFO")
        log_message("RESUMO FINAL DO PROCESSAMENTO", "INFO")
        log_message("=" * 50, "INFO")
        log_message(f"Total de exames: {total}", "INFO")
        log_message(f"✅ Processados com sucesso: {sucesso}", "SUCCESS")
        log_message(f"⏱️ Timeout (usuário não enviou): {timeout}", "WARNING")
        log_message(f"❌ Erros de processamento: {erros}", "ERROR")

        if erros > 0:
            log_message("\nDetalhes dos erros:", "ERROR")
            for r in resultados:
                if "erro" in r["status"]:
                    log_message(f"- {r['codigo']}: {r['detalhes']}", "ERROR")

        messagebox.showinfo(
            "Processamento Concluído",
            f"✅ Processamento finalizado!\n\n"
            f"Total: {total}\n"
            f"Sucesso: {sucesso}\n"
            f"Timeout: {timeout}\n"
            f"Erros: {erros}",
        )

    def run(self, params: dict):
        username = params.get("username")
        password = params.get("password")
        excel_file = params.get("excel_file")
        cancel_flag = params.get("cancel_flag")

        try:
            dados_exames = self.get_dados_exames(excel_file)
            if not dados_exames:
                messagebox.showerror("Erro", "Nenhum dado de exame encontrado na planilha.")
                return
            log_message(f"Encontrados {len(dados_exames)} exames para processar", "INFO")
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao ler o Excel: {e}")
            return

        url = os.getenv("SYSTEM_URL", "https://dap.pathoweb.com.br/login/auth")
        driver = None
        resultados = []

        try:
            driver = BrowserFactory.create_chrome()
            wait = WebDriverWait(driver, 20)

            log_message("Iniciando automação de conclusão com alteração e liberação...", "INFO")

            driver.get(url)
            wait.until(EC.presence_of_element_located((By.ID, "username")))

            username_field = driver.find_element(By.ID, "username")
            username_field.clear()
            username_field.send_keys(username)

            password_field = driver.find_element(By.ID, "password")
            password_field.clear()
            password_field.send_keys(password)

            submit_button = driver.find_element(By.CSS_SELECTOR, "button[type='submit']")
            submit_button.click()

            # Navegação módulo (mantém seu comportamento)
            log_message("Verificando se precisa navegar para módulo de exames...", "INFO")
            current_url = driver.current_url
            if current_url == "https://dap.pathoweb.com.br/" or "trocarModulo" in current_url:
                try:
                    modulo_link = wait.until(
                        EC.element_to_be_clickable((By.CSS_SELECTOR, "a[href='/site/trocarModulo?modulo=1']"))
                    )
                    modulo_link.click()
                    time.sleep(2)
                    log_message("✅ Navegação para módulo de exames realizada", "SUCCESS")
                except Exception:
                    driver.get("https://dap.pathoweb.com.br/moduloExame/index")
                    time.sleep(2)

            # Fechar modal se aparecer
            try:
                modal_close_button = driver.find_element(
                    By.CSS_SELECTOR, "#mensagemParaClienteModal .modal-footer button"
                )
                if modal_close_button.is_displayed():
                    modal_close_button.click()
            except Exception:
                pass

            log_message("✅ Login realizado com sucesso. Iniciando processamento dos exames.", "SUCCESS")

            for i, exame_data in enumerate(dados_exames, 1):
                if cancel_flag and cancel_flag.is_set():
                    log_message("Execução cancelada pelo usuário.", "WARNING")
                    break

                codigo = exame_data["codigo"]
                patologista = exame_data["patologista"]
                unimed = exame_data["unimed"]

                log_message(
                    f"\n➡️ Processando exame {i}/{len(dados_exames)}: {codigo} (patologista: {patologista}, unimed: {unimed})",
                    "INFO",
                )

                resultado = self.processar_exame(driver, wait, codigo, patologista, unimed)
                resultados.append(
                    {
                        "codigo": codigo,
                        "status": resultado["status"],
                        "detalhes": resultado.get("detalhes", ""),
                    }
                )

            self.mostrar_resumo_final(resultados)

        except Exception as e:
            log_message(f"❌ Erro durante a automação: {e}", "ERROR")
            messagebox.showerror("Erro", f"❌ Erro durante a automação:\n{str(e)[:200]}...")
        finally:
            log_message("✅ Execução finalizada - Browser permanece aberto", "SUCCESS")

def run(params: dict):
    module = ConclusaoComAlteracaoELiberacaoModule()
    module.run(params)
