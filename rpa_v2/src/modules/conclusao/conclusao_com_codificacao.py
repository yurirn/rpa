import os
import time
from tkinter import messagebox
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from dotenv import load_dotenv
from openpyxl import load_workbook

from src.core.browser_factory import BrowserFactory
from src.core.logger import log_message
from src.modules.base import BaseModule

load_dotenv()

class ConclusaoModule(BaseModule):
    def __init__(self):
        super().__init__(nome="Conclusão")

    def get_dados_exames(self, file_path: str) -> list:
        try:
            workbook = load_workbook(file_path)
            sheet = workbook.active
            dados = []
            ultima_mascara = None
            ultimo_codigo_procedimento = None

            # Lê da linha 2 em diante (linha 1 é cabeçalho)
            for row in range(2, sheet.max_row + 1):
                codigo = sheet[f'A{row}'].value
                mascara = sheet[f'B{row}'].value
                codigo_procedimento = sheet[f'C{row}'].value

                if codigo is not None:
                    codigo = str(codigo).strip()
                    valores_herdados = []

                    # Se não tem máscara, usa a última válida
                    if mascara is not None and str(mascara).strip():
                        mascara = str(mascara).strip()
                        ultima_mascara = mascara
                    else:
                        mascara = ultima_mascara
                        if mascara:
                            valores_herdados.append(f"máscara='{mascara}'")

                    # Se não tem patologista, usa o último válido
                    if codigo_procedimento is not None and str(codigo_procedimento).strip():
                        codigo_procedimento = str(codigo_procedimento).strip()
                        ultimo_codigo_procedimento = codigo_procedimento
                    else:
                        codigo_procedimento = ultimo_codigo_procedimento if ultimo_codigo_procedimento else ""
                        if codigo_procedimento:
                            valores_herdados.append(f"Código Procedimento='{codigo_procedimento}'")

                    # Log quando valores são herdados
                    if valores_herdados:
                        log_message(f"📋 Linha {row}: Exame {codigo} herdou valores: {', '.join(valores_herdados)}",
                                    "INFO")

                    dados.append({
                        'codigo': codigo,
                        'mascara': mascara,
                        'codigo_procedimento': codigo_procedimento
                    })

            workbook.close()
            return dados
        except Exception as e:
            raise Exception(f"Erro ao ler planilha: {e}")

    def verificar_sessao_browser(self, driver) -> bool:
        """Verifica se a sessão do browser ainda está ativa"""
        try:
            driver.current_url
            return True
        except Exception as e:
            if "invalid session id" in str(e).lower():
                log_message("❌ Sessão do browser perdida", "ERROR")
                return False
            return True

    def verificar_svg_conclusao(self, driver) -> bool:
        """Verifica se existe o SVG na etapa Conclusão"""
        try:
            estrategias_busca = [
                (
                    "data-id",
                    "//a[@data-id='C']"
                ),
                (
                    "data-stringjson",
                    "//a[contains(@data-stringjson, '\"C\"')]"
                ),
                (
                    "texto",
                    "//a[contains(normalize-space(translate(., 'ABCDEFGHIJKLMNOPQRSTUVWXYZÁÉÍÓÚÂÊÔÃÕÇ', "
                    "'abcdefghijklmnopqrstuvwxyzáéíóúâêôãõç')), 'conclus')]"
                )
            ]

            for identificador, xpath in estrategias_busca:
                try:
                    links = driver.find_elements(By.XPATH, xpath)
                    for link in links:
                        try:
                            classe_link = (link.get_attribute("class") or "").lower()
                            if "disabled" in classe_link or "inactive" in classe_link:
                                continue
                        except Exception:
                            pass

                        try:
                            texto_link = (link.text or "").lower()
                        except Exception:
                            texto_link = ""

                        if identificador == "texto" and "conclus" not in texto_link:
                            continue

                        svgs = link.find_elements(By.TAG_NAME, "svg")
                        if svgs:
                            log_message(f"✅ SVG encontrado na etapa Conclusão (método: {identificador})", "INFO")
                            return True
                except Exception:
                    continue

            # Método extra: procurar qualquer SVG arrow-right e verificar ancestrais com texto Conclusão
            try:
                svg_arrows = driver.find_elements(By.XPATH, "//svg[@data-icon='arrow-right']")
                for svg in svg_arrows:
                    try:
                        parent = svg.find_element(By.XPATH, "./ancestor::a[1]")
                        texto_parent = (parent.text or "").lower()
                        if "conclus" in texto_parent:
                            log_message("✅ SVG arrow-right encontrado próximo à Conclusão (método: ancestral)", "INFO")
                            return True
                    except Exception:
                        continue
            except Exception:
                pass

            log_message("⚠️ SVG não encontrado na etapa Conclusão", "WARNING")
            return False

        except Exception as e:
            log_message(f"Erro ao verificar SVG conclusão: {e}", "ERROR")
            return False

    def fechar_exame(self, driver, wait):
        """Clica no botão de fechar exame"""
        try:
            botao_fechar = wait.until(
                EC.element_to_be_clickable((By.ID, "fecharExameBarraFerramenta"))
            )
            botao_fechar.click()
            log_message("📁 Exame fechado (sem SVG na conclusão)", "INFO")

            # Aguardar retornar à tela principal
            try:
                # Verificar se voltou à tela principal aguardando o campo de código aparecer
                wait.until(EC.presence_of_element_located((By.ID, "inputSearchCodBarra")))
                log_message("✅ Retornou à tela principal após fechar exame", "INFO")
            except:
                log_message("⚠️ Pode não ter retornado à tela principal", "WARNING")
                # Tentar navegar de volta ao módulo se necessário
                try:
                    current_url = driver.current_url
                    if "modulo=1" not in current_url:
                        modulo_link = driver.find_element(By.CSS_SELECTOR, "a[href='/site/trocarModulo?modulo=1']")
                        modulo_link.click()
                        time.sleep(1.5)
                        log_message("🔄 Navegou de volta ao módulo de exames", "INFO")
                except:
                    pass

        except Exception as e:
            log_message(f"Erro ao fechar exame: {e}", "ERROR")

    def digitar_mascara_e_buscar(self, driver, wait, mascara):
        """Digita a máscara no campo buscaArvore e pressiona Enter"""
        try:
            # Aguardar o campo estar presente e visível com timeout maior
            log_message(f"🔍 Procurando campo buscaArvore...", "INFO")
            campo_busca = wait.until(EC.element_to_be_clickable((By.ID, "buscaArvore")))
            log_message(f"🔍 Campo buscaArvore encontrado e clicável", "INFO")

            # Verificar se o campo está visível
            if not campo_busca.is_displayed():
                log_message("⚠️ Campo buscaArvore não está visível", "WARNING")
                return

            # Digitar a máscara
            campo_busca.send_keys(mascara)
            log_message(f"✍️ Máscara '{mascara}' digitada no campo buscaArvore", "INFO")
            time.sleep(0.5)

            # Pressionar Enter
            campo_busca.send_keys(Keys.ENTER)
            log_message(f"⌨️ Enter pressionado após digitar máscara", "INFO")
            time.sleep(1)

        except Exception as e:
            log_message(f"Erro ao digitar máscara: {e}", "ERROR")
            # Tentar encontrar o campo de outra forma
            try:
                # Verificar se existe campo com classe específica
                campos_alternativos = driver.find_elements(By.XPATH, "//input[@class='btn-xs' and @type='text']")
                log_message(f"Encontrados {len(campos_alternativos)} campos alternativos", "INFO")

                if campos_alternativos:
                    campo_alternativo = campos_alternativos[0]
                    campo_alternativo.click()
                    campo_alternativo.clear()
                    campo_alternativo.send_keys(mascara)
                    campo_alternativo.send_keys(Keys.ENTER)
                    log_message(f"✅ Máscara digitada usando campo alternativo", "INFO")
                    return

                # Listar todos os inputs para debug
                inputs = driver.find_elements(By.TAG_NAME, "input")
                log_message(f"Total de inputs encontrados na página: {len(inputs)}", "INFO")
                for i, inp in enumerate(inputs[:10]):  # Apenas os primeiros 10
                    input_id = inp.get_attribute("id")
                    input_class = inp.get_attribute("class")
                    input_type = inp.get_attribute("type")
                    log_message(f"Input {i}: id='{input_id}', class='{input_class}', type='{input_type}'", "INFO")

            except Exception as debug_e:
                log_message(f"Erro no debug: {debug_e}", "ERROR")
            raise

    def salvar_conclusao(self, driver, wait):
        """Clica no botão Salvar"""
        try:
            # Aguardar o botão estar presente e clicável
            log_message("💾 Procurando botão Salvar...", "INFO")
            botao_salvar = wait.until(EC.element_to_be_clickable((By.ID, "salvarConcl")))
            log_message("💾 Botão Salvar encontrado e clicável", "INFO")

            # Verificar se o botão está visível
            if not botao_salvar.is_displayed():
                log_message("⚠️ Botão salvarConcl não está visível", "WARNING")
                return

            # Rolar até o botão para garantir visibilidade
            driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", botao_salvar)
            time.sleep(1)

            # Clicar no botão
            botao_salvar.click()
            log_message("💾 Clicou em Salvar", "INFO")
            time.sleep(1)

        except Exception as e:
            log_message(f"Erro ao salvar: {e}", "ERROR")
            # Tentar encontrar o botão de outra forma
            try:
                # Tentar por link com onclick específico
                botoes_onclick = driver.find_elements(By.XPATH, "//a[contains(@onclick, 'ajaxChangeSave')]")
                log_message(f"Encontrados {len(botoes_onclick)} botões com onclick ajaxChangeSave", "INFO")

                if botoes_onclick:
                    botao_onclick = botoes_onclick[0]
                    driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});",
                                          botao_onclick)
                    time.sleep(1)
                    botao_onclick.click()
                    log_message("💾 Clicou em Salvar usando onclick", "INFO")
                    return

                # Tentar por classe do botão
                botoes_classe = driver.find_elements(By.XPATH,
                                                     "//a[contains(@class, 'btn-primary') and contains(text(), 'Salvar')]")
                log_message(f"Encontrados {len(botoes_classe)} botões com classe btn-primary e texto Salvar", "INFO")

                if botoes_classe:
                    botao_classe = botoes_classe[0]
                    driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});",
                                          botao_classe)
                    time.sleep(1)
                    botao_classe.click()
                    log_message("💾 Clicou em Salvar usando classe", "INFO")
                    return

                # Listar todos os links/botões para debug
                links = driver.find_elements(By.TAG_NAME, "a")
                log_message(f"Total de links encontrados na página: {len(links)}", "INFO")
                for i, link in enumerate(links[:15]):  # Apenas os primeiros 15
                    link_id = link.get_attribute("id")
                    link_class = link.get_attribute("class")
                    link_text = link.text.strip()
                    link_onclick = link.get_attribute("onclick")
                    if (link_id and "salvar" in link_id.lower()) or \
                            (link_class and "salvar" in link_class.lower()) or \
                            (link_text and "salvar" in link_text.lower()) or \
                            (link_onclick and "save" in link_onclick.lower()):
                        log_message(
                            f"Link {i}: id='{link_id}', class='{link_class}', text='{link_text}', onclick='{link_onclick}'",
                            "INFO")

            except Exception as debug_e:
                log_message(f"Erro no debug de botões: {debug_e}", "ERROR")
            raise

    def enviar_proxima_etapa(self, driver, wait):
        """Clica no botão de enviar para próxima etapa"""
        try:
            botao_enviar = wait.until(
                EC.element_to_be_clickable((By.ID, "btn-enviar-proxima-etapa"))
            )
            botao_enviar.click()
            log_message("➡️ Clicou em Enviar para próxima etapa", "INFO")
            time.sleep(1.5)
        except Exception as e:
            log_message(f"Erro ao enviar para próxima etapa: {e}", "ERROR")
            raise

    def get_patologista_info(self, nome_patologista):
        """Retorna as informações do patologista (checkbox value e senha)"""
        # Mapeamento de patologistas (nome em maiúsculo -> (checkbox_value, senha))
        # NOTA: Os valores dos checkboxes precisam ser confirmados no sistema
        patologistas = {
            'GEORGE': ('2173', '1323'),
            'LEANDRO': ('73069', '1308'),  # Substitua XXXXX pelo valor correto do checkbox
            'MIRELLA': ('269762', '6523'),  # Substitua XXXXX pelo valor correto do checkbox
            'MARINA': ('269765', '1404'),  # Substitua XXXXX pelo valor correto do checkbox
            'ARYELA': ('306997', '1209'),  # Substitua XXXXX pelo valor correto do checkbox
        }

        nome_upper = nome_patologista.upper().strip()

        if nome_upper in patologistas:
            return patologistas[nome_upper]
        else:
            log_message(f"⚠️ Patologista '{nome_patologista}' não encontrado no mapeamento", "WARNING")
            return None

    def assinar_com_patologista(self, driver, wait, nome_patologista, checkbox_value, senha):
        """Assina com um patologista específico"""
        try:
            log_message(f"📝 Assinando com {nome_patologista}...", "INFO")

            # Encontrar e clicar no checkbox do patologista
            checkbox = wait.until(
                EC.element_to_be_clickable((By.XPATH, f"//input[@type='checkbox' and @value='{checkbox_value}']"))
            )
            checkbox.click()
            log_message(f"✅ Checkbox de {nome_patologista} marcado", "INFO")
            time.sleep(1)

            # Aguardar o campo de senha aparecer e digitar a senha
            campo_senha = wait.until(
                EC.presence_of_element_located((By.NAME, f"senha_{checkbox_value}"))
            )
            campo_senha.send_keys(senha)
            log_message(f"🔐 Senha de {nome_patologista} digitada", "INFO")
            time.sleep(1)

        except Exception as e:
            log_message(f"Erro ao assinar com {nome_patologista}: {e}", "ERROR")
            raise

    def processar_assinatura(self, driver, wait):
        """Processa a assinatura de acordo com as regras"""
        try:
            # Aguardar o modal de assinatura aparecer
            wait.until(EC.presence_of_element_located((By.ID, "assinatura")))
            log_message("📋 Modal de assinatura aberto", "INFO")

            # Obter informações do patologista
            info_patologista = self.get_patologista_info('GEORGE')
            if not info_patologista:
                raise Exception(f"Patologista GEORGE não encontrado no sistema")

            checkbox_patologista, senha_patologista = info_patologista

            # Sempre assina com o patologista primeiro
            self.assinar_com_patologista(driver, wait, 'Dr. George', checkbox_patologista, senha_patologista)

            # Clicar no botão Assinar
            botao_assinar = wait.until(
                EC.element_to_be_clickable((By.ID, "salvarAss"))
            )
            botao_assinar.click()
            log_message("✍️ Clicou em Assinar", "INFO")
            time.sleep(1.5)

        except Exception as e:
            log_message(f"Erro no processo de assinatura: {e}", "ERROR")
            raise

    def processar_visualizacao_laudos_final(self, driver, wait, dados_exames):
        """Processa a etapa final de visualização de laudos em lote"""
        try:
            log_message("\n" + "=" * 50, "INFO")
            log_message("INICIANDO PROCESSO DE VISUALIZAÇÃO DE LAUDOS", "INFO")
            log_message("=" * 50, "INFO")

            # 1. Marcar checkbox acumular
            self.marcar_checkbox_acumular(driver, wait)

            # 2. Acumular todos os exames no formulário
            self.acumular_exames_no_formulario(driver, wait, dados_exames)

            # 3. Selecionar todos os exames
            self.selecionar_todos_exames(driver, wait)

            # 4. Clicar no botão de ações
            self.clicar_botao_acoes(driver, wait)

            # 5. Clicar na opção Laudos
            self.clicar_opcao_laudos(driver, wait)

            log_message("\n" + "=" * 50, "SUCCESS")
            log_message("PROCESSO DE VISUALIZAÇÃO DE LAUDOS CONCLUÍDO", "SUCCESS")
            log_message("=" * 50 + "\n", "SUCCESS")

        except Exception as e:
            log_message(f"❌ Erro durante processo de visualização de laudos: {e}", "ERROR")
            raise

    def clicar_opcao_laudos(self, driver, wait):
        """Clica na opção 'Laudos' do dropdown"""
        try:
            log_message("📄 Clicando na opção 'Laudos'...", "INFO")

            # Encontrar o link "Laudos" no dropdown
            link_laudos = wait.until(
                EC.element_to_be_clickable(
                    (By.XPATH, "//a[@data-url='/moduloExame/visualizarLaudosAjax' and @data-index='2plus']"))
            )

            # Clicar no link
            link_laudos.click()
            log_message("✅ Opção 'Laudos' clicada", "SUCCESS")
            time.sleep(3)

            # Aguardar o popup abrir (se necessário)
            log_message("⏳ Aguardando processamento...", "INFO")
            time.sleep(2)

        except Exception as e:
            log_message(f"Erro ao clicar na opção Laudos: {e}", "ERROR")
            raise

    def clicar_botao_acoes(self, driver, wait):
        """Clica no botão de ações (engrenagem)"""
        try:
            log_message("⚙️ Clicando no botão Ações...", "INFO")

            # Encontrar o botão de ações
            botao_acoes = wait.until(
                EC.element_to_be_clickable(
                    (By.XPATH, "//a[@class='btn btn-default dropdown-toggle btn-sm' and @data-toggle='dropdown']"))
            )

            # Rolar até o botão
            driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", botao_acoes)
            time.sleep(1)

            # Clicar no botão
            botao_acoes.click()
            log_message("✅ Botão Ações clicado, dropdown aberto", "SUCCESS")
            time.sleep(1.5)

        except Exception as e:
            log_message(f"Erro ao clicar no botão Ações: {e}", "ERROR")
            raise

    def selecionar_todos_exames(self, driver, wait):
        """Clica no checkbox 'markAll' para selecionar todos os exames"""
        try:
            log_message("☑️ Selecionando todos os exames...", "INFO")

            # Verificar quantos exames foram acumulados
            try:
                tbody = driver.find_element(By.ID, "tabelaLocalizarExamesTbody")
                linhas = tbody.find_elements(By.TAG_NAME, "tr")
                log_message(f"📊 Encontrados {len(linhas)} exames na tabela", "INFO")

                if len(linhas) == 0:
                    raise Exception("Nenhum exame foi acumulado na tabela!")
            except Exception as e:
                log_message(f"⚠️ Erro ao verificar tabela: {e}", "WARNING")

            # Tentar múltiplos métodos para selecionar todos
            success = False

            # Método 1: Click no checkbox markAll via seletor CSS
            try:
                checkbox_mark_all = wait.until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, "input.markAll"))
                )

                # Rolar até o checkbox
                driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});",
                                      checkbox_mark_all)
                time.sleep(0.5)

                # Clicar
                checkbox_mark_all.click()
                time.sleep(1.5)
                log_message("✅ Método 1: Click no markAll executado", "SUCCESS")
                success = True
            except Exception as e1:
                log_message(f"⚠️ Método 1 falhou: {e1}", "WARNING")

                # Método 2: Click via JavaScript
                try:
                    driver.execute_script("""
                        var checkbox = document.querySelector('input.markAll');
                        if (checkbox) {
                            checkbox.click();
                        }
                    """)
                    time.sleep(1.5)
                    log_message("✅ Método 2: Click via JavaScript executado", "SUCCESS")
                    success = True
                except Exception as e2:
                    log_message(f"⚠️ Método 2 falhou: {e2}", "WARNING")

                    # Método 3: Selecionar todos manualmente
                    try:
                        checkboxes = driver.find_elements(By.CSS_SELECTOR, "input.escolhaExameLote")
                        log_message(f"📋 Selecionando {len(checkboxes)} exames manualmente...", "INFO")
                        for cb in checkboxes:
                            if not cb.is_selected():
                                cb.click()
                                time.sleep(0.2)
                        log_message("✅ Método 3: Seleção manual executada", "SUCCESS")
                        success = True
                    except Exception as e3:
                        log_message(f"❌ Método 3 falhou: {e3}", "ERROR")

            if success:
                # Verificar quantos foram selecionados
                try:
                    selecionados = driver.find_elements(By.CSS_SELECTOR, "input.escolhaExameLote:checked")
                    log_message(f"✔️ CONFIRMADO: {len(selecionados)} exames selecionados!", "SUCCESS")
                except:
                    pass
            else:
                raise Exception("Não foi possível selecionar os exames")

        except Exception as e:
            log_message(f"❌ Erro ao selecionar todos os exames: {e}", "ERROR")
            raise

    def marcar_checkbox_acumular(self, driver, wait):
        """Marca o checkbox 'acumular' na tela principal"""
        try:
            log_message("📌 Marcando checkbox 'acumular'...", "INFO")

            # Aguardar o checkbox estar presente
            checkbox = wait.until(EC.presence_of_element_located((By.ID, "acumular")))
            time.sleep(1)

            # Verificar se já está marcado (pela classe do wrapper do iCheck)
            try:
                wrapper = driver.find_element(By.XPATH,
                                              "//input[@id='acumular']/parent::div[contains(@class, 'icheckbox')]")
                wrapper_classes = wrapper.get_attribute("class")
                is_checked_visually = "checked" in wrapper_classes
            except:
                # Se não encontrar o wrapper, verificar pelo checkbox mesmo
                is_checked_visually = driver.execute_script("return document.getElementById('acumular').checked;")

            if not is_checked_visually:
                log_message("🖱️ Tentando marcar checkbox usando iCheck...", "INFO")

                # Tentar método 1: Trigger do iCheck via jQuery
                try:
                    driver.execute_script("""
                        $('#acumular').iCheck('check');
                    """)
                    time.sleep(2)
                    log_message("✅ Método 1: iCheck check() executado", "SUCCESS")
                except Exception as e1:
                    log_message(f"⚠️ Método 1 falhou: {e1}", "WARNING")

                    # Tentar método 2: Clicar no wrapper da div
                    try:
                        wrapper = driver.find_element(By.XPATH,
                                                      "//input[@id='acumular']/following-sibling::ins[@class='iCheck-helper']")
                        wrapper.click()
                        time.sleep(2)
                        log_message("✅ Método 2: Click no iCheck-helper executado", "SUCCESS")
                    except Exception as e2:
                        log_message(f"⚠️ Método 2 falhou: {e2}", "WARNING")

                        # Tentar método 3: Click direto no checkbox via JavaScript
                        try:
                            driver.execute_script("""
                                var checkbox = document.getElementById('acumular');
                                checkbox.click();
                            """)
                            time.sleep(2)
                            log_message("✅ Método 3: Click via JavaScript executado", "SUCCESS")
                        except Exception as e3:
                            log_message(f"❌ Método 3 falhou: {e3}", "ERROR")
                            raise Exception("Não foi possível marcar o checkbox acumular")

                # Verificar se foi marcado (pela classe visual do iCheck)
                try:
                    time.sleep(1)  # Aguardar um pouco mais para o iCheck atualizar
                    wrapper = driver.find_element(By.XPATH,
                                                  "//input[@id='acumular']/parent::div[contains(@class, 'icheckbox')]")
                    wrapper_classes = wrapper.get_attribute("class")
                    is_checked_final = "checked" in wrapper_classes

                    if is_checked_final:
                        log_message("✔️ CONFIRMADO: Checkbox 'acumular' está marcado visualmente!", "SUCCESS")
                    else:
                        # Se não tem a classe checked, mas o iCheck foi executado, assumir que está ok
                        log_message("⚠️ Checkbox pode estar marcado (iCheck executado com sucesso)", "WARNING")
                        log_message("▶️ Continuando processamento...", "INFO")
                except:
                    # Se não conseguir verificar, mas executou o comando, assumir que funcionou
                    log_message("✅ Comando de marcação executado, continuando...", "INFO")

            else:
                log_message("✅ Checkbox 'acumular' já estava marcado", "INFO")

        except Exception as e:
            log_message(f"❌ Erro crítico ao marcar checkbox acumular: {e}", "ERROR")
            raise

    def acumular_exames_no_formulario(self, driver, wait, dados_exames):
        """Acumula todos os exames no formulário digitando os códigos"""
        try:
            log_message(f"\n📝 Iniciando acumulação de {len(dados_exames)} exames...", "INFO")

            for i, exame_data in enumerate(dados_exames, 1):
                codigo = exame_data['codigo']
                log_message(f"➡️ Acumulando exame {i}/{len(dados_exames)}: {codigo}", "INFO")

                # Delay progressivo conforme mencionado pelo usuário
                # Quanto mais exames acumulados, mais lento fica o sistema
                delay_base = 1.0
                delay_progressivo = min(delay_base + (i * 0.2), 5.0)  # Máximo de 5 segundos

                tentativas = 0
                max_tentativas = 3

                while tentativas < max_tentativas:
                    try:
                        # Encontrar o campo de código de barras
                        campo_codigo = wait.until(
                            EC.element_to_be_clickable((By.ID, "inputSearchCodBarra"))
                        )

                        # Limpar o campo
                        campo_codigo.clear()
                        time.sleep(0.3)

                        # Digitar o código
                        campo_codigo.send_keys(codigo)
                        time.sleep(0.5)

                        # Pressionar Enter
                        campo_codigo.send_keys(Keys.ENTER)
                        log_message(f"✅ Código {codigo} digitado e Enter pressionado", "INFO")

                        # Aguardar o modal de carregamento desaparecer antes de continuar
                        self.aguardar_modal_carregamento_desaparecer(driver, wait, timeout=30)

                        # Aguardar delay progressivo para dar tempo ao sistema processar
                        log_message(f"⏳ Aguardando {delay_progressivo:.1f}s para sistema processar...", "INFO")
                        time.sleep(delay_progressivo)

                        # Verificar se o campo está realmente interagível antes de continuar
                        try:
                            campo_codigo = wait.until(
                                EC.element_to_be_clickable((By.ID, "inputSearchCodBarra"))
                            )
                            log_message(f"✅ Campo de código está interagível para próximo exame", "SUCCESS")
                            break  # Sucesso, sair do loop de tentativas
                        except Exception as e:
                            log_message(f"⚠️ Campo não está interagível ainda: {e}", "WARNING")
                            tentativas += 1
                            if tentativas < max_tentativas:
                                log_message(f"🔄 Tentativa {tentativas + 1}/{max_tentativas} em 3 segundos...",
                                            "WARNING")
                                time.sleep(3)
                            else:
                                log_message(f"⚠️ Máximo de tentativas atingido, continuando mesmo assim...", "WARNING")

                    except Exception as e:
                        tentativas += 1
                        log_message(f"❌ Erro na tentativa {tentativas}: {e}", "ERROR")
                        if tentativas < max_tentativas:
                            log_message(f"🔄 Tentando novamente em 5 segundos...", "WARNING")
                            time.sleep(5)
                        else:
                            log_message(f"❌ Máximo de tentativas atingido para exame {codigo}", "ERROR")
                            raise

                # Verificar se o exame foi adicionado na tabela
                try:
                    tbody = driver.find_element(By.ID, "tabelaLocalizarExamesTbody")
                    linhas = tbody.find_elements(By.TAG_NAME, "tr")
                    log_message(f"📊 Total de exames acumulados: {len(linhas)}", "INFO")
                except:
                    pass

            log_message("✅ Todos os exames foram acumulados no formulário", "SUCCESS")
            time.sleep(2)

        except Exception as e:
            log_message(f"Erro ao acumular exames: {e}", "ERROR")
            raise

    def run(self, params: dict):
        username = params.get("username")
        password = params.get("password")
        excel_file = params.get("excel_file")
        cancel_flag = params.get("cancel_flag")
        headless_mode = params.get("headless_mode", False)
        pular_para_laudos = params.get("pular_para_laudos", False)

        try:
            # Lê os dados dos exames da planilha (código e máscara)
            dados_exames = self.get_dados_exames(excel_file)
            if not dados_exames:
                messagebox.showerror("Erro", "Nenhum dado de exame encontrado na planilha.")
                return

            log_message(f"Encontrados {len(dados_exames)} exames para processar", "INFO")

            if pular_para_laudos:
                log_message("⚡ MODO RÁPIDO: Pulando processo de conclusão e indo direto para visualização de laudos", "WARNING")

        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao ler o Excel: {e}")
            return

        url = os.getenv("SYSTEM_URL", "https://dap.pathoweb.com.br/login/auth")
        driver = None
        resultados = []

        try:
            driver = BrowserFactory.create_chrome(headless=headless_mode)
            wait = WebDriverWait(driver, 20)

            log_message("Iniciando automação de conclusão...", "INFO")

            # Login
            log_message("Fazendo login...", "INFO")
            driver.get(url)

            # Aguardar página carregar completamente
            wait.until(EC.presence_of_element_located((By.ID, "username")))

            username_field = driver.find_element(By.ID, "username")
            username_field.clear()
            username_field.send_keys(username)

            password_field = driver.find_element(By.ID, "password")
            password_field.clear()
            password_field.send_keys(password)

            submit_button = driver.find_element(By.CSS_SELECTOR, "button[type='submit']")
            submit_button.click()

            log_message("Verificando se precisa navegar para módulo de exames...", "INFO")
            current_url = driver.current_url
            if current_url == "https://dap.pathoweb.com.br/" or "trocarModulo" in current_url:
                log_message("Detectada tela de seleção de módulos - navegando para módulo de exames...",
                            "INFO")
                try:
                    modulo_link = wait.until(
                        EC.element_to_be_clickable(
                            (By.CSS_SELECTOR, "a[href='/site/trocarModulo?modulo=1']")))
                    modulo_link.click()
                    time.sleep(2)
                    log_message("✅ Navegação para módulo de exames realizada", "SUCCESS")
                except Exception as e:
                    log_message(f"⚠️ Erro ao navegar para módulo: {e}", "WARNING")
                    # Tentar navegar diretamente pela URL como fallback
                    driver.get("https://dap.pathoweb.com.br/moduloExame/index")
                    time.sleep(2)
                    log_message("🔄 Navegação direta para módulo realizada", "INFO")

            elif "moduloExame" in current_url:
                log_message("✅ Já está no módulo de exames - pulando navegação", "SUCCESS")
            else:
                log_message(f"⚠️ URL inesperada detectada: {current_url}", "WARNING")
                # Tentar navegar diretamente como fallback
                driver.get("https://dap.pathoweb.com.br/moduloExame/index")
                time.sleep(2)
                log_message("🔄 Navegação direta para módulo realizada (fallback)", "INFO")

            # Fechar modal se aparecer
            try:
                modal_close_button = driver.find_element(By.CSS_SELECTOR,
                                                         "#mensagemParaClienteModal .modal-footer button")
                if modal_close_button.is_displayed():
                    modal_close_button.click()
            except Exception:
                pass

            log_message("✅ Login realizado com sucesso. Iniciando processamento dos exames.", "SUCCESS")

            if pular_para_laudos:
                log_message("\n" + "=" * 70, "INFO")
                log_message("⚡ MODO RÁPIDO ATIVADO - PULANDO PROCESSO DE CONCLUSÃO", "WARNING")
                log_message("=" * 70, "INFO")

                # Ir direto para visualização de laudos
                try:
                    self.processar_visualizacao_laudos_final(driver, wait, dados_exames)
                    log_message("✅ Visualização de laudos concluída com sucesso!", "SUCCESS")
                except Exception as laudos_error:
                    log_message(f"❌ Erro durante visualização de laudos: {laudos_error}", "ERROR")
                    messagebox.showerror("Erro", f"Erro durante visualização de laudos:\n{str(laudos_error)[:200]}")

                # Finalizar sem fazer mais nada
                return

            # Processar cada exame da planilha (modo normal)
            for i, exame_data in enumerate(dados_exames, 1):
                if cancel_flag and cancel_flag.is_set():
                    log_message("Execução cancelada pelo usuário.", "WARNING")
                    break

                codigo = exame_data['codigo']
                mascara = exame_data['mascara']
                codigo_procedimento = exame_data['codigo_procedimento']

                log_message(
                    f"\n➡️ Processando exame {i}/{len(dados_exames)}: {codigo} (máscara: {mascara}, codigo procedimento: {codigo_procedimento})",
                    "INFO")

                try:
                    # Verificar se o browser ainda está ativo
                    if not self.verificar_sessao_browser(driver):
                        log_message("🔄 Recriando browser devido à sessão perdida...", "WARNING")
                        try:
                            driver.quit()
                        except:
                            pass

                        # Recriar browser e fazer login novamente
                        driver = BrowserFactory.create_chrome(headless=headless_mode)
                        wait = WebDriverWait(driver, 20)

                        # Fazer login novamente
                        log_message("🔄 Fazendo login novamente...", "INFO")
                        driver.get(url)

                        # Aguardar página carregar completamente
                        wait.until(EC.presence_of_element_located((By.ID, "username")))
                        time.sleep(2)

                        username_field = driver.find_element(By.ID, "username")
                        username_field.clear()
                        username_field.send_keys(username)

                        password_field = driver.find_element(By.ID, "password")
                        password_field.clear()
                        password_field.send_keys(password)

                        submit_button = driver.find_element(By.CSS_SELECTOR, "button[type='submit']")
                        submit_button.click()

                        log_message("🔄 Navegando para módulo de exames novamente...", "INFO")

                        log_message("Verificando se precisa navegar para módulo de exames...", "INFO")
                        current_url = driver.current_url
                        if current_url == "https://dap.pathoweb.com.br/" or "trocarModulo" in current_url:
                            log_message("Detectada tela de seleção de módulos - navegando para módulo de exames...",
                                        "INFO")
                            try:
                                modulo_link = wait.until(
                                    EC.element_to_be_clickable(
                                        (By.CSS_SELECTOR, "a[href='/site/trocarModulo?modulo=1']")))
                                modulo_link.click()
                                time.sleep(2)
                                log_message("✅ Navegação para módulo de exames realizada", "SUCCESS")
                            except Exception as e:
                                log_message(f"⚠️ Erro ao navegar para módulo: {e}", "WARNING")
                                # Tentar navegar diretamente pela URL como fallback
                                driver.get("https://dap.pathoweb.com.br/moduloExame/index")
                                time.sleep(2)
                                log_message("🔄 Navegação direta para módulo realizada", "INFO")

                        elif "moduloExame" in current_url:
                            log_message("✅ Já está no módulo de exames - pulando navegação", "SUCCESS")
                        else:
                            log_message(f"⚠️ URL inesperada detectada: {current_url}", "WARNING")
                            # Tentar navegar diretamente como fallback
                            driver.get("https://dap.pathoweb.com.br/moduloExame/index")
                            time.sleep(2)
                            log_message("🔄 Navegação direta para módulo realizada (fallback)", "INFO")

                        # Fechar modal se aparecer
                        try:
                            modal_close_button = driver.find_element(By.CSS_SELECTOR,
                                                                     "#mensagemParaClienteModal .modal-footer button")
                            if modal_close_button.is_displayed():
                                modal_close_button.click()
                                time.sleep(1)
                        except Exception:
                            pass

                        log_message("✅ Browser recriado e login realizado novamente", "SUCCESS")

                    # Processar este exame específico
                    resultado = self.processar_exame(driver, wait, codigo, mascara, codigo_procedimento)
                    resultados.append({
                        'codigo': codigo,
                        'mascara': mascara,
                        'codigo_procedimento': codigo_procedimento,
                        'status': resultado['status'],
                        'detalhes': resultado.get('detalhes', '')
                    })

                except Exception as e:
                    log_message(f"❌ Erro ao processar exame {codigo}: {e}", "ERROR")
                    resultados.append({
                        'codigo': codigo,
                        'mascara': mascara,
                        'codigo_procedimento': codigo_procedimento,
                        'status': 'erro',
                        'detalhes': str(e)
                    })

            # Mostrar resumo final
            self.mostrar_resumo_final(resultados)

        except Exception as e:
            log_message(f"❌ Erro durante a automação: {e}", "ERROR")
            messagebox.showerror("Erro", f"❌ Erro durante a automação:\n{str(e)[:200]}...")
        finally:
            log_message("✅ Execução finalizada", "SUCCESS")
            driver.quit()

    def processar_exame(self, driver, wait, codigo, mascara, codigo_procedimento):
        """Processa um exame individual"""
        try:
            # Verificar se a sessão do browser ainda está ativa
            if not self.verificar_sessao_browser(driver):
                raise Exception("Sessão do browser perdida - necessário reiniciar")

            # Aguardar e encontrar o campo de código de barras
            log_message("Aguardando página carregar completamente...", "INFO")
            time.sleep(0.5)

            # Tentar diferentes formas de encontrar o campo
            campo_codigo = None

            # Método 1: Por ID
            try:
                campo_codigo = wait.until(EC.presence_of_element_located((By.ID, "inputSearchCodBarra")))
                log_message("✅ Campo encontrado pelo ID", "INFO")
            except:
                log_message("⚠️ Campo não encontrado pelo ID", "WARNING")

            # Método 2: Por atributos se o ID não funcionou
            if not campo_codigo:
                try:
                    campo_codigo = driver.find_element(By.XPATH, "//input[@placeholder='Leitor de código de barras']")
                    log_message("✅ Campo encontrado pelo placeholder", "INFO")
                except:
                    log_message("⚠️ Campo não encontrado pelo placeholder", "WARNING")

            # Método 3: Por nome se ainda não encontrou
            if not campo_codigo:
                try:
                    campo_codigo = driver.find_element(By.NAME, "barcode")
                    log_message("✅ Campo encontrado pelo name", "INFO")
                except:
                    log_message("⚠️ Campo não encontrado pelo name", "WARNING")

            # Se ainda não encontrou, listar todos os inputs para debug
            if not campo_codigo:
                log_message("❌ Campo não encontrado. Listando inputs disponíveis:", "ERROR")
                inputs = driver.find_elements(By.TAG_NAME, "input")
                for i, inp in enumerate(inputs):
                    input_id = inp.get_attribute("id") or "sem_id"
                    input_name = inp.get_attribute("name") or "sem_name"
                    input_placeholder = inp.get_attribute("placeholder") or "sem_placeholder"
                    input_type = inp.get_attribute("type") or "sem_type"
                    log_message(
                        f"Input {i}: id='{input_id}', name='{input_name}', placeholder='{input_placeholder}', type='{input_type}'",
                        "INFO")

                raise Exception("Campo de código de barras não encontrado")

            # Interagir com o campo usando os métodos já implementados
            self.interagir_com_campo_codigo(driver, campo_codigo, codigo)

            # Aguardar div de andamento aparecer
            return self.aguardar_e_processar_andamento(driver, wait, mascara, codigo_procedimento)

        except Exception as e:
            error_message = str(e)
            log_message(f"Erro ao processar exame {codigo}: {error_message}", "ERROR")

            # Verificar se é erro de sessão inválida
            if "invalid session id" in error_message.lower():
                log_message("❌ Erro de sessão inválida detectado", "ERROR")
                return {'status': 'erro_sessao', 'detalhes': 'Sessão do browser perdida'}

            # Screenshot do erro para outros tipos de erro
            try:
                screenshot_path = f"erro_exame_{codigo}_{int(time.time())}.png"
                driver.save_screenshot(screenshot_path)
                log_message(f"Screenshot do erro salvo em: {screenshot_path}", "INFO")
            except:
                pass
            return {'status': 'erro', 'detalhes': error_message}

    def interagir_com_campo_codigo(self, driver, campo_codigo, codigo):
        """Interage com o campo de código usando os métodos já implementados"""
        log_message("Campo de código encontrado, interagindo...", "INFO")

        # Garantir que o campo está visível
        driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", campo_codigo)
        time.sleep(1)

        # Verificar se o elemento está visível e habilitado
        is_displayed = campo_codigo.is_displayed()
        is_enabled = campo_codigo.is_enabled()
        log_message(f"Campo - Visível: {is_displayed}, Habilitado: {is_enabled}", "INFO")

        # Limpar o campo primeiro
        try:
            campo_codigo.clear()
            log_message("Campo limpo com sucesso", "INFO")
        except:
            driver.execute_script("arguments[0].value = '';", campo_codigo)
            log_message("Campo limpo com JavaScript", "INFO")

        time.sleep(0.5)

        # Digitar o código
        try:
            campo_codigo.send_keys(codigo)
            log_message(f"Código '{codigo}' digitado com sucesso", "INFO")
        except:
            driver.execute_script(f"arguments[0].value = '{codigo}';", campo_codigo)
            driver.execute_script("""
                var element = arguments[0];
                var event = new Event('input', { bubbles: true });
                element.dispatchEvent(event);
            """, campo_codigo)
            log_message(f"Código '{codigo}' digitado com JavaScript", "INFO")

        time.sleep(1)

        # Pressionar Enter
        try:
            campo_codigo.send_keys(Keys.ENTER)
            log_message("⌨️ Enter pressionado com sucesso", "INFO")
        except:
            driver.execute_script("""
                var element = arguments[0];
                var event = new KeyboardEvent('keydown', {
                    key: 'Enter',
                    code: 'Enter',
                    keyCode: 13,
                    bubbles: true
                });
                element.dispatchEvent(event);
            """, campo_codigo)
            log_message("⌨️ Enter pressionado com JavaScript", "INFO")

    def aguardar_e_processar_andamento(self, driver, wait, mascara,codigo_procedimento):
        """Aguarda a div de andamento e processa o exame"""
        log_message("Aguardando div de andamento do exame aparecer...", "INFO")

        # Aguardar mais tempo para o carregamento após digitar o código
        timeout_andamento = 30
        inicio = time.time()

        while time.time() - inicio < timeout_andamento:
            try:
                # Verificar se a div de andamento apareceu
                andamento_div = driver.find_element(By.ID, "divAndamentoExame")
                if andamento_div and andamento_div.is_displayed():
                    log_message("📋 Div de andamento do exame encontrada!", "SUCCESS")
                    break
            except:
                pass

            time.sleep(1)
            if int(time.time() - inicio) % 5 == 0:  # Log a cada 5 segundos
                log_message(f"⏳ Aguardando carregamento... ({int(time.time() - inicio)}s)", "INFO")
        else:
            log_message("⚠️ Div de andamento não apareceu no tempo esperado", "WARNING")
            return {'status': 'sem_andamento', 'detalhes': 'Exame não encontrado ou não carregou'}

        # Aguardar carregamento completo
        time.sleep(2)

        # Verificar se tem SVG na conclusão
        if self.verificar_svg_conclusao(driver):
            log_message("✅ SVG encontrado na etapa Conclusão - iniciando processo", "SUCCESS")
            return self.processar_conclusao_completa(driver, wait, mascara, codigo_procedimento)
        else:
            log_message("⚠️ SVG não encontrado na etapa Conclusão - fechando exame", "WARNING")
            self.fechar_exame(driver, wait)
            return {'status': 'sem_svg', 'detalhes': 'Exame não está na etapa de conclusão'}

    def processar_conclusao_completa(self, driver, wait, mascara, codigo_procedimento):
        """Processa a conclusão completa do exame"""
        try:
            # Digitar a máscara e buscar
            if mascara:
                self.digitar_mascara_e_buscar(driver, wait, mascara)
            else:
                log_message("⚠️ Nenhuma máscara encontrada, pulando busca", "WARNING")

            if codigo_procedimento:
                self.inserir_procedimento(driver, wait, codigo_procedimento)
            else:
                log_message("Nenhum código de procedimento encontrado, pulando inserção", "WARNING")

            self.salvar_conclusao(driver, wait)

            # Enviar para próxima etapa
            self.enviar_proxima_etapa(driver, wait)

            self.processar_assinatura(driver, wait)

            log_message("🎉 Processo de conclusão finalizado com sucesso!", "SUCCESS")
            return {'status': 'sucesso', 'detalhes': 'Conclusão processada e assinada'}

        except Exception as e:
            log_message(f"Erro durante processo de conclusão: {e}", "ERROR")
            return {'status': 'erro_conclusao', 'detalhes': str(e)}

    def inserir_procedimento(self, driver, wait, codigo_procedimento):
        """Insere o código do procedimento no formulário"""
        try:
            log_message(f"🔧 Iniciando inserção do código de procedimento: {codigo_procedimento}", "INFO")

            # Aguardar o formulário de procedimentos estar presente
            wait.until(EC.presence_of_element_located((By.ID, "procedimentosForm")))
            log_message("✅ Formulário de procedimentos encontrado", "INFO")
            time.sleep(1)

            # Encontrar a linha de novos procedimentos
            tr_novos = wait.until(
                EC.presence_of_element_located((By.ID, "novosProcedimentos"))
            )
            log_message("✅ Linha de novos procedimentos encontrada", "INFO")

            # Encontrar todos os td's dentro da tr
            tds = tr_novos.find_elements(By.TAG_NAME, "td")
            log_message(f"📋 Encontrados {len(tds)} elementos td", "INFO")

            if len(tds) < 3:
                raise Exception(f"Esperado pelo menos 3 td's, encontrados apenas {len(tds)}")

            # Pegar o terceiro td (índice 2)
            td_procedimento = tds[2]

            # Encontrar o link âncora dentro deste td
            ancora = td_procedimento.find_element(
                By.CSS_SELECTOR,
                "a.table-editable-ancora.autocomplete.autocompleteSetup"
            )
            log_message("✅ Âncora do procedimento encontrada", "INFO")

            # Rolar até o elemento para garantir visibilidade
            driver.execute_script(
                "arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});",
                ancora
            )
            time.sleep(1)

            # Clicar na âncora para habilitar o input
            ancora.click()
            log_message("🖱️ Clicou na âncora para habilitar edição", "INFO")
            time.sleep(1)

            # Aguardar o input aparecer e estar visível
            input_procedimento = wait.until(
                EC.visibility_of_element_located((By.ID, "procedimentoInput_novo"))
            )
            log_message("✅ Campo de input do procedimento está visível", "INFO")

            # Limpar o campo (pode ter "Vazio" como valor padrão)
            input_procedimento.clear()
            time.sleep(0.3)

            # Digitar o código do procedimento
            input_procedimento.send_keys(codigo_procedimento)
            log_message(f"✍️ Código do procedimento '{codigo_procedimento}' digitado", "INFO")
            time.sleep(1)

            # Pressionar Enter para selecionar
            input_procedimento.send_keys(Keys.ENTER)
            log_message("⌨️ Enter pressionado para confirmar procedimento", "INFO")
            time.sleep(2)

            # Aguardar o autocomplete processar (caso apareça lista de sugestões)
            try:
                # Verificar se o dropdown do autocomplete está visível
                dropdown = driver.find_element(By.CSS_SELECTOR, "ul.typeahead.dropdown-menu")
                if dropdown.is_displayed():
                    log_message("📋 Dropdown de autocomplete detectado, aguardando seleção...", "INFO")
                    time.sleep(1)
                    # Pressionar Enter novamente para selecionar o primeiro item se necessário
                    input_procedimento.send_keys(Keys.ENTER)
                    time.sleep(1)
            except:
                # Dropdown não apareceu ou já foi fechado
                pass

            log_message("✅ Código de procedimento inserido com sucesso", "SUCCESS")

        except Exception as e:
            log_message(f"❌ Erro ao inserir código de procedimento: {e}", "ERROR")

            # Tentar métodos alternativos
            try:
                log_message("🔄 Tentando método alternativo...", "WARNING")

                # Método alternativo: encontrar diretamente o input e forçar valor via JavaScript
                driver.execute_script(f"""
                    var input = document.getElementById('procedimentoInput_novo');
                    if (input) {{
                        input.style.display = 'inline-block';
                        input.value = '{codigo_procedimento}';
                        input.focus();
                        // Disparar eventos para simular digitação
                        var inputEvent = new Event('input', {{ bubbles: true }});
                        input.dispatchEvent(inputEvent);

                        var changeEvent = new Event('change', {{ bubbles: true }});
                        input.dispatchEvent(changeEvent);
                    }}
                """)
                time.sleep(1)

                # Pressionar Enter via JavaScript
                driver.execute_script("""
                    var input = document.getElementById('procedimentoInput_novo');
                    if (input) {
                        var event = new KeyboardEvent('keydown', {
                            key: 'Enter',
                            code: 'Enter',
                            keyCode: 13,
                            bubbles: true
                        });
                        input.dispatchEvent(event);
                    }
                """)
                time.sleep(2)

                log_message("✅ Método alternativo executado", "SUCCESS")

            except Exception as alt_error:
                log_message(f"❌ Método alternativo também falhou: {alt_error}", "ERROR")
                raise

    def mostrar_resumo_final(self, resultados):
        """Mostra o resumo final do processamento"""
        total = len(resultados)
        sucesso = len([r for r in resultados if r['status'] == 'sucesso'])
        sem_svg = len([r for r in resultados if r['status'] == 'sem_svg'])
        sem_andamento = len([r for r in resultados if r['status'] == 'sem_andamento'])
        erro_sessao = len([r for r in resultados if r['status'] == 'erro_sessao'])
        erros = len([r for r in resultados if 'erro' in r['status'] and r['status'] != 'erro_sessao'])

        log_message("\n" + "=" * 50, "INFO")
        log_message("RESUMO FINAL DO PROCESSAMENTO", "INFO")
        log_message("=" * 50, "INFO")
        log_message(f"Total de exames: {total}", "INFO")
        log_message(f"✅ Processados com sucesso: {sucesso}", "SUCCESS")
        log_message(f"⚠️ Sem SVG (não estão em conclusão): {sem_svg}", "WARNING")
        log_message(f"⚠️ Exames não encontrados: {sem_andamento}", "WARNING")
        log_message(f"🔄 Erros de sessão (browser perdido): {erro_sessao}", "WARNING")
        log_message(f"❌ Outros erros de processamento: {erros}", "ERROR")

        # Mostrar detalhes dos erros se houver
        erros_totais = erro_sessao + erros
        if erros_totais > 0:
            log_message("\nDetalhes dos erros:", "ERROR")
            for r in resultados:
                if 'erro' in r['status']:
                    log_message(f"- {r['codigo']}: {r['detalhes']}", "ERROR")

        messagebox.showinfo("Processamento Concluído",
                            f"✅ Processamento finalizado!\n\n"
                            f"Total: {total}\n"
                            f"Sucesso: {sucesso}\n"
                            f"Sem SVG: {sem_svg}\n"
                            f"Não encontrados: {sem_andamento}\n"
                            f"Erros de sessão: {erro_sessao}\n"
                            f"Outros erros: {erros}")


def run(params: dict):
    module = ConclusaoModule()
    module.run(params)
