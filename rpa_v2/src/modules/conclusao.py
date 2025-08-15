import os
import time
from tkinter import messagebox
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from dotenv import load_dotenv
from openpyxl import load_workbook

from src.core.browser_factory import BrowserFactory
from src.core.logger import log_message
from src.modules.base import BaseModule

load_dotenv()

class ConclusaoModule(BaseModule):
    def __init__(self):
        super().__init__(nome="Conclusão")

    def get_dados_exames(self, file_path: str) -> list:
        try:
            workbook = load_workbook(file_path)
            sheet = workbook.active
            dados = []
            ultima_mascara = None
            
            # Lê da linha 2 em diante (linha 1 é cabeçalho)
            for row in range(2, sheet.max_row + 1):
                codigo = sheet[f'A{row}'].value
                mascara = sheet[f'B{row}'].value
                
                if codigo is not None:
                    codigo = str(codigo).strip()
                    
                    # Se não tem máscara, usa a última válida
                    if mascara is not None and str(mascara).strip():
                        mascara = str(mascara).strip()
                        ultima_mascara = mascara
                    else:
                        mascara = ultima_mascara
                    
                    dados.append({
                        'codigo': codigo,
                        'mascara': mascara
                    })
            
            workbook.close()
            return dados
        except Exception as e:
            raise Exception(f"Erro ao ler planilha: {e}")

    def verificar_sessao_browser(self, driver) -> bool:
        """Verifica se a sessão do browser ainda está ativa"""
        try:
            driver.current_url
            return True
        except Exception as e:
            if "invalid session id" in str(e).lower():
                log_message("❌ Sessão do browser perdida", "ERROR")
                return False
            return True

    def verificar_svg_conclusao(self, driver) -> bool:
        """Verifica se existe o SVG na etapa Conclusão"""
        try:
            # Método 1: Procurar pelo link de Conclusão que contenha o SVG arrow-right
            try:
                conclusao_link = driver.find_element(
                    By.XPATH, 
                    "//a[@data-id='C' and contains(., 'Conclusão')]//svg[@data-icon='arrow-right']"
                )
                if conclusao_link:
                    log_message("✅ SVG arrow-right encontrado na etapa Conclusão", "INFO")
                    return True
            except:
                pass
            
            # Método 2: Verificar se o link de Conclusão está clicável/ativo
            try:
                conclusao_link = driver.find_element(By.XPATH, "//a[@data-id='C' and contains(., 'Conclusão')]")
                # Verificar se o link não tem classe que indica inativo
                classe_link = conclusao_link.get_attribute("class") or ""
                if "disabled" not in classe_link.lower() and "inactive" not in classe_link.lower():
                    # Verificar se existe SVG dentro do link
                    svgs = conclusao_link.find_elements(By.TAG_NAME, "svg")
                    if svgs:
                        log_message(f"✅ SVG encontrado na etapa Conclusão (método 2)", "INFO")
                        return True
            except:
                pass
                
            # Método 3: Verificar qualquer SVG com arrow-right próximo à Conclusão
            try:
                svg_arrows = driver.find_elements(By.XPATH, "//svg[@data-icon='arrow-right']")
                for svg in svg_arrows:
                    # Verificar se o SVG está próximo ao texto "Conclusão"
                    parent = svg.find_element(By.XPATH, "..")
                    if "conclusão" in parent.text.lower():
                        log_message("✅ SVG arrow-right encontrado próximo à Conclusão (método 3)", "INFO")
                        return True
            except:
                pass
            
            log_message("⚠️ SVG não encontrado na etapa Conclusão", "WARNING")
            return False
            
        except Exception as e:
            log_message(f"Erro ao verificar SVG conclusão: {e}", "ERROR")
            return False

    def fechar_exame(self, driver, wait):
        """Clica no botão de fechar exame"""
        try:
            botao_fechar = wait.until(
                EC.element_to_be_clickable((By.ID, "fecharExameBarraFerramenta"))
            )
            botao_fechar.click()
            log_message("📁 Exame fechado (sem SVG na conclusão)", "INFO")
            
            # Aguardar retornar à tela principal
            try:
                # Verificar se voltou à tela principal aguardando o campo de código aparecer
                wait.until(EC.presence_of_element_located((By.ID, "inputSearchCodBarra")))
                log_message("✅ Retornou à tela principal após fechar exame", "INFO")
            except:
                log_message("⚠️ Pode não ter retornado à tela principal", "WARNING")
                # Tentar navegar de volta ao módulo se necessário
                try:
                    current_url = driver.current_url
                    if "modulo=1" not in current_url:
                        modulo_link = driver.find_element(By.CSS_SELECTOR, "a[href='/site/trocarModulo?modulo=1']")
                        modulo_link.click()
                        time.sleep(1.5)
                        log_message("🔄 Navegou de volta ao módulo de exames", "INFO")
                except:
                    pass
                    
        except Exception as e:
            log_message(f"Erro ao fechar exame: {e}", "ERROR")

    def digitar_mascara_e_buscar(self, driver, wait, mascara):
        """Digita a máscara no campo buscaArvore e pressiona Enter"""
        try:
            # Aguardar o campo estar presente e visível com timeout maior
            log_message(f"🔍 Procurando campo buscaArvore...", "INFO")
            campo_busca = wait.until(EC.element_to_be_clickable((By.ID, "buscaArvore")))
            log_message(f"🔍 Campo buscaArvore encontrado e clicável", "INFO")
            
            # Verificar se o campo está visível
            if not campo_busca.is_displayed():
                log_message("⚠️ Campo buscaArvore não está visível", "WARNING")
                return
            
            # Rolar até o campo para garantir visibilidade
            driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", campo_busca)
            time.sleep(1)
            
            # Focar no campo primeiro
            campo_busca.click()
            time.sleep(0.5)
            
            # Digitar a máscara
            campo_busca.send_keys(mascara)
            log_message(f"✍️ Máscara '{mascara}' digitada no campo buscaArvore", "INFO")
            time.sleep(1)
            
            # Pressionar Enter
            campo_busca.send_keys(Keys.ENTER)
            log_message(f"⌨️ Enter pressionado após digitar máscara", "INFO")
            time.sleep(1)
            
        except Exception as e:
            log_message(f"Erro ao digitar máscara: {e}", "ERROR")
            # Tentar encontrar o campo de outra forma
            try:
                # Verificar se existe campo com classe específica
                campos_alternativos = driver.find_elements(By.XPATH, "//input[@class='btn-xs' and @type='text']")
                log_message(f"Encontrados {len(campos_alternativos)} campos alternativos", "INFO")
                
                if campos_alternativos:
                    campo_alternativo = campos_alternativos[0]
                    campo_alternativo.click()
                    campo_alternativo.clear()
                    campo_alternativo.send_keys(mascara)
                    campo_alternativo.send_keys(Keys.ENTER)
                    log_message(f"✅ Máscara digitada usando campo alternativo", "INFO")
                    return
                
                # Listar todos os inputs para debug
                inputs = driver.find_elements(By.TAG_NAME, "input")
                log_message(f"Total de inputs encontrados na página: {len(inputs)}", "INFO")
                for i, inp in enumerate(inputs[:10]):  # Apenas os primeiros 10
                    input_id = inp.get_attribute("id")
                    input_class = inp.get_attribute("class")
                    input_type = inp.get_attribute("type")
                    log_message(f"Input {i}: id='{input_id}', class='{input_class}', type='{input_type}'", "INFO")
                    
            except Exception as debug_e:
                log_message(f"Erro no debug: {debug_e}", "ERROR")
            raise

    def salvar_conclusao(self, driver, wait):
        """Clica no botão Salvar"""
        try:
            # Aguardar o botão estar presente e clicável
            log_message("💾 Procurando botão Salvar...", "INFO")
            botao_salvar = wait.until(EC.element_to_be_clickable((By.ID, "salvarConcl")))
            log_message("💾 Botão Salvar encontrado e clicável", "INFO")
            
            # Verificar se o botão está visível
            if not botao_salvar.is_displayed():
                log_message("⚠️ Botão salvarConcl não está visível", "WARNING")
                return
            
            # Rolar até o botão para garantir visibilidade
            driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", botao_salvar)
            time.sleep(1)
            
            # Clicar no botão
            botao_salvar.click()
            log_message("💾 Clicou em Salvar", "INFO")
            time.sleep(1)
            
        except Exception as e:
            log_message(f"Erro ao salvar: {e}", "ERROR")
            # Tentar encontrar o botão de outra forma
            try:
                # Tentar por link com onclick específico
                botoes_onclick = driver.find_elements(By.XPATH, "//a[contains(@onclick, 'ajaxChangeSave')]")
                log_message(f"Encontrados {len(botoes_onclick)} botões com onclick ajaxChangeSave", "INFO")
                
                if botoes_onclick:
                    botao_onclick = botoes_onclick[0]
                    driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", botao_onclick)
                    time.sleep(1)
                    botao_onclick.click()
                    log_message("💾 Clicou em Salvar usando onclick", "INFO")
                    return
                
                # Tentar por classe do botão
                botoes_classe = driver.find_elements(By.XPATH, "//a[contains(@class, 'btn-primary') and contains(text(), 'Salvar')]")
                log_message(f"Encontrados {len(botoes_classe)} botões com classe btn-primary e texto Salvar", "INFO")
                
                if botoes_classe:
                    botao_classe = botoes_classe[0]
                    driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", botao_classe)
                    time.sleep(1)
                    botao_classe.click()
                    log_message("💾 Clicou em Salvar usando classe", "INFO")
                    return
                
                # Listar todos os links/botões para debug
                links = driver.find_elements(By.TAG_NAME, "a")
                log_message(f"Total de links encontrados na página: {len(links)}", "INFO")
                for i, link in enumerate(links[:15]):  # Apenas os primeiros 15
                    link_id = link.get_attribute("id")
                    link_class = link.get_attribute("class")
                    link_text = link.text.strip()
                    link_onclick = link.get_attribute("onclick")
                    if (link_id and "salvar" in link_id.lower()) or \
                       (link_class and "salvar" in link_class.lower()) or \
                       (link_text and "salvar" in link_text.lower()) or \
                       (link_onclick and "save" in link_onclick.lower()):
                        log_message(f"Link {i}: id='{link_id}', class='{link_class}', text='{link_text}', onclick='{link_onclick}'", "INFO")
                        
            except Exception as debug_e:
                log_message(f"Erro no debug de botões: {debug_e}", "ERROR")
            raise

    def enviar_proxima_etapa(self, driver, wait):
        """Clica no botão de enviar para próxima etapa"""
        try:
            botao_enviar = wait.until(
                EC.element_to_be_clickable((By.ID, "btn-enviar-proxima-etapa"))
            )
            botao_enviar.click()
            log_message("➡️ Clicou em Enviar para próxima etapa", "INFO")
            time.sleep(1.5)
        except Exception as e:
            log_message(f"Erro ao enviar para próxima etapa: {e}", "ERROR")
            raise

    def assinar_com_george(self, driver, wait):
        """Faz o processo de assinatura com Dr. George"""
        try:
            # Aguardar o modal de assinatura aparecer
            wait.until(EC.presence_of_element_located((By.ID, "assinatura")))
            log_message("📋 Modal de assinatura aberto", "INFO")
            
            # Encontrar e clicar no checkbox do Dr. George (value="2173")
            checkbox_george = wait.until(
                EC.element_to_be_clickable((By.XPATH, "//input[@type='checkbox' and @value='2173']"))
            )
            checkbox_george.click()
            log_message("✅ Checkbox do Dr. George marcado", "INFO")
            time.sleep(1)
            
            # Aguardar o campo de senha aparecer e digitar a senha
            campo_senha = wait.until(
                EC.presence_of_element_located((By.NAME, "senha_2173"))
            )
            campo_senha.send_keys("1323")
            log_message("🔐 Senha digitada", "INFO")
            time.sleep(2)
            
            # Clicar no botão Assinar
            botao_assinar = wait.until(
                EC.element_to_be_clickable((By.ID, "salvarAss"))
            )
            botao_assinar.click()
            log_message("✍️ Clicou em Assinar", "INFO")
            time.sleep(1.5)
            
        except Exception as e:
            log_message(f"Erro no processo de assinatura: {e}", "ERROR")
            raise

    def run(self, params: dict):
        username = params.get("username")
        password = params.get("password")
        excel_file = params.get("excel_file")
        cancel_flag = params.get("cancel_flag")
        
        try:
            # Lê os dados dos exames da planilha (código e máscara)
            dados_exames = self.get_dados_exames(excel_file)
            if not dados_exames:
                messagebox.showerror("Erro", "Nenhum dado de exame encontrado na planilha.")
                return
            
            log_message(f"Encontrados {len(dados_exames)} exames para processar", "INFO")
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao ler o Excel: {e}")
            return

        url = os.getenv("SYSTEM_URL", "https://pathoweb.com.br/login/auth")
        driver = None
        resultados = []
        
        try:
            driver = BrowserFactory.create_chrome()
            wait = WebDriverWait(driver, 20)
            
            log_message("Iniciando automação de conclusão...", "INFO")
            
            # Login
            log_message("Fazendo login...", "INFO")
            driver.get(url)
            
            # Aguardar página carregar completamente
            wait.until(EC.presence_of_element_located((By.ID, "username")))
            
            username_field = driver.find_element(By.ID, "username")
            username_field.clear()
            username_field.send_keys(username)
            
            password_field = driver.find_element(By.ID, "password")
            password_field.clear()
            password_field.send_keys(password)
            
            submit_button = driver.find_element(By.CSS_SELECTOR, "button[type='submit']")
            submit_button.click()
            
            log_message("Navegando para módulo de exames...", "INFO")
            
            # Navegar para o módulo de exames (módulo 1)
            modulo_link = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "a[href='/site/trocarModulo?modulo=1']")))
            modulo_link.click()
            time.sleep(2)
            
            # Fechar modal se aparecer
            try:
                modal_close_button = driver.find_element(By.CSS_SELECTOR, "#mensagemParaClienteModal .modal-footer button")
                if modal_close_button.is_displayed():
                    modal_close_button.click()
            except Exception:
                pass

            log_message("✅ Login realizado com sucesso. Iniciando processamento dos exames.", "SUCCESS")
            
            # Processar cada exame da planilha
            for i, exame_data in enumerate(dados_exames, 1):
                if cancel_flag and cancel_flag.is_set():
                    log_message("Execução cancelada pelo usuário.", "WARNING")
                    break
                
                codigo = exame_data['codigo']
                mascara = exame_data['mascara']
                
                log_message(f"\n➡️ Processando exame {i}/{len(dados_exames)}: {codigo} (máscara: {mascara})", "INFO")
                
                try:
                    # Verificar se o browser ainda está ativo
                    if not self.verificar_sessao_browser(driver):
                        log_message("🔄 Recriando browser devido à sessão perdida...", "WARNING")
                        try:
                            driver.quit()
                        except:
                            pass
                        
                        # Recriar browser e fazer login novamente
                        driver = BrowserFactory.create_chrome()
                        wait = WebDriverWait(driver, 20)
                        
                        # Fazer login novamente
                        log_message("🔄 Fazendo login novamente...", "INFO")
                        driver.get(url)
                        
                        # Aguardar página carregar completamente
                        wait.until(EC.presence_of_element_located((By.ID, "username")))
                        time.sleep(2)
                        
                        username_field = driver.find_element(By.ID, "username")
                        username_field.clear()
                        username_field.send_keys(username)
                        
                        password_field = driver.find_element(By.ID, "password")
                        password_field.clear()
                        password_field.send_keys(password)
                        
                        submit_button = driver.find_element(By.CSS_SELECTOR, "button[type='submit']")
                        submit_button.click()
                        
                        log_message("🔄 Navegando para módulo de exames novamente...", "INFO")
                        
                        # Navegar para o módulo de exames (módulo 1)
                        modulo_link = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "a[href='/site/trocarModulo?modulo=1']")))
                        modulo_link.click()
                        time.sleep(2.5)
                        
                        # Fechar modal se aparecer
                        try:
                            modal_close_button = driver.find_element(By.CSS_SELECTOR, "#mensagemParaClienteModal .modal-footer button")
                            if modal_close_button.is_displayed():
                                modal_close_button.click()
                                time.sleep(1)
                        except Exception:
                            pass
                        
                        log_message("✅ Browser recriado e login realizado novamente", "SUCCESS")
                    
                    # Processar este exame específico
                    resultado = self.processar_exame(driver, wait, codigo, mascara)
                    resultados.append({
                        'codigo': codigo,
                        'mascara': mascara,
                        'status': resultado['status'],
                        'detalhes': resultado.get('detalhes', '')
                    })
                    
                except Exception as e:
                    log_message(f"❌ Erro ao processar exame {codigo}: {e}", "ERROR")
                    resultados.append({
                        'codigo': codigo,
                        'mascara': mascara,
                        'status': 'erro',
                        'detalhes': str(e)
                    })
            
            # Mostrar resumo final
            self.mostrar_resumo_final(resultados)
            
        except Exception as e:
            log_message(f"❌ Erro durante a automação: {e}", "ERROR")
            messagebox.showerror("Erro", f"❌ Erro durante a automação:\n{str(e)[:200]}...")
        finally:
            if driver:
                try:
                    driver.quit()
                    log_message("Browser fechado", "INFO")
                except Exception as quit_error:
                    log_message(f"Erro ao fechar browser: {quit_error}", "WARNING")

    def processar_exame(self, driver, wait, codigo, mascara):
        """Processa um exame individual"""
        try:
            # Verificar se a sessão do browser ainda está ativa
            if not self.verificar_sessao_browser(driver):
                raise Exception("Sessão do browser perdida - necessário reiniciar")
            
            # Aguardar e encontrar o campo de código de barras
            log_message("Aguardando página carregar completamente...", "INFO")
            time.sleep(0.5)
            
            # Tentar diferentes formas de encontrar o campo
            campo_codigo = None
            
            # Método 1: Por ID
            try:
                campo_codigo = wait.until(EC.presence_of_element_located((By.ID, "inputSearchCodBarra")))
                log_message("✅ Campo encontrado pelo ID", "INFO")
            except:
                log_message("⚠️ Campo não encontrado pelo ID", "WARNING")
            
            # Método 2: Por atributos se o ID não funcionou
            if not campo_codigo:
                try:
                    campo_codigo = driver.find_element(By.XPATH, "//input[@placeholder='Leitor de código de barras']")
                    log_message("✅ Campo encontrado pelo placeholder", "INFO")
                except:
                    log_message("⚠️ Campo não encontrado pelo placeholder", "WARNING")
            
            # Método 3: Por nome se ainda não encontrou
            if not campo_codigo:
                try:
                    campo_codigo = driver.find_element(By.NAME, "barcode")
                    log_message("✅ Campo encontrado pelo name", "INFO")
                except:
                    log_message("⚠️ Campo não encontrado pelo name", "WARNING")
            
            # Se ainda não encontrou, listar todos os inputs para debug
            if not campo_codigo:
                log_message("❌ Campo não encontrado. Listando inputs disponíveis:", "ERROR")
                inputs = driver.find_elements(By.TAG_NAME, "input")
                for i, inp in enumerate(inputs):
                    input_id = inp.get_attribute("id") or "sem_id"
                    input_name = inp.get_attribute("name") or "sem_name"
                    input_placeholder = inp.get_attribute("placeholder") or "sem_placeholder"
                    input_type = inp.get_attribute("type") or "sem_type"
                    log_message(f"Input {i}: id='{input_id}', name='{input_name}', placeholder='{input_placeholder}', type='{input_type}'", "INFO")
                
                raise Exception("Campo de código de barras não encontrado")
            
            # Interagir com o campo usando os métodos já implementados
            self.interagir_com_campo_codigo(driver, campo_codigo, codigo)
            
            # Aguardar div de andamento aparecer
            return self.aguardar_e_processar_andamento(driver, wait, mascara)
                
        except Exception as e:
            error_message = str(e)
            log_message(f"Erro ao processar exame {codigo}: {error_message}", "ERROR")
            
            # Verificar se é erro de sessão inválida
            if "invalid session id" in error_message.lower():
                log_message("❌ Erro de sessão inválida detectado", "ERROR")
                return {'status': 'erro_sessao', 'detalhes': 'Sessão do browser perdida'}
            
            # Screenshot do erro para outros tipos de erro
            try:
                screenshot_path = f"erro_exame_{codigo}_{int(time.time())}.png"
                driver.save_screenshot(screenshot_path)
                log_message(f"Screenshot do erro salvo em: {screenshot_path}", "INFO")
            except:
                pass
            return {'status': 'erro', 'detalhes': error_message}

    def interagir_com_campo_codigo(self, driver, campo_codigo, codigo):
        """Interage com o campo de código usando os métodos já implementados"""
        log_message("Campo de código encontrado, interagindo...", "INFO")
        
        # Garantir que o campo está visível
        driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", campo_codigo)
        time.sleep(1)
        
        # Verificar se o elemento está visível e habilitado
        is_displayed = campo_codigo.is_displayed()
        is_enabled = campo_codigo.is_enabled()
        log_message(f"Campo - Visível: {is_displayed}, Habilitado: {is_enabled}", "INFO")
        
        # Limpar o campo primeiro
        try:
            campo_codigo.clear()
            log_message("Campo limpo com sucesso", "INFO")
        except:
            driver.execute_script("arguments[0].value = '';", campo_codigo)
            log_message("Campo limpo com JavaScript", "INFO")
        
        time.sleep(0.5)
        
        # Digitar o código
        try:
            campo_codigo.send_keys(codigo)
            log_message(f"Código '{codigo}' digitado com sucesso", "INFO")
        except:
            driver.execute_script(f"arguments[0].value = '{codigo}';", campo_codigo)
            driver.execute_script("""
                var element = arguments[0];
                var event = new Event('input', { bubbles: true });
                element.dispatchEvent(event);
            """, campo_codigo)
            log_message(f"Código '{codigo}' digitado com JavaScript", "INFO")
        
        time.sleep(1)
        
        # Pressionar Enter
        try:
            campo_codigo.send_keys(Keys.ENTER)
            log_message("⌨️ Enter pressionado com sucesso", "INFO")
        except:
            driver.execute_script("""
                var element = arguments[0];
                var event = new KeyboardEvent('keydown', {
                    key: 'Enter',
                    code: 'Enter',
                    keyCode: 13,
                    bubbles: true
                });
                element.dispatchEvent(event);
            """, campo_codigo)
            log_message("⌨️ Enter pressionado com JavaScript", "INFO")

    def aguardar_e_processar_andamento(self, driver, wait, mascara):
        """Aguarda a div de andamento e processa o exame"""
        log_message("Aguardando div de andamento do exame aparecer...", "INFO")
        
        # Aguardar mais tempo para o carregamento após digitar o código
        timeout_andamento = 30
        inicio = time.time()
        
        while time.time() - inicio < timeout_andamento:
            try:
                # Verificar se a div de andamento apareceu
                andamento_div = driver.find_element(By.ID, "divAndamentoExame")
                if andamento_div and andamento_div.is_displayed():
                    log_message("📋 Div de andamento do exame encontrada!", "SUCCESS")
                    break
            except:
                pass
            
            time.sleep(1)
            if int(time.time() - inicio) % 5 == 0:  # Log a cada 5 segundos
                log_message(f"⏳ Aguardando carregamento... ({int(time.time() - inicio)}s)", "INFO")
        else:
            log_message("⚠️ Div de andamento não apareceu no tempo esperado", "WARNING")
            return {'status': 'sem_andamento', 'detalhes': 'Exame não encontrado ou não carregou'}
        
        # Aguardar carregamento completo
        time.sleep(2)
        
        # Verificar se tem SVG na conclusão
        if self.verificar_svg_conclusao(driver):
            log_message("✅ SVG encontrado na etapa Conclusão - iniciando processo", "SUCCESS")
            return self.processar_conclusao_completa(driver, wait, mascara)
        else:
            log_message("⚠️ SVG não encontrado na etapa Conclusão - fechando exame", "WARNING")
            self.fechar_exame(driver, wait)
            return {'status': 'sem_svg', 'detalhes': 'Exame não está na etapa de conclusão'}

    def processar_conclusao_completa(self, driver, wait, mascara):
        """Processa a conclusão completa do exame"""
        try:
            # Digitar a máscara e buscar
            if mascara:
                self.digitar_mascara_e_buscar(driver, wait, mascara)
            else:
                log_message("⚠️ Nenhuma máscara encontrada, pulando busca", "WARNING")
            
            # Salvar
            self.salvar_conclusao(driver, wait)
            
            # Enviar para próxima etapa
            self.enviar_proxima_etapa(driver, wait)
            
            # Assinar com Dr. George
            self.assinar_com_george(driver, wait)
            
            log_message("🎉 Processo de conclusão finalizado com sucesso!", "SUCCESS")
            return {'status': 'sucesso', 'detalhes': 'Conclusão processada e assinada'}
            
        except Exception as e:
            log_message(f"Erro durante processo de conclusão: {e}", "ERROR")
            return {'status': 'erro_conclusao', 'detalhes': str(e)}

    def mostrar_resumo_final(self, resultados):
        """Mostra o resumo final do processamento"""
        total = len(resultados)
        sucesso = len([r for r in resultados if r['status'] == 'sucesso'])
        sem_svg = len([r for r in resultados if r['status'] == 'sem_svg'])
        sem_andamento = len([r for r in resultados if r['status'] == 'sem_andamento'])
        erro_sessao = len([r for r in resultados if r['status'] == 'erro_sessao'])
        erros = len([r for r in resultados if 'erro' in r['status'] and r['status'] != 'erro_sessao'])
        
        log_message("\n" + "="*50, "INFO")
        log_message("RESUMO FINAL DO PROCESSAMENTO", "INFO")
        log_message("="*50, "INFO")
        log_message(f"Total de exames: {total}", "INFO")
        log_message(f"✅ Processados com sucesso: {sucesso}", "SUCCESS")
        log_message(f"⚠️ Sem SVG (não estão em conclusão): {sem_svg}", "WARNING")
        log_message(f"⚠️ Exames não encontrados: {sem_andamento}", "WARNING")
        log_message(f"🔄 Erros de sessão (browser perdido): {erro_sessao}", "WARNING")
        log_message(f"❌ Outros erros de processamento: {erros}", "ERROR")
        
        # Mostrar detalhes dos erros se houver
        erros_totais = erro_sessao + erros
        if erros_totais > 0:
            log_message("\nDetalhes dos erros:", "ERROR")
            for r in resultados:
                if 'erro' in r['status']:
                    log_message(f"- {r['codigo']}: {r['detalhes']}", "ERROR")
        
        messagebox.showinfo("Processamento Concluído", 
            f"✅ Processamento finalizado!\n\n"
            f"Total: {total}\n"
            f"Sucesso: {sucesso}\n"
            f"Sem SVG: {sem_svg}\n"
            f"Não encontrados: {sem_andamento}\n"
            f"Erros de sessão: {erro_sessao}\n"
            f"Outros erros: {erros}")

def run(params: dict):
    module = ConclusaoModule()
    module.run(params)
